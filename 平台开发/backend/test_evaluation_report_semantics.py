import io
import os
import sqlite3
import sys
import tempfile
import unittest
from contextlib import contextmanager
from datetime import datetime, timedelta

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(__file__))
import app as app_module


class EvaluationReportSemanticsTest(unittest.TestCase):
    def setUp(self):
        handle = tempfile.NamedTemporaryFile(suffix='.db', delete=False)
        handle.close()
        self.db_path = handle.name
        self.original_get_db = app_module.get_db
        self.original_tokens = dict(app_module._tokens)
        self.original_cache = dict(app_module._site_ids_cache)

        @contextmanager
        def temporary_db():
            db = sqlite3.connect(self.db_path)
            db.row_factory = sqlite3.Row
            try:
                yield db
                db.commit()
            finally:
                db.close()

        app_module.get_db = temporary_db
        app_module._tokens.clear()
        app_module._site_ids_cache.clear()
        app_module._tokens['admin-token'] = {
            'id': 1, 'role': 'admin', 'roles': ['admin'], 'real_name': '管理员',
        }
        app_module._tokens['reviewer-token'] = {
            'id': 3, 'role': 'reviewer', 'roles': ['reviewer'], 'real_name': '审核员',
        }

        now = datetime.now()
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        previous_month = month_start - timedelta(days=1)
        closed_at = month_start + timedelta(days=1, hours=10)
        open_created = month_start + timedelta(days=2, hours=9)
        with temporary_db() as db:
            db.executescript('''
                CREATE TABLE users (
                    id INTEGER PRIMARY KEY, real_name TEXT, role TEXT, phone TEXT,
                    status TEXT, username TEXT
                );
                CREATE TABLE user_roles (user_id INTEGER, role TEXT);
                CREATE TABLE user_sites (user_id INTEGER, site_id INTEGER);
                CREATE TABLE sites (
                    id INTEGER PRIMARY KEY, name TEXT, manager TEXT, status TEXT, type TEXT
                );
                CREATE TABLE work_orders (
                    id INTEGER PRIMARY KEY, order_no TEXT, site_id INTEGER, title TEXT,
                    event_type TEXT, level TEXT, assignee TEXT, status TEXT, created_at TEXT,
                    check_in_time TEXT, resolved_at TEXT, sla_deadline TEXT, description TEXT
                );
                CREATE TABLE insp_plans (id INTEGER PRIMARY KEY, assignee_id INTEGER);
                CREATE TABLE insp_plan_items (
                    id INTEGER PRIMARY KEY, plan_id INTEGER, site_id INTEGER, completed_at TEXT,
                    reviewer_id INTEGER, review_time TEXT
                );
                CREATE TABLE param_thresholds (metric TEXT PRIMARY KEY, low REAL, high REAL);
                CREATE TABLE sensor_data (site_id INTEGER, metric TEXT, value REAL, recorded_at TEXT);
                CREATE TABLE alerts (site_id INTEGER, created_at TEXT, status TEXT, metric TEXT, level TEXT);
                INSERT INTO users VALUES (1, '管理员', 'admin', '', 'active', 'admin');
                INSERT INTO users VALUES (2, '运维甲', 'operator', '', 'active', 'operator');
                INSERT INTO users VALUES (3, '审核员', 'reviewer', '', 'active', 'reviewer');
                INSERT INTO users VALUES (4, '运维乙', 'operator', '', 'active', 'operator2');
                INSERT INTO user_roles VALUES (1, 'admin');
                INSERT INTO user_roles VALUES (2, 'operator');
                INSERT INTO user_roles VALUES (3, 'reviewer');
                INSERT INTO user_roles VALUES (4, 'operator');
                INSERT INTO sites VALUES (1, '零样本站', '运维甲', 'normal', 'water_quality');
                INSERT INTO sites VALUES (2, '越权站点', '运维乙', 'normal', 'water_quality');
                INSERT INTO user_sites VALUES (3, 1);
                INSERT INTO user_sites VALUES (2, 1);
                INSERT INTO user_sites VALUES (4, 2);
                INSERT INTO param_thresholds VALUES ('ph', 6, 9);
            ''')
            db.execute('''INSERT INTO work_orders VALUES (1,?,?,?,?,?,?,?,?,?,?,?,?)''', (
                'WO-CROSS-MONTH', 1, '跨期办结', '设备故障', 'normal', '运维甲', 'closed',
                previous_month.strftime('%Y-%m-%d 09:00:00'),
                closed_at.strftime('%Y-%m-%d 09:00:00'),
                closed_at.strftime('%Y-%m-%d %H:%M:%S'),
                closed_at.strftime('%Y-%m-%d 18:00:00'), '',
            ))
            db.execute('''INSERT INTO work_orders VALUES (2,?,?,?,?,?,?,?,?,?,?,?,?)''', (
                'WO-OPEN', 1, '本月处理中', '设备故障', 'normal', '运维甲', 'in_progress',
                open_created.strftime('%Y-%m-%d %H:%M:%S'), None, None,
                (open_created + timedelta(days=2)).strftime('%Y-%m-%d %H:%M:%S'), '',
            ))
            db.execute('''INSERT INTO work_orders VALUES (3,?,?,?,?,?,?,?,?,?,?,?,?)''', (
                'WO-OUT-OF-SCOPE', 2, '外站工单', '设备故障', 'normal', '运维乙', 'in_progress',
                open_created.strftime('%Y-%m-%d %H:%M:%S'), None, None,
                (open_created + timedelta(days=2)).strftime('%Y-%m-%d %H:%M:%S'), '',
            ))
        self.client = app_module.app.test_client()

    def tearDown(self):
        app_module.get_db = self.original_get_db
        app_module._tokens.clear()
        app_module._tokens.update(self.original_tokens)
        app_module._site_ids_cache.clear()
        app_module._site_ids_cache.update(self.original_cache)
        os.unlink(self.db_path)

    @staticmethod
    def headers():
        return {'Authorization': 'Bearer admin-token'}

    def test_closed_workorder_uses_resolved_month_and_open_uses_created_month(self):
        response = self.client.get('/api/evaluation/personnel?period=month', headers=self.headers())
        self.assertEqual(response.status_code, 200, response.json)
        operator = next(row for row in response.json['list'] if row['real_name'] == '运维甲')
        self.assertEqual(operator['wo_total'], 2)
        self.assertEqual(operator['wo_closed'], 1)
        self.assertEqual(operator['wo_closed_rate'], 50.0)
        self.assertEqual(operator['closed_sla_sample'], 1)
        self.assertEqual(operator['closed_sla_met'], 1)
        self.assertEqual(operator['closed_sla_breach'], 0)
        self.assertEqual(operator['on_time_rate'], 100.0)
        self.assertEqual(operator['open_overdue'], 1)

    def test_exports_include_identity_layout_and_real_expected_health_counts(self):
        evaluation = self.client.get('/api/export/evaluation?period=month', headers=self.headers())
        self.assertEqual(evaluation.status_code, 200)
        evaluation_book = load_workbook(io.BytesIO(evaluation.data))
        evaluation_sheet = evaluation_book['人员评估']
        self.assertEqual(evaluation_sheet['A1'].value, '人员运维绩效评估')
        self.assertEqual(evaluation_sheet['A2'].value, '统计周期')
        self.assertEqual(evaluation_sheet.freeze_panes, 'A7')
        self.assertTrue(evaluation_sheet.auto_filter.ref.startswith('A7:'))
        self.assertEqual(evaluation_sheet['H7'].value, '已关单SLA样本')
        self.assertEqual(evaluation_sheet['K7'].value, '开放已逾期')
        evaluation_book.close()
        evaluation.close()

        report = self.client.get('/api/export/ops-report?period=quarter', headers=self.headers())
        self.assertEqual(report.status_code, 200)
        report_book = load_workbook(io.BytesIO(report.data))
        overview = report_book['概览']
        self.assertIn('水质智慧运维报告', overview['A1'].value)
        self.assertEqual(overview['A2'].value, '统计周期')
        health = report_book['站点健康度']
        self.assertEqual(health.freeze_panes, 'A7')
        self.assertEqual(health['C8'].value, 0)
        self.assertEqual(health['F8'].value, 0)
        self.assertEqual(health['H8'].value, '未启用监测')
        personnel = report_book['人员绩效']
        self.assertEqual(personnel['H7'].value, '已关单SLA样本')
        self.assertEqual(personnel['K7'].value, '开放已逾期')
        report_book.close()
        report.close()

    def test_reviewer_report_is_limited_to_assigned_sites(self):
        headers = {'Authorization': 'Bearer reviewer-token'}
        personnel = self.client.get('/api/evaluation/personnel?period=month', headers=headers)
        self.assertEqual(personnel.status_code, 200, personnel.json)
        self.assertEqual(personnel.json['scope_label'], '零样本站')
        self.assertEqual([row['real_name'] for row in personnel.json['list']], ['运维甲'])

        response = self.client.get('/api/export/ops-report?period=quarter', headers=headers)
        self.assertEqual(response.status_code, 200, response.json if response.is_json else None)
        workbook = load_workbook(io.BytesIO(response.data))
        self.assertEqual(workbook['概览']['B4'].value, '零样本站')
        self.assertEqual(workbook['概览']['B7'].value, 1)
        site_names = [cell.value for cell in workbook['站点健康度']['A'][7:]]
        self.assertEqual(site_names, ['零样本站'])
        workorder_numbers = [cell.value for cell in workbook['工单明细']['A'][7:]]
        self.assertNotIn('WO-OUT-OF-SCOPE', workorder_numbers)
        workbook.close()
        response.close()


if __name__ == '__main__':
    unittest.main()

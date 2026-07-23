"""
水利运维智慧运营平台 - 投标演示后端
Flask RESTful API + SQLite + APScheduler

==================== 完整API端点文档 ====================

【系统与概览】
  GET  /api/health
       健康检查。返回 {'status':'ok','time':'...'}

  GET  /api/dashboard/summary
       仪表盘汇总。返回总览数据、最新告警TOP5、待处理工单TOP5、今日巡检统计

  GET  /api/data/overview
       数据概览。返回站点数/在线数、设备数/在线数、活跃告警数、进行中工单数

【站点管理】
  GET  /api/sites
       所有站点列表。返回站点基本信息及设备数量

  GET  /api/sites/<site_id>
       单个站点详情。返回站点信息、设备列表、活跃告警数、进行中工单数

【实时数据】
  GET  /api/data/realtime
       各站点最新传感器数据。返回每个站点的最新一条数据（metric/value/unit/time）

  GET  /api/data/site/<site_id>?limit=50
       指定站点最近N条数据（默认50条）。返回metric/value/unit/recorded_at数组

【告警管理】
  GET  /api/alerts?status=pending&limit=50
       告警列表。支持按status筛选（pending/acknowledged/resolved），按等级priority排序

  POST /api/alerts/<alert_id>/acknowledge
       确认告警。将告警状态改为acknowledged

  GET  /api/alerts/statistics
       告警统计。返回total、by_level(red/orange/yellow/blue)、by_status(pending/acknowledged/resolved)

【工单管理】
  GET  /api/workorders?status=pending&limit=50
       工单列表。支持按status筛选，返回工单及关联站点名

  POST /api/workorders
       创建工单。请求体JSON: {site_id,source,event_type,level,title,description,images,assignee}
       自动生成工单号和SLA截止时间，返回{'success':True,'order_no':'WO-...'}

  PUT  /api/workorders/<order_no>/status
       更新工单状态。请求体JSON: {status,remark?,satisfaction?}
       状态流转：pending->accepted->dispatched->in_progress->reviewing->acceptance->closed

  GET  /api/workorders/statistics
       工单统计。返回total、by_status各状态计数、today_new、today_closed

【巡检管理】
  GET  /api/inspections
       巡检计划列表。返回计划及完成进度(total_items/completed_items)

  POST /api/inspections
       创建巡检计划。请求体JSON: {plan_name,site_id,type,start_date,end_date,check_items?}
       自动生成检查任务项，返回{'success':True,'plan_id':N}

  GET  /api/inspections/<plan_id>/tasks
       巡检任务列表。返回指定计划下所有检查项

  PUT  /api/inspections/tasks/<task_id>
       提交巡检结果。请求体JSON: {result,photo?,gps_lat?,gps_lng?,check_time?,remark?}

  GET  /api/inspections/statistics
       巡检统计。返回计划数/完成数、任务数/完成数、异常数

【热线管理】
  GET  /api/hotline/events?limit=50
       热线事件列表。返回热线来电记录

  POST /api/hotline/events
       登记热线事件。请求体JSON: {caller_name,caller_phone,event_type,description,location,operator}

  POST /api/hotline/events/<event_id>/convert
       热线转工单。请求体JSON: {level,assignee}
       自动生成工单并更新热线事件状态，返回{'success':True,'order_no':'WO-...'}

【天气数据】 -- NEW
  GET  /api/weather
       天气数据。返回当前天气（温度/湿度/风速/风向/降水量/气压/天气类型）、
       未来24小时逐小时预报数组、天气预警列表（暴雨/大风/高温）

【水质监测】 -- NEW
  GET  /api/water-quality?site_id=<可选>
       水质监测数据。返回供水站/水库的水质指标（浊度/pH/余氯/氨氮/COD），
       每个指标含当前值、记录时间和7日均值对比。支持按site_id筛选

【设备监控】 -- NEW
  GET  /api/devices/status
       设备状态汇总。返回设备总数/在线数/离线数、各类型设备统计、离线设备详情列表

【数据质量】 -- NEW
  GET  /api/data-quality
       数据质量报告。返回今日数据到达率/完整率/及时率、异常站点列表、最近24小时质量趋势
"""
import os
import json
import sqlite3
import random
import time
import threading
import hashlib
import secrets
import tempfile
from datetime import datetime, timedelta
from contextlib import contextmanager

from flask import Flask, jsonify, request, g, send_from_directory, send_file
from flask_cors import CORS
from apscheduler.schedulers.background import BackgroundScheduler
from inspection_rules import validate_submission_photos
import os, uuid, urllib.request, urllib.error, json as _json
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

app = Flask(__name__, static_folder=None)  # 禁用默认static，手动控制
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024  # 限制请求体最大20MB

# 告警级别中文映射（用于升级描述，避免把 red/orange/yellow 英文写入工单）
ALERT_LEVEL_LABEL = {
    'red': '紧急', 'orange': '重要', 'yellow': '一般', 'blue': '提示',
    'urgent': '紧急', 'critical': '严重', 'normal': '正常',
}

# 微信订阅消息推送配置（移动端告警信息 / 审批结果订阅消息下发）
# 占位值：正式发布前请在微信小程序后台获取并填入真实值
WX_APPID = 'wx1b28df61adae8ca1'  # 微信小程序 AppID（正式号，用户提供）
WX_APPSECRET = os.environ.get('WX_APPSECRET', '')  # 部署环境注入，避免提交凭据
WX_TMPL_ALERT = 'x_KtbMzoSIbxpUZGf040r9uvuNqd9pfhOynKaT72Ub4'    # 订阅消息模板 ID —— 告警信息
WX_TMPL_APPROVE = '4MrY8lzIXYyujudoJGsG7gka5X_ySpxg5eVKVqC__mw'  # 订阅消息模板 ID —— 审批结果
#
# ┌─ 订阅消息模板关键词约定（须与微信后台「申请的模板」关键词一一对应）──────────┐
# │ 微信订阅消息的 data 字段 key 由所选模板的关键词类型+序号决定，常见类型：  │
# │   thingN(事项≤20字) / phraseN(提示词≤5字) / character_stringN(字符≤32)  │
# │   timeN(24h制时间) / dateN(日期) / numberN(数字) / amountN(金额)        │
# │                                                                              │
# │ 推荐在微信后台按如下关键词结构申请两套模板（顺序与下列 data key 一致）：    │
# │   ① 告警信息（WX_TMPL_ALERT）：                                          │
# │        thing1           告警内容（如「溶解氧异常偏低」）                    │
# │        character_string1 站点编码                                           │
# │        phrase2          告警等级（红/橙/黄/蓝，≤5字）                    │
# │        time3            告警时间（2026年07月19日 12:30:00）               │
# │   ② 审批结果（WX_TMPL_APPROVE）：                                        │
# │        character_string1 工单编号                                          │
# │        thing2           审批结果（核验通过/核验退回/已完成，≤20字）        │
# │        time3            审批时间                                            │
# │                                                                              │
# │ 若你申请的模板关键词类型/数量不同，只需同步修改下面两个 push 函数里的     │
# │ data 字典 key 与取值即可，其余推送收发逻辑无需改动。                       │
# └────────────────────────────────────────────────────────────────────────────┘

# 指标中文映射（用于后端直接生成中文标题，避免把英文 metric 原值写入工单/通知）
METRIC_LABEL = {
    'ammonia': '氨氮',
    'dissolved_oxygen': '溶解氧',
    'device_status': '设备状态',
    'ph': 'pH',
    'codmn': '高锰酸盐',
    'temperature': '水温',
    'turbidity': '浊度',
    'conductivity': '电导率',
    'water_level': '水位',
    'flow_rate': '流量',
    'rainfall': '雨量',
    'evaporation': '蒸发',
    'soil_moisture': '墒情',
    'groundwater': '地下水',
    'chlorophyll': '叶绿素',
    'blue_green_algae': '蓝绿藻',
    'total_nitrogen': '总氮',
    'total_phosphorus': '总磷',
    'codcr': '化学需氧量',
    'bod5': '五日生化需氧量',
    'ss': '悬浮物',
    'do': '溶解氧',
}
CORS(app)

DB_PATH = os.path.join(os.path.dirname(__file__), 'data', 'water.db')
scheduler = BackgroundScheduler()
scheduler.start()

# ===================== 状态跟踪（每个站点的当前监测值趋势） =====================

_site_state = {}
def get_site_trend(site_id, metric, base, var, min_v=None, max_v=None):
    """生成有趋势的传感器数据：当前值 = 前值 + 随机趋势，避免纯随机跳变"""
    key = (site_id, metric)
    if key not in _site_state:
        _site_state[key] = base
    was = _site_state[key]
    drift = random.uniform(-var, var)
    # 0.1%概率注入异常突变（10倍漂移），用于触发异常检测
    if random.random() < 0.001:
        drift *= 10
    val = round(was + drift, 2)
    if min_v is not None:
        val = max(min_v, val)
    if max_v is not None:
        val = min(max_v, val)
    _site_state[key] = val
    return val

# ===================== Database =====================

@contextmanager
def get_db():
    db = sqlite3.connect(DB_PATH, timeout=3, check_same_thread=False)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA busy_timeout=3000")
    db.execute("PRAGMA synchronous=NORMAL")
    db.execute("PRAGMA cache_size=-8000")
    try:
        yield db
    finally:
        db.close()

def migrate_spare_parts_inventory_columns():
    """兼容旧库：为 spare_parts_inventory 补全 manufacturer / model 字段，并对常见备件回填默认值"""
    default_specs = {
        'SP-PH-001': ('Hach', 'PHC10101'),
        'SP-DO-001': ('Hach', 'LDO10101'),
        'SP-TB-001': ('Hach', 'LPV441.99.00002'),
        'SP-COND-001': ('Hach', 'CDC40101'),
        'SP-PUMP-001': ('格兰富', 'KP150'),
        'SP-FILTER-001': ('Millipore', '0.45μm'),
        'SP-CABLE-001': ('Belden', 'RS485-2×0.5'),
        'SP-NH3-001': ('Hach', 'LCW420'),
        'SP-BAT-001': ('松下', '18650-3.7V'),
        'BJ-001': ('Honeywell', 'HPT-100'),
        'BJ-002': ('Campbell', 'TB4-L'),
        'BJ-003': ('英利', 'YL-20M'),
        'BJ-004': ('理士', 'DJW12-7.0'),
        'BJ-005': ('宏电', 'H5110'),
        'BJ-006': ('移远', 'M26'),
        'BJ-007': ('国产', 'L型304'),
        'BJ-008': ('雷迅', 'ASP-S20'),
        'BJ-009': ('起帆', 'RVVP4×0.5'),
        'BJ-010': ('国产', 'Φ25硅胶'),
        'BJ-011': ('Sensirion', 'SHT30'),
        'BJ-012': ('Gill', 'WindSonic'),
    }
    with get_db() as db:
        try:
            cols = [row['name'] for row in db.execute("PRAGMA table_info(spare_parts_inventory)").fetchall()]
            if 'manufacturer' not in cols:
                db.execute("ALTER TABLE spare_parts_inventory ADD COLUMN manufacturer TEXT DEFAULT ''")
                print('[Migrate] spare_parts_inventory 新增 manufacturer 字段')
            if 'model' not in cols:
                db.execute("ALTER TABLE spare_parts_inventory ADD COLUMN model TEXT DEFAULT ''")
                print('[Migrate] spare_parts_inventory 新增 model 字段')
            updated = 0
            for part_code, (mfr, model) in default_specs.items():
                cur = db.execute(
                    "UPDATE spare_parts_inventory SET manufacturer=?, model=? WHERE part_code=? AND (manufacturer='' OR model='')",
                    (mfr, model, part_code))
                updated += cur.rowcount
            if updated:
                print(f'[Migrate] 回填 {updated} 条备件 manufacturer/model')
            db.commit()
        except Exception as e:
            print(f'[Migrate] spare_parts_inventory 字段迁移跳过: {e}')

def migrate_workorder_flow_columns():
    """兼容旧库：为工单流程卡控补全字段
    - work_orders: 到场签到 check_in_lat/lng/time/user
    - vehicle_applications: 关联 site_id / work_order_no（移动端极简用车申请）
    - spare_part_requests: 关联 work_order_no（移动端工单关联备件申请）
    """
    with get_db() as db:
        try:
            wo_cols = [r['name'] for r in db.execute("PRAGMA table_info(work_orders)").fetchall()]
            for col, ctype in (('check_in_lat', 'REAL'), ('check_in_lng', 'REAL'),
                               ('check_in_time', 'TEXT'), ('check_in_user', 'TEXT')):
                if col not in wo_cols:
                    db.execute(f"ALTER TABLE work_orders ADD COLUMN {col} {ctype}")
            va_cols = [r['name'] for r in db.execute("PRAGMA table_info(vehicle_applications)").fetchall()]
            for col in ('site_id', 'work_order_no'):
                if col not in va_cols:
                    db.execute(f"ALTER TABLE vehicle_applications ADD COLUMN {col} TEXT")
            spr_cols = [r['name'] for r in db.execute("PRAGMA table_info(spare_part_requests)").fetchall()]
            if 'work_order_no' not in spr_cols:
                db.execute("ALTER TABLE spare_part_requests ADD COLUMN work_order_no TEXT DEFAULT ''")
            if 'spare_part_id' not in spr_cols:
                db.execute("ALTER TABLE spare_part_requests ADD COLUMN spare_part_id INTEGER")
            db.commit()
            print('[Migrate] 工单流程卡控字段已确保（签到/用车关联/备件关联）')
        except Exception as e:
            print(f'[Migrate] 工单流程字段迁移跳过: {e}')

def migrate_vehicle_applications_nullable():
    """兼容旧库：用车申请表 vehicle_id/applicant_id/start_at/end_at 改为可空，
    以支持移动端“仅填事由”的极简申请形态（重建表以解除 NOT NULL 约束）。"""
    with get_db() as db:
        try:
            cols = [r['name'] for r in db.execute("PRAGMA table_info(vehicle_applications)").fetchall()]
            # 仅当存在 NOT NULL 约束时才重建
            notnull = db.execute("PRAGMA table_info(vehicle_applications)").fetchall()
            if all(c['notnull'] == 0 for c in notnull if c['name'] in ('vehicle_id', 'applicant_id', 'start_at', 'end_at')):
                print('[Migrate] vehicle_applications 已为可空，跳过重建')
                return
            db.execute("""CREATE TABLE vehicle_applications_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                vehicle_id INTEGER,
                applicant_id INTEGER,
                start_at TIMESTAMP,
                end_at TIMESTAMP,
                destination TEXT DEFAULT NULL,
                reason TEXT DEFAULT NULL,
                status TEXT DEFAULT 'pending',
                approver_id INTEGER,
                approved_at TIMESTAMP,
                reject_reason TEXT,
                created_at TIMESTAMP DEFAULT (datetime('now','localtime')),
                site_id TEXT,
                work_order_no TEXT
            )""")
            db.execute(f"""INSERT INTO vehicle_applications_new
                SELECT {','.join(cols)} FROM vehicle_applications""")
            db.execute("DROP TABLE vehicle_applications")
            db.execute("ALTER TABLE vehicle_applications_new RENAME TO vehicle_applications")
            db.commit()
            print('[Migrate] vehicle_applications 已重建为可空（保留原数据）')
        except Exception as e:
            print(f'[Migrate] vehicle_applications 重建跳过: {e}')

def migrate_plan_schedules():
    """巡检计划调度层（周/月/季/年统一）：
    1. 创建 plan_schedules 表（不叫 inspection_schedules，该表名已被"检查项排程"占用）
    2. insp_plans 补 plan_schedule_id 列（执行任务溯源到调度计划）
    3. 旧 weekly_inspection_plans 数据幂等迁移（plan_data 兼容 "1"/"周一" 两种键）
    """
    with get_db() as db:
        db.execute("""
            CREATE TABLE IF NOT EXISTS plan_schedules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                schedule_type TEXT NOT NULL DEFAULT 'weekly',
                period_start DATE NOT NULL,
                period_end DATE NOT NULL,
                plan_data TEXT NOT NULL DEFAULT '{}',
                vehicle_days TEXT DEFAULT '{}',
                spare_parts TEXT DEFAULT '[]',
                work_order_ids TEXT DEFAULT '[]',
                status TEXT DEFAULT 'draft',
                approver_id INTEGER,
                submitted_at TIMESTAMP,
                approved_at TIMESTAMP,
                reject_reason TEXT,
                version INTEGER DEFAULT 1,
                change_reason TEXT,
                previous_plan_data TEXT,
                remarks TEXT,
                tasks_generated INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (user_id) REFERENCES users(id)
            )""")
        for col_sql in [
            "ALTER TABLE insp_plans ADD COLUMN plan_schedule_id INTEGER",
            "ALTER TABLE plan_schedules ADD COLUMN previous_vehicle_days TEXT",
            "ALTER TABLE plan_schedules ADD COLUMN coverage_exception_reason TEXT",
            "ALTER TABLE plan_schedules ADD COLUMN validation_snapshot TEXT",
            "ALTER TABLE insp_plans ADD COLUMN schedule_version INTEGER DEFAULT 1",
            "ALTER TABLE insp_plans ADD COLUMN plan_snapshot TEXT",
            "ALTER TABLE insp_plan_items ADD COLUMN execution_status TEXT DEFAULT 'active'",
        ]:
            try:
                db.execute(col_sql)
            except Exception:
                pass
        # 调度计划的资源仅在审批后预留；实际领用/退回由现场执行确认，不能把预申报当作出库。
        db.execute("""CREATE TABLE IF NOT EXISTS plan_resource_reservations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            schedule_id INTEGER NOT NULL,
            part_id INTEGER NOT NULL,
            planned_quantity INTEGER NOT NULL DEFAULT 0,
            reserved_quantity INTEGER NOT NULL DEFAULT 0,
            issued_quantity INTEGER NOT NULL DEFAULT 0,
            used_quantity INTEGER NOT NULL DEFAULT 0,
            returned_quantity INTEGER NOT NULL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'reserved',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS plan_schedule_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            schedule_id INTEGER NOT NULL,
            version INTEGER NOT NULL,
            event_type TEXT NOT NULL,
            operator_id INTEGER,
            payload TEXT DEFAULT '{}',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )""")
        db.commit()
        # ---- 旧周计划数据迁移（幂等：同用户同周期已存在则跳过）----
        try:
            import json as _json
            _day_cn = {'周一': 0, '周二': 1, '周三': 2, '周四': 3, '周五': 4, '周六': 5, '周日': 6}
            old_rows = db.execute("SELECT * FROM weekly_inspection_plans").fetchall()
            migrated = 0
            for old in old_rows:
                dup = db.execute(
                    "SELECT id FROM plan_schedules WHERE user_id=? AND period_start=? AND schedule_type='weekly' LIMIT 1",
                    (old['user_id'], old['week_start'])).fetchone()
                if dup:
                    continue
                # plan_data: {"1":[site_ids]} 或 {"周一":[site_ids]} → {"2026-07-13":{"sites":[...],"notes":""}}
                try:
                    old_pd = _json.loads(old['plan_data']) if old['plan_data'] else {}
                except Exception:
                    old_pd = {}
                new_pd = {}
                try:
                    ws = datetime.strptime(str(old['week_start'])[:10], '%Y-%m-%d')
                except Exception:
                    ws = None
                for key, site_ids in (old_pd or {}).items():
                    if not isinstance(site_ids, list):
                        continue
                    offset = None
                    if str(key).isdigit():
                        offset = int(key) - 1  # "1"=周一
                    elif key in _day_cn:
                        offset = _day_cn[key]
                    if offset is None or ws is None:
                        continue
                    d = (ws + timedelta(days=offset)).strftime('%Y-%m-%d')
                    new_pd[d] = {"sites": site_ids, "notes": ""}
                # 旧单车辆 → vehicle_days（有安排的日期都用这辆车）
                vehicle_days = {}
                if old['vehicle_id']:
                    for d, day_data in new_pd.items():
                        if day_data["sites"]:
                            vehicle_days[d] = old['vehicle_id']
                status = old['status'] if old['status'] in ('draft', 'submitted', 'approved', 'rejected', 'archived') else 'draft'
                db.execute("""
                    INSERT INTO plan_schedules
                        (user_id, schedule_type, period_start, period_end, plan_data, vehicle_days,
                         status, approver_id, submitted_at, approved_at, remarks, tasks_generated)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (old['user_id'], 'weekly', old['week_start'],
                     (ws + timedelta(days=6)).strftime('%Y-%m-%d') if ws else old['week_start'],
                     _json.dumps(new_pd, ensure_ascii=False), _json.dumps(vehicle_days),
                     status, old['approver_id'], old['submitted_at'], old['approved_at'],
                     old['remarks'] or '', 1 if status == 'approved' else 0))
                migrated += 1
            if migrated:
                db.commit()
                print(f'[Migrate] 旧周计划迁移完成（{migrated} 条 → plan_schedules）')
        except Exception as e:
            print(f'[Migrate] 旧周计划迁移跳过: {e}')

def migrate_reagent_qc():
    """试剂质控（更换后跑标样验证）：
    1. reagent_inventory 补 qc_status 列（pending=待质控/passed=通过/failed=不通过）。
       默认 'passed'——存量数据视为已验证，避免历史库存全部被标为待质控。
    2. 创建 reagent_qc_records 表记录每次标样质控结果。"""
    with get_db() as db:
        db.execute("""
            CREATE TABLE IF NOT EXISTS reagent_qc_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                reagent_id INTEGER NOT NULL,
                standard_value REAL,
                measured_value REAL,
                deviation REAL,
                passed INTEGER NOT NULL DEFAULT 0,
                fail_action TEXT DEFAULT '',
                operator TEXT DEFAULT '',
                operator_id INTEGER,
                qc_time TEXT,
                remark TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            )""")
        for col_sql in [
            "ALTER TABLE reagent_inventory ADD COLUMN qc_status TEXT DEFAULT 'passed'",
        ]:
            try:
                db.execute(col_sql)
            except Exception:
                pass
        db.commit()

def init_db():
    with get_db() as db:
        db.executescript('''
            CREATE TABLE IF NOT EXISTS sites (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                type TEXT NOT NULL,
                lat REAL, lng REAL,
                district TEXT DEFAULT '',
                river TEXT DEFAULT '',
                basin TEXT DEFAULT '',
                address TEXT DEFAULT '',
                elevation REAL,
                build_date TEXT DEFAULT '',
                status TEXT DEFAULT 'online',
                manager TEXT, phone TEXT,
                last_heartbeat TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS sensor_data (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                metric TEXT NOT NULL,
                value REAL NOT NULL,
                unit TEXT,
                threshold_low REAL,
                threshold_high REAL,
                threshold_critical REAL,
                recorded_at TEXT NOT NULL,
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            CREATE TABLE IF NOT EXISTS param_thresholds (
                metric TEXT PRIMARY KEY,
                label TEXT,
                unit TEXT,
                low REAL,
                high REAL,
                critical_low REAL,
                critical_high REAL
            );

            CREATE TABLE IF NOT EXISTS alerts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                metric TEXT NOT NULL,
                value REAL NOT NULL,
                level TEXT NOT NULL,  -- blue/yellow/orange/red
                message TEXT NOT NULL,
                status TEXT DEFAULT 'pending',  -- pending/acknowledged/resolved
                created_at TEXT DEFAULT (datetime('now','localtime')),
                resolved_at TEXT,
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            CREATE TABLE IF NOT EXISTS work_orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_no TEXT UNIQUE NOT NULL,
                site_id INTEGER,
                source TEXT NOT NULL,  -- auto/patrol/hotline/superior
                event_type TEXT NOT NULL,
                level TEXT NOT NULL,  -- normal/urgent/critical
                title TEXT NOT NULL,
                description TEXT,
                images TEXT,
                assignee TEXT,
                status TEXT DEFAULT 'pending',  -- pending/accepted/dispatched/in_progress/reviewing/acceptance/closed
                sla_deadline TEXT,
                resolved_at TEXT,
                remark TEXT,
                satisfaction INTEGER,
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            CREATE TABLE IF NOT EXISTS inspection_plans (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                plan_name TEXT NOT NULL,
                site_id INTEGER NOT NULL,
                type TEXT NOT NULL,  -- daily/weekly/monthly/special
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                status TEXT DEFAULT 'pending',
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );
            -- Migrate: add period column if not exists
            -- This ALTER TABLE is executed separately below, not in executescript

            CREATE TABLE IF NOT EXISTS inspection_tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                plan_id INTEGER NOT NULL,
                site_id INTEGER NOT NULL,
                inspector TEXT,
                check_item TEXT NOT NULL,
                result TEXT,  -- normal/abnormal/na
                photo TEXT,
                gps_lat REAL, gps_lng REAL,
                check_time TEXT,
                remark TEXT,
                FOREIGN KEY (plan_id) REFERENCES inspection_plans(id),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            CREATE TABLE IF NOT EXISTS plan_sites (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                plan_id INTEGER NOT NULL,
                site_id INTEGER NOT NULL,
                FOREIGN KEY (plan_id) REFERENCES inspection_plans(id),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            CREATE TABLE IF NOT EXISTS hotline_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                caller_name TEXT,
                caller_phone TEXT,
                event_type TEXT NOT NULL,
                description TEXT NOT NULL,
                location TEXT,
                status TEXT DEFAULT 'registered',  -- registered/dispatched/closed
                related_order_no TEXT,
                operator TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS device_shadows (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                device_code TEXT UNIQUE NOT NULL,
                device_name TEXT NOT NULL,
                device_type TEXT,
                device_model TEXT DEFAULT '',
                manufacturer TEXT DEFAULT '',
                install_date TEXT DEFAULT '',
                status TEXT DEFAULT 'online',
                battery REAL,
                voltage REAL DEFAULT 0,
                last_data_time TEXT,
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            -- 设备回收记录表
            CREATE TABLE IF NOT EXISTS device_recycle (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                device_id INTEGER NOT NULL,
                device_code TEXT NOT NULL,
                device_name TEXT NOT NULL,
                device_type TEXT,
                site_id INTEGER,
                site_name TEXT,
                recycle_date TEXT NOT NULL,
                reason TEXT DEFAULT '',
                destination TEXT DEFAULT '',
                operator TEXT DEFAULT '',
                remark TEXT DEFAULT '',
                status TEXT DEFAULT 'recycled',
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (device_id) REFERENCES device_shadows(id)
            );

            -- 天气数据表 (新增)
            CREATE TABLE IF NOT EXISTS weather_data (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                temperature REAL,
                humidity REAL,
                wind_speed REAL,
                wind_direction TEXT,
                precipitation REAL,
                pressure REAL,
                weather_type TEXT,
                warning_info TEXT,
                recorded_at TEXT DEFAULT (datetime('now','localtime'))
            );

            -- 运维计划表
            CREATE TABLE IF NOT EXISTS maintenance_plans (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                plan_name TEXT NOT NULL,
                category TEXT NOT NULL,
                frequency TEXT NOT NULL,
                due_date TEXT,
                status TEXT DEFAULT 'pending',
                assignee TEXT,
                completed_at TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            -- 运维计划模板表
            CREATE TABLE IF NOT EXISTS maintenance_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category TEXT NOT NULL,
                sub_category TEXT NOT NULL,
                title TEXT NOT NULL,
                frequency TEXT NOT NULL,
                description TEXT,
                standard TEXT,
                check_items TEXT,
                photo_required INTEGER DEFAULT 0,
                estimated_hours REAL DEFAULT 1,
                sort_order INTEGER DEFAULT 0
            );

            -- 数据到报表
            CREATE TABLE IF NOT EXISTS data_arrival (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                date TEXT NOT NULL,
                metric TEXT NOT NULL,
                expected_count INTEGER DEFAULT 0,
                actual_count INTEGER DEFAULT 0,
                arrival_rate REAL DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            -- 水位校验表
            CREATE TABLE IF NOT EXISTS water_level_checks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                manual_level REAL,
                telemetry_level REAL,
                diff REAL,
                status TEXT DEFAULT 'normal',
                adjust_action TEXT,
                operator TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            -- 时间线事件表
            CREATE TABLE IF NOT EXISTS timeline_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_type TEXT NOT NULL,
                source_id INTEGER NOT NULL,
                event_type TEXT NOT NULL,
                operator TEXT,
                remark TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );

            -- 通知表（巡检计划通知、工单通知等实时消息）
            CREATE TABLE IF NOT EXISTS notifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                source_type TEXT NOT NULL,
                source_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                content TEXT DEFAULT '',
                is_read INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (user_id) REFERENCES users(id)
            );

            -- 用户表（登录系统）
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'operator',
                real_name TEXT NOT NULL,
                phone TEXT DEFAULT '',
                status TEXT DEFAULT 'active',
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );

            -- 用户-站点分配（多对多）
            CREATE TABLE IF NOT EXISTS user_sites (
                user_id INTEGER NOT NULL,
                site_id INTEGER NOT NULL,
                PRIMARY KEY (user_id, site_id),
                FOREIGN KEY (user_id) REFERENCES users(id),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );
            -- 巡检方案母表
            CREATE TABLE IF NOT EXISTS inspection_schemes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                period TEXT NOT NULL,
                name TEXT NOT NULL,
                status TEXT DEFAULT 'active',
                updated_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id),
                UNIQUE(site_id, period)
            );

            -- 方案检查项明细
            CREATE TABLE IF NOT EXISTS inspection_scheme_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                scheme_id INTEGER NOT NULL,
                category TEXT DEFAULT '',
                check_item TEXT NOT NULL,
                sort_order INTEGER DEFAULT 0,
                is_required INTEGER DEFAULT 1,
                FOREIGN KEY (scheme_id) REFERENCES inspection_schemes(id) ON DELETE CASCADE
            );

            -- 备件库存表
            CREATE TABLE IF NOT EXISTS spare_parts_inventory (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                part_code TEXT UNIQUE NOT NULL,
                part_name TEXT NOT NULL,
                manufacturer TEXT DEFAULT '',
                model TEXT DEFAULT '',
                category TEXT DEFAULT '其他',
                unit TEXT DEFAULT '个',
                quantity INTEGER DEFAULT 0,
                min_quantity INTEGER DEFAULT 5,
                site_id INTEGER,
                remark TEXT DEFAULT '',
                updated_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            -- 校准模板（用于移动端设备校验多字段表单）
            CREATE TABLE IF NOT EXISTS calibration_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                device_type TEXT NOT NULL,
                template_name TEXT NOT NULL,
                fields TEXT NOT NULL,
                calculations TEXT,
                thresholds TEXT,
                category TEXT DEFAULT '',
                sort_order INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );

            -- 巡检跳过记录
            CREATE TABLE IF NOT EXISTS inspection_skip_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                plan_id INTEGER NOT NULL,
                task_id INTEGER,
                site_id INTEGER NOT NULL,
                check_item TEXT NOT NULL,
                reason TEXT DEFAULT '',
                skip_type TEXT DEFAULT 'user',
                skip_count INTEGER DEFAULT 1,
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (plan_id) REFERENCES inspection_plans(id),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            -- 巡检照片类型配置
            CREATE TABLE IF NOT EXISTS inspection_photo_types (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                plan_id INTEGER,
                site_type TEXT DEFAULT '',
                photo_type TEXT NOT NULL,
                label TEXT NOT NULL,
                min_count INTEGER DEFAULT 1,
                sort_order INTEGER DEFAULT 0
            );

            -- 备件申请表
            CREATE TABLE IF NOT EXISTS spare_part_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_no TEXT UNIQUE NOT NULL,
                site_id INTEGER NOT NULL,
                applicant TEXT NOT NULL,
                part_name TEXT NOT NULL,
                spare_part_id INTEGER,
                quantity INTEGER NOT NULL DEFAULT 1,
                reason TEXT DEFAULT '',
                status TEXT DEFAULT 'pending',
                approver TEXT DEFAULT '',
                approval_comment TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now','localtime')),
                updated_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            -- 库存变更流水表
            CREATE TABLE IF NOT EXISTS inventory_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                part_id INTEGER NOT NULL,
                type TEXT NOT NULL CHECK(type IN ('in','out')),
                quantity INTEGER NOT NULL,
                ref_type TEXT DEFAULT '',
                ref_id INTEGER,
                operator TEXT DEFAULT '',
                remark TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (part_id) REFERENCES spare_parts_inventory(id)
            );

            CREATE TABLE IF NOT EXISTS operation_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                module TEXT NOT NULL DEFAULT '',
                action TEXT NOT NULL DEFAULT '',
                target_type TEXT NOT NULL DEFAULT '',
                target_id    INTEGER DEFAULT 0,
                operator TEXT DEFAULT '',
                operator_id  INTEGER DEFAULT 0,
                details TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );

            -- 告警规则引擎配置（前台可编辑阈值）
            CREATE TABLE IF NOT EXISTS alert_rule_config (
                id TEXT PRIMARY KEY,
                metric TEXT NOT NULL,
                metric_label TEXT NOT NULL,
                description TEXT DEFAULT '',
                enabled INTEGER DEFAULT 1,
                flow_type TEXT DEFAULT 'auto',
                unit TEXT DEFAULT '',
                thresholds TEXT DEFAULT '{}',
                is_reversed INTEGER DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS data_sources (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                source_type TEXT NOT NULL DEFAULT 'api',
                protocol TEXT DEFAULT 'HTTP',
                url TEXT NOT NULL,
                auth_type TEXT DEFAULT 'none',
                auth_config TEXT DEFAULT '{}',
                sync_interval INTEGER DEFAULT 60,
                status TEXT DEFAULT 'inactive',
                last_sync TEXT,
                remark TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );

            -- ===== 巡检V2：三层架构 =====
            -- 方案模板（站点无关，通用定义）
            CREATE TABLE IF NOT EXISTS inspection_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                template_name TEXT NOT NULL,
                category TEXT NOT NULL,
                frequency TEXT NOT NULL,
                description TEXT DEFAULT '',
                status TEXT DEFAULT 'active',
                sort_order INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );

            -- 模板检查项
            CREATE TABLE IF NOT EXISTS inspection_template_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                template_id INTEGER NOT NULL,
                item_name TEXT NOT NULL,
                category TEXT DEFAULT '',
                frequency_level TEXT DEFAULT 'mid',
                photo_required INTEGER DEFAULT 0,
                sort_order INTEGER DEFAULT 0,
                FOREIGN KEY (template_id) REFERENCES inspection_templates(id) ON DELETE CASCADE
            );

            -- 巡检配置（站点类型 → 模板匹配规则）
            CREATE TABLE IF NOT EXISTS inspection_configs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_type TEXT NOT NULL,
                device_types TEXT DEFAULT '',
                template_id INTEGER NOT NULL,
                is_active INTEGER DEFAULT 1,
                remark TEXT DEFAULT '',
                FOREIGN KEY (template_id) REFERENCES inspection_templates(id)
            );

            -- 检查项排程（每项每站独立调度）
            CREATE TABLE IF NOT EXISTS inspection_schedules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                template_id INTEGER NOT NULL,
                template_item_id INTEGER NOT NULL,
                frequency TEXT NOT NULL,
                next_due_date TEXT NOT NULL,
                last_completed_at TEXT,
                status TEXT DEFAULT 'active',
                cycle_count INTEGER DEFAULT 0,
                FOREIGN KEY (site_id) REFERENCES sites(id),
                FOREIGN KEY (template_id) REFERENCES inspection_templates(id),
                FOREIGN KEY (template_item_id) REFERENCES inspection_template_items(id)
            );

            -- 巡检计划V2
            CREATE TABLE IF NOT EXISTS insp_plans (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                plan_name TEXT NOT NULL,
                assignee TEXT NOT NULL,
                assignee_id INTEGER,
                period TEXT NOT NULL,
                generate_date TEXT NOT NULL,
                status TEXT DEFAULT 'draft',
                completion_rate REAL DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (assignee_id) REFERENCES users(id)
            );

            -- 计划检查项
            CREATE TABLE IF NOT EXISTS insp_plan_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                plan_id INTEGER NOT NULL,
                site_id INTEGER NOT NULL,
                schedule_id INTEGER,
                template_id INTEGER,
                item_name TEXT NOT NULL,
                category TEXT DEFAULT '',
                frequency TEXT DEFAULT '',
                result TEXT,
                photo_urls TEXT,
                gps_lat REAL, gps_lng REAL,
                check_time TEXT,
                remark TEXT DEFAULT '',
                calibrator TEXT,
                calibration_values TEXT,
                completed_at TEXT,
                FOREIGN KEY (plan_id) REFERENCES insp_plans(id) ON DELETE CASCADE,
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            -- 巡检提醒配置
            CREATE TABLE IF NOT EXISTS inspection_reminders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                remind_days_before INTEGER DEFAULT 1,
                remind_method TEXT DEFAULT 'notification',
                overdue_escalation INTEGER DEFAULT 0,
                escalation_days INTEGER DEFAULT 3,
                is_active INTEGER DEFAULT 1
            );

            -- 收藏的历史计划（整计划快照，供下次复用）
            CREATE TABLE IF NOT EXISTS plan_favorites (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                plan_id INTEGER,
                name TEXT NOT NULL,
                snapshot TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (user_id) REFERENCES users(id)
            );

            -- 巡检到站打卡记录
            CREATE TABLE IF NOT EXISTS inspection_checkins (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                site_name TEXT,
                user_id INTEGER,
                user_name TEXT,
                check_time TEXT NOT NULL,
                lat REAL, lng REAL,
                photo_url TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id),
                FOREIGN KEY (user_id) REFERENCES users(id)
            );

            -- 统一文件管理
            CREATE TABLE IF NOT EXISTS files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT NOT NULL,
                stored_path TEXT NOT NULL,
                file_type TEXT DEFAULT '',
                mime_type TEXT DEFAULT '',
                file_size INTEGER DEFAULT 0,
                md5_hash TEXT DEFAULT '',
                uploader_id INTEGER,
                source_type TEXT DEFAULT '',
                source_id INTEGER,
                is_deleted INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );

            -- 运维影像附件集中管理
            CREATE TABLE IF NOT EXISTS operation_attachments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT NOT NULL,
                stored_path TEXT NOT NULL,
                thumbnail_path TEXT DEFAULT '',
                file_type TEXT DEFAULT 'image',
                mime_type TEXT DEFAULT '',
                file_size INTEGER DEFAULT 0,
                description TEXT DEFAULT '',
                source_type TEXT DEFAULT '',   -- workorder/inspection/calibration/reagent/vehicle/maintenance/patrol/test（统一归口：所有影像/文件均入本表）
                source_id INTEGER DEFAULT 0,
                site_id INTEGER,
                uploader_id INTEGER,
                uploader_name TEXT DEFAULT '',
                gps_lat REAL,
                gps_lng REAL,
                taken_at TEXT,
                category TEXT DEFAULT '',      -- 现场照片/仪器照片/环境照片/签字确认/其他
                is_deleted INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            -- 试剂使用及更换记录
            CREATE TABLE IF NOT EXISTS reagent_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                reagent_name TEXT NOT NULL,
                reagent_type TEXT DEFAULT '',
                usage_date TEXT,
                replacement_date TEXT,
                operator TEXT DEFAULT '',
                notes TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            -- 时序数据分层：原始表（7天热数据）
            CREATE TABLE IF NOT EXISTS sensor_data_raw (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                metric TEXT NOT NULL,
                value REAL,
                recorded_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_sdr_site_metric_time ON sensor_data_raw(site_id, metric, recorded_at);
            CREATE INDEX IF NOT EXISTS idx_sdr_recorded_at ON sensor_data_raw(recorded_at);

            -- 时序数据分层：小时聚合（12月温数据）
            CREATE TABLE IF NOT EXISTS sensor_data_hourly (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                metric TEXT NOT NULL,
                hour TEXT NOT NULL,
                avg_value REAL, min_value REAL, max_value REAL,
                sample_count INTEGER DEFAULT 0
            );
            CREATE UNIQUE INDEX IF NOT EXISTS idx_sdh_site_metric_hour ON sensor_data_hourly(site_id, metric, hour);

            -- 时序数据分层：日聚合（永久冷数据）
            CREATE TABLE IF NOT EXISTS sensor_data_daily (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                metric TEXT NOT NULL,
                date TEXT NOT NULL,
                avg_value REAL, min_value REAL, max_value REAL,
                sample_count INTEGER DEFAULT 0
            );
            CREATE UNIQUE INDEX IF NOT EXISTS idx_sdd_site_metric_date ON sensor_data_daily(site_id, metric, date);
        ''')
        # 兼容已有数据库：尝试添加列，忽略已存在的错误
        for col_sql in [
            "ALTER TABLE alerts ADD COLUMN urge_count INTEGER DEFAULT 0",
            "ALTER TABLE alerts ADD COLUMN last_urged_at TEXT",
            "ALTER TABLE alerts ADD COLUMN related_order_no TEXT",
            "ALTER TABLE alerts ADD COLUMN response_deadline TEXT",
            "ALTER TABLE maintenance_plans ADD COLUMN urge_count INTEGER DEFAULT 0",
            "ALTER TABLE maintenance_plans ADD COLUMN last_urged_at TEXT",
            "ALTER TABLE maintenance_plans ADD COLUMN review_status TEXT DEFAULT NULL",
            "ALTER TABLE maintenance_plans ADD COLUMN review_comment TEXT DEFAULT NULL",
            "ALTER TABLE maintenance_plans ADD COLUMN template_id INTEGER",
            "ALTER TABLE maintenance_plans ADD COLUMN sub_category TEXT",
            "ALTER TABLE maintenance_plans ADD COLUMN check_results TEXT",
            "ALTER TABLE maintenance_plans ADD COLUMN remark TEXT",
            "ALTER TABLE inspection_plans ADD COLUMN period TEXT DEFAULT 'once'",
            "ALTER TABLE inspection_plans ADD COLUMN description TEXT DEFAULT ''",
            "ALTER TABLE inspection_plans ADD COLUMN scheme_id INTEGER",
            "ALTER TABLE device_shadows ADD COLUMN voltage REAL DEFAULT 0",
            "ALTER TABLE device_shadows ADD COLUMN device_model TEXT DEFAULT ''",
            "ALTER TABLE device_shadows ADD COLUMN manufacturer TEXT DEFAULT ''",
            "ALTER TABLE device_shadows ADD COLUMN install_date TEXT DEFAULT ''",
            "ALTER TABLE sites ADD COLUMN basin TEXT DEFAULT ''",
            "ALTER TABLE sites ADD COLUMN address TEXT DEFAULT ''",
            "ALTER TABLE sites ADD COLUMN elevation REAL",
            "ALTER TABLE sites ADD COLUMN build_date TEXT DEFAULT ''",
            # 移动巡检方案相关字段
            "ALTER TABLE inspection_scheme_items ADD COLUMN frequency_level TEXT DEFAULT 'mid'",
            "ALTER TABLE inspection_tasks ADD COLUMN photo_urls TEXT",
            "ALTER TABLE inspection_tasks ADD COLUMN calibrator TEXT",
            "ALTER TABLE inspection_tasks ADD COLUMN calibration_values TEXT",
            "ALTER TABLE inspection_tasks ADD COLUMN photo_required INTEGER DEFAULT 1",
            # === 数据自洽性修复：新增关联字段 ===
            "ALTER TABLE work_orders ADD COLUMN related_alert_id INTEGER",
            "ALTER TABLE work_orders ADD COLUMN metric TEXT DEFAULT ''",
            "ALTER TABLE work_orders ADD COLUMN used_parts TEXT DEFAULT ''",
            "ALTER TABLE hotline_events ADD COLUMN site_id INTEGER",
            "ALTER TABLE insp_plans ADD COLUMN vehicle_id INTEGER",
            "ALTER TABLE insp_plans ADD COLUMN submitted_at TEXT",
            "ALTER TABLE insp_plans ADD COLUMN approver_id INTEGER",
            "ALTER TABLE insp_plans ADD COLUMN approve_comment TEXT",
            "ALTER TABLE inspection_template_items ADD COLUMN need_review INTEGER DEFAULT 0",
            "ALTER TABLE inspection_template_items ADD COLUMN max_photos INTEGER DEFAULT 0",
            "ALTER TABLE inspection_template_items ADD COLUMN inspection_standard TEXT DEFAULT ''",
            # === 闭环联动：告警↔数据审核↔工单 关联字段 ===
            "ALTER TABLE alerts ADD COLUMN review_id INTEGER",
            "ALTER TABLE alerts ADD COLUMN resolve_reason TEXT",
            "ALTER TABLE data_reviews ADD COLUMN alert_id INTEGER",
            "ALTER TABLE data_reviews ADD COLUMN resolved_by_order_id INTEGER",
            "ALTER TABLE data_reviews ADD COLUMN sla_deadline TEXT",
            "ALTER TABLE work_orders ADD COLUMN conclusion TEXT",
            # === 影像资料归档字段 ===
            "ALTER TABLE operation_attachments ADD COLUMN archived INTEGER DEFAULT 0",
            "ALTER TABLE operation_attachments ADD COLUMN archived_at TEXT",
            "ALTER TABLE operation_attachments ADD COLUMN archived_by INTEGER",
            "ALTER TABLE operation_attachments ADD COLUMN archive_reason TEXT DEFAULT ''",
            # === 影像资料识别/归类字段（今日水印相机解析，v1 轻量方案）===
            "ALTER TABLE operation_attachments ADD COLUMN watermark_text TEXT DEFAULT ''",
            "ALTER TABLE operation_attachments ADD COLUMN recognized_category TEXT DEFAULT ''",
            "ALTER TABLE operation_attachments ADD COLUMN match_status TEXT DEFAULT 'auto'",
            "ALTER TABLE operation_attachments ADD COLUMN match_confidence REAL DEFAULT NULL",
            "ALTER TABLE operation_attachments ADD COLUMN review_required INTEGER DEFAULT 0",
            "ALTER TABLE operation_attachments ADD COLUMN extra_json TEXT DEFAULT ''",
            # === 照片类型配置扩展 ===
            "ALTER TABLE photo_requirements ADD COLUMN category TEXT DEFAULT ''",
            "ALTER TABLE photo_requirements ADD COLUMN watermark_keyword TEXT DEFAULT ''",
            # insp_plan_items 签到打卡字段
            "ALTER TABLE insp_plan_items ADD COLUMN check_in_time TEXT",
            "ALTER TABLE insp_plan_items ADD COLUMN check_out_time TEXT",
            "ALTER TABLE insp_plan_items ADD COLUMN part_consumed TEXT DEFAULT ''",
            # insp_plan_items 拍照清单 + 审核字段（历史库外手动添加，此处补入迁移保证重建库不缺列）
            "ALTER TABLE insp_plan_items ADD COLUMN required_photos INTEGER DEFAULT 0",
            "ALTER TABLE insp_plan_items ADD COLUMN actual_photos INTEGER DEFAULT 0",
            "ALTER TABLE insp_plan_items ADD COLUMN review_status INTEGER DEFAULT 0",
            "ALTER TABLE insp_plan_items ADD COLUMN review_comment TEXT DEFAULT ''",
            "ALTER TABLE insp_plan_items ADD COLUMN reviewer_id INTEGER DEFAULT NULL",
            "ALTER TABLE insp_plan_items ADD COLUMN review_time TEXT DEFAULT ''",
            "ALTER TABLE insp_plan_items ADD COLUMN location_address TEXT DEFAULT ''",
            # 移动端：用户微信 openid（用于订阅消息按站点群发）
            "ALTER TABLE users ADD COLUMN openid TEXT DEFAULT ''",
            # 迁移 plan_sites 数据
        ]:
            try:
                db.execute(col_sql)
            except:
                pass
        # 从 inspection_plans.site_id 迁移到 plan_sites
        try:
            existing = db.execute("SELECT COUNT(*) FROM plan_sites").fetchone()[0]
            if existing == 0:
                db.execute("""
                    INSERT OR IGNORE INTO plan_sites (plan_id, site_id)
                    SELECT id, site_id FROM inspection_plans WHERE site_id IS NOT NULL
                """)
        except Exception:
            pass
        # 迁移：为已有设备填充型号和厂商
        try:
            empty_model = db.execute("SELECT COUNT(*) FROM device_shadows WHERE device_model='' OR device_model IS NULL").fetchone()[0]
            if empty_model > 0:
                type_model_map = {
                    'rainfall_gauge': ('SL3-1', '上海气象仪器厂'),
                    'electronic_rainfall': ('RG-50', '南京水文仪器有限公司'),
                    'radar_water_level': ('RWL-200', '深圳水情科技'),
                    'pressure_water_level': ('PWL-10', '南京水文仪器有限公司'),
                    'flow_meter': ('LS25-1', '长沙水文仪器厂'),
                    'hydro_collector': ('HWP-1', '南京水文仪器有限公司'),
                    'current_meter': ('LS25-1', '长沙水文仪器厂'),
                    'rainfall_meter': ('SL3-1', '上海气象仪器厂'),
                    'water_level_meter': ('SWJ-1A', '南京水文仪器有限公司'),
                    'soil_moisture_sensor': ('TDR-300', '北京农业物联网'),
                    'soil_temperature': ('PT100-A', '北京农业物联网'),
                    'evaporation_pan': ('EVP-1', '南京水文仪器有限公司'),
                    'weather_screen': ('ENV-200', '北京农业物联网'),
                    'anemometer': ('WS-100', '上海气象仪器厂'),
                    'groundwater_level': ('GWL-2', '南京水文仪器有限公司'),
                    'water_quality_monitor': ('WQ-100', '杭州环保科技'),
                    'video_surveillance': ('IPC-500', '海康威视'),
                    'security_alarm': ('SA-100', '深圳安防科技'),
                    'env_sensor': ('ENV-200', '北京农业物联网'),
                    'sensor': ('GEN-1', '通用厂商'),
                    'comm': ('RTU-200', '深圳水情科技'),
                    'power': ('UPS-1000', '深圳电源科技'),
                }
                for dtype, (model, mfr) in type_model_map.items():
                    db.execute("UPDATE device_shadows SET device_model=?, manufacturer=? WHERE device_type=? AND (device_model='' OR device_model IS NULL)", (model, mfr, dtype))
                # 为没有install_date的设备生成安装日期
                db.execute("UPDATE device_shadows SET install_date='2019-06-15' WHERE (install_date='' OR install_date IS NULL) AND id % 3 = 0")
                db.execute("UPDATE device_shadows SET install_date='2020-03-20' WHERE (install_date='' OR install_date IS NULL) AND id % 3 = 1")
                db.execute("UPDATE device_shadows SET install_date='2021-09-10' WHERE (install_date='' OR install_date IS NULL) AND id % 3 = 2")
        except Exception:
            pass
        # 迁移：为已有站点填充流域、地址、高程、建站日期
        try:
            empty_basin = db.execute("SELECT COUNT(*) FROM sites WHERE (basin='' OR basin IS NULL)").fetchone()[0]
            if empty_basin > 0:
                basin_map = {'hydrology': '赣江', 'water_level': '赣江', 'rainfall': '抚河', 'soil_moisture': '信江', 'evaporation': '鄱阳湖', 'groundwater': '赣江', 'station_yard': '赣江', 'reservoir': '修河'}
                for stype, basin in basin_map.items():
                    db.execute("UPDATE sites SET basin=? WHERE type=? AND (basin='' OR basin IS NULL)", (basin, stype))
                db.execute("UPDATE sites SET address=district WHERE (address='' OR address IS NULL) AND district IS NOT NULL AND district != ''")
                db.execute("UPDATE sites SET elevation=ROUND(10.0 + (id % 70), 1) WHERE elevation IS NULL")
                db.execute("UPDATE sites SET build_date='2015-06-01' WHERE (build_date='' OR build_date IS NULL) AND id % 4 = 0")
                db.execute("UPDATE sites SET build_date='2017-03-15' WHERE (build_date='' OR build_date IS NULL) AND id % 4 = 1")
                db.execute("UPDATE sites SET build_date='2019-09-20' WHERE (build_date='' OR build_date IS NULL) AND id % 4 = 2")
                db.execute("UPDATE sites SET build_date='2021-01-10' WHERE (build_date='' OR build_date IS NULL) AND id % 4 = 3")
        except Exception:
            pass
        # 添加关键索引以支持大数据量查询
        for idx_sql in [
            "CREATE INDEX IF NOT EXISTS idx_sd_site_time ON sensor_data(site_id, recorded_at DESC)",
            "CREATE INDEX IF NOT EXISTS idx_sd_metric_time ON sensor_data(metric, recorded_at DESC)",
            "CREATE INDEX IF NOT EXISTS idx_al_site_status ON alerts(site_id, status)",
            "CREATE INDEX IF NOT EXISTS idx_al_status_time ON alerts(status, created_at DESC)",
            "CREATE INDEX IF NOT EXISTS idx_insp_sch_site_due ON inspection_schedules(site_id, next_due_date)",
            "CREATE INDEX IF NOT EXISTS idx_insp_sch_tpl_item ON inspection_schedules(template_id, template_item_id)",
            "CREATE INDEX IF NOT EXISTS idx_insp_pi_plan ON insp_plan_items(plan_id)",
            "CREATE INDEX IF NOT EXISTS idx_insp_pi_site ON insp_plan_items(site_id)",
            "CREATE INDEX IF NOT EXISTS idx_insp_cfg_type ON inspection_configs(site_type)",
        ]:
            try:
                db.execute(idx_sql)
            except Exception:
                pass
        db.commit()

# ===================== Seed Data =====================

_NC_CENTER = (28.68, 115.86)  # 南昌市中心

def _gen_nanchang_sites():
    """生成南昌市300+水文监测站点"""
    sites = []
    # 赣江干流关键节点（南→北）
    ganjiang = [
        (28.20, 115.95), (28.24, 115.93), (28.28, 115.90), (28.32, 115.88),
        (28.36, 115.86), (28.40, 115.85), (28.44, 115.84), (28.48, 115.85),
        (28.52, 115.87), (28.56, 115.90), (28.60, 115.93), (28.64, 115.96),
        (28.68, 115.99), (28.72, 116.02), (28.76, 116.04), (28.80, 116.06),
        (28.84, 116.08), (28.88, 116.10), (28.92, 116.12),
    ]
    # 抚河关键节点（东南→西北）
    fuhe = [
        (28.30, 116.20), (28.34, 116.16), (28.38, 116.12), (28.42, 116.08),
        (28.46, 116.04), (28.50, 116.00), (28.54, 115.96),
    ]
    # 鄱阳湖沿岸（东北）
    poyang = [
        (28.90, 116.08), (28.95, 116.12), (29.00, 116.16), (29.05, 116.20),
        (29.10, 116.18), (29.15, 116.14),
    ]
    # 区县域坐标范围
    districts = {
        '西湖': (28.65, 115.87), '东湖': (28.69, 115.89), '青山湖': (28.70, 115.95),
        '青云谱': (28.63, 115.92), '新建': (28.70, 115.82), '南昌县': (28.55, 115.95),
        '进贤': (28.38, 116.24), '安义': (28.85, 115.55), '湾里': (28.72, 115.73),
    }
    sid = 0
    # 生成雨量站（120个）：沿河高密度+区域均匀
    for i, (blat, blng) in enumerate(ganjiang):
        for j in range(4):  # 每段4个
            lat = blat + random.uniform(-0.04, 0.04)
            lng = blng + random.uniform(-0.04, 0.04)
            sid += 1; dname = random.choice(list(districts.keys()))
            sites.append((f'YL-{dname[:2].upper()}-{sid:03d}', f'{dname}雨量站{sid}', 'rainfall', round(lat,4), round(lng,4), dname, '赣江'))
    # 沿抚河补充
    for i, (blat, blng) in enumerate(fuhe):
        for j in range(3):
            lat = blat + random.uniform(-0.03, 0.03)
            lng = blng + random.uniform(-0.03, 0.03)
            sid += 1; dname = random.choice(['南昌县','进贤'])
            sites.append((f'YL-{dname[:2].upper()}-{sid:03d}', f'{dname}雨量站{sid}', 'rainfall', round(lat,4), round(lng,4), dname, '抚河'))
    # 市区低密度补充到120个
    while len([s for s in sites if s[2]=='rainfall']) < 120:
        lat = random.uniform(28.45, 28.95); lng = random.uniform(115.50, 116.40)
        # 避免离已有站点太近
        too_close = any(abs(s[3]-lat)+abs(s[4]-lng)<0.06 for s in sites)
        if not too_close:
            sid += 1
            dname = min(districts.keys(), key=lambda d: abs(lat-districts[d][0])+abs(lng-districts[d][1]))
            sites.append((f'YL-{dname[:2].upper()}-{sid:03d}', f'{dname}雨量站{sid}', 'rainfall', round(lat,4), round(lng,4), dname, ''))

    # 生成水位站（90个）：赣江沿岸高密度
    for i, (blat, blng) in enumerate(ganjiang):
        for j in range(3):
            lat = blat + random.uniform(-0.015, 0.015)
            lng = blng + random.uniform(-0.015, 0.015)
            sid += 1; dname = random.choice(list(districts.keys()))
            sites.append((f'SW-{dname[:2].upper()}-{sid:03d}', f'{dname}水位站{sid}', 'water_level', round(lat,4), round(lng,4), dname, '赣江'))
    # 抚河补充
    for i, (blat, blng) in enumerate(fuhe):
        for j in range(2):
            lat = blat + random.uniform(-0.015, 0.015)
            lng = blng + random.uniform(-0.015, 0.015)
            sid += 1; dname = random.choice(['南昌县','进贤'])
            sites.append((f'SW-{dname[:2].upper()}-{sid:03d}', f'{dname}水位站{sid}', 'water_level', round(lat,4), round(lng,4), dname, '抚河'))
    # 鄱阳湖补充
    for i, (blat, blng) in enumerate(poyang):
        sid += 1
        sites.append((f'SW-PY-{sid:03d}', f'鄱阳水位站{sid}', 'water_level', round(blat+random.uniform(-0.02,0.02),4), round(blng+random.uniform(-0.02,0.02),4), '进贤', '鄱阳湖'))

    # 生成水文站（45个）：关键断面
    key_points = ganjiang[::2] + fuhe[::2] + poyang[::2]
    for i, (blat, blng) in enumerate(key_points):
        for j in range(2 if i < 8 else 1):
            lat = blat + random.uniform(-0.01, 0.01); lng = blng + random.uniform(-0.01, 0.01)
            sid += 1; dname = random.choice(list(districts.keys()))
            river = '赣江' if i < len(ganjiang)//2 else ('抚河' if i < len(ganjiang)//2+len(fuhe)//2 else '鄱阳湖')
            sites.append((f'HW-{dname[:2].upper()}-{sid:03d}', f'{dname}水文站{sid}', 'hydrology', round(lat,4), round(lng,4), dname, river))

    # 生成墒情站（30个）：农田/灌区
    farm_areas = [(28.45,115.85),(28.50,115.90),(28.55,115.92),(28.60,115.88),(28.65,115.80),
                  (28.70,115.78),(28.75,115.85),(28.40,116.10),(28.45,116.05),(28.35,116.15)]
    for i in range(30):
        if i < len(farm_areas):
            lat, lng = farm_areas[i]
        else:
            lat = random.uniform(28.35,28.80); lng = random.uniform(115.60,116.20)
        lat += random.uniform(-0.02,0.02); lng += random.uniform(-0.02,0.02)
        sid += 1; dname = min(districts.keys(), key=lambda d: abs(lat-districts[d][0])+abs(lng-districts[d][1]))
        sites.append((f'SQ-{dname[:2].upper()}-{sid:03d}', f'{dname}墒情站{sid}', 'soil_moisture', round(lat,4), round(lng,4), dname, ''))

    # 生成蒸发站（15个）：空旷地带
    open_areas = [(28.72,115.73),(28.80,115.60),(28.60,115.70),(28.50,115.75),
                  (28.90,115.90),(28.75,116.00),(28.40,115.88),(28.55,115.65),
                  (28.70,115.55),(28.85,115.70),(28.45,115.78),(28.62,115.82),
                  (28.92,115.95),(28.52,115.68),(28.78,115.92)]
    for i, (lat, lng) in enumerate(open_areas):
        sid += 1; dname = min(districts.keys(), key=lambda d: abs(lat-districts[d][0])+abs(lng-districts[d][1]))
        sites.append((f'ZF-{dname[:2].upper()}-{sid:03d}', f'{dname}蒸发站{sid}', 'evaporation', round(lat+random.uniform(-0.01,0.01),4), round(lng+random.uniform(-0.01,0.01),4), dname, ''))

    return sites

def seed_data():
    """种子数据：真实站点 + 设备 + 工单 + 热线事件（仅首次运行）"""
    with get_db() as db:
        count = db.execute("SELECT COUNT(*) FROM sites").fetchone()[0]
        if count > 0:
            print("[Seed] 站点数据已存在，跳过站点/设备/工单种子数据")
            return

        # === 235个真实站点导入 ===
        import json as _json
        json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'site_data.json')
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                all_sites = _json.load(f)
        else:
            all_sites = _gen_nanchang_sites()
        for s in all_sites:
            lat = s['lat'] or random.uniform(28.4, 29.2)
            lng = s['lng'] or random.uniform(115.5, 116.5)
            basin_map = {'hydrology': '赣江', 'water_level': '赣江', 'rainfall': '抚河', 'soil_moisture': '信江', 'evaporation': '鄱阳湖', 'groundwater': '赣江', 'station_yard': '赣江', 'reservoir': '修河'}
            basin = basin_map.get(s['type'], '赣江')
            addr = s.get('address', '') or s.get('note', '') or f"江西省南昌市{s.get('address', '')}"
            elev = round(random.uniform(10, 80), 1)
            build_year = random.randint(2005, 2020)
            build_date = f"{build_year}-{random.randint(1,12):02d}-{random.randint(1,28):02d}"
            db.execute(
                "INSERT INTO sites (code,name,type,lat,lng,district,basin,address,elevation,build_date) VALUES (?,?,?,?,?,?,?,?,?,?)",
                (s['code'], s['name'], s['type'], lat, lng, s.get('address',''), basin, addr, elev, build_date)
            )
        print(f"[Seed] 生成 {len(all_sites)} 个站点")

        # === 分配负责人（同一行政区划分给同一运维人员，与人员管理一致） ===
        real_users = db.execute("SELECT username, real_name FROM users WHERE role='operator' ORDER BY id").fetchall()
        real_names = [u['real_name'] for u in real_users]
        if not real_names: real_names = ['张建国','黎明','王刚','赵洪']
        all_rows = db.execute("SELECT id, district FROM sites ORDER BY district, id").fetchall()
        mgr_map = {}; mgr_idx = 0
        for row in all_rows:
            dist = row['district'] or ''
            if dist not in mgr_map:
                mgr_map[dist] = real_names[mgr_idx % len(real_names)]
                mgr_idx += 1
            db.execute("UPDATE sites SET manager=?, phone=? WHERE id=?",
                       (mgr_map[dist], f'1{random.randint(30,39)}0000{random.randint(1000,9999)}', row['id']))

        # === 设备生成（每站按类型配设备） ===
        type_devices = {
            'rainfall': [('翻斗式雨量计','rainfall_gauge','SL3-1','上海气象仪器厂'),('电子雨量计','electronic_rainfall','RG-50','南京水文仪器有限公司')],
            'water_level': [('雷达水位计','radar_water_level','RWL-200','深圳水情科技'),('压力式水位计','pressure_water_level','PWL-10','南京水文仪器有限公司'),('流速计','flow_meter','LS25-1','长沙水文仪器厂')],
            'hydrology': [('水文综合采集仪','hydro_collector','HWP-1','南京水文仪器有限公司'),('流速仪','current_meter','LS25-1','长沙水文仪器厂'),('雨量计','rainfall_meter','SL3-1','上海气象仪器厂'),('水位计','water_level_meter','SWJ-1A','南京水文仪器有限公司')],
            'soil_moisture': [('土壤水分传感器','soil_moisture_sensor','TDR-300','北京农业物联网'),('土壤温度计','soil_temperature','PT100-A','北京农业物联网')],
            'evaporation': [('蒸发皿','evaporation_pan','EVP-1','南京水文仪器有限公司'),('气象百叶箱','weather_screen','ENV-200','北京农业物联网'),('风速仪','anemometer','WS-100','上海气象仪器厂')],
            'groundwater': [('地下水位计','groundwater_level','GWL-2','南京水文仪器有限公司'),('水质在线监测仪','water_quality_monitor','WQ-100','杭州环保科技')],
            'station_yard': [('视频监控','video_surveillance','IPC-500','海康威视'),('安防报警','security_alarm','SA-100','深圳安防科技'),('环境传感器','env_sensor','ENV-200','北京农业物联网')],
            'water_quality': [('多参数水质分析仪','multi_param_analyzer','WQA-200','哈希'),('pH计','ph_meter','PHG-2088','上海雷磁'),('溶解氧传感器','do_sensor','DOG-3082','上海雷磁'),('浊度仪','turbidity_meter','TURB-3000','哈希'),('氨氮分析仪','ammonia_analyzer','NH3N-2000','聚光科技'),('COD分析仪','cod_analyzer','CODcr-2000','聚光科技'),('数据采集传输终端','dtu','DTU-WQ01','厦门四信')],
        }
        all_sites_db = db.execute("SELECT id, code, type FROM sites ORDER BY id").fetchall()
        for site in all_sites_db:
            devs = type_devices.get(site['type'], [('通用传感器','generic','GEN-1','通用厂商')])
            for i, (dname, dtype, dmodel, dmfr) in enumerate(devs):
                install_date = f"20{18 + (site['id'] % 6):02d}-{(site['id'] % 12) + 1:02d}-{(site['id'] % 28) + 1:02d}"
                db.execute(
                    "INSERT INTO device_shadows (site_id,device_code,device_name,device_type,device_model,manufacturer,install_date,status,battery,voltage) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (site['id'], f"{site['code']}-{i+1:02d}{dtype[:4].upper()}", dname, dtype, dmodel, dmfr, install_date,
                     'online', round(random.uniform(60,100), 0),
                     round(random.uniform(11.5, 14.2), 1))
                )

        # 试剂使用及更换记录种子数据（水质监测站）
        reagent_templates = [
            ('pH缓冲液', '校准试剂', 'pH计校准用'),
            ('氨氮试剂A', '分析试剂', '氨氮分析仪显色剂'),
            ('氨氮试剂B', '分析试剂', '氨氮分析仪氧化剂'),
            ('COD消解液', '分析试剂', '高锰酸盐指数消解用'),
            ('总磷试剂', '分析试剂', '总磷分析仪显色剂'),
            ('溶解氧膜', '耗材', '溶解氧传感器膜头更换'),
            ('浊度标准液', '校准试剂', '浊度仪校准用'),
        ]
        water_quality_sites = db.execute("SELECT id FROM sites WHERE type='water_quality'").fetchall()
        for site in water_quality_sites[:10]:  # 前10个水质站
            for j, (rname, rtype, rnote) in enumerate(reagent_templates[:3 + (site['id'] % 4)]):
                days_ago = 30 + j * 45 + (site['id'] % 20)
                usage_date = (datetime.now() - timedelta(days=days_ago)).strftime('%Y-%m-%d')
                replacement_date = (datetime.now() - timedelta(days=days_ago - 15)).strftime('%Y-%m-%d') if days_ago > 15 else '—'
                operator = ['刘娜', '王强', '黄丽'][j % 3]
                db.execute(
                    "INSERT INTO reagent_records (site_id, reagent_name, reagent_type, usage_date, replacement_date, operator, notes) VALUES (?,?,?,?,?,?,?)",
                    (site['id'], rname, rtype, usage_date, replacement_date, operator, rnote)
                )

        # 工单种子数据（取前几个站ID）
        sample_ids = [r['id'] for r in db.execute("SELECT id FROM sites ORDER BY id LIMIT 5").fetchall()]
        orders = [
            (f'WO-20260618-{i+1:03d}', sample_ids[i] if i < len(sample_ids) else sample_ids[0],
             'auto','设备故障','normal','水位计数据中断','设备持续30分钟无数据上报','', '张建国','dispatched','2026-06-18 16:00','2026-06-18 08:30') for i in range(3)
        ]
        for o in orders:
            db.execute(
                "INSERT INTO work_orders (order_no,site_id,source,event_type,level,title,description,images,assignee,status,sla_deadline,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                o
            )

        # 热线事件
        hotlines = [
            ('张先生','13900001100','水质问题','家里的自来水发黄，已经持续两天了','城南街道阳光小区','registered', '', '李敏','2026-06-10 14:05'),
            ('李女士','13900002200','设施损坏','河道护栏被撞坏，存在安全隐患','滨江路新华桥东侧','dispatched','WO-20260611-004','李敏','2026-06-10 16:30'),
            ('陈先生','13900003300','违规举报','有人在河道内非法采砂','滨江堤防B段下游','registered', '', '王芳','2026-06-11 08:15'),
            ('匿名','', '水位异常','东湖水位这两天涨得很快，担心漫堤','东湖公园湖区','registered', '', '王芳','2026-06-11 09:30'),
        ]
        for h in hotlines:
            db.execute(
                "INSERT INTO hotline_events (caller_name,caller_phone,event_type,description,location,status,related_order_no,operator,created_at) VALUES (?,?,?,?,?,?,?,?,?)",
                h
            )

        # 备件库存种子数据
        inv_cnt = db.execute("SELECT COUNT(*) FROM spare_parts_inventory").fetchone()[0]
        if inv_cnt == 0:
            spare_parts = [
                ('BJ-001', '水位计传感器', 'Honeywell', 'HPT-100', '传感器', '个', 30, 5),
                ('BJ-002', '雨量筒翻斗', 'Campbell', 'TB4-L', '传感器', '个', 15, 5),
                ('BJ-003', '太阳能板(20W)', '英利', 'YL-20M', '电源', '块', 8, 3),
                ('BJ-004', '蓄电池(12V)', '理士', 'DJW12-7.0', '电源', '个', 12, 5),
                ('BJ-005', '数据采集终端RTU', '宏电', 'H5110', '通信', '台', 5, 2),
                ('BJ-006', 'GPRS通信模块', '移远', 'M26', '通信', '个', 10, 3),
                ('BJ-007', '不锈钢水位计支架', '国产', 'L型304', '结构', '套', 6, 3),
                ('BJ-008', '防雷模块', '雷迅', 'ASP-S20', '电源', '个', 20, 5),
                ('BJ-009', '信号线缆(10m)', '起帆', 'RVVP4×0.5', '线缆', '根', 25, 10),
                ('BJ-010', '水位计密封圈', '国产', 'Φ25硅胶', '其他', '个', 50, 10),
                ('BJ-011', '温湿度传感器', 'Sensirion', 'SHT30', '传感器', '个', 8, 3),
                ('BJ-012', '风速风向仪', 'Gill', 'WindSonic', '传感器', '台', 3, 2),
            ]
            for pc, pn, mfr, model, cat, unit, qty, minq in spare_parts:
                db.execute(
                    "INSERT INTO spare_parts_inventory (part_code,part_name,manufacturer,model,category,unit,quantity,min_quantity) VALUES (?,?,?,?,?,?,?,?)",
                    (pc, pn, mfr, model, cat, unit, qty, minq)
                )

        # 备件申请种子数据（演示用）
        req_cnt = db.execute("SELECT COUNT(*) FROM spare_part_requests").fetchone()[0]
        if req_cnt == 0:
            from datetime import datetime, timedelta
            now = datetime.now()
            sample_reqs = [
                (1, '系统管理员', '水位计传感器', 2, '水位计数据异常，需更换', 'approved', '系统管理员', '同意更换', (now - timedelta(hours=2)).strftime('%Y-%m-%d %H:%M:%S')),
                (2, '系统管理员', '太阳能板(20W)', 1, '太阳能板破损', 'approved', '系统管理员', '已核实，批准', (now - timedelta(hours=5)).strftime('%Y-%m-%d %H:%M:%S')),
                (3, '运维人员', 'GPRS通信模块', 2, '通信模块频繁断连', 'pending', '', '', (now - timedelta(minutes=30)).strftime('%Y-%m-%d %H:%M:%S')),
                (4, '运维人员', '防雷模块', 3, '汛期前补充', 'pending', '', '', (now - timedelta(hours=1)).strftime('%Y-%m-%d %H:%M:%S')),
                (5, '系统管理员', '数据采集终端RTU', 1, 'RTU老化需更换', 'rejected', '系统管理员', '库存不足，暂缓采购', (now - timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')),
            ]
            for idx, (sid, applicant, pname, qty, reason, status, approver, comment, ctime) in enumerate(sample_reqs):
                rno = f"BJ-{now.strftime('%Y%m%d')}-{idx+1:03d}"
                db.execute(
                    "INSERT INTO spare_part_requests (request_no,site_id,applicant,part_name,quantity,reason,status,approver,approval_comment,created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (rno, sid, applicant, pname, qty, reason, status, approver, comment, ctime)
                )
                # 已批准的申请记录扣减库存流水
                if status == 'approved':
                    inv = db.execute("SELECT id, quantity FROM spare_parts_inventory WHERE part_name LIKE ? LIMIT 1", (f"%{pname}%",)).fetchone()
                    if inv:
                        new_qty = max(0, inv['quantity'] - qty)
                        db.execute("UPDATE spare_parts_inventory SET quantity=? WHERE id=?", (new_qty, inv['id']))
                        db.execute("INSERT INTO inventory_logs (part_id,type,quantity,ref_type,ref_id,operator,remark) VALUES (?,'out',?,'request',?,?,?)",
                            (inv['id'], qty, 0, '系统管理员', f"种子数据：{rno}"))

        db.commit()
        print("[Seed] Database seeded with initial data.")

def seed_inspections():
    """巡检种子数据（依据《运维事项.pdf》内容分布，独立判断，可重复运行）"""
    with get_db() as db:
        cnt = db.execute("SELECT COUNT(*) FROM inspection_plans").fetchone()[0]
        if cnt == 0:
            all_sites = db.execute("SELECT id, name, type FROM sites ORDER BY id").fetchall()
            if len(all_sites) < 5: return
            # 按站点类型分组
            s_by_type = {}
            for s in all_sites:
                s_by_type.setdefault(s['type'], []).append(s)
            now_str = datetime.now().strftime('%Y-%m-%d')

            # ========== 计划定义 (依据运维事项.pdf) ==========
            # 格式: (名称, 站点类型, 频次, 持续天数, 状态, 每类型选站数, 检查项列表)
            plans = [
                # ====== 水文站 (hydrology) ======
                # 每日：水位日常观测（PDF：水位项目日常巡查-驻测站每日8时、18-20时）
                ('水位日常观测', 'hydrology', 'daily', 1, 'in_progress', 4,
                 ['观测基本水尺读数并记录', '校对遥测水位及时间', '检查清洗水尺', '水位设备无异常检查', '填记水位巡查表并拍照存档']),
                # 每周：站院环境维护（PDF：驻测站站院环境维护-每周打扫）
                ('站院环境维护', 'hydrology', 'weekly', 7, 'pending', 3,
                 ['水位井打扫', '站院地面、窗台、设备清洁', '墙面天花板无污迹蜘蛛网', '草地灌木修剪打扫']),
                # 每月：设施设备维护（PDF：设施设备维护-每月检查清洗水尺、爬梯、护栏）
                ('设施设备检查', 'hydrology', 'monthly', 30, 'pending', 2,
                 ['检查清洗水尺', '爬梯牢固度检查', '护栏牢固度检查', '设施设备巡查表填写', '异常维修拍照存档报中心站网监测科']),
                # 每月：观测场管理（PDF：观测场管理-每月2次）
                ('观测场管理', 'hydrology', 'monthly', 30, 'pending', 2,
                 ['降蒸观测场草地维护', '站院草皮高度低于20cm', '杂草杂物清理']),
                # 每月：断面环境管理（PDF：断面环境管理-每月断面检查清理）
                ('断面环境管理', 'hydrology', 'monthly', 30, 'pending', 2,
                 ['基本水尺码头清理淤泥杂草', '停船码头清理淤泥杂草', '流速仪测流断面清理', '基本水尺底部淤泥杂草清理']),
                # 每月：安全检查（PDF：安全检查-每月一次）
                ('安全检查', 'hydrology', 'monthly', 30, 'pending', 2,
                 ['测验设施设备检查', '安全环境检查', '站房检查', '灭火器检查', '安全器材检查', '安全检查记录填记']),
                # 每月：发电机保养（PDF：发电机保养及维护-每月1次检查）
                ('发电机保养', 'hydrology', 'monthly', 30, 'pending', 1,
                 ['检查机油', '检查线路及各部件', '发电运行不少于30分钟并记录', '汛前汛后更换机油保养维护', '备足燃料及机油']),
                # 不定期(季度)：缆道日常巡检（PDF：缆道日常巡检-测流时）
                ('缆道巡检', 'hydrology', 'quarterly', 90, 'pending', 2,
                 ['行主索检查维护', '循环索检查维护', '拉线卡头检查', '工作索毛刺断骨检查拍照留底',
                  '锚碇位移检查', '锚碇周围土壤裂纹崩塌检查', '导向轮游轮行车架运转检查', '绞车运转检查']),
                # 每半年：综合检查
                ('半年综合检查', 'hydrology', 'halfyear', 180, 'pending', 2,
                 ['水文缆道全面检修', '备用电源充放电测试', '通信系统切换测试', '所有传感器校准']),
                # 每年：年度大修（PDF：断面环境管理-每年汛前汛后全面清理 + 发电机-汛前汛后保养）
                ('年度检修', 'hydrology', 'yearly', 365, 'completed', 1,
                 ['汛前流速仪测流断面全面清理', '汛前缆道铁塔四周全面清理',
                  '汛后断面全面清理', '发电机更换机油', '发电机各部件全面检查',
                  '全站水文年鉴资料整编', '年度总结报告编制']),

                # ====== 水位站 (water_level) ======
                # 每日：水位日常观测（PDF：水位项目日常巡查-巡测站每日8-10时）
                ('水位日常观测', 'water_level', 'daily', 1, 'in_progress', 4,
                 ['观测基本水尺读数并记录', '校对遥测水位及时间', '检查水位设备无异常', '水位巡查表填记并拍照存档']),
                # 每月：站房维护（PDF：巡测站站房维护-每月2次）
                ('站房维护', 'water_level', 'monthly', 30, 'pending', 2,
                 ['站房全面打扫', '地面窗台设备清洁', '墙面天花板无污迹蜘蛛网', '保持干净整洁']),
                # 每月：设施设备维护（PDF：设施设备维护-每月）
                ('设施设备检查', 'water_level', 'monthly', 30, 'pending', 2,
                 ['检查清洗水尺', '水位设备检查', '爬梯牢固度检查', '设施设备巡查表填写']),
                # 每月：安全检查（PDF：安全检查-每月一次）
                ('安全检查', 'water_level', 'monthly', 30, 'pending', 2,
                 ['测验设施设备检查', '安全环境检查', '灭火器检查', '安全器材检查', '安全检查记录填记']),
                # 每半年：综合检查
                ('半年综合检查', 'water_level', 'halfyear', 180, 'pending', 2,
                 ['水准点校核', '水尺零高测量', 'RTU主板检查', '通信系统切换测试', '所有传感器校准']),

                # ====== 雨量站 (rainfall) ======
                # 每日：数据检查
                ('雨量日常检查', 'rainfall', 'daily', 1, 'pending', 4,
                 ['雨量数据检查', '通信状态检查', '电源电压检查']),
                # 每月：雨量项目日常巡检（PDF：雨量项目日常巡检-每月1次）
                ('雨量项目巡检', 'rainfall', 'monthly', 30, 'pending', 2,
                 ['数据采集终端外观检查', '数据读取和上报检查', '终端内部状态检查',
                  '供电设备检查', '布线检查', '雨量筒外观检查',
                  '雨量筒器口水平检查', '雨量筒气泡居中检查', '雨量筒运行状态检查',
                  '雨量采集准确性核对', '站点周边环境清理']),
                # 每季度：注水试验（PDF：雨量项目-每季度注水试验）
                ('雨量注水试验', 'rainfall', 'quarterly', 90, 'pending', 2,
                 ['注入5-10mm清洗湿润过水部件', '翻斗运转灵活性检查', '信号输出正常检查',
                  '清除翻斗存留水量', '每次注水三次不少于12.5mm',
                  '测量误差不大于±4%为合格', '记录存盘']),
                # 每半年：综合检查
                ('半年综合检查', 'rainfall', 'halfyear', 180, 'pending', 2,
                 ['雨量器全套校准', '通信系统测试', '备份电池检查', '机箱密封检查']),
                # 每年：年度校准（PDF：每年汛前自动蒸发注水实验部分涉及）
                ('年度校准', 'rainfall', 'yearly', 365, 'completed', 1,
                 ['雨量资料整编', '雨量器更换评估', '年度校准报告编制']),

                # ====== 蒸发站 (evaporation) ======
                # 每日：数据检查
                ('蒸发日常检查', 'evaporation', 'daily', 1, 'pending', 3,
                 ['蒸发量数据检查', '水面状态观察', '通信状态检查']),
                # 每月：蒸发项目日常巡检（PDF：蒸发项目日常巡检-每月不少于1次）
                ('蒸发项目巡检', 'evaporation', 'monthly', 30, 'pending', 2,
                 ['自动蒸发设备遥测终端现场巡检', '数据采集和传输终端外观检查',
                  '终端内部状态检查', '供电设备检查', '布线检查']),
                # 每月：蒸发器换水（PDF：一个月至少换水一次）
                ('蒸发器换水', 'evaporation', 'monthly', 30, 'pending', 1,
                 ['蒸发器换水', '水圈清洁保持无泥沙杂草杂物青苔', '取用能代表当地自然水体的水']),
                # 每半年：渗漏检查（PDF：每半年需进行一次渗漏检查）
                ('蒸发渗漏检查', 'evaporation', 'halfyear', 180, 'pending', 1,
                 ['8时关闭蒸发皿阀门', '人工量测蒸发皿1日蒸发量', '通过邻站对比判断是否漏水',
                  '同步观测自记值判断输水管道或静水桶是否漏水',
                  '每日合理性检查-蒸发异常偏大时需进行渗漏检查']),
                # 每半年：综合检查
                ('半年综合检查', 'evaporation', 'halfyear', 180, 'pending', 2,
                 ['蒸发器全套标定', '通信系统测试', '数据对比分析']),
                # 每年：注水实验（PDF：每年汛前对自动蒸发进行注水实验）
                ('蒸发注水实验', 'evaporation', 'yearly', 365, 'pending', 1,
                 ['选择无雨日早晨或黄昏进行注水实验', '使用雨杯量取注入水量',
                  '分别注入0.1mm至4mm梯度水量', '等待1-2分钟待液位稳定后读取',
                  '同时人工测针测记蒸发器液位', '统计计算各项误差',
                  '一代伟思折算系数0.868，二代伟思折算系数0.909']),

                # ====== 墒情站 (soil_moisture) ======
                # 每日：数据检查
                ('墒情日常检查', 'soil_moisture', 'daily', 1, 'pending', 3,
                 ['墒情数据检查', '通信状态检查', '电源状态检查']),
                # 每季度：墒情站日常巡查（PDF：每季度对基本站巡查不少于1次）
                ('墒情站巡查', 'soil_moisture', 'quarterly', 90, 'pending', 2,
                 ['机箱内干净整洁检查', '清理周边杂草', '保持整洁无积水', '进行数据校测并做好记录',
                  '干旱天气按规范做好取土检验工作']),
                # 每半年：综合检查
                ('半年综合检查', 'soil_moisture', 'halfyear', 180, 'pending', 1,
                 ['传感器埋设状态检查', '数据对比分析', '机箱密封检查']),

                # ====== 地下水监测站 (groundwater) ======
                # 每日：数据监控（PDF：数据监控及台账建立-实时查看地下水数据到报率）
                ('地下水日常监测', 'groundwater', 'daily', 1, 'pending', 3,
                 ['地下水数据检查', '通信状态检查', '电源状态检查']),
                # 每月：设备巡检（PDF：设施设备维护-每月检查）
                ('地下水设备巡检', 'groundwater', 'monthly', 30, 'pending', 2,
                 ['数据采集终端检查', '供电设备检查', '浮子式水位计运行检查',
                  '压力式水位计运行检查', '机箱密封检查', '周边环境清理']),
                # 每季度：巡查（PDF：墒情站巡查参考-每季度不少于1次）
                ('地下水站巡查', 'groundwater', 'quarterly', 90, 'pending', 2,
                 ['机箱内干净整洁检查', '清理周边杂草', '保持整洁无积水',
                  '数据校测并做好记录', '传感器运行状态检查']),
                # 每半年：综合检查
                ('半年综合检查', 'groundwater', 'halfyear', 180, 'pending', 1,
                 ['传感器全套校准', '通信系统切换测试', '数据对比分析']),

                # ====== 站院 (station_yard) ======
                # 每周：站院环境维护（PDF：驻测站站院环境维护-每周打扫）
                ('站院环境维护', 'station_yard', 'weekly', 7, 'pending', 2,
                 ['站院地面清洁', '窗台设备清洁', '墙面天花板无污迹蜘蛛网',
                  '草地灌木修剪', '遇重大活动增加维护次数']),
                # 每月：设施设备维护（PDF：设施设备维护-每月）
                ('设施设备检查', 'station_yard', 'monthly', 30, 'pending', 2,
                 ['检查清洗水尺', '设施设备全面检查', '爬梯牢固度检查',
                  '护栏牢固度检查', '设施设备巡查表填写', '异常维修拍照存档']),
                # 每月：安全检查（PDF：安全检查-每月一次）
                ('安全检查', 'station_yard', 'monthly', 30, 'pending', 2,
                 ['测验设施设备检查', '安全环境检查', '站房检查',
                  '灭火器检查', '安全器材检查', '安全检查记录填记']),
                # 每半年：综合检查
                ('半年综合检查', 'station_yard', 'halfyear', 180, 'pending', 1,
                 ['设施设备全面检修', '安全环境综合评估', '通信系统测试']),
            ]

            # ========== 生成计划 ==========
            for pname, stype, freq, days, status, sel_cnt, check_items in plans:
                sites_of_type = s_by_type.get(stype, [])
                if not sites_of_type:
                    continue
                # 按站点打包：将同类型站点分批，每批生成一个计划（一批含多个站点）
                chunk_size = stype in ('station_yard','reservoir') and 5 or 10
                selected = sites_of_type
                for chunk_idx in range(0, len(selected), chunk_size):
                    chunk = selected[chunk_idx:chunk_idx + chunk_size]
                    if not chunk: continue
                    site_ids = [s['id'] for s in chunk]
                    first_site = chunk[0]
                    batch_num = chunk_idx // chunk_size + 1
                    plan_label = pname
                    if len(selected) > chunk_size:
                        plan_label = f'{pname}({batch_num})'
                    end_dt = datetime.now() + timedelta(days=days)
                    start_date = now_str
                    end_date = end_dt.strftime('%Y-%m-%d')
                    cur = db.execute(
                        "INSERT INTO inspection_plans (plan_name,site_id,type,start_date,end_date,status) VALUES (?,?,?,?,?,?)",
                        (f'{plan_label}', first_site['id'], freq, start_date, end_date, status)
                    )
                    pid = cur.lastrowid
                    for sid in site_ids:
                        db.execute("INSERT OR IGNORE INTO plan_sites (plan_id,site_id) VALUES (?,?)", (pid, sid))
                    for sid in site_ids:
                        for item in check_items:
                            result = 'normal' if status == 'completed' else None
                            remark = '一切正常' if status == 'completed' else None
                            check_time = (start_date + ' 09:00') if status == 'completed' else None
                            db.execute(
                                "INSERT INTO inspection_tasks (plan_id,site_id,check_item,result,remark,check_time) VALUES (?,?,?,?,?,?)",
                                (pid, sid, item, result, remark, check_time)
                            )
                    # 对 in_progress 的计划，部分任务已完成
                    if status == 'in_progress':
                        # 取前一半的站点的所有任务标记为已完成
                        half_sites = site_ids[:max(1, len(site_ids)//2)]
                        for sid in half_sites:
                            tasks = db.execute(
                                "SELECT id FROM inspection_tasks WHERE plan_id=? AND site_id=?",
                                (pid, sid)
                            ).fetchall()
                            for r in tasks:
                                db.execute(
                                    "UPDATE inspection_tasks SET result='normal', remark='运行正常', check_time=? WHERE id=?",
                                    (start_date + ' 08:30', r['id'])
                                )
        db.commit()
        print("[Seed] Inspection plans seeded.")

def seed_alerts():
    """历史告警种子数据（仅首次）"""
    with get_db() as db:
        acnt = db.execute("SELECT COUNT(*) FROM alerts").fetchone()[0]
        if acnt == 0:
            # 从各类型站点取前几个生成告警
            sid_map = {}
            # 种子数据不再生成阈值类告警，异常告警由定时数据生成时自动产生
            pass

def seed_abnormal_scenarios():
    """注入15种异常场景数据，用于全流程演示（可重复运行，通过标记避免重复）"""
    with get_db() as db:
        # 检查是否已注入
        marker = db.execute("SELECT id FROM timeline_events WHERE event_type='abnormal_scenarios_seeded' LIMIT 1").fetchone()
        if marker:
            print("[Seed] 异常场景数据已存在，跳过")
            return

        now = datetime.now()
        now_str = now.strftime('%Y-%m-%d %H:%M:%S')

        # 获取各类站点ID
        hydro_sites = db.execute("SELECT id, name FROM sites WHERE type='hydrology' ORDER BY id LIMIT 10").fetchall()
        wl_sites = db.execute("SELECT id, name FROM sites WHERE type='water_level' ORDER BY id LIMIT 10").fetchall()
        rain_sites = db.execute("SELECT id, name FROM sites WHERE type='rainfall' ORDER BY id LIMIT 10").fetchall()
        soil_sites = db.execute("SELECT id, name FROM sites WHERE type='soil_moisture' ORDER BY id LIMIT 5").fetchall()
        all_sites = db.execute("SELECT id, name, type FROM sites ORDER BY id LIMIT 20").fetchall()

        if len(all_sites) < 10:
            print("[Seed] 站点数据不足，跳过异常场景注入")
            return

        print("[Seed] 开始注入15种异常场景...")

        # === 场景1：站点离线（预设离线站点5,108,193已在generate_sensor_data中处理） ===
        # 额外设置几个站点为离线状态
        offline_site_ids = [5, 108, 193]
        for sid in offline_site_ids:
            db.execute("UPDATE sites SET status='offline' WHERE id=?", (sid,))
            db.execute("UPDATE device_shadows SET status='offline', last_data_time=NULL WHERE site_id=?", (sid,))
        print(f"  [场景1] 设置 {len(offline_site_ids)} 个站点离线")

        # === 场景2：数据突变告警 ===
        if hydro_sites:
            sid = hydro_sites[0]['id']
            # 注入一个突变值到sensor_data
            db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
                       (sid, 'water_level', 15.8, 'm', now_str))
            create_alert_internal(db, sid, 'data_spike', 15.8, 'red',
                f'数据异常陡增：水位 15.80m（均值4.20m，变化276%）')
        print("  [场景2] 数据突变告警已创建")

        # === 场景3：数据冻结告警 ===
        if wl_sites:
            sid = wl_sites[1]['id'] if len(wl_sites) > 1 else wl_sites[0]['id']
            # 注入6条相同值的数据
            for i in range(6):
                t = (now - timedelta(minutes=i*5)).strftime('%Y-%m-%d %H:%M:%S')
                db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
                           (sid, 'water_level', 3.45, 'm', t))
            create_alert_internal(db, sid, 'data_freeze', 3.45, 'yellow',
                f'数据冻结：水位连续6条记录值一致（3.45），传感器可能故障')
        print("  [场景3] 数据冻结告警已创建")

        # === 场景4：数据延迟/缺失告警 ===
        if rain_sites:
            sid = rain_sites[2]['id'] if len(rain_sites) > 2 else rain_sites[0]['id']
            create_alert_internal(db, sid, 'data_gap', 180, 'yellow',
                f'数据延迟：降雨量已有180分钟未更新')
        print("  [场景4] 数据缺失告警已创建")

        # === 场景5：设备离线告警 ===
        if soil_sites:
            sid = soil_sites[0]['id']
            dev = db.execute("SELECT id FROM device_shadows WHERE site_id=? LIMIT 1", (sid,)).fetchone()
            if dev:
                db.execute("UPDATE device_shadows SET status='offline', last_data_time=NULL WHERE id=?", (dev['id'],))
                create_alert_internal(db, sid, 'device_status', 0, 'yellow',
                    f'设备离线: 土壤水分传感器 · {all_sites[0]["name"]}')
        print("  [场景5] 设备离线告警已创建")

        # === 场景6：告警未确认（pending状态，已在上面创建中体现） ===
        print("  [场景6] 多条pending状态告警已存在")

        # === 场景7：告警转工单（创建已关联的告警+工单对） ===
        if len(all_sites) > 5:
            sid = all_sites[5]['id']
            # 创建告警
            db.execute("""
                INSERT INTO alerts (site_id,metric,value,level,message,status,flow_type,flow_status)
                VALUES (?,?,?,?,?,?,?,?)
            """, (sid, 'device_status', 0, 'yellow', '设备离线: 雷达水位计', 'pending', 'auto', 'converted'))
            alert_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            # 创建关联工单
            order_no = f"WO-{now.strftime('%Y%m%d')}-701"
            db.execute("""
                INSERT INTO work_orders (order_no,site_id,source,event_type,level,title,description,assignee,status,sla_deadline,related_alert_id)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """, (order_no, sid, 'auto', '设备故障', 'normal', '[自动] 设备离线: 雷达水位计',
                  '设备离线: 雷达水位计', '张建国', 'in_progress',
                  (now + timedelta(hours=72)).strftime('%Y-%m-%d %H:%M'), alert_id))
            db.execute("UPDATE alerts SET related_order_no=?, flow_status='converted' WHERE id=?", (order_no, alert_id))
        print("  [场景7] 告警-工单关联对已创建")

        # === 场景8：工单超时（SLA超期） ===
        if len(all_sites) > 6:
            sid = all_sites[6]['id']
            past_deadline = (now - timedelta(hours=5)).strftime('%Y-%m-%d %H:%M')
            order_no = f"WO-{now.strftime('%Y%m%d')}-801"
            db.execute("""
                INSERT INTO work_orders (order_no,site_id,source,event_type,level,title,description,assignee,status,sla_deadline,created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """, (order_no, sid, 'auto', '设备故障', 'urgent', '水位计数据中断',
                  '设备持续2小时无数据上报', '张建国', 'in_progress', past_deadline,
                  (now - timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')))
        print("  [场景8] SLA超时工单已创建")

        # === 场景9：工单长时间未更新 ===
        if len(all_sites) > 7:
            sid = all_sites[7]['id']
            order_no = f"WO-{now.strftime('%Y%m%d')}-901"
            db.execute("""
                INSERT INTO work_orders (order_no,site_id,source,event_type,level,title,description,assignee,status,sla_deadline,created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """, (order_no, sid, 'patrol', '巡检维修', 'normal', '护栏损坏修复',
                  '巡检发现河道护栏损坏', '黎明', 'in_progress',
                  (now + timedelta(hours=48)).strftime('%Y-%m-%d %H:%M'),
                  (now - timedelta(days=2)).strftime('%Y-%m-%d %H:%M:%S')))
        print("  [场景9] 长时间未更新工单已创建")

        # === 场景10：巡检异常 ===
        if len(all_sites) > 3:
            sid = all_sites[3]['id']
            # 创建巡检计划
            db.execute("""
                INSERT INTO inspection_plans (plan_name,site_id,type,start_date,end_date,status,period)
                VALUES (?,?,?,?,?,?,?)
            """, (f"异常巡检-{now.strftime('%Y%m%d')}", sid, 'daily',
                  now.strftime('%Y-%m-%d'), now.strftime('%Y-%m-%d'), 'in_progress', 'daily'))
            plan_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            # 创建异常巡检任务
            db.execute("""
                INSERT INTO inspection_tasks (plan_id,site_id,task_name,result,remark)
                VALUES (?,?,?,?,?,?)
            """, (plan_id, sid, '水位计校验', 'abnormal',
                  (now - timedelta(hours=2)).strftime('%Y-%m-%d %H:%M:%S'), '水位计读数偏差超过5cm'))
            # 触发告警
            create_alert_internal(db, sid, 'inspection', 0, 'yellow',
                '巡检异常：水位计校验 - 水位计读数偏差超过5cm')
        print("  [场景10] 巡检异常告警已创建")

        # === 场景11：巡检计划逾期未完成 ===
        if len(all_sites) > 8:
            sid = all_sites[8]['id']
            past_date = (now - timedelta(days=3)).strftime('%Y-%m-%d')
            db.execute("""
                INSERT INTO inspection_plans (plan_name,site_id,type,start_date,end_date,status,period)
                VALUES (?,?,?,?,?,?,?)
            """, ('逾期未完成巡检-演示', sid, 'weekly', past_date, past_date, 'pending', 'weekly'))
            plan_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            # 创建未完成的任务
            for item in ['设备外观检查', '数据采集器校验', '通信模块检查']:
                db.execute("""
                    INSERT INTO inspection_tasks (plan_id,site_id,task_name)
                    VALUES (?,?,?)
                """, (plan_id, sid, item))
        print("  [场景11] 逾期未完成巡检计划已创建")

        # === 场景12：备件库存不足 ===
        # 将某个备件库存降到0以下
        low_part = db.execute("SELECT id FROM spare_parts_inventory WHERE part_code='BJ-005'").fetchone()
        if low_part:
            db.execute("UPDATE spare_parts_inventory SET quantity=0, min_quantity=2 WHERE id=?", (low_part['id'],))
        print("  [场景12] 备件库存不足已设置")

        # === 场景13：备件申请待审批 ===
        # 已在seed_data中创建，这里确保有pending状态的申请
        pending_req = db.execute("SELECT COUNT(*) FROM spare_part_requests WHERE status='pending'").fetchone()[0]
        if pending_req == 0:
            rno = f"BJ-{now.strftime('%Y%m%d')}-999"
            db.execute("""
                INSERT INTO spare_part_requests (request_no,site_id,applicant,part_name,quantity,reason,status)
                VALUES (?,?,?,?,?,?,?)
            """, (rno, all_sites[0]['id'], '运维人员', '雷达水位计', 1, '设备故障需更换', 'pending'))
        print("  [场景13] 备件申请待审批已确认")

        # === 场景14：热线事件未处理 ===
        # 创建一个新的未处理热线事件
        db.execute("""
            INSERT INTO hotline_events (caller_name,caller_phone,event_type,description,location,status,operator,created_at)
            VALUES (?,?,?,?,?,?,?,?)
        """, ('赵先生', '13900009900', '水位异常', '河道水位上涨迅速，疑似上游水库泄洪',
              '赣江下游段', 'registered', '李敏',
              (now - timedelta(minutes=30)).strftime('%Y-%m-%d %H:%M:%S')))
        print("  [场景14] 未处理热线事件已创建")

        # === 场景15：热线转工单后工单未完成 ===
        # 创建热线事件+关联工单
        db.execute("""
            INSERT INTO hotline_events (caller_name,caller_phone,event_type,description,location,status,related_order_no,site_id,operator,created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, ('孙女士', '13900008800', '设施损坏', '堤防护坡出现塌陷', '城南堤防段',
              'dispatched', f"WO-{now.strftime('%Y%m%d')}-1501", all_sites[2]['id'], '王芳',
              (now - timedelta(hours=3)).strftime('%Y-%m-%d %H:%M:%S')))
        order_no = f"WO-{now.strftime('%Y%m%d')}-1501"
        db.execute("""
            INSERT INTO work_orders (order_no,site_id,source,event_type,level,title,description,assignee,status,sla_deadline,created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (order_no, all_sites[2]['id'], 'hotline', '设施维修', 'urgent',
              '[热线] 堤防护坡塌陷', '堤防护坡出现塌陷，需紧急修复', '王刚', 'in_progress',
              (now + timedelta(hours=4)).strftime('%Y-%m-%d %H:%M'),
              (now - timedelta(hours=3)).strftime('%Y-%m-%d %H:%M:%S')))
        print("  [场景15] 热线转工单（未完成）已创建")

        # 注入标记
        db.execute("""
            INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark)
            VALUES ('system',0,'abnormal_scenarios_seeded','系统','15种异常场景数据注入完成')
        """)
        db.commit()
        print("[Seed] 15种异常场景数据注入完成！")

def seed_maintenance():
    """运维计划种子数据（仅首次）"""
    with get_db() as db:
        mcnt = db.execute("SELECT COUNT(*) FROM maintenance_plans").fetchone()[0]
        if mcnt == 0:
            sites = db.execute("SELECT id, name FROM sites LIMIT 50").fetchall()
            categories = [
                ('站院环境维护','environment','weekly'),
                ('站房维护','facility','biweekly'),
                ('设施设备维护','facility','monthly'),
                ('观测场管理','observation','biweekly'),
                ('断面环境管理','section','monthly'),
                ('安全检查','safety','monthly'),
                ('发电机保养','generator','monthly'),
            ]
            now = datetime.now()
            import itertools
            plan_id_counter = itertools.count(1)
            for site in sites:
                for cat_name, cat_key, freq in categories:
                    if freq == 'weekly':
                        due = now + timedelta(days=random.randint(0,7))
                    elif freq == 'biweekly':
                        due = now + timedelta(days=random.randint(0,14))
                    elif freq == 'monthly':
                        due = now + timedelta(days=random.randint(0,30))
                    else:
                        due = now + timedelta(days=random.randint(0,90))
                    # 60%概率已处理，40%待处理
                    status = 'completed' if random.random() < 0.6 else 'pending'
                    completed_at = due.strftime('%Y-%m-%d') if status == 'completed' else None
                    db.execute(
                        "INSERT INTO maintenance_plans (site_id,plan_name,category,frequency,due_date,status,assignee,completed_at) VALUES (?,?,?,?,?,?,?,?)",
                        (site['id'], f"{site['name']}{cat_name}", cat_key, freq,
                         due.strftime('%Y-%m-%d'), status, '管理员', completed_at)
                    )
            db.commit()
            print(f"[Seed] {len(sites)*len(categories)} maintenance plans seeded.")

def seed_maintenance_templates():
    """预置标准化运维模板（仅首次）"""
    templates = [
        # === 日常维护类（站院环境）===
        ('日常维护','environment','驻测站站院环境维护（每周）','weekly',
         '对水位井、站院、大门口进行全面的打扫，确保干净整洁',
         '地面、窗台、设备等干净整洁，墙面、天花板无污迹、蜘蛛网、昆虫等',
         '[{"id":"c1","label":"水位井区域全面打扫"},{"id":"c2","label":"站院地面及大门口清洁"},{"id":"c3","label":"设备表面及窗台擦拭"},{"id":"c4","label":"墙面天花板检查（无污迹/蜘蛛网）"},{"id":"c5","label":"站院草地灌木修剪维护"}]',
         1, 2, 1),
        ('日常维护','observation','观测场草地维护（每半月）','biweekly',
         '对降蒸观测场草地进行维护，草皮高度符合规范要求',
         '降蒸观测场、站院草皮高度低于20cm，遇重大活动增加维护次数',
         '[{"id":"c1","label":"草地修剪（草高<20cm）"},{"id":"c2","label":"杂草清理"},{"id":"c3","label":"场地平整度检查"},{"id":"c4","label":"巡测站站房全面打扫"}]',
         1, 2, 2),
        ('日常维护','section','断面环境管理（每月+汛后）','monthly',
         '测流断面、水尺断面、码头清理杂草杂木淤泥，确保断面整洁',
         '断面无积水、无淤泥、无杂草、无杂物',
         '[{"id":"c1","label":"测流断面上下游各5米杂草清理"},{"id":"c2","label":"缆道铁塔四周清理"},{"id":"c3","label":"基本水尺断面上下游各10米清理"},{"id":"c4","label":"水尺码头/停船码头淤泥清理"},{"id":"c5","label":"比降断面水尺道路清理（汛期）"},{"id":"c6","label":"洪水退水及时清理"}]',
         1, 3, 3),
        # === 日常管理类（水位观测）===
        ('日常管理','water_level','水位项目日常巡查（每日/每周）','weekly',
         '观测基本水尺读数并记录，校对遥测水位及时间，检查清洗水尺设备',
         '人工与遥测水位相差≥0.02m时需复核报送调整；驻测站每日2次，巡测站每日1次',
         '[{"id":"c1","label":"基本水尺读数记录"},{"id":"c2","label":"遥测水位及时间校对"},{"id":"c3","label":"偏差检测（≥0.02m报送水情科）"},{"id":"c4","label":"水尺清洗检查"},{"id":"c5","label":"水位设备运行检查"},{"id":"c6","label":"填写水位巡查表并拍照存档"}]',
         1, 0.5, 4),
        ('日常管理','facility','设施设备巡查（每月）','monthly',
         '检查清洗水尺，对设施设备、爬梯、护栏牢固度进行全面检查',
         '填写设施设备巡查表，异常维修拍照存档并报中心站网监测科',
         '[{"id":"c1","label":"水尺清洗检查"},{"id":"c2","label":"爬梯牢固度检查"},{"id":"c3","label":"护栏牢固度检查"},{"id":"c4","label":"设施设备外观检查"},{"id":"c5","label":"异常维修拍照存档"},{"id":"c6","label":"上报中心站网监测科"}]',
         1, 2, 5),
        ('日常管理','safety','安全检查（每月）','monthly',
         '对测验设施设备、安全环境、站房、灭火器、安全器材进行全面安全检查',
         '填记安全检查记录表，存在安全隐患需及时告知鄱阳湖水文水资源监测中心',
         '[{"id":"c1","label":"灭火器压力及有效期检查"},{"id":"c2","label":"安全器材完好性检查"},{"id":"c3","label":"站房结构安全检查"},{"id":"c4","label":"电气线路检查"},{"id":"c5","label":"填写安全检查记录表"},{"id":"c6","label":"安全隐患告知中心"}]',
         1, 1.5, 6),
        ('日常管理','generator','发电机保养维护（每月+汛前汛后）','monthly',
         '每月检查机油线路并运行≥30分钟；每年汛前汛后更换机油及线路保养',
         '发电机运行正常，备足燃料及机油，记录运行时间',
         '[{"id":"c1","label":"机油液位检查"},{"id":"c2","label":"线路及各部件检查"},{"id":"c3","label":"发电运行≥30分钟并记录"},{"id":"c4","label":"燃料及机油储备检查"},{"id":"c5","label":"汛前/汛后更换机油保养"}]',
         1, 1.5, 7),
        # === 设备仪器维护类 ===
        ('设备仪器维护','rainfall','雨量项目日常巡检（每月）','monthly',
         '遥测雨量器现场运行维护巡检，含数据采集终端、供电设备、雨量筒检查',
         '每季度进行注水试验（≥12.5mm，误差≤±4%），特大暴雨后及时检查',
         '[{"id":"c1","label":"数据采集终端外观及状态检查"},{"id":"c2","label":"供电设备检查"},{"id":"c3","label":"布线检查"},{"id":"c4","label":"雨量筒外观/器口水平检查"},{"id":"c5","label":"环境清理"},{"id":"c6","label":"季度注水试验（误差≤±4%）"}]',
         1, 2, 8),
        ('设备仪器维护','evaporation','蒸发项目日常巡检（每月）','monthly',
         '自动蒸发设备遥测终端现场运行维护巡检及换水',
         '每月不少于1次巡测，每半年渗漏检查，每月至少换水一次保持清洁',
         '[{"id":"c1","label":"自动蒸发设备遥测终端巡检"},{"id":"c2","label":"蒸发器换水（保持清洁）"},{"id":"c3","label":"水圈清洁及环境维护"},{"id":"c4","label":"渗漏检查（每半年）"},{"id":"c5","label":"数据合理性检查"},{"id":"c6","label":"汛前自动注水实验"}]',
         1, 1.5, 9),
        ('设备仪器维护','cableway','缆道日常巡检（测流时）','seasonal',
         '测流时对主索、循环索、锚碇、导向轮、绞车等进行检查维护',
         '检查锚碇位移、钢丝绳夹头松紧、绞车运转；异常拍照留底并通知甲方',
         '[{"id":"c1","label":"主索/循环索检查维护"},{"id":"c2","label":"拉线/卡头检查（异常通知甲方）"},{"id":"c3","label":"工作索毛刺断骨拍照留底"},{"id":"c4","label":"锚碇位移/土壤裂纹检查"},{"id":"c5","label":"导向轮/游轮运转检查"},{"id":"c6","label":"绞车运转检查"},{"id":"c7","label":"钢丝绳夹头/生锈/排水检查"}]',
         1, 3, 10),
        ('设备仪器维护','soil_moisture','墒情站日常巡查（季度）','seasonal',
         '对墒情基本站进行巡查，保持整洁、数据校测',
         '每季度对基本站巡查不少于1次，保持机箱内干净整洁，清理周边杂草；干旱天气取土检验',
         '[{"id":"c1","label":"机箱内部清洁"},{"id":"c2","label":"周边杂草清理"},{"id":"c3","label":"无积水检查"},{"id":"c4","label":"数据校测记录"},{"id":"c5","label":"辅助站取土烘干法检验（干旱触发）"}]',
         0, 1.5, 11),
    ]
    with get_db() as db:
        cnt = db.execute("SELECT COUNT(*) FROM maintenance_templates").fetchone()[0]
        if cnt == 0:
            for t in templates:
                db.execute(
                    "INSERT INTO maintenance_templates (category,sub_category,title,frequency,description,standard,check_items,photo_required,estimated_hours,sort_order) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    t
                )
            db.commit()
            print(f"[Seed] {len(templates)} maintenance templates seeded.")

def seed_water_quality_templates():
    """水质站点巡检模板种子数据"""
    with get_db() as db:
        # 检查是否已存在
        existing = db.execute("SELECT COUNT(*) FROM inspection_templates WHERE category='水质监测'").fetchone()[0]
        if existing > 0:
            # 确保水质排程存在
            wq_sch = db.execute("""
                SELECT COUNT(*) FROM inspection_schedules s
                JOIN inspection_templates t ON s.template_id = t.id
                WHERE t.category='水质监测'
            """).fetchone()[0]
            if wq_sch == 0:
                _init_v2_schedules(db)
            return

        now = datetime.now()
        today = now.strftime('%Y-%m-%d')

        # === 创建水质监测模板 ===
        # 每周巡检模板
        weekly_items = [
            ('进出站点拍照定位打卡', '站房环境', 1, 0, 1, '对着站房门拍摄，清晰显示站房全貌及门牌标识，自动记录GPS坐标', 1),
            ('站房及配件设施', '站房环境', 1, 0, 4, '站房内部环境全景、配电箱、空调/除湿设备、温湿度计', 2),
            ('采水系统', '设备运维', 1, 0, 1, '拍摄采水管路、取水口、预处理单元，确认无渗漏、无堵塞', 3),
            ('消防设施及检查登记', '站房环境', 1, 0, 1, '拍摄灭火器及消防检查卡，确认压力指针在绿区、有效期未超', 4),
            ('高锰酸盐指数仪器质控', '质控校准', 1, 1, 2, '质控样测定结果界面+数据记录页，偏差须在±10%内', 5),
            ('氨氮仪器质控', '质控校准', 1, 1, 2, '质控样测定结果界面+数据记录页，偏差须在±10%内', 6),
            ('总磷仪器质控', '质控校准', 1, 1, 2, '质控样测定结果界面+数据记录页，偏差须在±10%内', 7),
            ('总氮仪器质控', '质控校准', 1, 1, 2, '质控样测定结果界面+数据记录页，偏差须在±10%内', 8),
            ('五参数仪器质控', '质控校准', 1, 1, 2, 'pH/电导率/DO/浊度/温度五参数标液核查结果+记录页', 9),
            ('运维维护登记本', '台账登记', 1, 0, 1, '运维登记本当前页，确认最近一周记录填写完整', 10),
            ('质控登记本', '台账登记', 1, 0, 1, '质控登记本当前页，确认数据已规范记录', 11),
            ('废液处理登记本', '台账登记', 1, 0, 1, '废液产生量、转移量、处理时间记录完整', 12),
        ]

        monthly_items = [
            ('电表读数记录', '站房环境', 1, 0, 1, '拍摄电表读数界面，记录当月用电量', 1),
            ('仪器月度校准', '质控校准', 1, 1, 4, '高锰酸盐指数/氨氮/总磷/总氮四台仪器校准结果各1张，校准曲线r≥0.999', 2),
        ]

        templates_data = [
            ('水质站点每周巡检模板', '水质监测', 'weekly', '水质站点每周例行巡检', weekly_items),
            ('水质站点每月巡检模板', '水质监测', 'monthly', '水质站点每月深度巡检', monthly_items),
        ]

        template_ids = {}

        for tpl_name, category, frequency, desc, items in templates_data:
            db.execute("""
                INSERT INTO inspection_templates (template_name, category, frequency, description, sort_order)
                VALUES (?,?,?,?,?)
            """, (tpl_name, category, frequency, desc, len(template_ids) + 1))
            tid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            template_ids[tpl_name] = tid

            for item_name, item_cat, photo_req, need_rev, max_photos, standard, sort_ord in items:
                db.execute("""
                    INSERT INTO inspection_template_items
                    (template_id, item_name, category, photo_required, need_review, max_photos, inspection_standard, sort_order)
                    VALUES (?,?,?,?,?,?,?,?)
                """, (tid, item_name, item_cat, photo_req, need_rev, max_photos, standard, sort_ord))

        print(f"[InspV2] 水质监测 {len(templates_data)} 个模板已创建")

        # === 创建配置规则 ===
        water_quality_sites = db.execute("SELECT id FROM sites WHERE type='water_quality'").fetchall()
        config_count = 0
        for site in water_quality_sites:
            for tpl_name, tid in template_ids.items():
                # 以站点ID作为site_type创建独立配置（因为每个站点可能不同）
                db.execute("""
                    INSERT INTO inspection_configs (site_type, device_types, template_id, is_active)
                    VALUES (?,?,?,1)
                """, ('water_quality', '', tid))
                config_count += 1

        print(f"[InspV2] 水质监测 {config_count} 条配置规则已创建")

        # === 初始化排程 ===
        _init_v2_schedules(db)

        db.commit()
        print("[InspV2] 水质监测排程已初始化")


def seed_param_thresholds():
    """监测参数正常阈值表（用于数据健康度「超限」判定与告警阈值回退）。
    幂等：INSERT OR IGNORE，重复启动不覆盖既有配置。"""
    rows = [
        # metric,          label,     unit,   low,  high,  critical_low, critical_high
        ('ph',              'pH',      '',     6.0,  9.0,   5.0,          10.0),
        ('ammonia',         '氨氮',     'mg/L', 0.0,  2.0,   0.0,          3.0),
        ('dissolved_oxygen', '溶解氧',   'mg/L', 2.0,  20.0,  1.0,          25.0),
        ('cod',             '化学需氧量', 'mg/L', 0.0,  40.0,  0.0,          60.0),
        ('total_nitrogen',  '总氮',     'mg/L', 0.0,  2.0,   0.0,          3.0),
        ('total_phosphorus', '总磷',     'mg/L', 0.0,  0.4,   0.0,          0.6),
        ('turbidity',       '浊度',     'NTU',  0.0,  100.0, 0.0,          200.0),
        ('water_temp',      '水温',     '°C',   0.0,  35.0,  -2.0,         40.0),
    ]
    with get_db() as db:
        for r in rows:
            db.execute(
                "INSERT OR IGNORE INTO param_thresholds (metric, label, unit, low, high, critical_low, critical_high) "
                "VALUES (?,?,?,?,?,?,?)", r)
        db.commit()
    print(f"[Seed] 参数阈值表已就绪（{len(rows)} 项）")


def seed_inspection_v2():
    """巡检V2三层架构种子数据：方案模板 + 巡检配置 + 排程初始化"""
    with get_db() as db:
        cnt = db.execute("SELECT COUNT(*) FROM inspection_templates").fetchone()[0]
        if cnt > 0:
            print("[InspV2] 模板已存在，跳过")
            # 确保排程已初始化
            sch_cnt = db.execute("SELECT COUNT(*) FROM inspection_schedules").fetchone()[0]
            if sch_cnt == 0:
                _init_v2_schedules(db)
            seed_water_quality_templates()
            # 确保配置存在
            cfg_cnt = db.execute("SELECT COUNT(*) FROM inspection_configs").fetchone()[0]
            if cfg_cnt == 0:
                print("[InspV2] 配置为空，重新填充巡检配置")
                _seed_inspection_configs(db)
            return

        now = datetime.now()
        today = now.strftime('%Y-%m-%d')

        # === 第一层：方案模板 ===
        templates_data = [
            # (template_name, category, frequency, description, items)
            # items = [(item_name, category, frequency_level, photo_required, sort_order), ...]
            ('水位观测日常方案', '水位观测', 'daily', '每日水位观测及设备的日常巡查', [
                ('基本水尺读数观测记录', '水位观测', 'high', 1, 1),
                ('遥测水位及时间校对', '水位观测', 'high', 0, 2),
                ('人工与遥测水位偏差检测', '水位观测', 'high', 0, 3),
                ('水尺清洗检查', '水位观测', 'mid', 1, 4),
                ('水位设备运行检查', '水位观测', 'mid', 0, 5),
                ('填记水位巡查表并拍照存档', '水位观测', 'high', 1, 6),
            ]),
            ('水位观测月度方案', '水位观测', 'monthly', '月度水位观测设备深度检查', [
                ('水位计精度校验', '水位观测', 'mid', 0, 1),
                ('水位数据完整性审查', '水位观测', 'mid', 0, 2),
                ('传感器线缆检查', '水位观测', 'low', 0, 3),
                ('备品备件储备检查', '水位观测', 'low', 0, 4),
            ]),
            ('雨量监测日常方案', '雨量监测', 'daily', '每日雨量监测设备巡查', [
                ('雨量筒外观及水平检查', '雨量监测', 'mid', 1, 1),
                ('翻斗灵活性检查', '雨量监测', 'high', 0, 2),
                ('数据采集终端状态检查', '雨量监测', 'high', 0, 3),
                ('供电设备检查', '雨量监测', 'mid', 0, 4),
            ]),
            ('雨量监测季度方案', '雨量监测', 'quarterly', '季度雨量监测设备深度维护', [
                ('注水试验(≥12.5mm误差≤±4%)', '雨量监测', 'low', 0, 1),
                ('雨量筒内部清洁', '雨量监测', 'mid', 1, 2),
                ('特大暴雨后设备全面检查', '雨量监测', 'low', 1, 3),
            ]),
            ('蒸发监测月度方案', '蒸发监测', 'monthly', '月度蒸发监测设备维护', [
                ('自动蒸发设备遥测终端巡检', '蒸发监测', 'mid', 0, 1),
                ('蒸发器换水保持清洁', '蒸发监测', 'mid', 1, 2),
                ('水圈清洁及环境维护', '蒸发监测', 'mid', 1, 3),
                ('数据合理性检查', '蒸发监测', 'mid', 0, 4),
                ('渗漏检查(半年期)', '蒸发监测', 'low', 0, 5),
            ]),
            ('蒸发监测半年方案', '蒸发监测', 'semi_annual', '半年度蒸发监测深度检查', [
                ('蒸发系统全面渗漏检查', '蒸发监测', 'low', 0, 1),
                ('汛前自动注水实验', '蒸发监测', 'low', 1, 2),
            ]),
            ('站院环境周方案', '站院环境', 'weekly', '每周站院环境维护', [
                ('水位井/站院/大门口全面打扫', '站院环境', 'high', 1, 1),
                ('设备表面及窗台擦拭', '站院环境', 'mid', 1, 2),
                ('墙面天花板检查(无污迹/蜘蛛网)', '站院环境', 'low', 0, 3),
                ('草地灌木修剪维护', '站院环境', 'mid', 1, 4),
                ('巡测站站房全面打扫', '站院环境', 'mid', 1, 5),
                ('观测场草地维护(草高<20cm)', '站院环境', 'mid', 1, 6),
            ]),
            ('站院环境月方案', '站院环境', 'monthly', '月度站院深度清洁维护', [
                ('站房深度清洁(含窗户/天花板)', '站院环境', 'mid', 1, 1),
                ('仪器设备全面擦拭', '站院环境', 'mid', 1, 2),
                ('排水沟清理', '站院环境', 'low', 0, 3),
                ('站院安全隐患排查', '站院环境', 'low', 0, 4),
            ]),
            ('设施设备巡查方案', '设施设备', 'monthly', '月度设施设备全面检查', [
                ('水尺清洗检查', '设施设备', 'mid', 1, 1),
                ('爬梯/护栏牢固度全面检查', '设施设备', 'mid', 0, 2),
                ('设施设备外观检查', '设施设备', 'mid', 1, 3),
                ('异常维修与拍照存档', '设施设备', 'mid', 1, 4),
                ('上报中心站网监测科', '设施设备', 'low', 0, 5),
            ]),
            ('安全检查月方案', '安全检查', 'monthly', '月度安全全面检查', [
                ('测验设施设备安全环境检查', '安全防护', 'mid', 0, 1),
                ('灭火器压力及有效期检查', '安全防护', 'mid', 1, 2),
                ('安全器材完好性检查', '安全防护', 'mid', 0, 3),
                ('站房结构安全及电气线路检查', '安全防护', 'mid', 0, 4),
                ('填写安全检查记录表', '安全防护', 'mid', 1, 5),
                ('安全隐患及时告知中心', '安全防护', 'low', 0, 6),
            ]),
            ('发电机保养方案', '发电机', 'quarterly', '季度发电机保养维护', [
                ('发电机维护保养(更换机油/线路/备足燃料)', '发电机', 'low', 1, 1),
                ('机油液位检查', '发电机', 'mid', 0, 2),
                ('线路及各部件检查', '发电机', 'mid', 0, 3),
                ('发电运行≥30分钟并记录', '发电机', 'mid', 1, 4),
                ('燃料及机油储备检查', '发电机', 'mid', 0, 5),
            ]),
            ('缆道系统巡查方案', '缆道系统', 'monthly', '月度缆道系统检查维护', [
                ('行主索/循环索检查维护', '缆道系统', 'mid', 1, 1),
                ('拉线/卡头检查(异常通知甲方)', '缆道系统', 'mid', 1, 2),
                ('工作索毛刺断骨拍照留底', '缆道系统', 'mid', 1, 3),
                ('锚碇位移/土壤裂纹检查', '缆道系统', 'mid', 0, 4),
                ('导向轮/游轮/行车架运转检查', '缆道系统', 'mid', 0, 5),
                ('绞车运转检查', '缆道系统', 'mid', 0, 6),
                ('钢丝绳夹头/生锈/排水检查', '缆道系统', 'low', 0, 7),
            ]),
            ('断面环境季度方案', '断面环境', 'quarterly', '季度断面环境清理维护', [
                ('测流断面上下游各5米清理杂草杂木', '断面环境', 'mid', 1, 1),
                ('缆道铁塔四周清理', '断面环境', 'mid', 1, 2),
                ('基本水尺断面上下游各10米清理', '断面环境', 'mid', 1, 3),
                ('水尺码头/停船码头清理淤泥杂草', '断面环境', 'mid', 1, 4),
                ('比降断面水尺道路清理', '断面环境', 'low', 0, 5),
            ]),
            ('墒情监测日常方案', '墒情监测', 'daily', '每日墒情监测设备巡查', [
                ('机箱内部清洁', '墒情监测', 'mid', 1, 1),
                ('周边杂草清理', '墒情监测', 'mid', 0, 2),
                ('无积水检查', '墒情监测', 'mid', 0, 3),
                ('数据校测记录', '墒情监测', 'high', 0, 4),
            ]),
        ]

        template_ids = {}  # template_name -> id
        item_ids = {}      # (template_name, item_name) -> id

        for tpl_name, category, frequency, desc, items in templates_data:
            db.execute("""
                INSERT INTO inspection_templates (template_name, category, frequency, description, sort_order)
                VALUES (?,?,?,?,?)
            """, (tpl_name, category, frequency, desc, len(template_ids) + 1))
            tid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            template_ids[tpl_name] = tid
            for item_name, item_cat, freq_level, photo_req, sort_ord in items:
                db.execute("""
                    INSERT INTO inspection_template_items (template_id, item_name, category, frequency_level, photo_required, sort_order)
                    VALUES (?,?,?,?,?,?)
                """, (tid, item_name, item_cat, freq_level, photo_req, sort_ord))
                iid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
                item_ids[(tpl_name, item_name)] = iid

        print(f"[InspV2] {len(templates_data)} 个方案模板已创建")

        # === 第二层：巡检配置（站点类型 → 模板匹配） ===
        config_map = {
            'hydrology': [
                '水位观测日常方案', '水位观测月度方案',
                '站院环境周方案', '站院环境月方案',
                '设施设备巡查方案', '安全检查月方案',
                '发电机保养方案', '缆道系统巡查方案',
            ],
            'water_level': [
                '水位观测日常方案', '水位观测月度方案',
                '站院环境周方案', '站院环境月方案',
                '设施设备巡查方案', '安全检查月方案',
            ],
            'rainfall': [
                '雨量监测日常方案', '雨量监测季度方案',
                '站院环境周方案', '站院环境月方案',
                '设施设备巡查方案', '安全检查月方案',
            ],
            'evaporation': [
                '蒸发监测月度方案', '蒸发监测半年方案',
                '站院环境周方案', '站院环境月方案',
                '设施设备巡查方案', '安全检查月方案',
            ],
            'soil_moisture': [
                '墒情监测日常方案',
                '站院环境周方案', '站院环境月方案',
                '设施设备巡查方案',
            ],
            'groundwater': [
                '水位观测日常方案',
                '站院环境月方案',
                '设施设备巡查方案',
            ],
            'station_yard': [
                '站院环境周方案', '站院环境月方案',
                '安全检查月方案',
            ],
        }

        config_count = 0
        for site_type, tpl_names in config_map.items():
            for tpl_name in tpl_names:
                tid = template_ids.get(tpl_name)
                if tid:
                    db.execute("""
                        INSERT INTO inspection_configs (site_type, device_types, template_id, is_active)
                        VALUES (?,?,?,1)
                    """, (site_type, '', tid))
                    config_count += 1

        print(f"[InspV2] {config_count} 条巡检配置规则已创建")

        # === 初始化排程 ===
        _init_v2_schedules(db)

        # === 默认提醒配置 ===
        db.execute("""
            INSERT INTO inspection_reminders (remind_days_before, remind_method, overdue_escalation, escalation_days, is_active)
            VALUES (1, 'notification', 1, 3, 1)
        """)
        # === 水质监测模板 ===
        seed_water_quality_templates()

        db.commit()
        print("[InspV2] 默认提醒配置已创建")
        print("[InspV2] 三层架构种子数据初始化完成！")

def _init_v2_schedules(db):
    """根据巡检配置为所有站点初始化排程"""
    import random
    now = datetime.now()
    today = now.strftime('%Y-%m-%d')

    # 频次对应的天数
    freq_days = {
        'daily': 1, 'weekly': 7, 'monthly': 30,
        'quarterly': 90, 'semi_annual': 180, 'annual': 365,
    }

    # 获取所有站点
    sites = db.execute("SELECT id, type FROM sites").fetchall()
    if not sites:
        return

    # 获取所有配置
    configs = db.execute("""
        SELECT ic.site_type, ic.template_id, it.template_name, it.frequency
        FROM inspection_configs ic
        JOIN inspection_templates it ON ic.template_id = it.id
        WHERE ic.is_active = 1
    """).fetchall()

    # 按站点类型分组配置
    type_configs = {}
    for cfg in configs:
        st = cfg['site_type']
        if st not in type_configs:
            type_configs[st] = []
        type_configs[st].append(cfg)

    schedule_count = 0
    for site in sites:
        site_type = site['type']
        matched = type_configs.get(site_type, [])
        if not matched:
            # 通用配置：至少给站院环境周方案
            matched = type_configs.get('station_yard', [])

        for cfg in matched:
            tpl_id = cfg['template_id']
            frequency = cfg['frequency']
            # 获取该模板的所有检查项
            items = db.execute("""
                SELECT id, item_name FROM inspection_template_items
                WHERE template_id = ?
            """, (tpl_id,)).fetchall()

            for item in items:
                # 检查是否已有排程
                existing = db.execute("""
                    SELECT id FROM inspection_schedules
                    WHERE site_id=? AND template_item_id=?
                """, (site['id'], item['id'])).fetchone()
                if existing:
                    continue

                # 计算初始 next_due_date（随机偏移模拟历史执行）
                fd = freq_days.get(frequency, 30)
                # 部分项已执行过几个周期
                past_cycles = random.randint(0, 3)
                due_date = now - timedelta(days=past_cycles * fd) + timedelta(days=fd)
                # 确保不超过今天太多
                if due_date > now + timedelta(days=fd):
                    due_date = now + timedelta(days=random.randint(0, fd))

                db.execute("""
                    INSERT INTO inspection_schedules (site_id, template_id, template_item_id, frequency, next_due_date, cycle_count)
                    VALUES (?,?,?,?,?,?)
                """, (site['id'], tpl_id, item['id'], frequency,
                      due_date.strftime('%Y-%m-%d'), past_cycles))
                schedule_count += 1

    db.commit()
    print(f"[InspV2] {schedule_count} 条排程记录已初始化")

def _seed_inspection_configs(db):
    """为所有已知站点类型填充巡检配置（不重复）"""
    existing = db.execute("SELECT site_type, template_id FROM inspection_configs").fetchall()
    existing_keys = {(r['site_type'], r['template_id']) for r in existing}
    inserted = 0
    # 获取所有模板
    tpls = db.execute("SELECT id, template_name, category FROM inspection_templates").fetchall()
    tpl_by_name = {r['template_name']: r for r in tpls}
    # 站点类型→模板清单（覆盖已知所有类型）
    config_map = {
        'water_quality': ['水质站点每周巡检模板', '水质站点每月巡检模板'],
        'manual_station': ['水质站点每周巡检模板', '水质站点每月巡检模板'],
        'drinking_source': ['水质站点每周巡检模板'],
        'cross_boundary': ['水质站点每周巡检模板', '水质站点每月巡检模板'],
        'groundwater': ['水位观测日常方案', '水位观测月度方案', '设施设备巡查方案'],
        'station_yard': ['站院环境周方案', '站院环境月方案', '安全检查月方案'],
        'hydrology': ['水位观测日常方案', '水位观测月度方案', '站院环境周方案', '站院环境月方案',
                      '设施设备巡查方案', '安全检查月方案', '发电机保养方案', '缆道系统巡查方案'],
        'water_level': ['水位观测日常方案', '水位观测月度方案', '站院环境周方案', '站院环境月方案',
                        '设施设备巡查方案', '安全检查月方案'],
        'rainfall': ['雨量监测日常方案', '雨量监测季度方案', '站院环境周方案', '站院环境月方案',
                     '设施设备巡查方案', '安全检查月方案'],
        'evaporation': ['蒸发监测月度方案', '蒸发监测半年方案', '站院环境周方案', '站院环境月方案',
                        '设施设备巡查方案', '安全检查月方案'],
        'soil_moisture': ['墒情监测日常方案', '站院环境周方案', '站院环境月方案', '设施设备巡查方案'],
    }
    for site_type, tpl_names in config_map.items():
        for tpl_name in tpl_names:
            tpl = tpl_by_name.get(tpl_name)
            if not tpl:
                continue
            key = (site_type, tpl['id'])
            if key not in existing_keys:
                db.execute("""
                    INSERT INTO inspection_configs (site_type, device_types, template_id, is_active)
                    VALUES (?,?,?,1)
                """, (site_type, '', tpl['id']))
                inserted += 1
    if inserted > 0:
        db.commit()
        print(f"[InspV2] 新增 {inserted} 条巡检配置")
    else:
        print(f"[InspV2] 巡检配置已存在，跳过")

# ===================== Simulator =====================

# 各河流警戒水位配置
RIVER_THRESHOLDS = {
    '赣江': {'high': 22.0, 'critical': 23.5, 'base': 18.5},
    '抚河': {'high': 32.0, 'critical': 33.5, 'base': 30.0},
    '鄱阳湖': {'high': 18.5, 'critical': 19.8, 'base': 16.5},
    '': {'high': 15.0, 'critical': 16.5, 'base': 13.0},  # 城区默认
}

# 站点类型对应的监测指标（水质9参数）
TYPE_METRICS = {
    'water_quality': [
        ('ph','',9.0,6.0),
        ('dissolved_oxygen','mg/L',None,5.0),
        ('codmn','mg/L',6.0,None),
        ('ammonia','mg/L',1.0,None),
        ('total_phosphorus','mg/L',0.2,None),
        ('total_nitrogen','mg/L',1.0,None),
        ('turbidity','NTU',20,None),
        ('conductivity','μS/cm',None,None),
        ('water_temp','°C',None,None),
    ],
    'manual_station': [
        ('ph','',9.0,6.0),
        ('dissolved_oxygen','mg/L',None,5.0),
        ('codmn','mg/L',6.0,None),
        ('ammonia','mg/L',1.0,None),
        ('total_phosphorus','mg/L',0.2,None),
        ('total_nitrogen','mg/L',1.0,None),
    ],
    'drinking_source': [
        ('ph','',8.5,6.5),
        ('dissolved_oxygen','mg/L',None,5.0),
        ('codmn','mg/L',4.0,None),
        ('ammonia','mg/L',0.5,None),
        ('total_phosphorus','mg/L',0.1,None),
        ('turbidity','NTU',5,None),
        ('conductivity','μS/cm',None,None),
    ],
    'cross_boundary': [
        ('ph','',9.0,6.0),
        ('dissolved_oxygen','mg/L',None,5.0),
        ('codmn','mg/L',6.0,None),
        ('ammonia','mg/L',1.0,None),
        ('total_phosphorus','mg/L',0.2,None),
        ('total_nitrogen','mg/L',1.0,None),
        ('turbidity','NTU',20,None),
    ],
    'groundwater': [
        ('ph','',8.5,6.5),
        ('ammonia','mg/L',0.5,None),
        ('turbidity','NTU',10,None),
        ('conductivity','μS/cm',None,None),
    ],
}

def _generate_site_data(site, db, now):
    """为单个站点生成传感器数据，并检测异常"""
    sid = site['id']; stype = site['type']
    river = site['river'] or ''
    th = RIVER_THRESHOLDS.get(river, RIVER_THRESHOLDS[''])
    base_wl = th['base']
    metrics_gen = []  # 记录已生成的指标，供异常检测使用

    if stype in ('water_quality','manual_station','drinking_source','cross_boundary','groundwater'):
        # 水质监测核心指标（基于GB 3838-2002 III类水标准模拟）
        ph = get_site_trend(sid,'ph',7.5,0.12,5.5,9.0)
        do = get_site_trend(sid,'do',6.5,0.3,3.0,10.0)
        codmn = get_site_trend(sid,'codmn',3.5,0.6,0.5,8.0)
        ammonia = get_site_trend(sid,'ammonia',0.25,0.04,0.01,1.2)
        tp = get_site_trend(sid,'tp',0.08,0.01,0.005,0.3)
        tn = get_site_trend(sid,'tn',0.5,0.06,0.01,1.5)
        turbidity = get_site_trend(sid,'turbidity',3.0,0.5,0.1,25)
        conductivity = get_site_trend(sid,'cond',350,15,100,800)
        water_temp = get_site_trend(sid,'water_temp',22,1.2,2,35)
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
            (sid,'ph',ph,'',now))
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
            (sid,'dissolved_oxygen',do,'mg/L',now))
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
            (sid,'codmn',codmn,'mg/L',now))
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
            (sid,'ammonia',ammonia,'mg/L',now))
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
            (sid,'total_phosphorus',tp,'mg/L',now))
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
            (sid,'total_nitrogen',tn,'mg/L',now))
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
            (sid,'turbidity',turbidity,'NTU',now))
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
            (sid,'conductivity',conductivity,'μS/cm',now))
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
            (sid,'water_temp',water_temp,'°C',now))
        metrics_gen = [('ph', ph), ('dissolved_oxygen', do), ('codmn', codmn),
                       ('ammonia', ammonia), ('total_phosphorus', tp), ('total_nitrogen', tn),
                       ('turbidity', turbidity), ('conductivity', conductivity), ('water_temp', water_temp)]

    # 对每个指标进行异常检测
    for metric, val in metrics_gen:
        detect_site_anomalies(db, sid, stype, metric, val, now)

def auto_resolve_alerts(db, site_id):
    """检查未办结告警对应站点的数据是否已恢复，是则自动办结"""
    try:
        unresolved = db.execute(
            "SELECT id, metric FROM alerts WHERE site_id=? AND status IN ('pending','acknowledged') AND flow_type='auto'",
            (site_id,)
        ).fetchall()
        for a in unresolved:
            if a['metric'] == 'device_status':
                continue  # 设备状态告警需人工确认
            # 检查是否有最近1小时的数据
            has_data = db.execute(
                "SELECT COUNT(*) FROM sensor_data WHERE site_id=? AND recorded_at >= datetime('now','-1 hour')",
                (site_id,)
            ).fetchone()[0]
            if has_data > 0:
                db.execute("UPDATE alerts SET status='resolved', resolved_at=datetime('now','localtime') WHERE id=?", (a['id'],))
    except Exception as e:
        print(f'[AutoResolve] error site={site_id}: {e}')

# ---- 告警规则引擎缓存（避免每条数据都查DB）----
_alert_rules_cache = None
_alert_rules_cache_time = 0

def _load_alert_rules(db):
    """加载告警规则配置，缓存60秒"""
    global _alert_rules_cache, _alert_rules_cache_time
    now = __import__('time').time()
    if _alert_rules_cache is not None and (now - _alert_rules_cache_time) < 60:
        return _alert_rules_cache
    import json as _j
    rules = {}
    rows = db.execute('SELECT * FROM alert_rule_config WHERE enabled=1').fetchall()
    for r in rows:
        d = dict(r)
        try: d['thresholds'] = _j.loads(d['thresholds']) if d.get('thresholds') else {}
        except: d['thresholds'] = {}
        rules[d['metric']] = d
    _alert_rules_cache = rules
    _alert_rules_cache_time = now
    return rules

def detect_site_anomalies(db, site_id, site_type, metric, current_value, recorded_at):
    """检测单个站点指标的异常情况：突变、冻结、缺失"""
    METRIC_CN = {
        'codmn':'高锰酸盐指数','ammonia':'氨氮','total_phosphorus':'总磷','total_nitrogen':'总氮',
        'water_temp':'水温','dissolved_oxygen':'溶解氧','ph':'pH','turbidity':'浊度',
        'conductivity':'电导率','temperature':'气温',
        'data_spike':'数据突变','data_freeze':'数据冻结','data_gap':'数据缺失',
        'device_status':'设备状态',
    }
    metric_cn = METRIC_CN.get(metric, metric)

    # ===== 水质指标阈值超标检测（基于GB 3838-2002 III类水标准） =====
    WQ_THRESHOLDS = {
        'ph':              {'min': 6.0, 'max': 9.0, 'unit': ''},
        'dissolved_oxygen':{'min': 5.0, 'max': None,'unit': 'mg/L'},
        'codmn':           {'min': None,'max': 6.0,  'unit': 'mg/L'},
        'ammonia':         {'min': None,'max': 1.0,  'unit': 'mg/L'},
        'total_phosphorus':{'min': None,'max': 0.2,  'unit': 'mg/L'},
        'total_nitrogen':  {'min': None,'max': 1.0,  'unit': 'mg/L'},
        'turbidity':       {'min': None,'max': 20,   'unit': 'NTU'},
        'conductivity':    {'min': None,'max': None,  'unit': 'μS/cm'},
    }
    if metric in WQ_THRESHOLDS:
        th = WQ_THRESHOLDS[metric]
        exceeded = False
        level = 'yellow'
        msg = ''
        if th['min'] is not None and current_value < th['min']:
            exceeded = True
            level = 'orange' if current_value < th['min'] * 0.85 else 'yellow'
            msg = f'{metric_cn}偏低：{current_value:.2f}{th["unit"]}（阈值≥{th["min"]}{th["unit"]}）'
        if th['max'] is not None and current_value > th['max']:
            exceeded = True
            level = 'red' if current_value > th['max'] * 1.2 else 'orange'
            msg = f'{metric_cn}超标：{current_value:.2f}{th["unit"]}（阈值≤{th["max"]}{th["unit"]}）'
        # pH双向超标用更严重的等级
        if metric == 'ph' and exceeded:
            if current_value < 5.5 or current_value > 9.5:
                level = 'red'
            elif current_value < 6.0 or current_value > 9.0:
                level = 'orange'
        if exceeded:
            create_alert_internal(db, site_id, metric, current_value, level, msg)

    try:
        # 在检测新异常之前，先检查已有的未办结告警是否可自动恢复
        auto_resolve_alerts(db, site_id)
        # 自动解除已有data_gap误报：站点恢复数据后自动办结告警
        try:
            existing_gaps = db.execute(
                "SELECT id FROM alerts WHERE site_id=? AND metric='data_gap' AND status IN ('pending','acknowledged')",
                (site_id,)
            ).fetchall()
            if existing_gaps:
                recent_data = db.execute(
                    "SELECT COUNT(*) FROM sensor_data WHERE site_id=? AND recorded_at >= datetime('now','-1 hour')",
                    (site_id,)
                ).fetchone()[0]
                if recent_data > 0:
                    for gap in existing_gaps:
                        db.execute(
                            "UPDATE alerts SET status='resolved', resolved_at=datetime('now','localtime') WHERE id=?",
                            (gap['id'],)
                        )
                        db.execute(
                            "INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                            ('alert', gap['id'], 'resolved', '系统', '数据已自动恢复，告警自动办结')
                        )
        except Exception:
            pass
        # 排除自然波动大的指标（pH、溶解氧、电导率受环境影响正常波动）
        EXCLUDE_SPIKE = {'ph','dissolved_oxygen','conductivity','water_temp'}
        recent = db.execute(
            "SELECT value, recorded_at FROM sensor_data WHERE site_id=? AND metric=? ORDER BY recorded_at DESC LIMIT 12",
            (site_id, metric)
        ).fetchall()
        if len(recent) < 4:
            return
        values = [r['value'] for r in recent]
        timestamps = [r['recorded_at'] for r in recent]
        latest = values[0]

        # 1. 数据冻结检测：连续N条完全相同（N从规则引擎读取，默认6条）
        rule_freeze = _load_alert_rules(db).get('data_freeze', {})
        freeze_n = int((rule_freeze.get('thresholds') or {}).get('yellow', 6))
        if len(values) >= freeze_n and metric not in EXCLUDE_SPIKE:
            frozen = len(set(round(v, 4) for v in values[:freeze_n])) == 1
            if frozen:
                create_alert_internal(db, site_id, 'data_freeze', latest, 'yellow',
                    f'数据冻结：{metric_cn}连续{freeze_n}条记录值一致（{latest}），传感器可能故障')
                return

        # 2. 突变检测（排除自然波动指标）
        # 要求至少8条历史数据，确保趋势稳定后再检测（避免重启后误报）
        if len(values) >= 8 and metric not in EXCLUDE_SPIKE:
            prev_vals = values[1:8]
            mean = sum(prev_vals) / len(prev_vals)
            # 均值为0或接近0时跳过
            if abs(mean) < 0.001:
                return
            # 计算百分比变化
            pct_change = abs(latest - mean) / abs(mean)
            # 标准差检测
            std = (sum((v - mean)**2 for v in prev_vals) / len(prev_vals))**0.5
            if std < abs(mean) * 0.005:
                std = abs(mean) * 0.005
            z_score = abs(latest - mean) / std
            # 要求：变化幅度 > 30% 且 偏离 > 8σ，同时绝对变化值大于指标特定阈值
            min_abs_change = {'ph': 0.8, 'dissolved_oxygen': 1.5, 'ammonia': 0.15,
                              'codmn': 2, 'total_phosphorus': 0.05, 'total_nitrogen': 0.3,
                              'turbidity': 5, 'conductivity': 50, 'water_temp': 3}
            abs_change = abs(latest - mean)
            min_abs = min_abs_change.get(metric, abs(mean) * 0.35)
            # 从规则引擎读取突变百分比阈值（默认30%），同时保留Z-score和绝对变化值双重校验
            rule_spike = _load_alert_rules(db).get('data_spike', {})
            spike_pct = (rule_spike.get('thresholds') or {}).get('yellow', 30)
            if pct_change * 100 > spike_pct and z_score > 8 and abs_change > min_abs:
                direction = '陡增' if latest > mean else '陡降'
                level = 'red' if z_score > 10 else 'orange'
                create_alert_internal(db, site_id, 'data_spike', latest, level,
                    f'数据异常{direction}：{metric_cn} {latest:.2f}（均值{mean:.2f}，变化{pct_change*100:.0f}%）')

        # 3. 数据缺失检测（阈值从规则引擎读取，默认60分钟）
        if len(timestamps) >= 2:
            try:
                t1 = datetime.strptime(str(timestamps[0])[:19], '%Y-%m-%d %H:%M:%S')
                t0 = datetime.strptime(str(timestamps[1])[:19], '%Y-%m-%d %H:%M:%S')
                gap_min = (t1 - t0).total_seconds() / 60
                rule_gap = _load_alert_rules(db).get('data_gap', {})
                threshold = int((rule_gap.get('thresholds') or {}).get('yellow', 60))
                if gap_min > threshold:
                    create_alert_internal(db, site_id, 'data_gap', gap_min, 'yellow',
                        f'数据延迟：{metric_cn}已有{gap_min:.0f}分钟未更新')
            except Exception:
                pass
    except Exception as e:
        print(f'[Anomaly] 检测异常失败 site={site_id} metric={metric}: {e}')

def _scheduler_db():
    """专用调度器数据库连接（超时5秒，异步写入，避免阻塞API）"""
    db = sqlite3.connect(DB_PATH, timeout=5, check_same_thread=False)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA busy_timeout=5000")
    db.execute("PRAGMA synchronous=OFF")
    return db

def generate_sensor_data():
    """每30秒生成模拟传感器数据"""
    PRESET_OFFLINE = {5}  # 预设离线站点（仅江桥水质监测站），跳过状态更新
    db = None
    try:
        db = _scheduler_db()
    except Exception as e:
        print(f'[Sim] 调度器连接失败: {e}')
        return
    try:
        sites = db.execute("SELECT * FROM sites").fetchall()
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        for idx, site in enumerate(sites):
            sid = site['id']
            # 预设离线站点跳过所有数据生成，保持种子数据一致性
            if sid in PRESET_OFFLINE:
                # 预设离线站点：确保设备和站点状态为offline
                db.execute("UPDATE device_shadows SET status='offline' WHERE site_id=?", (sid,))
                db.execute("UPDATE sites SET status='offline' WHERE id=?", (sid,))
                try:
                    db.commit()
                except Exception:
                    pass
                continue
            try:
                _generate_site_data(site, db, now)
            except Exception as e:
                if 'database is locked' in str(e):
                    print(f'[Sim] DB locked, skip site {sid}')
                else:
                    print(f'[Sim] site {sid} error: {e}')

            # === 数据自洽性修复：设备/站点状态同步 ===
            # 0.5%概率设备离线（模拟真实场景），99.5%在线
            devices = db.execute("SELECT id FROM device_shadows WHERE site_id=?", (sid,)).fetchall()
            any_online = False
            for dev in devices:
                dev_status = 'offline' if random.random() < 0.005 else 'online'
                if dev_status == 'online':
                    any_online = True
                db.execute("UPDATE device_shadows SET status=?, last_data_time=? WHERE id=?",
                           (dev_status, now if dev_status == 'online' else None, dev['id']))

            # 站点状态根据设备状态决定：全部离线→站点离线，否则在线
            site_status = 'online' if any_online else 'offline'
            db.execute("UPDATE sites SET status=?, last_heartbeat=? WHERE id=?",
                       (site_status, now if site_status == 'online' else None, sid))

            # 每个站点单独提交，释放写锁，让API请求能快速插入
            try:
                db.commit()
            except Exception as e:
                print(f'[Sim] commit fail site {sid}: {e}')

        # === 数据自洽性修复：长时间无数据的设备自动标记离线 ===
        rule_dev = _load_alert_rules(db).get('device_status', {})
        dev_timeout = int((rule_dev.get('thresholds') or {}).get('yellow', 30))
        threshold_time = (datetime.now() - timedelta(minutes=dev_timeout)).strftime('%Y-%m-%d %H:%M:%S')
        stale_devices = db.execute(
            "SELECT id, site_id FROM device_shadows WHERE last_data_time < ? AND status != 'offline'",
            (threshold_time,)
        ).fetchall()
        for dev in stale_devices:
            db.execute("UPDATE device_shadows SET status='offline' WHERE id=?", (dev['id'],))
            # 检查该站点是否所有设备都离线
            online_count = db.execute(
                "SELECT COUNT(*) FROM device_shadows WHERE site_id=? AND status='online'",
                (dev['site_id'],)
            ).fetchone()[0]
            if online_count == 0:
                db.execute("UPDATE sites SET status='offline' WHERE id=?", (dev['site_id'],))
        if stale_devices:
            db.commit()

        # === 数据到报率模拟（在同一个连接中完成） ===
        today = datetime.now().strftime('%Y-%m-%d')
        for site in sites:
            metrics_map = {
                'water_quality': 'codmn',
                'manual_station': 'ammonia',
                'drinking_source': 'ph',
                'cross_boundary': 'total_nitrogen',
                'groundwater': 'ammonia',
            }
            m = metrics_map.get(site['type'])
            if not m: continue
            is_miss = random.random() < 0.08
            existing = db.execute(
                "SELECT id, expected_count, actual_count FROM data_arrival WHERE site_id=? AND date=? AND metric=?",
                (site['id'], today, m)
            ).fetchone()
            if existing:
                exp = existing['expected_count'] + 1
                act = existing['actual_count'] + (0 if is_miss else 1)
                rate = round(act / exp * 100, 1)
                db.execute("UPDATE data_arrival SET expected_count=?, actual_count=?, arrival_rate=? WHERE id=?",
                           (exp, act, rate, existing['id']))
            else:
                db.execute("INSERT INTO data_arrival (site_id,date,metric,expected_count,actual_count,arrival_rate) VALUES (?,?,?,1,?,?)",
                           (site['id'], today, m, 0 if is_miss else 1, 100 if not is_miss else 0))

        # === 天气数据 ===
        # 启动时尝试获取实时天气（不插入模拟数据，让 API 请求时自动刷新）
        fetch_real_weather()
        
        # === 离线设备告警 ===
        offline_devices = db.execute("""
            SELECT d.site_id, d.device_name, d.device_code, s.name as site_name
            FROM device_shadows d JOIN sites s ON d.site_id=s.id
            WHERE d.status='offline'
        """).fetchall()
        for dev in offline_devices:
            # 该站点已存在未闭环离线工单则跳过，避免模拟器每 30s 反复触发告警/工单
            if db.execute(
                "SELECT 1 FROM work_orders WHERE site_id=? AND metric IN ('device_status','data_gap') AND status NOT IN ('closed','resolved') LIMIT 1",
                (dev['site_id'],)
            ).fetchone():
                continue
            create_alert_internal(db, dev['site_id'], 'device_status', 0, 'yellow',
                f"设备离线: {dev['device_name']} ({dev['device_code']}) · {dev['site_name']}")
    except Exception as e:
        print(f'[Sim] 数据生成异常: {e}')
    finally:
        if db:
            try:
                db.close()
            except:
                pass

def migrate_alerts_messages():
    """迁移旧告警消息中的英文指标名为中文"""
    METRIC_EN_CN = {
        'codmn':'高锰酸盐指数','ammonia':'氨氮','total_phosphorus':'总磷','total_nitrogen':'总氮',
        'water_temp':'水温','dissolved_oxygen':'溶解氧','ph':'pH','turbidity':'浊度',
        'conductivity':'电导率','temperature':'气温',
        'data_spike':'数据突变','data_freeze':'数据冻结','data_gap':'数据缺失',
        'device_status':'设备状态'
    }
    with get_db() as db:
        for en, cn in METRIC_EN_CN.items():
            db.execute("UPDATE alerts SET message=REPLACE(message,?,?) WHERE message LIKE ?",
                       (en, cn, '%'+en+'%'))
        db.commit()
        fixed = db.execute("SELECT changes()").fetchone()[0]
        if fixed:
            print(f"[Migrate] 已修正 {fixed} 条告警消息中的英文指标名")

def migrate_alert_flow():
    """迁移告警表：新增 flow_type / flow_status / tracking 字段"""
    with get_db() as db:
        for col_sql in [
            "ALTER TABLE alerts ADD COLUMN flow_type TEXT DEFAULT 'manual'",
            "ALTER TABLE alerts ADD COLUMN flow_status TEXT DEFAULT 'pending_review'",
            "ALTER TABLE alerts ADD COLUMN tracking_count INTEGER DEFAULT 0",
        ]:
            try:
                db.execute(col_sql)
            except Exception:
                pass
        # 所有未设 flow_type 的告警统一为 manual（create_alert_internal 自行管理 auto 类型）
        db.execute("UPDATE alerts SET flow_type='manual', flow_status='pending_review' WHERE flow_type IS NULL")
        # 已有 related_order_no 的设置 converted
        db.execute("UPDATE alerts SET flow_status='converted' WHERE related_order_no IS NOT NULL AND related_order_no != ''")
        # 修复：之前被错误自动转化的 data_gap/device_status 告警 → 重置为手动复核
        # （仅限没有关联工单的，有工单的保持已完结状态）
        db.execute("UPDATE alerts SET flow_type='manual', flow_status='pending_review', status='pending' WHERE metric IN ('data_gap','device_status') AND flow_type='auto' AND (related_order_no IS NULL OR related_order_no='')")
        db.commit()
        # 统计修复的告警数
        fixed = db.execute("SELECT COUNT(*) as c FROM alerts WHERE flow_type='manual' AND flow_status='pending_review' AND status='pending' AND metric IN ('data_gap','device_status')").fetchone()['c']
        if fixed:
            print(f"[Migrate] 已重置 {fixed} 条 device_status/data_gap 告警为手动复核模式")
        print("[Migrate] alert_flow 迁移完成: flow_type/flow_status 字段已添加并初始化")

def _auto_convert_alert(db, alert_id, site_id, alert_level, message, metric):
    """A级告警自动转工单（防重复：同站点离线类已有未闭环工单则跳过，避免列表重复）"""
    # 离线类告警（设备离线/数据缺失）视为同一现场事件，按站点去重
    OFFLINE_METRICS = ('device_status', 'data_gap')
    if metric in OFFLINE_METRICS:
        dup = db.execute(
            "SELECT order_no FROM work_orders WHERE site_id=? AND metric IN ('device_status','data_gap') AND status NOT IN ('closed','resolved') LIMIT 1",
            (site_id,)
        ).fetchone()
        if dup:
            # 已存在未闭环工单：仅关联告警，不再新建，消除重复工单
            db.execute("UPDATE alerts SET flow_status='converted', related_order_no=?, status='pending' WHERE id=?",
                       (dup['order_no'], alert_id))
            return
    now = datetime.now()
    order_no = f"WO-{now.strftime('%Y%m%d')}-{random.randint(100,999)}"
    order_level = 'critical' if alert_level == 'red' else ('urgent' if alert_level == 'orange' else 'normal')
    sla_hours = {'normal': 72, 'urgent': 24, 'critical': 2}.get(order_level, 72)
    sla_deadline = (now + timedelta(hours=sla_hours)).strftime('%Y-%m-%d %H:%M')

    # 自动派单：根据站点责任人（user_sites 唯一真相源），忽略前端任意指派
    site = db.execute("SELECT manager FROM sites WHERE id=?", (site_id,)).fetchone()
    assignee = _station_operator(site_id) if site else ''

    # description 与 title 均仅取单条干净消息（按 ' | ' 取首段），杜绝多段拼接导致的重复/乱码
    clean_desc = (message or '').split(' | ')[0] if message else ''
    db.execute("""
        INSERT INTO work_orders (order_no,site_id,source,event_type,metric,level,title,description,assignee,status,sla_deadline,related_alert_id)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        order_no, site_id, 'auto', '告警自动转工单', metric or '',
        order_level, f"[自动] {clean_desc}", clean_desc,
        assignee, 'pending', sla_deadline, alert_id
    ))
    # 更新告警状态：保持 pending 可见，标记已流转
    db.execute("UPDATE alerts SET flow_status='converted', related_order_no=?, status='pending' WHERE id=?",
               (order_no, alert_id))
    # 时间线
    db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
               ('alert', alert_id, 'auto_converted', '系统', f'自动转工单 {order_no}'))
    db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
               ('order', 0, 'auto_created', '系统', f'告警{alert_id}自动转工单-{order_no}'))

# =============================================================================
# 微信订阅消息推送（移动端）
# 设计原则：最佳努力（best-effort）—— 任何微信接口异常均静默捕获，绝不阻断
# 主业务流程（告警生成 / 工单审批）。未配置 AppSecret 或模板 ID 时直接跳过。
# 推送对象：按站点群发——查该站点已绑定 openid 的用户逐个下发（因 work_orders
# 的 assignee 存的是 real_name，不可靠，故统一按 user_sites 站点级投递，与告警一致）。
# =============================================================================
_WX_TOKEN_CACHE = {'token': '', 'expire_at': 0}

def _wx_get_access_token():
    """获取并缓存微信 access_token（默认 7200s 有效期，提前 60s 续期）"""
    global _WX_TOKEN_CACHE
    now = time.time()
    if _WX_TOKEN_CACHE['token'] and _WX_TOKEN_CACHE['expire_at'] > now + 60:
        return _WX_TOKEN_CACHE['token']
    if not WX_APPID or not WX_APPSECRET:
        return ''
    try:
        url = ("https://api.weixin.qq.com/cgi-bin/token"
                "?grant_type=client_credential&appid=%s&secret=%s" % (WX_APPID, WX_APPSECRET))
        req = urllib.request.Request(url, headers={'User-Agent': 'water-ops'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = _json.loads(resp.read().decode('utf-8'))
        if 'access_token' in data:
            _WX_TOKEN_CACHE['token'] = data['access_token']
            _WX_TOKEN_CACHE['expire_at'] = now + int(data.get('expires_in', 7200))
            return data['access_token']
        print('[WX] 获取 access_token 失败: %s' % data)
    except Exception as e:
        print('[WX] 获取 access_token 异常: %s' % e)
    return ''

def _wx_code2openid(code):
    """用 wx.login 拿到的 code 换取用户 openid"""
    if not WX_APPID or not WX_APPSECRET or not code:
        return ''
    try:
        url = ("https://api.weixin.qq.com/sns/jscode2session"
                "?appid=%s&secret=%s&js_code=%s&grant_type=authorization_code"
                % (WX_APPID, WX_APPSECRET, code))
        req = urllib.request.Request(url, headers={'User-Agent': 'water-ops'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = _json.loads(resp.read().decode('utf-8'))
        if 'openid' in data:
            return data['openid']
        print('[WX] code2openid 失败: %s' % data)
    except Exception as e:
        print('[WX] code2openid 异常: %s' % e)
    return ''

def _wx_push_subscribe(openid, template_id, data):
    """向单个 openid 下发订阅消息，成功返回 True；参数缺失或失败返回 False"""
    if not openid or not template_id:
        return False
    token = _wx_get_access_token()
    if not token:
        return False
    try:
        url = "https://api.weixin.qq.com/cgi-bin/message/subscribe/send?access_token=%s" % token
        payload = _json.dumps({
            'touser': openid,
            'template_id': template_id,
            'data': data,
        }).encode('utf-8')
        req = urllib.request.Request(url, data=payload,
                                    headers={'Content-Type': 'application/json'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            result = _json.loads(resp.read().decode('utf-8'))
        if result.get('errcode', 0) == 0:
            return True
        print('[WX] 订阅下发失败 openid=%s: %s' % (openid, result))
    except Exception as e:
        print('[WX] 订阅下发异常 openid=%s: %s' % (openid, e))
    return False

def _wx_push_to_site_users(site_id, template_id, data):
    """按站点群发：查该站点已绑定 openid 的用户，逐个下发订阅消息"""
    if not site_id or not template_id:
        return
    try:
        with get_db() as db:
            rows = db.execute(
                "SELECT DISTINCT u.openid FROM users u "
                "JOIN user_sites us ON us.user_id=u.id "
                "WHERE us.site_id=? AND u.openid IS NOT NULL AND u.openid!=''",
                (site_id,)
            ).fetchall()
        for r in rows:
            _wx_push_subscribe(r['openid'], template_id, data)
    except Exception as e:
        print('[WX] 站点群发异常 site_id=%s: %s' % (site_id, e))

def _wx_push_alert(site_id, metric, value, level, message):
    """新告警产生时，向该站点负责用户推送「告警信息」订阅消息

    推送字段结构与微信「告警信息」订阅模板关键词对应（见上方配置注释）：
        thing1=告警内容 / character_string1=站点编码 / phrase2=告警等级 / time3=告警时间
    """
    if not WX_TMPL_ALERT:
        return
    # 告警等级/指标中文（内联映射，避免依赖局部变量）
    level_cn = {'red': '红色', 'orange': '橙色', 'yellow': '黄色', 'blue': '蓝色'}.get(level, level or '')
    metric_cn = METRIC_CN.get(metric, metric or '')
    # 站点名称：优先用站点 name，fallback code / site_id
    site_name = ''
    try:
        with get_db() as db:
            row = db.execute("SELECT code, name FROM sites WHERE id=?", (site_id,)).fetchone()
        if row:
            site_name = (row['name'] or row['code'] or str(site_id))[:20]
    except Exception:
        site_name = str(site_id)[:20]
    # 告警内容：指标类(metric_cn) + 说明；特殊告警类型(data_spike/data_freeze/data_gap/device_status)的 message 已语义完整，直接取 message 避免重复
    SPECIAL_ALERT_METRICS = {'data_spike', 'data_freeze', 'data_gap', 'device_status'}
    if metric in SPECIAL_ALERT_METRICS and message:
        content = message[:20]
    elif message:
        content = (metric_cn + '：' + message)[:20]
    else:
        content = (metric_cn or '水质') + '异常'
    # 告警推送字段：按微信「监测预警信息通知」模板关键词
    #   thing1=预警接收人 / thing2=预警内容 / time3=预警时间 / thing7=监测点名称
    data = {
        'thing1': {'value': '站点负责人'},
        'thing2': {'value': content[:20]},
        'time3': {'value': datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')},
        'thing7': {'value': site_name},
    }
    _wx_push_to_site_users(site_id, WX_TMPL_ALERT, data)

def _wx_push_approve_result(site_id, order_no, result_cn):
    """工单审批结果（通过 / 退回 / 完成）时，向该站点负责用户推送「审批结果」订阅消息

    推送字段结构与微信「注册审核结果通知」模板关键词对应：
        name1=姓名 / phrase3=审核结果 / thing6=审核详情
    """
    if not WX_TMPL_APPROVE or not site_id:
        return
    data = {
        'name1': {'value': '运维人员'},
        'phrase3': {'value': (result_cn or '')[:5]},
        'thing6': {'value': (order_no or '')[:20]},
    }
    _wx_push_to_site_users(site_id, WX_TMPL_APPROVE, data)


def create_alert_internal(db, site_id, metric, value, level, message):
    """创建告警——同站点合并为一条告警（不同异常追加消息），去重同站点同metric"""
    LEVEL_PRIORITY = {'yellow':0, 'orange':1, 'red':2}
    # 判断流转类型（A级：data_gap/device_status → 自动转工单；B级：其他 → 人工复核）
    A_LEVEL_METRICS = {'data_gap', 'device_status'}
    is_auto = metric in A_LEVEL_METRICS
    flow_type = 'auto' if is_auto else 'manual'
    flow_status = 'pending' if is_auto else 'pending_review'

    # 强化去重：同site+同metric+同level+未办结 且120分钟内 → 更新tracking_count，不新建
    dedup_window = "datetime('now','localtime','-120 minutes')"
    existing = db.execute(
        f"SELECT id, tracking_count FROM alerts WHERE site_id=? AND metric=? AND level=? AND status IN ('pending','acknowledged') AND created_at > {dedup_window}",
        (site_id, metric, level if level else 'yellow')
    ).fetchone()
    if existing:
        new_count = (existing['tracking_count'] or 0) + 1
        db.execute(
            "UPDATE alerts SET tracking_count=?, message=?, created_at=datetime('now','localtime') WHERE id=?",
            (new_count, message, existing['id'])
        )
        db.commit()
        return existing['id']

    # 同站点同metric精确去重（120分钟窗口）——计数累加
    same = db.execute(
        f"SELECT id, tracking_count FROM alerts WHERE site_id=? AND metric=? AND status!='resolved' AND created_at > {dedup_window}",
        (site_id, metric)
    ).fetchone()
    if same:
        new_tracking = (same['tracking_count'] or 0) + 1
        db.execute("UPDATE alerts SET tracking_count=?, value=? WHERE id=?",
                   (new_tracking, value, same['id']))
        # 同一测项第3次触发 → 自动升级为A级
        if new_tracking >= 2:
            db.execute("UPDATE alerts SET flow_type='auto', flow_status='pending' WHERE id=?",
                       (same['id'],))
            alert_row = db.execute("SELECT * FROM alerts WHERE id=?", (same['id'],)).fetchone()
            if alert_row and alert_row['flow_type'] == 'auto' and alert_row['flow_status'] == 'pending':
                _auto_convert_alert(db, same['id'], site_id, alert_row['level'],
                                    alert_row['message'], metric)
        return same['id']

    # 检查同站点其他metric的未办结告警（120分钟内合并到同一条告警）
    existing = db.execute(
        f"SELECT id, message, level, flow_type, flow_status FROM alerts WHERE site_id=? AND status!='resolved' AND created_at > {dedup_window} ORDER BY id DESC LIMIT 1",
        (site_id,)
    ).fetchone()
    if existing:
        new_level = level
        if LEVEL_PRIORITY.get(existing['level'], 0) > LEVEL_PRIORITY.get(level, 0):
            new_level = existing['level']
        # 仅当新消息未包含于已有消息时追加，避免模拟器轮询反复拼接导致消息膨胀
        if message and message not in existing['message']:
            new_message = existing['message'] + ' | ' + message
        else:
            new_message = existing['message']

        # 合并后包含A级metric → 整体视为A级
        merged_flow_type = existing['flow_type']
        merged_flow_status = existing['flow_status']
        if is_auto or existing['flow_type'] == 'auto':
            merged_flow_type = 'auto'
            merged_flow_status = 'pending'

        db.execute(
            "UPDATE alerts SET message=?, level=?, value=?, flow_type=?, flow_status=? WHERE id=?",
            (new_message, new_level, value, merged_flow_type, merged_flow_status, existing['id'])
        )

        # 合并后为A级且原未转工单 → 立即自动转工单
        if merged_flow_type == 'auto' and existing['flow_status'] in ('pending', 'pending_review'):
            # 传递单条 message（非拼接 blob）；重复工单由 _auto_convert_alert 内站点去重拦截
            _auto_convert_alert(db, existing['id'], site_id, new_level, message, metric)
    else:
        db.execute(
            "INSERT INTO alerts (site_id,metric,value,level,message,flow_type,flow_status) VALUES (?,?,?,?,?,?,?)",
            (site_id, metric, value, level, message, flow_type, flow_status)
        )
        new_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        # 新创建的A级告警 → 立即自动转工单
        if is_auto:
            _auto_convert_alert(db, new_id, site_id, level, message, metric)
        # 推送订阅消息给站点负责用户（最佳努力，不阻断主流程）
        try:
            _wx_push_alert(site_id, metric, value, level, message)
        except Exception as e:
            print('[WX] 告警推送异常: %s' % e)
        return new_id

    return existing['id']

# ===================== 登录认证系统 =====================

# Token存储
_tokens = {}
# 站点范围缓存：token -> [site_id,...]，避免每个/api/请求都查 user_sites（消除并发读时的锁等待卡顿）
_site_ids_cache = {}

def _hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def seed_users():
    with get_db() as db:
        cnt = db.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        if cnt > 0:
            print("[Auth] 用户已存在，跳过")
            return
        users = [
            ('admin', 'admin123', 'admin', '系统管理员', '13800000000'),
            ('zhangsan', 'yw123456', 'operator', '张建国', '13800000001'),
            ('lisi', 'yw123456', 'operator', '黎明', '13800000002'),
            ('wangwu', 'yw123456', 'operator', '王刚', '13800000003'),
            ('zhaoliu', 'yw123456', 'operator', '赵洪', '13800000004'),
        ]
        for u in users:
            db.execute("INSERT INTO users (username,password_hash,role,real_name,phone) VALUES (?,?,?,?,?)",
                       (u[0], _hash_pw(u[1]), u[2], u[3], u[4]))
        print("[Auth] 5个用户已创建")
        all_ids = [r['id'] for r in db.execute("SELECT id FROM sites").fetchall()]
        for sid in all_ids:
            db.execute("INSERT OR IGNORE INTO user_sites (user_id,site_id) VALUES (?,?)", (1, sid))
        assignments = [(2, 1, 70), (3, 71, 140), (4, 141, 210), (5, 211, 267)]
        for uid, start_id, end_id in assignments:
            for sid in range(start_id, end_id + 1):
                if sid <= max(all_ids):
                    db.execute("INSERT OR IGNORE INTO user_sites (user_id,site_id) VALUES (?,?)", (uid, sid))
        db.commit()
        print("[Auth] 站点分配完成")


def seed_vehicles():
    """车辆演示数据"""
    with get_db() as db:
        cnt = db.execute("SELECT COUNT(*) FROM vehicles").fetchone()[0]
        if cnt > 0:
            print("[Vehicle] 车辆数据已存在，跳过")
            return
        vehicles = [
            ('赣A·X0001', 'SUV', 5, 38520, 45000),
            ('赣A·X0002', '皮卡', 5, 51200, 60000),
            ('赣A·X0003', '面包车', 7, 28300, 35000),
            ('赣A·X0004', '轿车', 5, 42100, 50000),
        ]
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for plate, model, seats, mileage, next_maint in vehicles:
            db.execute(
                "INSERT INTO vehicles (plate_no, model, seats, status, current_mileage, next_maintenance_mileage, last_maintenance_at) VALUES (?,?,?,'idle',?,?,?)",
                (plate, model, seats, mileage, next_maint, now))
        # 创建2个历史用车申请
        from datetime import timedelta as _td
        for i, (uid, dst, reason) in enumerate([
            (2, '星子水站', '周巡检'), (3, '蛤蟆石站', '紧急抢修'),
        ]):
            day = (datetime.now() - _td(days=i+1)).strftime('%Y-%m-%d')
            db.execute(
                "INSERT INTO vehicle_applications (vehicle_id, applicant_id, start_at, end_at, destination, reason, status) VALUES (?,?,?,?,?,?,'approved')",
                (i+1, uid, f'{day} 08:00:00', f'{day} 17:00:00', dst, reason))
        db.commit()
        print("[Vehicle] 4辆车 + 2条历史申请已创建")

def login_required(f):
    """认证中间件：从Authorization头中提取token，注入g.current_user和g.user_sites"""
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        auth = request.headers.get('Authorization', '')
        token = auth.replace('Bearer ', '').strip() if auth.startswith('Bearer ') else ''
        if not token or token not in _tokens:
            return jsonify({'error': '未登录或登录已过期', 'code': 'AUTH_REQUIRED'}), 401
        user = _tokens[token]
        g.current_user = user
        with get_db() as db:
            rows = db.execute("SELECT site_id FROM user_sites WHERE user_id=?", (user['id'],)).fetchall()
        g.user_site_ids = [r['site_id'] for r in rows]
        return f(*args, **kwargs)
    return wrapper


from telemetry import create_telemetry_blueprint
app.register_blueprint(create_telemetry_blueprint(get_db, login_required))
from operations_baseline import create_operations_baseline_blueprint
app.register_blueprint(create_operations_baseline_blueprint(get_db, login_required))

def _filter_site_ids():
    """返回当前用户可见的site_id列表（管理员或无站点绑定返回None=全部）"""
    site_ids = getattr(g, 'user_site_ids', None)
    if not site_ids:  # None 或空列表都返回 None（全部可见）
        return None
    return site_ids


def _station_operator(site_id):
    """返回站点责任运维人员姓名（user_sites 为唯一真相源）；无则空串。
    管理员永不作为 assignee；忽略任意前端传入的指派。"""
    with get_db() as db:
        row = db.execute(
            "SELECT u.real_name FROM user_sites us JOIN users u ON u.id=us.user_id "
            "WHERE us.site_id=? AND u.role='operator' LIMIT 1", (site_id,)).fetchone()
        return row['real_name'] if row else ''


# ===================== 全局 API 鉴权门禁 =====================
@app.before_request
def global_api_auth():
    """全局 API 鉴权：除白名单外，所有 /api/ 请求必须携带有效 Bearer token。
    一次性封堵所有"未加 @login_required"的匿名端点；同时注入 g.current_user / g.user_site_ids，
    供各路由及 _filter_site_ids 复用（含此前无鉴权的列表端点与待办审核）。"""
    # 非 API 路径（前端页面、静态资源）放行
    if not request.path.startswith('/api/'):
        return
    # 预检请求放行（CORS）
    if request.method == 'OPTIONS':
        return
    # 白名单：登录、健康检查
    if request.path in ('/api/auth/login', '/api/health'):
        return
    # 校验 token 并注入用户上下文
    auth = request.headers.get('Authorization', '')
    token = auth.replace('Bearer ', '').strip() if auth.startswith('Bearer ') else ''
    if not token or token not in _tokens:
        return jsonify({'error': '未登录或登录已过期', 'code': 'AUTH_REQUIRED'}), 401
    user = _tokens[token]
    g.current_user = user
    # 站点范围按 token 缓存，避免每个请求都查库（并发读时撞 SQLite 写锁会卡顿）
    cached = _site_ids_cache.get(token)
    if cached is None:
        with get_db() as db:
            rows = db.execute("SELECT site_id FROM user_sites WHERE user_id=?", (user['id'],)).fetchall()
        cached = [r['site_id'] for r in rows]
        _site_ids_cache[token] = cached
    g.user_site_ids = cached
    return None


@app.before_request
def retire_legacy_inspection_plan_writes():
    """旧 inspection-v2 仅保留历史读取/审核兼容，禁止继续生成第二条计划主链。"""
    if request.method in ('POST', 'PUT', 'PATCH', 'DELETE') and request.path.startswith('/api/inspection-v2/plans'):
        return jsonify({
            'error': '旧版巡检计划链路已停用，请使用“计划调度”创建和审批巡检任务',
            'code': 'LEGACY_PLAN_RETIRED',
        }), 410
    if request.method in ('POST', 'PUT', 'PATCH', 'DELETE') and request.path.startswith('/api/inspection-v2/schedules'):
        return jsonify({
            'error': '旧版巡检排程已停用，请使用“计划调度”',
            'code': 'LEGACY_PLAN_RETIRED',
        }), 410

# ===================== 通知系统辅助函数 =====================

def _create_notification(user_id, source_type, source_id, title, content='', db=None):
    """创建通知（内部函数）。db 可传入以复用现有连接（避免嵌套写锁）"""
    own = db is None
    if own:
        db = get_db()
    try:
        db.execute(
            "INSERT INTO notifications (user_id, source_type, source_id, title, content) VALUES (?,?,?,?,?)",
            (user_id, source_type, source_id, title, content)
        )
        if own:
            db.commit()
    except Exception:
        if own:
            try: db.rollback()
            except Exception: pass
        raise


# ===================== 闭环联动辅助函数 =====================
# 统一的"处置结论"枚举（前后端共用同一套 key，中文标签见前端 constants.js）
REVIEW_CONCLUSION_FALSE = {'false_alarm', 'normal_deviation', 'environmental_factor', 'equipment_maintenance', 'other', 'fixed'}


def _link_review_alert(db, site_id, metric):
    """把同一 (site_id, metric) 下尚未关联的告警与数据审核项互相绑定。"""
    rev = db.execute(
        "SELECT id FROM data_reviews WHERE site_id=? AND metric=? AND alert_id IS NULL ORDER BY recorded_at DESC LIMIT 1",
        (site_id, metric)).fetchone()
    al = db.execute(
        "SELECT id FROM alerts WHERE site_id=? AND metric=? AND review_id IS NULL AND status NOT IN ('resolved','closed') ORDER BY created_at DESC LIMIT 1",
        (site_id, metric)).fetchone()
    if rev and al:
        db.execute("UPDATE data_reviews SET alert_id=? WHERE id=?", (al['id'], rev['id']))
        db.execute("UPDATE alerts SET review_id=? WHERE id=?", (rev['id'], al['id']))


def _resolve_linked_alert(db, site_id, metric, conclusion, review_id=None, order_no=None):
    """找到关联/同 (site, metric) 的未办结告警并办结，reason=conclusion。"""
    if review_id:
        al = db.execute(
            "SELECT id FROM alerts WHERE review_id=? AND status NOT IN ('resolved','closed')",
            (review_id,)).fetchone()
    else:
        al = db.execute(
            "SELECT id FROM alerts WHERE site_id=? AND metric=? AND status NOT IN ('resolved','closed') ORDER BY created_at DESC LIMIT 1",
            (site_id, metric)).fetchone()
    if not al:
        return
    aid = al['id']
    db.execute(
        "UPDATE alerts SET status='resolved', resolved_at=datetime('now','localtime'), resolve_reason=? WHERE id=?",
        (conclusion, aid))
    db.execute(
        "INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES ('alert',?,'resolved','系统',?)",
        (aid, f'关联数据审核判定为{conclusion}，告警自动办结'))


def _archive_linked_review(db, site_id, metric, conclusion, order_no=None, review_id=None):
    """找到关联/同 (site, metric) 的未归档数据审核项并归档。"""
    if review_id:
        rv = db.execute(
            "SELECT id FROM data_reviews WHERE id=? AND status!='archived'",
            (review_id,)).fetchone()
    else:
        rv = db.execute(
            "SELECT id FROM data_reviews WHERE site_id=? AND metric=? AND status!='archived' ORDER BY recorded_at DESC LIMIT 1",
            (site_id, metric)).fetchone()
    if not rv:
        return
    rid = rv['id']
    upd = "UPDATE data_reviews SET manual_result='rejected', manual_reason=?, status='archived', reviewed_at=datetime('now','localtime')"
    params = [conclusion]
    if order_no:
        upd += ", resolved_by_order_id=?"
        params.append(order_no)
    upd += " WHERE id=?"
    params.append(rid)
    db.execute(upd, params)


def _notify_review_l3(db, review_row):
    """数据审核进入 L3 人工复核时，推送通知给审核员/管理员。"""
    try:
        site_name = (db.execute("SELECT name FROM sites WHERE id=?", (review_row['site_id'],)).fetchone() or {}).get('name', '')
        lvl = db.execute(
            "SELECT level FROM alerts WHERE site_id=? AND metric=? AND status NOT IN ('resolved','closed') ORDER BY created_at DESC LIMIT 1",
            (review_row['site_id'], review_row['metric'])).fetchone()
        level_cn = {'blue': '一般关注', 'yellow': '一般告警', 'orange': '较重告警', 'red': '紧急告警'}.get(lvl['level'], '') if lvl else ''
        title = f'数据审核待人工复核（L3）· {site_name}'
        content = f"{review_row['metric']} 疑似异常{(' · ' + level_cn) if level_cn else ''}"
        for u in db.execute("SELECT id FROM users WHERE role IN ('admin','reviewer','inspector')").fetchall():
            _create_notification(u['id'], 'data_review', review_row['id'], title, content, db=db)
    except Exception as e:
        print(f'[L3 notify] 失败: {e}')


def _notify_inspection_plan(plan_id, plan_name, site_id, event):
    """巡检计划事件通知相关站点负责人"""
    with get_db() as db:
        # 查站点负责人
        site = db.execute("SELECT name, manager FROM sites WHERE id=?", (site_id,)).fetchone()
        if not site: return
        manager = site['manager'] or ''
        if not manager: return
        # 根据负责人姓名找到对应用户
        user = db.execute("SELECT id FROM users WHERE real_name=? AND role='operator'", (manager,)).fetchone()
        if user:
            title = f'巡检计划{event}' if event != 'completed' else '巡检计划已完成'
            content = f'{site["name"]}-{plan_name}'
            _create_notification(user['id'], 'inspection', plan_id, title, content)
        # 管理员也收到通知
        admin = db.execute("SELECT id FROM users WHERE role='admin' LIMIT 1").fetchone()
        if admin:
            _create_notification(admin['id'], 'inspection', plan_id, f'巡检计划{event}', f'{site["name"]}-{plan_name}')


# ===================== API Routes =====================

@app.route('/api/health')
def health():
    return jsonify({'status': 'ok', 'time': datetime.now().isoformat()})

# --- SL651状态查询 ---
@app.route('/api/sl651/status')
def sl651_status():
    """查看SL651接收器状态（运行中/连接数/映射配置）"""
    try:
        from sl651_server import get_mapping, _connections
        mapping = get_mapping()
        return jsonify({
            'enabled': True,
            'port': 5005,
            'connections': len(_connections),
            'connections_list': list(_connections),
            'debug_mode': mapping.get('debug_mode', False),
            'mapped_metrics': list(mapping.get('metric_mapping', {}).keys()),
            'station_overrides': list(mapping.get('per_station', {}).keys()),
        })
    except ImportError:
        return jsonify({'enabled': False, 'port': 5005})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sl651/reload', methods=['POST'])
def sl651_reload():
    """热重载映射配置"""
    try:
        from sl651_server import reload_mapping
        reload_mapping()
        return jsonify({'success': True, 'message': '映射配置已热重载'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# --- Sites ---
@app.route('/api/sites')
def get_sites_simple():
    allowed = _filter_site_ids()  # 管理员(无绑定)/None → 全部；操作员 → 仅本人站点
    site_clause = ''
    params = []
    if allowed is not None:
        site_clause = f' AND s.id IN ({",".join("?" * len(allowed))})'
        params = list(allowed)
    with get_db() as db:
        rows = db.execute(f"""
            SELECT s.id, s.code, s.name, s.type, s.gps_lat as lat, s.gps_lng as lng, s.district, s.address, s.river,
                   s.manager, s.phone, s.last_heartbeat, s.created_at,
                   COUNT(d.id) as device_count,
                   SUM(CASE WHEN d.status='offline' THEN 1 ELSE 0 END) as offline_count,
                   CASE WHEN SUM(CASE WHEN d.status='offline' THEN 1 ELSE 0 END) > 0 THEN 'offline' ELSE 'online' END as status
            FROM sites s LEFT JOIN device_shadows d ON s.id=d.site_id
            WHERE s.id >= 0{site_clause}
            GROUP BY s.id ORDER BY s.id
        """, params).fetchall()
        result = []
        for r in rows:
            rd = dict(r)
            result.append({'id': rd['id'], 'name': rd['name'], 'lat': rd['lat'], 'lng': rd['lng'],
                          'code': rd['code'], 'type': rd['type'], 'device_count': rd['device_count'],
                          'status': rd['status'],
                          'district': rd.get('district') or '',
                          'address': rd.get('address') or '',
                          'manager': rd.get('manager') or ''})
        # 批量补齐每个站点最新监测值（驾驶舱卡片展示阈值色阶用）
        site_ids = [d['id'] for d in result]
        if site_ids:
            placeholders = ','.join('?' * len(site_ids))
            latest_rows = db.execute(f"""
                SELECT s.site_id, s.metric AS latest_metric, s.value AS latest_value,
                       s.unit AS latest_unit, s.recorded_at AS latest_time
                FROM (SELECT site_id, MAX(id) AS max_id FROM sensor_data
                      WHERE site_id IN ({placeholders}) GROUP BY site_id) m
                JOIN sensor_data s ON s.id = m.max_id
            """, site_ids).fetchall()
            latest_by_site = {row['site_id']: dict(row) for row in latest_rows}
            for d in result:
                lv = latest_by_site.get(d['id'], {})
                d['latest_metric'] = lv.get('latest_metric') or ''
                d['latest_value'] = round(lv['latest_value'], 2) if lv.get('latest_value') is not None else None
                d['latest_unit'] = lv.get('latest_unit') or ''
                d['latest_time'] = lv.get('latest_time') or ''
        return jsonify(result)

@app.route('/api/thresholds')
@login_required
def get_thresholds():
    """返回全部水质指标阈值配置（色阶数据源）：metric/label/unit/low/high/critical_low/critical_high。"""
    with get_db() as db:
        rows = db.execute("SELECT metric, label, unit, low, high, critical_low, critical_high FROM param_thresholds ORDER BY metric").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/v2/sites')
@login_required
def get_sites_v2():
    return jsonify({'test': 'hello', 'sites': []})

@app.route('/api/sites/<int:site_id>')
def get_site(site_id):
    allowed = _filter_site_ids()
    if allowed is not None and site_id not in allowed:
        return jsonify({'error': '无权限访问该站点'}), 403
    with get_db() as db:
        site = db.execute("SELECT * FROM sites WHERE id=?", (site_id,)).fetchone()
        if not site:
            return jsonify({'error': 'not found'}), 404
        devices = db.execute("SELECT * FROM device_shadows WHERE site_id=?", (site_id,)).fetchall()
        alerts_count = db.execute("SELECT COUNT(*) as c FROM alerts WHERE site_id=? AND status='pending'", (site_id,)).fetchone()['c']
        orders_count = db.execute("SELECT COUNT(*) as c FROM work_orders WHERE site_id=? AND status NOT IN ('closed')", (site_id,)).fetchone()['c']
        site_dict = dict(site)
        # Calculate status from devices, not from sites.status
        offline_devices = [d for d in devices if d['status'] == 'offline']
        site_dict['status'] = 'offline' if len(offline_devices) > 0 else 'online'
        site_dict['devices'] = [dict(d) for d in devices]
        site_dict['active_alerts'] = alerts_count
        site_dict['open_orders'] = orders_count
        # 水库额外信息
        reservoir_extra = {
            1: {'capacity': 1280, 'flood_level': 49.5, 'critical_level': 51.5, 'normal_level': 48.0},
            2: {'capacity': 860, 'flood_level': 48.0, 'critical_level': 50.0, 'normal_level': 47.0},
        }
        if site['type'] == 'reservoir':
            extra = reservoir_extra.get(site['id'], {})
            site_dict['capacity'] = extra.get('capacity')
            site_dict['flood_level'] = extra.get('flood_level')
            site_dict['critical_level'] = extra.get('critical_level')
            site_dict['normal_level'] = extra.get('normal_level')
        return jsonify(site_dict)

@app.route('/api/sites/archive/upload-calibration', methods=['POST'])
def upload_site_calibration():
    """上传站点校准附件（照片/PDF/文档），统一归口 operation_attachments（source_type='calibration'）。"""
    file = request.files.get('file')
    site_id = request.form.get('site_id', type=int)
    allowed = _filter_site_ids()
    if allowed is not None and (not site_id or site_id not in allowed):
        return jsonify({'error': '无权限操作非本人站点档案'}), 403
    if not file or not site_id:
        return jsonify({'error': '请选择文件并指定站点'}), 400
    ext = os.path.splitext(file.filename or '.jpg')[1].lower() or '.jpg'
    img_exts = {'.png', '.jpg', '.jpeg', '.gif', '.webp', '.bmp'}
    if ext not in img_exts:
        return jsonify({'error': '仅支持图片格式'}), 400
    file_data = file.read()
    if len(file_data) > 20 * 1024 * 1024:
        return jsonify({'error': '文件大小超过20MB限制'}), 400
    import json as _json
    cal_type = request.form.get('cal_type', '校准照片')
    result = request.form.get('result', '合格')
    valid_until = request.form.get('valid_until', '')
    fname = str(uuid.uuid4())[:12] + ext
    now = datetime.now()
    subdir = now.strftime('calibration/%Y/%m')
    stored_dir = os.path.join(UPLOAD_DIR, subdir)
    os.makedirs(stored_dir, exist_ok=True)
    stored_path = os.path.join(stored_dir, fname)
    with open(stored_path, 'wb') as f:
        f.write(file_data)
    url = f'/uploads/{subdir}/{fname}'
    extra = _json.dumps({'cal_type': cal_type, 'result': result, 'valid_until': valid_until}, ensure_ascii=False)
    # 智能识别：按文件名/说明关键词匹配照片类型配置（水印/场景自动归类，接入审核链）
    _match = match_photo_requirement('', site_id, file.filename or '', f'{cal_type}（{result}）')
    _req_id = _match['requirement_id'] if _match else None
    _rec_cat = _match['recognized_category'] if _match else ''
    _match_status = _match['match_status'] if _match else 'manual'
    _match_conf = _match['match_confidence'] if _match else None
    _review_required = _match['review_required'] if _match else 0
    with get_db() as db:
        cur = db.execute("""INSERT INTO operation_attachments
            (filename, stored_path, file_type, mime_type, file_size, description,
             source_type, source_id, site_id, category, extra_json, created_at,
             watermark_text, recognized_category, match_status, match_confidence,
             review_required, requirement_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (file.filename or fname, url, 'file', ext, len(file_data),
             f'{cal_type}（{result}）', 'calibration', site_id, site_id, '校准照片', extra,
             now.strftime('%Y-%m-%d %H:%M:%S'),
             '', _rec_cat, _match_status, _match_conf,
             _review_required, _req_id))
        new_id = cur.lastrowid
        db.commit()
    return jsonify({'id': new_id, 'url': url, 'success': True})

@app.route('/api/sites/<int:site_id>/archive')
@login_required
def get_site_archive(site_id):
    """站点档案：聚合基本信息、设备、故障记录、巡检记录等"""
    try:
        with get_db() as db:
            site = db.execute("SELECT * FROM sites WHERE id=?", (site_id,)).fetchone()
            if not site:
                return jsonify({'error': '站点不存在'}), 404

            site_dict = dict(site)
            site_type = site_dict.get('type', '')
    
            # 设备列表（从 device_shadows 取，与 device 表同步）
            devices = db.execute("SELECT * FROM device_shadows WHERE site_id=?", (site_id,)).fetchall()
            site_dict['equipment'] = [dict(d) for d in devices]
    
            # 故障记录 → 从 alerts 表取真实告警（时序上最近50条）
            fault_alerts = db.execute(
                "SELECT id, message, level, status, created_at FROM alerts WHERE site_id=? ORDER BY created_at DESC LIMIT 50",
                (site_id,)
            ).fetchall()
            fault_records = [
                {
                    'id': f['id'],
                    'date': f['created_at'],
                    'title': f['message'][:60],
                    'event': f['level'],
                    'description': f['message'],
                    'detail': f['message'],
                    'severity': f['level'],
                    'operator': '系统',
                }
                for f in [dict(r) for r in fault_alerts]
            ]
            site_dict['fault_records'] = fault_records
    
            # 设备更换/维护记录 → 从 device_shadows 的 install_date 推断 + 告警转换
            dev_replacements = db.execute(
                "SELECT id, device_name, device_type, install_date, status FROM device_shadows WHERE site_id=? ORDER BY install_date DESC",
                (site_id,)
            ).fetchall()
            replacement_records = [
                {
                    'id': r['id'],
                    'date': r['install_date'] or '—',
                    'old_equipment': '—',
                    'new_equipment': r['device_name'],
                    'reason': f'初始安装/{r.get("status","运行")}',
                    'operator': '—',
                }
                for r in [dict(r) for r in dev_replacements]
            ]
            site_dict['replacement_records'] = replacement_records
    
            # 巡检记录 → 从 insp_plans + insp_plan_items 取
            insp_plans = db.execute(
                "SELECT DISTINCT p.id, p.plan_name, p.period, p.status, p.created_at, p.generate_date "
                "FROM insp_plans p "
                "JOIN insp_plan_items i ON i.plan_id = p.id "
                "WHERE i.site_id = ? "
                "ORDER BY p.created_at DESC LIMIT 50",
                (site_id,)
            ).fetchall()
            site_dict['inspection_records'] = [
                {
                    'id': p['id'],
                    'date': p.get('generate_date') or p.get('created_at', ''),
                    'type': {'weekly':'每周巡检','monthly':'每月巡检'}.get(p.get('period',''), p.get('period','—')),
                    'result': {'draft':'草稿','active':'执行中','completed':'已完成'}.get(p.get('status',''), p.get('status','—')),
                    'issues': p.get('plan_name', '') or '—',
                    'inspector': '—',
                }
                for p in [dict(r) for r in insp_plans]
            ]
    
            # 校准报告：优先真实上传（operation_attachments, source_type='calibration'），无则 mock 兜底
            import json as _json
            real_cals = db.execute(
                "SELECT id, filename, stored_path, extra_json, created_at FROM operation_attachments "
                "WHERE source_type='calibration' AND source_id=? AND is_deleted=0 ORDER BY created_at DESC",
                (site_id,)
            ).fetchall()
            if real_cals:
                site_dict['calibration_reports'] = [
                    {
                        'id': c['id'],
                        'date': (c['created_at'] or '')[:10],
                        'type': (_json.loads(c['extra_json']).get('cal_type', '校准照片') if c['extra_json'] else '校准照片'),
                        'result': (_json.loads(c['extra_json']).get('result', '') if c['extra_json'] else ''),
                        'valid_until': (_json.loads(c['extra_json']).get('valid_until', '') if c['extra_json'] else ''),
                        'file': {'url': c['stored_path'], 'name': c['filename'] or (c['stored_path'].split('/')[-1])},
                    }
                    for c in [dict(r) for r in real_cals]
                ]
            else:
                site_dict['calibration_reports'] = _generate_mock_calibration_reports(
                    site_type, site_dict.get('name', ''), site_dict['equipment']
                )
    
            # 试剂使用及更换记录
            reagent_records = db.execute(
                "SELECT * FROM reagent_records WHERE site_id=? ORDER BY created_at DESC",
                (site_id,)
            ).fetchall()
            site_dict['reagent_records'] = [
                {
                    'id': r['id'],
                    'reagent_name': r['reagent_name'],
                    'reagent_type': r['reagent_type'] or '—',
                    'usage_date': r['usage_date'] or '—',
                    'replacement_date': r['replacement_date'] or '—',
                    'operator': r['operator'] or '—',
                    'notes': r['notes'] or '—',
                }
                for r in reagent_records
            ]
    
            # 历史记录
            history = []
            for f in site_dict['fault_records']:
                history.append({'date': f['date'], 'event': f['title'], 'operator': f['operator']})
            for r in site_dict['replacement_records']:
                history.append({'date': r['date'], 'event': "设备更换: " + r['old_equipment'], 'operator': r['operator']})
            history.sort(key=lambda x: x['date'] or '', reverse=True)
            site_dict['history_records'] = history[:20]
    
            return jsonify(site_dict)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def _generate_mock_fault_records(site_type, site_name):
    """生成模拟故障记录"""
    from datetime import datetime, timedelta
    now = datetime.now()
    templates = {
        'water_quality': [
            {'title': 'pH计读数异常', 'severity': 'high', 'desc': 'pH计读数持续偏低，经检查为电极老化，已更换新电极'},
            {'title': '氨氮分析仪试剂不足', 'severity': 'medium', 'desc': '氨氮分析仪试剂余量低于20%，已补充试剂并重新校准'},
            {'title': '采水泵运行异常', 'severity': 'medium', 'desc': '采水泵噪音增大，流量下降，已清理过滤器并润滑轴承'},
            {'title': '通信模块信号弱', 'severity': 'low', 'desc': '4G通信模块信号强度低于阈值，已调整天线位置'},
        ],
    }
    records = templates.get(site_type, templates['water_quality'])
    result = []
    for i, t in enumerate(records):
        days_ago = (i + 1) * 15 + i * 7
        dt = now - timedelta(days=days_ago)
        result.append({
            'id': f'mock_fault_{i+1}',
            'date': dt.strftime('%Y-%m-%d'),
            'title': t['title'],
            'event': t['title'],
            'description': t['desc'],
            'detail': t['desc'],
            'severity': t['severity'],
            'operator': '张工' if i % 2 == 0 else '李工',
        })
    return result


def _generate_mock_replacement_records(site_type, site_name):
    """生成模拟设备更换记录"""
    from datetime import datetime, timedelta
    now = datetime.now()
    templates = {
        'water_quality': [
            {'old': 'PHG-2088 pH计（2020款）', 'new': 'PHG-2099 pH计（2024款）', 'reason': '电极老化，测量精度下降'},
            {'old': 'NH3N-2000 氨氮分析仪', 'new': 'NH3N-3000 氨氮分析仪', 'reason': '设备升级，提高测量灵敏度'},
        ],
    }
    records = templates.get(site_type, templates['water_quality'])
    result = []
    for i, t in enumerate(records):
        days_ago = (i + 1) * 120 + i * 60
        dt = now - timedelta(days=days_ago)
        result.append({
            'id': f'mock_replace_{i+1}',
            'date': dt.strftime('%Y-%m-%d'),
            'old_equipment': t['old'],
            'new_equipment': t['new'],
            'reason': t['reason'],
            'operator': '王工' if i % 2 == 0 else '赵工',
        })
    return result


def _generate_mock_calibration_reports(site_type, site_name, equipment):
    """生成模拟校准报告，匹配站点类型和设备类型"""
    from datetime import datetime, timedelta
    now = datetime.now()

    # 设备类型→校准类型映射
    device_calibration_map = {
        'ph_meter': 'pH计校准',
        'do_sensor': '溶解氧传感器校准',
        'turbidity_meter': '浊度计校准',
        'ammonia_analyzer': '氨氮分析仪校准',
        'codmn_analyzer': '高锰酸盐指数分析仪校准',
        'tp_analyzer': '总磷分析仪校准',
        'tn_analyzer': '总氮分析仪校准',
        'conductivity_meter': '电导率仪校准',
        'multi_param_analyzer': '多参数分析仪校准',
        'submersible_pump': '潜水泵检查',
        'sample_float': '采样浮筒检查',
        'dtu': '数据采集传输仪校验',
    }

    # 校准结果模板
    calibration_results = [
        {
            'date': (now - timedelta(days=30)).strftime('%Y-%m-%d'),
            'type': '年度校准',
            'result': '合格，各项指标均在允许误差范围内',
            'valid_until': (now + timedelta(days=335)).strftime('%Y-%m-%d'),
            'file': {'url': '#', 'name': f'{site_name}_年度校准报告_{now.year}.pdf'},
        },
        {
            'date': (now - timedelta(days=180)).strftime('%Y-%m-%d'),
            'type': '半年度核查',
            'result': '合格，传感器精度满足规范要求',
            'valid_until': (now + timedelta(days=155)).strftime('%Y-%m-%d'),
            'file': {'url': '#', 'name': f'{site_name}_半年度核查报告_{now.year}.pdf'},
        },
    ]

    # 为每个设备生成专项校准记录
    for eq in (equipment or [])[:3]:
        eq_type = eq.get('device_type', '')
        cal_type = device_calibration_map.get(eq_type, '设备校准')
        days_ago = 60 + len(calibration_results) * 30
        calibration_results.append({
            'date': (now - timedelta(days=days_ago)).strftime('%Y-%m-%d'),
            'type': cal_type,
            'result': '合格',
            'valid_until': (now + timedelta(days=365 - days_ago)).strftime('%Y-%m-%d'),
            'file': {'url': '#', 'name': f'{eq.get("device_name", eq_type)}_校准证书.pdf'},
        })

    return calibration_results


def _generate_mock_recycle_records():
    """生成模拟设备回收记录"""
    from datetime import datetime, timedelta
    now = datetime.now()

    # 模拟回收记录模板
    templates = [
        {
            'device_code': 'DEV-2019-001',
            'device_name': 'SL3-1 翻斗雨量计',
            'device_type': 'rainfall_gauge',
            'site_name': '江桥水文站',
            'reason': '设备老化，精度下降',
            'destination': 'scrap',
            'operator': '张工',
            'days_ago': 180,
        },
        {
            'device_code': 'DEV-2020-015',
            'device_name': 'UHZ-40 水位计',
            'device_type': 'water_level_meter',
            'site_name': '南矶山水位站',
            'reason': '升级为雷达水位计',
            'destination': 'replace',
            'operator': '李工',
            'days_ago': 120,
        },
        {
            'device_code': 'DEV-2021-008',
            'device_name': 'HC-600 数据采集器',
            'device_type': 'hydro_collector',
            'site_name': '泉岭雨量站',
            'reason': '通信模块故障',
            'destination': 'repair',
            'operator': '王工',
            'days_ago': 90,
        },
        {
            'device_code': 'DEV-2020-023',
            'device_name': 'PWL-200 压力式水位计',
            'device_type': 'pressure_water_level',
            'site_name': '廖南墒情站',
            'reason': '传感器漂移严重',
            'destination': 'scrap',
            'operator': '赵工',
            'days_ago': 60,
        },
        {
            'device_code': 'DEV-2022-005',
            'device_name': 'LS25-3A 流速仪',
            'device_type': 'current_meter',
            'site_name': '邓埠水文站',
            'reason': '定期维护更换',
            'destination': 'return',
            'operator': '张工',
            'days_ago': 45,
        },
        {
            'device_code': 'DEV-2021-012',
            'device_name': 'RG-50 雨量计',
            'device_type': 'rainfall_gauge',
            'site_name': '聂城水文站',
            'reason': '翻斗卡滞无法修复',
            'destination': 'scrap',
            'operator': '李工',
            'days_ago': 30,
        },
        {
            'device_code': 'DEV-2023-002',
            'device_name': 'VL-30 雷达水位计',
            'device_type': 'radar_water_level',
            'site_name': '新祺周水文站',
            'reason': '天线损坏',
            'destination': 'repair',
            'operator': '王工',
            'days_ago': 15,
        },
    ]

    records = []
    for i, t in enumerate(templates):
        dt = now - timedelta(days=t['days_ago'])
        records.append({
            'id': i + 1,
            'device_id': i + 1,
            'device_code': t['device_code'],
            'device_name': t['device_name'],
            'device_type': t['device_type'],
            'site_id': i + 1,
            'site_name': t['site_name'],
            'recycle_date': dt.strftime('%Y-%m-%d'),
            'reason': t['reason'],
            'destination': t['destination'],
            'operator': t['operator'],
            'remark': '',
            'status': 'completed',
            'created_at': dt.strftime('%Y-%m-%d %H:%M:%S'),
            'work_order_no': f'WO-2026-{1000 + i}',
        })

    return records

@app.route('/api/site/status/<int:site_id>')
@login_required
def site_status(site_id):
    """统一站点状态查询：聚合站点信息+告警+数据健康+最新数据"""
    with get_db() as db:
        site = db.execute("SELECT * FROM sites WHERE id=?", (site_id,)).fetchone()
        if not site:
            return jsonify({'error': 'not found'}), 404
        site_dict = dict(site)
        devices = db.execute("SELECT * FROM device_shadows WHERE site_id=?", (site_id,)).fetchall()
        offline_devices = [d for d in devices if d['status'] == 'offline']
        site_dict['devices'] = [dict(d) for d in devices]
        site_dict['status'] = 'offline' if offline_devices else 'online'
        pending_alerts = db.execute(
            "SELECT * FROM alerts WHERE site_id=? AND status!='resolved' ORDER BY created_at DESC", (site_id,)
        ).fetchall()
        site_dict['active_alerts'] = [dict(a) for a in pending_alerts]
        site_dict['alert_count'] = len(pending_alerts)
        orders_count = db.execute(
            "SELECT COUNT(*) as c FROM work_orders WHERE site_id=? AND status NOT IN ('closed')", (site_id,)
        ).fetchone()['c']
        site_dict['open_orders'] = orders_count
        latest_row = db.execute(
            "SELECT metric, value, unit, recorded_at FROM sensor_data WHERE site_id=? ORDER BY id DESC LIMIT 1", (site_id,)
        ).fetchone()
        if latest_row:
            site_dict['latest_metric'] = latest_row['metric']
            site_dict['latest_value'] = round(latest_row['value'], 2)
            site_dict['latest_unit'] = latest_row['unit']
            site_dict['latest_time'] = latest_row['recorded_at']
        else:
            site_dict['latest_metric'] = ''
            site_dict['latest_value'] = None
            site_dict['latest_unit'] = ''
            site_dict['latest_time'] = ''
        wl_row = db.execute(
            "SELECT value, recorded_at FROM sensor_data WHERE site_id=? AND metric='water_level' ORDER BY id DESC LIMIT 1", (site_id,)
        ).fetchone()
        if wl_row:
            site_dict['wl_value'] = round(wl_row['value'], 2)
            site_dict['wl_time'] = wl_row['recorded_at']
        has_alert = len(pending_alerts) > 0
        lv = site_dict.get('latest_value')
        lm = site_dict.get('latest_metric')
        if has_alert:
            site_dict['data_health'] = 'alert'
            site_dict['data_health_reason'] = '有未办结告警'
        elif site_dict['status'] == 'offline':
            site_dict['data_health'] = 'abnormal'
            site_dict['data_health_reason'] = '设备离线'
        elif lv is None and lm:
            site_dict['data_health'] = 'abnormal'
            site_dict['data_health_reason'] = '数据缺失'
        elif lm and (lv > 1000 or lv < 0):
            site_dict['data_health'] = 'abnormal'
            site_dict['data_health_reason'] = '数据异常'
        else:
            site_dict['data_health'] = 'normal'
            site_dict['data_health_reason'] = ''
        # 传感器数据时间维度健康度检查
        try:
            last_sensor = db.execute(
                "SELECT MAX(recorded_at) FROM sensor_data WHERE site_id=?", (site_id,)
            ).fetchone()[0]
            if last_sensor:
                from datetime import datetime as _dt
                last_time = _dt.strptime(last_sensor, '%Y-%m-%d %H:%M:%S')
                hours_ago = (_dt.now() - last_time).total_seconds() / 3600
                if hours_ago > 24:
                    site_dict['sensor_health'] = 'stale'
                    site_dict['sensor_health_reason'] = f'传感器数据已{int(hours_ago)}小时未更新'
                elif hours_ago > 2:
                    site_dict['sensor_health'] = 'delayed'
                    site_dict['sensor_health_reason'] = f'传感器数据延迟{int(hours_ago)}小时'
                else:
                    site_dict['sensor_health'] = 'normal'
        except Exception:
            site_dict['sensor_health'] = 'unknown'
        river = site_dict.get('river', '')
        th = RIVER_THRESHOLDS.get(river, RIVER_THRESHOLDS[''])
        site_dict['wl_threshold_high'] = th['high']
        site_dict['wl_threshold_critical'] = th['critical']
        if wl_row:
            wv = wl_row['value']
            if wv > th['critical']:
                site_dict['wl_status'] = '危急'
            elif wv > th['high']:
                site_dict['wl_status'] = '告警'
            else:
                site_dict['wl_status'] = '正常'
        else:
            site_dict['wl_status'] = '--'
        return jsonify(site_dict)

# --- Sensor Data ---
@app.route('/api/data/realtime')
@login_required
def realtime_data():
    """各站点最新一条数据（优化：一次查询，不用N+1）"""
    site_ids = _filter_site_ids()
    with get_db() as db:
        # 一次查询获取所有站点的最新传感器数据（使用MAX(id)保证每站一条，比GROUP BY快10倍）
        latest = {}
        try:
            latest_rows = db.execute("""
                SELECT sd.site_id, sd.metric, sd.value, sd.unit, sd.recorded_at
                FROM sensor_data sd
                WHERE sd.id IN (SELECT MAX(id) FROM sensor_data GROUP BY site_id)
            """).fetchall()
            for r in latest_rows:
                latest[r['site_id']] = r
        except:
            pass
        site_sql = """SELECT s.id, s.code, s.name, s.type, s.gps_lat as lat, s.gps_lng as lng,
                   CASE WHEN COUNT(d.id) = 0 THEN 'online'
                        WHEN SUM(CASE WHEN d.status='offline' THEN 1 ELSE 0 END) > 0 THEN 'offline'
                        ELSE 'online' END as status
            FROM sites s LEFT JOIN device_shadows d ON s.id=d.site_id"""
        site_params = []
        if site_ids is not None:
            placeholders = ','.join('?' * len(site_ids))
            site_sql += f" WHERE s.id IN ({placeholders})"
            site_params = site_ids
        site_sql += " GROUP BY s.id"
        sites = db.execute(site_sql, site_params).fetchall()
        result = []
        # 额外查询水位站的最新水位数据
        wl_latest = {}
        try:
            wl_rows = db.execute("""
                SELECT sd.site_id, sd.value, sd.recorded_at
                FROM sensor_data sd
                WHERE sd.id IN (
                    SELECT MAX(id) FROM sensor_data WHERE metric='water_level' GROUP BY site_id
                )
            """).fetchall()
            for r in wl_rows:
                wl_latest[r['site_id']] = r
        except:
            pass
        for s in sites:
            row = latest.get(s['id'])
            site_dict = dict(s)
            site_dict['latest_value'] = round(row['value'],2) if row else 0
            site_dict['latest_metric'] = row['metric'] if row else ''
            site_dict['latest_unit'] = row['unit'] if row else ''
            site_dict['latest_time'] = row['recorded_at'] if row else ''
            # 水位站单独附加水位数据
            wl_row = wl_latest.get(s['id'])
            if wl_row:
                site_dict['wl_value'] = round(wl_row['value'],2)
                site_dict['wl_time'] = wl_row['recorded_at']
            result.append(site_dict)
        return jsonify(result)

@app.route('/api/data/site/<int:site_id>')
def site_data(site_id):
    """站点最近2小时数据"""
    limit = request.args.get('limit', 50, type=int)
    with get_db() as db:
        rows = db.execute(
            "SELECT metric, value, unit, recorded_at FROM sensor_data WHERE site_id=? ORDER BY recorded_at DESC LIMIT ?",
            (site_id, limit)
        ).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route('/api/data/site/<int:site_id>/trend')
@login_required
def site_data_trend(site_id):
    """站点历史数据趋势（用于曲线图），支持按指标和时间范围筛选"""
    metric = request.args.get('metric', '')
    hours = request.args.get('hours', 24, type=int)
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    limit = request.args.get('limit', 2000, type=int)
    with get_db() as db:
        q = "SELECT metric, value, unit, recorded_at FROM sensor_data WHERE site_id=?"
        params = [site_id]
        if metric:
            q += " AND metric=?"
            params.append(metric)
        if date_from:
            q += " AND recorded_at>=?"
            params.append(date_from)
        elif hours and hours > 0:
            from datetime import datetime, timedelta
            cutoff = (datetime.now() - timedelta(hours=hours)).strftime('%Y-%m-%d %H:%M:%S')
            q += " AND recorded_at>=?"
            params.append(cutoff)
        if date_to:
            q += " AND recorded_at<=?"
            params.append(date_to)
        q += " ORDER BY recorded_at ASC LIMIT ?"
        params.append(limit)
        rows = db.execute(q, params).fetchall()
    # 按指标分组
    grouped = {}
    for r in rows:
        m = r['metric']
        if m not in grouped:
            grouped[m] = []
        grouped[m].append({
            'value': round(r['value'], 2) if r['value'] is not None else None,
            'unit': r['unit'],
            'recorded_at': r['recorded_at']
        })
    return jsonify({
        'site_id': site_id,
        'metrics': list(grouped.keys()),
        'series': grouped,
        'total_points': len(rows)
    })
@app.route('/api/data/overview')
@login_required
def data_overview():
    site_ids = _filter_site_ids()
    with get_db() as db:
        if site_ids is not None:
            placeholders = ','.join('?' * len(site_ids))
            total_sites = db.execute(f"SELECT COUNT(*) as c FROM sites WHERE id IN ({placeholders})", site_ids).fetchone()['c']
            online_sites = db.execute(f"SELECT COUNT(*) as c FROM sites WHERE status='online' AND id IN ({placeholders})", site_ids).fetchone()['c']
            device_total = db.execute(f"SELECT COUNT(*) as c FROM device_shadows WHERE site_id IN ({placeholders})", site_ids).fetchone()['c']
            device_online = db.execute(f"SELECT COUNT(*) as c FROM device_shadows WHERE status='online' AND site_id IN ({placeholders})", site_ids).fetchone()['c']
            active_alerts = db.execute(f"SELECT COUNT(*) as c FROM alerts WHERE status='pending' AND site_id IN ({placeholders})", site_ids).fetchone()['c']
            open_orders = db.execute(f"SELECT COUNT(*) as c FROM work_orders WHERE status NOT IN ('closed') AND site_id IN ({placeholders})", site_ids).fetchone()['c']
        else:
            total_sites = db.execute("SELECT COUNT(*) as c FROM sites").fetchone()['c']
            online_sites = db.execute("SELECT COUNT(*) as c FROM sites WHERE status='online'").fetchone()['c']
            device_total = db.execute("SELECT COUNT(*) as c FROM device_shadows").fetchone()['c']
            device_online = db.execute("SELECT COUNT(*) as c FROM device_shadows WHERE status='online'").fetchone()['c']
            active_alerts = db.execute("SELECT COUNT(*) as c FROM alerts WHERE status='pending'").fetchone()['c']
            open_orders = db.execute("SELECT COUNT(*) as c FROM work_orders WHERE status NOT IN ('closed')").fetchone()['c']
        return jsonify({
            'total_sites': total_sites, 'online_sites': online_sites,
            'device_total': device_total, 'device_online': device_online,
            'active_alerts': active_alerts, 'open_orders': open_orders
        })

# --- Alerts ---
@app.route('/api/alerts')
@login_required
def get_alerts():
    site_ids = _filter_site_ids()
    status = request.args.get('status', '')
    limit = request.args.get('limit', 50, type=int)
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    site_filter = request.args.get('site_id', '')  # 可选，按站点过滤
    with get_db() as db:
        # 自动办结：已关联工单且工单已闭环的告警 → 自动 resolved
        resolved_ids = db.execute("""
            SELECT a.id FROM alerts a
            WHERE a.status='pending' AND a.flow_status='converted' AND a.related_order_no IS NOT NULL AND a.related_order_no != ''
            AND a.related_order_no IN (SELECT order_no FROM work_orders WHERE status='closed')
        """).fetchall()
        if resolved_ids:
            ids = [r['id'] for r in resolved_ids]
            ph = ','.join('?' * len(ids))
            db.execute(f"UPDATE alerts SET status='resolved', resolved_at=datetime('now','localtime') WHERE id IN ({ph})", ids)
            for rid in ids:
                db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                    ('alert', rid, 'resolved', '系统', '关联工单已闭环，告警自动办结'))
            db.commit()
            print(f"[AutoResolve] 自动办结 {len(ids)} 条告警（关联工单已闭环）")
        q = """
            SELECT a.*, s.name as site_name, s.code as site_code
            FROM alerts a LEFT JOIN sites s ON a.site_id=s.id
            WHERE 1=1
        """
        params = []
        if site_ids is not None:
            placeholders = ','.join('?' * len(site_ids))
            q += f" AND a.site_id IN ({placeholders})"
            params.extend(site_ids)
        if site_filter:
            q += " AND a.site_id=?"
            params.append(int(site_filter))
        if status:
            q += " AND a.status=?"
            params.append(status)
        if date_from:
            q += " AND a.created_at>=?"
            params.append(date_from)
        if date_to:
            q += " AND a.created_at<=?"
            params.append(date_to + ' 23:59:59')
        q += " ORDER BY CASE a.level WHEN 'red' THEN 1 WHEN 'orange' THEN 2 WHEN 'yellow' THEN 3 ELSE 4 END, a.created_at DESC LIMIT ?"
        params.append(limit)
        return jsonify([dict(r) for r in db.execute(q, params).fetchall()])

@app.route('/api/alerts/<int:alert_id>/acknowledge', methods=['POST'])
def acknowledge_alert(alert_id):
    data = request.get_json(silent=True) or {}
    operator = data.get('operator', '系统')
    with get_db() as db:
        db.execute("UPDATE alerts SET status='acknowledged' WHERE id=?", (alert_id,))
        # 记录时间线
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('alert', alert_id, 'acknowledged', operator, '确认告警'))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/alerts/<int:alert_id>/resolve', methods=['POST'])
def resolve_alert(alert_id):
    """办结告警，支持办结原因（reason）"""
    data = request.get_json(silent=True) or {}
    operator = data.get('operator', '系统')
    reason = data.get('reason', '办结告警')
    remark = data.get('remark', '')
    # reason可选值: 误报 / 仪器正常偏差 / 已自动恢复 / 已人工处理 / 自定义
    full_remark = reason + (' - ' + remark if remark else '')
    with get_db() as db:
        db.execute("UPDATE alerts SET status='resolved', resolved_at=datetime('now','localtime') WHERE id=?", (alert_id,))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('alert', alert_id, 'resolved', operator, full_remark))
        # 闭环：直接办结告警时，一并归档关联数据审核项
        ar = db.execute("SELECT review_id FROM alerts WHERE id=?", (alert_id,)).fetchone()
        if ar and ar['review_id']:
            _archive_linked_review(db, None, None, reason, review_id=ar['review_id'])
        db.commit()
        summary = db.execute("SELECT COUNT(*) FROM alerts WHERE status='pending'").fetchone()[0]
        return jsonify({'success': True, 'summary': {'alerts_pending': summary}})

@app.route('/api/alerts/<int:alert_id>/ack-resolve', methods=['POST'])
def ack_resolve_alert(alert_id):
    """一键确认并办结（跳过已确认状态，直接pending→resolved）"""
    data = request.get_json(silent=True) or {}
    operator = data.get('operator', '系统')
    remark = data.get('remark', '一键办结')
    with get_db() as db:
        db.execute("UPDATE alerts SET status='resolved', resolved_at=datetime('now','localtime') WHERE id=? AND status='pending'", (alert_id,))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('alert', alert_id, 'acknowledged', operator, '确认告警'))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('alert', alert_id, 'resolved', operator, remark))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/alerts/<int:alert_id>/urge', methods=['POST'])
def urge_alert(alert_id):
    """告警督办，支持时限、督办人、督办意见"""
    data = request.get_json(silent=True) or {}
    operator = data.get('operator', '系统')
    remark = data.get('opinion', data.get('remark', '督办告警'))
    deadline = data.get('deadline', '')
    supervisor = data.get('supervisor', '')
    cooperator = data.get('cooperator', '')
    # 将额外信息拼入remark
    extra = []
    if supervisor: extra.append('督办人:'+supervisor)
    if deadline: extra.append('限办:'+deadline)
    if cooperator: extra.append('协办:'+cooperator)
    full_remark = remark + (' | ' + '; '.join(extra) if extra else '')
    # 更新数据库中的response_deadline字段
    with get_db() as db:
        db.execute("UPDATE alerts SET urge_count=COALESCE(urge_count,0)+1, last_urged_at=datetime('now','localtime') WHERE id=?", (alert_id,))
        if deadline:
            db.execute("UPDATE alerts SET response_deadline=? WHERE id=?", (deadline, alert_id))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('alert', alert_id, 'urged', supervisor or operator, full_remark))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/alerts/<int:alert_id>/undo-acknowledge', methods=['POST'])
def undo_acknowledge_alert(alert_id):
    """撤销告警确认，将状态改回pending"""
    data = request.get_json(silent=True) or {}
    operator = data.get('operator', '系统')
    remark = data.get('remark', '撤销确认')
    with get_db() as db:
        db.execute("UPDATE alerts SET status='pending', resolved_at=NULL WHERE id=?", (alert_id,))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('alert', alert_id, 'undo_acknowledge', operator, remark))
        db.commit()
        return jsonify({'success': True})

# === 告警流转（A级自动转 / B级人工复核）===

@app.route('/api/alerts/pending-review', methods=['GET'])
@login_required
def get_pending_review_alerts():
    """获取所有待复核的B级告警及其已等待时间"""
    site_ids = _filter_site_ids()
    with get_db() as db:
        q = """
            SELECT a.id, a.site_id, a.metric, a.level, a.message, a.created_at,
                   s.name as site_name, s.code as site_code,
                   ROUND((julianday('now','localtime') - julianday(a.created_at)) * 24 * 60) as wait_minutes
            FROM alerts a LEFT JOIN sites s ON a.site_id=s.id
            WHERE a.flow_type='manual' AND a.flow_status='pending_review'
        """
        params = []
        if site_ids is not None:
            ph = ','.join('?' * len(site_ids))
            q += f" AND a.site_id IN ({ph})"
            params.extend(site_ids)
        q += " ORDER BY a.created_at ASC"
        rows = db.execute(q, params).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route('/api/alerts/<int:alert_id>/confirm-convert', methods=['POST'])
@login_required
def confirm_convert_alert(alert_id):
    """B级告警人工复核确认转工单或关闭"""
    data = request.get_json(silent=True) or {}
    action = data.get('action', 'convert')  # 'convert' 或 'dismiss'
    operator = data.get('operator', g.current_user.get('real_name', '系统'))
    with get_db() as db:
        alert = db.execute("SELECT * FROM alerts WHERE id=?", (alert_id,)).fetchone()
        if not alert:
            return jsonify({'error': '告警不存在'}), 404
        if alert['flow_status'] in ('converted', 'dismissed'):
            return jsonify({'error': '该告警已处理，无法重复操作'}), 400
        if action == 'dismiss':
            remark_txt = data.get('remark', '').strip() or '人工复核后关闭'
            db.execute("UPDATE alerts SET flow_status='dismissed', status='resolved', resolved_at=datetime('now','localtime') WHERE id=?", (alert_id,))
            db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                       ('alert', alert_id, 'dismissed', operator, remark_txt))
            db.commit()
            summary = db.execute("SELECT COUNT(*) FROM alerts WHERE status='pending'").fetchone()[0]
            return jsonify({'success': True, 'action': 'dismissed', 'summary': {'alerts_pending': summary}})
        else:
            # 转工单
            now = datetime.now()
            order_no = f"WO-{now.strftime('%Y%m%d')}-{random.randint(100,999)}"
            order_level = 'critical' if alert['level'] == 'red' else ('urgent' if alert['level'] == 'orange' else 'normal')
            sla_hours = {'normal': 72, 'urgent': 24, 'critical': 2}.get(order_level, 72)
            sla_deadline = (now + timedelta(hours=sla_hours)).strftime('%Y-%m-%d %H:%M')
            site = db.execute("SELECT manager FROM sites WHERE id=?", (alert['site_id'],)).fetchone()
            assignee = _station_operator(alert['site_id']) if site else ''
            db.execute("""
                INSERT INTO work_orders (order_no,site_id,source,event_type,level,title,description,assignee,status,sla_deadline)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (
                order_no, alert['site_id'], 'alert_convert', '告警复核转工单',
                order_level, f"[复核] {alert['message']}", alert['message'],
                assignee, 'in_progress', sla_deadline
            ))
            db.execute("UPDATE alerts SET flow_status='converted', related_order_no=?, status='pending' WHERE id=?",
                       (order_no, alert_id))
            db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                       ('alert', alert_id, 'manual_converted', operator, f'人工复核转工单 {order_no}'))
            # 自动流转时间线
            db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                       ('order', 0, 'accepted', '系统', f'工单{order_no} → 已受理（自动）'))
            db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                       ('order', 0, 'dispatched', assignee or '系统', f'工单{order_no} → 已派发（自动）'))
            db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                       ('order', 0, 'in_progress', '系统', f'工单{order_no} → 处置中（自动）'))
            db.commit()
            summary = db.execute("SELECT COUNT(*) FROM alerts WHERE status='pending'").fetchone()[0]
            return jsonify({'success': True, 'order_no': order_no, 'summary': {'alerts_pending': summary}})

@app.route('/api/alerts/<int:alert_id>/convert-order', methods=['POST'])
def convert_alert_to_order(alert_id):
    """告警转工单"""
    data = request.get_json(silent=True) or {}
    operator = data.get('operator', '系统')
    with get_db() as db:
        alert = db.execute("SELECT * FROM alerts WHERE id=?", (alert_id,)).fetchone()
        if not alert:
            return jsonify({'error': 'not found'}), 404
        if alert['flow_status'] in ('converted', 'dismissed'):
            return jsonify({'error': '该告警已处理，无法重复操作'}), 400
        now = datetime.now()
        order_no = f"WO-{now.strftime('%Y%m%d')}-{random.randint(100,999)}"
        level = data.get('level', alert['level'])
        if level == 'red':
            order_level = 'critical'
        elif level == 'orange':
            order_level = 'urgent'
        else:
            order_level = 'normal'
        sla_hours = {'normal': 72, 'urgent': 24, 'critical': 2}.get(order_level, 72)
        sla_deadline = (now + timedelta(hours=sla_hours)).strftime('%Y-%m-%d %H:%M')
        site = db.execute("SELECT manager FROM sites WHERE id=?", (alert['site_id'],)).fetchone()
        assignee = _station_operator(alert['site_id']) if site else ''
        db.execute("""
            INSERT INTO work_orders (order_no,site_id,source,event_type,level,title,description,assignee,status,sla_deadline)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (
            order_no, alert['site_id'], 'auto', '告警转工单',
            order_level, f"[告警转] {alert['message']}", alert['message'],
            assignee, 'in_progress', sla_deadline
        ))
        # 更新告警关联工单号
        db.execute("UPDATE alerts SET related_order_no=?, flow_status='converted', status='pending' WHERE id=?", (order_no, alert_id))
        # 记录时间线
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('alert', alert_id, 'converted', operator, f'转工单 {order_no}'))
        # 自动流转时间线
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('order', 0, 'accepted', '系统', f'工单{order_no} → 已受理（自动）'))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('order', 0, 'dispatched', assignee or '系统', f'工单{order_no} → 已派发（自动）'))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('order', 0, 'in_progress', '系统', f'工单{order_no} → 处置中（自动）'))
        db.commit()
        summary = db.execute("SELECT COUNT(*) FROM alerts WHERE status='pending'").fetchone()[0]
        return jsonify({'success': True, 'order_no': order_no, 'summary': {'alerts_pending': summary}})

@app.route('/api/alerts/batch', methods=['POST'])
def batch_alert_operations():
    """告警批量操作: acknowledge/resolve/urge/convert"""
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    action = data.get('action', '')
    operator = data.get('operator', '系统')
    if not ids or not action:
        return jsonify({'error': 'ids and action required'}), 400
    with get_db() as db:
        if action == 'acknowledge':
            placeholders = ','.join(['?'] * len(ids))
            db.execute(f"UPDATE alerts SET status='acknowledged' WHERE id IN ({placeholders})", ids)
            for aid in ids:
                db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                           ('alert', aid, 'acknowledged', operator, '批量确认'))
        elif action == 'resolve':
            reason = data.get('reason', '批量办结')
            placeholders = ','.join(['?'] * len(ids))
            db.execute(f"UPDATE alerts SET status='resolved', resolved_at=datetime('now','localtime') WHERE id IN ({placeholders}) AND status='pending'", ids)
            for aid in ids:
                db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                           ('alert', aid, 'resolved', operator, reason))
        elif action == 'urge':
            remark = data.get('remark', '批量督办')
            for aid in ids:
                db.execute("UPDATE alerts SET urge_count=COALESCE(urge_count,0)+1, last_urged_at=datetime('now','localtime') WHERE id=?", (aid,))
                db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                           ('alert', aid, 'urged', operator, remark))
        elif action == 'convert':
            for aid in ids:
                alert = db.execute("SELECT * FROM alerts WHERE id=?", (aid,)).fetchone()
                if not alert:
                    continue
                now = datetime.now()
                order_no = f"WO-{now.strftime('%Y%m%d')}-{random.randint(100,999)}"
                level = alert['level']
                order_level = 'critical' if level == 'red' else ('urgent' if level == 'orange' else 'normal')
                sla_hours = {'normal': 72, 'urgent': 24, 'critical': 2}.get(order_level, 72)
                sla_deadline = (now + timedelta(hours=sla_hours)).strftime('%Y-%m-%d %H:%M')
                db.execute("""
                    INSERT INTO work_orders (order_no,site_id,source,event_type,level,title,description,status,sla_deadline)
                    VALUES (?,?,?,?,?,?,?,?,?)
                """, (order_no, alert['site_id'], 'auto', '告警批量转工单', order_level,
                      f"[告警转] {alert['message']}", alert['message'], 'in_progress', sla_deadline))
                db.execute("UPDATE alerts SET related_order_no=?, flow_status='converted', status='pending' WHERE id=?", (order_no, aid))
                db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                           ('alert', aid, 'converted', operator, f'批量转工单 {order_no}'))
                # 自动流转时间线
                db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                           ('order', 0, 'accepted', '系统', f'工单{order_no} → 已受理（自动）'))
                db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                           ('order', 0, 'dispatched', '系统', f'工单{order_no} → 已派发（自动）'))
                db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                           ('order', 0, 'in_progress', '系统', f'工单{order_no} → 处置中（自动）'))
        else:
            return jsonify({'error': f'unknown action: {action}'}), 400
        db.commit()
        summary = db.execute("SELECT COUNT(*) FROM alerts WHERE status='pending'").fetchone()[0]
        return jsonify({'success': True, 'count': len(ids), 'summary': {'alerts_pending': summary}})

@app.route('/api/timeline')
def get_timeline():
    """时间线查询，可按来源过滤"""
    source_type = request.args.get('source_type', '')
    source_id = request.args.get('source_id', '', type=int) if request.args.get('source_id') else None
    limit = request.args.get('limit', 50, type=int)
    with get_db() as db:
        q = "SELECT * FROM timeline_events WHERE 1=1"
        params = []
        if source_type:
            q += " AND source_type=?"
            params.append(source_type)
        if source_id is not None:
            q += " AND source_id=?"
            params.append(source_id)
        q += " ORDER BY created_at DESC LIMIT ?"
        params.append(limit)
        return jsonify([dict(r) for r in db.execute(q, params).fetchall()])

@app.route('/api/alerts/statistics')
@login_required
def alert_statistics():
    site_ids = _filter_site_ids()
    status = request.args.get('status', '')
    status_where = ''
    params = []
    if status:
        status_where = ' WHERE status=?'
        params.append(status)
    with get_db() as db:
        total = db.execute(f"SELECT COUNT(*) as c FROM alerts{status_where}", params).fetchone()['c']
        by_level = {}
        for lv in ['red','orange','yellow','blue']:
            lv_params = params + [lv]
            by_level[lv] = db.execute(f"SELECT COUNT(*) as c FROM alerts{status_where + ' AND level=?' if status else ' WHERE level=?'}", lv_params).fetchone()['c']
        by_status = {}
        if not status:
            for st in ['pending','acknowledged','resolved']:
                by_status[st] = db.execute("SELECT COUNT(*) as c FROM alerts WHERE status=?",(st,)).fetchone()['c']
        else:
            by_status[status] = total
        # 待复核告警统计
        pending_review = db.execute("SELECT COUNT(*) as c FROM alerts WHERE flow_type='manual' AND flow_status='pending_review'").fetchone()['c']
        auto_converted = db.execute("SELECT COUNT(*) as c FROM alerts WHERE flow_type='auto' AND flow_status='converted'").fetchone()['c']
        return jsonify({'total':total, 'by_level':by_level, 'by_status':by_status,
                        'pending_review': pending_review, 'auto_converted': auto_converted})

# --- Simulate Alert (for demo/rule engine) ---
@app.route('/api/alerts/simulate', methods=['POST'])
@login_required
def simulate_alert():
    data = request.get_json()
    site_id = data.get('site_id')
    metric = data.get('metric', 'data_spike')
    value = data.get('value', 0)
    level = data.get('level', 'blue')
    msg = data.get('message', f'[模拟] 站点 {site_id} 触发 {metric} 告警')
    if not site_id:
        return jsonify({'error': '缺少 site_id'}), 400
    with get_db() as db:
        site = db.execute("SELECT name FROM sites WHERE id=?", (site_id,)).fetchone()
        site_name = site['name'] if site else f'站点{site_id}'
        cur = db.execute(
            "INSERT INTO alerts (site_id, metric, value, level, message, status) VALUES (?,?,?,?,?,?)",
            (site_id, metric, value, level, f'[模拟] {site_name} {msg}', 'pending')
        )
        alert_id = cur.lastrowid
        # Also create a timeline event
        db.execute(
            "INSERT INTO timeline_events (event_type, ref_id, ref_type, site_id, message, created_at) VALUES (?,?,?,?,?,datetime('now','localtime'))",
            ('alert_generated', alert_id, 'alert', site_id, f'模拟触发{level}级告警: {metric}={value}', )
        )
        return jsonify({'id': alert_id, 'site_name': site_name, 'level': level, 'message': msg})

# --- Work Orders ---
@app.route('/api/workorders')
@login_required
def get_workorders():
    status = request.args.get('status', '')
    limit = request.args.get('limit', 50, type=int)
    site_ids = _filter_site_ids()
    with get_db() as db:
        q = """
            SELECT w.*, s.name as site_name
            FROM work_orders w LEFT JOIN sites s ON w.site_id=s.id
            WHERE 1=1
        """
        params = []
        if site_ids is not None:
            ph = ','.join('?' * len(site_ids))
            q += f" AND w.site_id IN ({ph})"
            params.extend(site_ids)
        if status:
            q += " AND w.status=?"
            params.append(status)
        q += " ORDER BY w.created_at DESC LIMIT ?"
        params.append(limit)
        return jsonify([dict(r) for r in db.execute(q, params).fetchall()])

@app.route('/api/workorders', methods=['POST'])
def create_workorder():
    """创建工单 — 直接进入处置中（跳过待受理/已受理/已派发，系统自动完成）"""
    data = request.get_json(force=True, silent=True)
    if not data:
        return jsonify({'error': '无效的请求数据'}), 400

    # ---- 字段规范化：移动端可能写入非标准值，统一映射为Web端标准值 ----
    # 来源规范化
    _source_map = {
        'inspection': 'inspection', '巡检': 'inspection',
        'patrol': 'patrol', '巡查': 'patrol',
        'auto': 'auto', '自动': 'auto',
        'manual': 'manual', '人工': 'manual',
        'superior': 'superior', '上级': 'superior',
        'hotline': 'hotline', '热线': 'hotline',
        'alert_convert': 'alert_convert', '告警转工单': 'alert_convert',
        'alert_auto': 'alert_auto',
    }
    raw_source = (data.get('source') or 'manual').strip().lower()
    data['source'] = _source_map.get(raw_source, raw_source if raw_source in _source_map.values() else 'manual')

    # 级别规范化
    _level_map = {
        'normal': 'normal', '一般': 'normal', 'blue': 'normal', 'yellow': 'normal',
        'medium': 'normal',
        'urgent': 'urgent', '紧急': 'urgent', 'orange': 'urgent',
        'critical': 'critical', '重大': 'critical', 'red': 'critical',
    }
    raw_level = (data.get('level') or 'normal').strip().lower()
    data['level'] = _level_map.get(raw_level, raw_level if raw_level in _level_map.values() else 'normal')

    # ---- 自动填充负责人：从当前登录用户获取 ----
    assignee = (data.get('assignee') or '').strip()
    if not assignee:
        # 尝试从token中提取当前用户
        auth = request.headers.get('Authorization', '')
        token = auth.replace('Bearer ', '').strip() if auth.startswith('Bearer ') else ''
        user = _tokens.get(token)
        if user:
            assignee = user.get('username') or user.get('name') or ''

    # ---- 去重：同站点+同来源+相似标题的未关闭工单已存在则返回已有工单 ----
    title = (data.get('title') or '').strip()
    site_id = data.get('site_id')
    source = data.get('source', 'manual')
    if site_id and title:
        with get_db() as db:
            existing = db.execute(
                "SELECT order_no, title FROM work_orders WHERE site_id=? AND source=? AND status NOT IN ('closed') ORDER BY id DESC LIMIT 5",
                (site_id, source)
            ).fetchall()
            for ex in existing:
                # 标题相似度判断：前20个字符相同即视为重复
                if ex['title'] and ex['title'][:20] == title[:20]:
                    return jsonify({'success': True, 'order_no': ex['order_no'], 'duplicate': True})

    max_retries = 3
    for attempt in range(max_retries):
        try:
            with get_db() as db:
                now = datetime.now()
                order_no = f"WO-{now.strftime('%Y%m%d')}-{random.randint(100,999)}"
                sla_hours = {'normal': 72, 'urgent': 24, 'critical': 2}.get(data.get('level','normal'), 72)
                sla_deadline = (now + timedelta(hours=sla_hours)).strftime('%Y-%m-%d %H:%M')
                # 直接创建为 in_progress（待受理→已受理→已派发→处置中，系统瞬间完成）
                db.execute("""
                    INSERT INTO work_orders (order_no,site_id,source,event_type,level,title,description,images,assignee,status,sla_deadline)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    order_no, data.get('site_id'), data.get('source','manual'),
                    data.get('event_type',''), data.get('level','normal'),
                    title, data.get('description',''),
                    data.get('images',''), assignee,
                    'in_progress', sla_deadline
                ))
                # 时间线记录：记录完整的自动流转链路
                db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                           ('order', 0, 'accepted', '系统', f'工单{order_no} → 已受理（自动）'))
                db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                           ('order', 0, 'dispatched', '系统' if not assignee else assignee, f'工单{order_no} → 已派发（自动）'))
                db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                           ('order', 0, 'in_progress', '系统', f'工单{order_no} → 处置中（自动）'))
                db.commit()
                return jsonify({'success': True, 'order_no': order_no})
        except Exception as e:
            if 'database is locked' in str(e) and attempt < max_retries - 1:
                import time as _t
                _t.sleep(0.3 * (attempt + 1))  # 退避: 0.3s, 0.6s
                continue
            return jsonify({'error': str(e)}), 500

@app.route('/api/workorders/<order_no>/status', methods=['PUT'])
def update_workorder_status(order_no):
    data = request.get_json(silent=True) or {}
    new_status = data.get('status')
    # 关单/核验通过属管理员审批决策（纵深防御：即便绕过前端直调也须管理员）
    if new_status == 'closed' and (not g.get('current_user') or g.get('current_user', {}).get('role') != 'admin'):
        return jsonify({'success': False, 'error': '关单需管理员权限'}), 403
    valid_transitions = {
        'pending': ['accepted'],
        'accepted': ['in_progress', 'dispatched'],
        'dispatched': ['in_progress'],
        'in_progress': ['reviewing', 'accepted'],
        'reviewing': ['closed', 'in_progress'],
    }
    with get_db() as db:
        cur = db.execute("SELECT status, related_alert_id, used_parts, site_id, check_in_time FROM work_orders WHERE order_no=?", (order_no,)).fetchone()
        if not cur:
            return jsonify({'error': 'not found'}), 404
        if new_status not in valid_transitions.get(cur['status'], []):
            return jsonify({'error': f'invalid transition from {cur["status"]} to {new_status}'}), 400
        # 流程门禁（移动端强制）：到场签到后方可开始处置，防办公室内远程空转
        if data.get('client') == 'mobile' and new_status == 'in_progress' and not cur['check_in_time']:
            return jsonify({'error': '请先到场签到后再开始处置'}), 400
        updates = ["status=?"]
        params = [new_status]
        if new_status == 'closed':
            updates.append("resolved_at=datetime('now','localtime')")
        if 'remark' in data:
            updates.append("remark=?")
            params.append(data['remark'])
        if 'conclusion' in data:
            updates.append("conclusion=?")
            params.append(data['conclusion'])
        if 'satisfaction' in data:
            updates.append("satisfaction=?")
            params.append(data['satisfaction'])
        if 'images' in data:
            updates.append("images=?")
            params.append(data['images'])
        params.append(order_no)
        db.execute(f"UPDATE work_orders SET {','.join(updates)} WHERE order_no=?", params)
        # 时间线记录
        operator = data.get('operator', '系统')
        status_cn = {'pending':'待受理','accepted':'已受理','dispatched':'已派发','in_progress':'处置中','reviewing':'审核中','closed':'已完成'}
        event_label = status_cn.get(new_status, new_status)
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('order', 0, new_status, operator, f'工单{order_no} → {event_label}'))

        # === 数据自洽性修复：工单关闭时联动更新 ===
        if new_status == 'closed':
            # 1. 关联告警自动办结（带回写结论）
            if cur['related_alert_id']:
                concl = data.get('conclusion') or 'false_alarm'
                db.execute(
                    "UPDATE alerts SET status='resolved', resolved_at=datetime('now','localtime'), resolve_reason=? WHERE id=? AND status != 'resolved'",
                    (concl, cur['related_alert_id'])
                )
                db.execute(
                    "INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                    ('alert', cur['related_alert_id'], 'resolved', '系统', f'关联工单{order_no}完成，告警自动办结（{concl}）')
                )
                # 闭环：关联数据审核项一并归档
                a = db.execute("SELECT site_id, metric FROM alerts WHERE id=?", (cur['related_alert_id'],)).fetchone()
                if a:
                    _archive_linked_review(db, a['site_id'], a['metric'], concl, order_no=order_no)
            # 2. 关联热线事件自动关闭
            db.execute(
                "UPDATE hotline_events SET status='closed' WHERE related_order_no=? AND status != 'closed'",
                (order_no,)
            )
            # 3. 备件库存扣减（used_parts为JSON格式: [{part_id, quantity}]）
            if cur['used_parts']:
                try:
                    import json as _json
                    parts = _json.loads(cur['used_parts'])
                    for part in parts:
                        part_id = part.get('part_id')
                        qty = part.get('quantity', 0)
                        if part_id and qty > 0:
                            db.execute(
                                "UPDATE spare_parts_inventory SET quantity=MAX(0, quantity-?), updated_at=datetime('now','localtime') WHERE id=?",
                                (qty, part_id)
                            )
                            db.execute(
                                "INSERT INTO inventory_logs (part_id,type,quantity,ref_type,ref_id,operator,remark) VALUES (?,'out',?,'workorder',?,?,'工单领料')",
                                (part_id, qty, 0, order_no)
                            )
                except Exception as e:
                    print(f'[WO] 备件扣减失败: {e}')

        db.commit()
        # 审批结果推送（最佳努力，不阻断主流程）
        try:
            if new_status == 'closed':
                _wx_push_approve_result(cur['site_id'], order_no, '已完成')
        except Exception as e:
            print('[WX] 关单推送异常: %s' % e)
        return jsonify({'success': True, 'status': new_status})

# --- Work Order Verification ---
@app.route('/api/workorders/<order_no>/submit-review', methods=['POST'])
@login_required
def submit_workorder_review(order_no):
    data = request.get_json(silent=True) or {}
    with get_db() as db:
        cur = db.execute("SELECT status, images FROM work_orders WHERE order_no=?", (order_no,)).fetchone()
        if not cur:
            return jsonify({'error': '工单不存在'}), 404
        if cur['status'] != 'in_progress':
            return jsonify({'error': f'当前状态 {cur["status"]} 不允许提交核验'}), 400
        # 随请求带入影像则先落库（移动端可能在此一并提交）
        if 'images' in data:
            db.execute("UPDATE work_orders SET images=? WHERE order_no=?", (data['images'], order_no))
        # 流程门禁（移动端强制）：提交核验前须至少有 1 张处置影像
        if data.get('client') == 'mobile' and not (cur['images'] or data.get('images')):
            return jsonify({'error': '请先上传至少一张处置影像后再提交核验'}), 400
        db.execute("UPDATE work_orders SET status='reviewing' WHERE order_no=?", (order_no,))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('order', 0, 'submit_review', '系统', f'工单{order_no} 提交核验'))
        db.commit()
        return jsonify({'success': True, 'status': 'reviewing'})

@app.route('/api/workorders/<order_no>/approve', methods=['POST'])
@login_required
def approve_workorder(order_no):
    # 核验通过属管理员审批决策（纵深防御：即便绕过前端直调也须管理员）
    if (not g.get('current_user') or g.get('current_user', {}).get('role') != 'admin'):
        return jsonify({'success': False, 'error': '核验通过需管理员权限'}), 403
    data = request.get_json(silent=True) or {}
    with get_db() as db:
        cur = db.execute("SELECT status, related_alert_id, used_parts, site_id FROM work_orders WHERE order_no=?", (order_no,)).fetchone()
        if not cur:
            return jsonify({'error': '工单不存在'}), 404
        if cur['status'] != 'reviewing':
            return jsonify({'error': f'当前状态 {cur["status"]} 不允许核验通过'}), 400
        db.execute("UPDATE work_orders SET status='closed', resolved_at=datetime('now','localtime') WHERE order_no=?", (order_no,))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('order', 0, 'approved', '系统', f'工单{order_no} 核验通过'))
        # === 数据自洽性修复：工单关闭时联动更新 ===
        if cur['related_alert_id']:
            concl = data.get('conclusion') or 'false_alarm'
            db.execute(
                "UPDATE alerts SET status='resolved', resolved_at=datetime('now','localtime'), resolve_reason=? WHERE id=? AND status != 'resolved'",
                (concl, cur['related_alert_id'])
            )
            a = db.execute("SELECT site_id, metric FROM alerts WHERE id=?", (cur['related_alert_id'],)).fetchone()
            if a:
                _archive_linked_review(db, a['site_id'], a['metric'], concl, order_no=order_no)
        db.execute(
            "UPDATE hotline_events SET status='closed' WHERE related_order_no=? AND status != 'closed'",
            (order_no,)
        )
        if cur['used_parts']:
            try:
                import json as _json
                parts = _json.loads(cur['used_parts'])
                for part in parts:
                    part_id = part.get('part_id')
                    qty = part.get('quantity', 0)
                    if part_id and qty > 0:
                        db.execute(
                            "UPDATE spare_parts_inventory SET quantity=MAX(0, quantity-?), updated_at=datetime('now','localtime') WHERE id=?",
                            (qty, part_id)
                        )
                        db.execute(
                            "INSERT INTO inventory_logs (part_id,type,quantity,ref_type,ref_id,operator,remark) VALUES (?,'out',?,'workorder',?,?,'工单领料')",
                            (part_id, qty, 0, order_no)
                        )
            except Exception as e:
                print(f'[WO] 备件扣减失败: {e}')
        db.commit()
        # 审批通过推送（最佳努力，不阻断主流程）
        try:
            _wx_push_approve_result(cur['site_id'], order_no, '核验通过')
        except Exception as e:
            print('[WX] 审批通过推送异常: %s' % e)
        return jsonify({'success': True, 'status': 'closed'})

@app.route('/api/workorders/<order_no>/reject', methods=['POST'])
@login_required
def reject_workorder(order_no):
    # 核验退回属管理员审批决策（纵深防御：即便绕过前端直调也须管理员）
    if (not g.get('current_user') or g.get('current_user', {}).get('role') != 'admin'):
        return jsonify({'success': False, 'error': '核验退回需管理员权限'}), 403
    with get_db() as db:
        cur = db.execute("SELECT status, site_id FROM work_orders WHERE order_no=?", (order_no,)).fetchone()
        if not cur:
            return jsonify({'error': '工单不存在'}), 404
        if cur['status'] != 'reviewing':
            return jsonify({'error': f'当前状态 {cur["status"]} 不允许退回'}), 400
        db.execute("UPDATE work_orders SET status='in_progress' WHERE order_no=?", (order_no,))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('order', 0, 'rejected', '系统', f'工单{order_no} 核验退回'))
        db.commit()
        # 审批退回推送（最佳努力，不阻断主流程）
        try:
            _wx_push_approve_result(cur['site_id'], order_no, '核验退回')
        except Exception as e:
            print('[WX] 审批退回推送异常: %s' % e)
        return jsonify({'success': True, 'status': 'in_progress'})

@app.route('/api/workorders/<order_no>/used-parts', methods=['PUT'])
@login_required
def update_workorder_used_parts(order_no):
    """更新工单使用的备件列表（JSON格式: [{part_id, quantity}]）"""
    data = request.get_json(silent=True) or {}
    used_parts = data.get('used_parts', [])
    with get_db() as db:
        cur = db.execute("SELECT id, status FROM work_orders WHERE order_no=?", (order_no,)).fetchone()
        if not cur:
            return jsonify({'error': '工单不存在'}), 404
        if cur['status'] == 'closed':
            return jsonify({'error': '工单已关闭，无法修改备件'}), 400
        import json as _json
        db.execute("UPDATE work_orders SET used_parts=? WHERE order_no=?",
                   (_json.dumps(used_parts, ensure_ascii=False), order_no))
        db.commit()
        return jsonify({'success': True})


# ===== 工单处置拍照模板（按 event_type 分类） =====
WORKORDER_PHOTO_TEMPLATES = {
    '设备故障': [
        {'item_name': '故障设备照片', 'max_photos': 1, 'need_review': 1, 'inspection_standard': '拍摄故障设备全景，清晰显示设备编号和故障部位'},
        {'item_name': '故障部位特写', 'max_photos': 1, 'need_review': 1, 'inspection_standard': '对故障部位进行特写拍摄，标注故障现象'},
        {'item_name': '维修过程记录', 'max_photos': 1, 'need_review': 0, 'inspection_standard': '拍摄维修操作过程'},
        {'item_name': '维修完成确认', 'max_photos': 1, 'need_review': 1, 'inspection_standard': '维修完成后拍摄设备运行状态'},
    ],
    '设备维护': [
        {'item_name': '维护前设备状态', 'max_photos': 1, 'need_review': 0, 'inspection_standard': '维护前拍摄设备当前状态'},
        {'item_name': '维护操作过程', 'max_photos': 1, 'need_review': 0, 'inspection_standard': '记录维护操作过程'},
        {'item_name': '维护后设备状态', 'max_photos': 1, 'need_review': 0, 'inspection_standard': '维护完成后拍摄设备运行状态'},
    ],
    '数据异常': [
        {'item_name': '现场仪表读数', 'max_photos': 1, 'need_review': 1, 'inspection_standard': '拍摄现场仪表显示数据'},
        {'item_name': '数据采集终端界面', 'max_photos': 1, 'need_review': 1, 'inspection_standard': '拍摄数据采集终端当前界面'},
        {'item_name': '处理过程记录', 'max_photos': 1, 'need_review': 0, 'inspection_standard': '记录异常处理过程'},
        {'item_name': '处理完成确认', 'max_photos': 1, 'need_review': 1, 'inspection_standard': '确认数据恢复正常后拍摄'},
    ],
    '质控校准': [
        {'item_name': '标准溶液/标样照片', 'max_photos': 1, 'need_review': 1, 'inspection_standard': '拍摄使用的标准溶液或标样'},
        {'item_name': '校准操作过程', 'max_photos': 1, 'need_review': 0, 'inspection_standard': '记录校准操作过程'},
        {'item_name': '校准结果界面', 'max_photos': 1, 'need_review': 1, 'inspection_standard': '拍摄校准结果界面，校准曲线r≥0.999'},
    ],
    '备件更换': [
        {'item_name': '旧备件照片', 'max_photos': 1, 'need_review': 0, 'inspection_standard': '拍摄更换下来的旧备件'},
        {'item_name': '新备件照片', 'max_photos': 1, 'need_review': 0, 'inspection_standard': '拍摄新备件型号和规格'},
        {'item_name': '更换完成确认', 'max_photos': 1, 'need_review': 1, 'inspection_standard': '更换完成后的设备状态'},
    ],
    '站房环境': [
        {'item_name': '问题环境照片', 'max_photos': 1, 'need_review': 0, 'inspection_standard': '拍摄存在问题的环境区域'},
        {'item_name': '处理过程记录', 'max_photos': 1, 'need_review': 0, 'inspection_standard': '记录环境问题处理过程'},
        {'item_name': '处理完成确认', 'max_photos': 1, 'need_review': 1, 'inspection_standard': '处理完成后的环境状态'},
    ],
    '告警自动转工单': [
        {'item_name': '现场核实照片', 'max_photos': 1, 'need_review': 1, 'inspection_standard': '到达现场后拍摄仪器仪表当前读数'},
        {'item_name': '问题排查过程', 'max_photos': 1, 'need_review': 0, 'inspection_standard': '记录问题排查过程'},
        {'item_name': '处理结果照片', 'max_photos': 1, 'need_review': 1, 'inspection_standard': '处理完成后拍摄正常状态作为凭证'},
    ],
    'default': [
        {'item_name': '现场照片', 'max_photos': 1, 'need_review': 0, 'inspection_standard': '拍摄现场情况'},
        {'item_name': '处置过程', 'max_photos': 1, 'need_review': 0, 'inspection_standard': '记录处置过程'},
        {'item_name': '完成确认', 'max_photos': 1, 'need_review': 0, 'inspection_standard': '处置完成确认'},
    ],
}

def get_workorder_photo_template(event_type):
    """根据工单 event_type 获取拍照模板，匹配不到返回 default"""
    for key in WORKORDER_PHOTO_TEMPLATES:
        if key in event_type or event_type in key:
            return WORKORDER_PHOTO_TEMPLATES[key]
    return WORKORDER_PHOTO_TEMPLATES['default']

@app.route('/api/workorder-photo-templates', methods=['GET'])
def api_workorder_photo_templates():
    """返回所有工单拍照模板定义"""
    return jsonify(WORKORDER_PHOTO_TEMPLATES)

@app.route('/api/workorders/<order_no>/photo-templates', methods=['GET'])
def api_workorder_photo_template_for_order(order_no):
    """返回指定工单的拍照模板"""
    with get_db() as db:
        wo = db.execute("SELECT * FROM work_orders WHERE order_no=?", (order_no,)).fetchone()
        if not wo:
            return jsonify({'error': '工单不存在'}), 404
        wo_dict = dict(wo)
        event_type = wo_dict.get('event_type', '')
        template = get_workorder_photo_template(event_type)
        return jsonify({'event_type': event_type, 'template': template})

@app.route('/api/workorders/<order_no>/photos', methods=['GET'])
def api_workorder_photos(order_no):
    """获取工单已上传的照片及进度"""
    with get_db() as db:
        wo = db.execute("SELECT * FROM work_orders WHERE order_no=?", (order_no,)).fetchone()
        if not wo:
            return jsonify({'error': '工单不存在'}), 404
        wo_dict = dict(wo)
        # 从 operation_attachments 获取关联照片
        atts = db.execute("""
            SELECT * FROM operation_attachments
            WHERE source_type='workorder' AND source_id=? AND is_deleted=0
            ORDER BY created_at DESC
        """, (wo['id'],)).fetchall()
        photos = [dict(a) for a in atts]
        
        # 获取拍照模板 + 统计进度
        event_type = wo_dict.get('event_type', '')
        template = get_workorder_photo_template(event_type)
        total_required = sum(t['max_photos'] for t in template)
        
        # 按检查项分组统计
        item_progress = []
        for t in template:
            item_photos = [p for p in photos if p.get('category') == t['item_name']]
            item_progress.append({
                'item_name': t['item_name'],
                'max_photos': t['max_photos'],
                'actual_photos': len(item_photos),
                'need_review': t.get('need_review', 0),
                'photos': item_photos,
            })
        
        return jsonify({
            'order_no': order_no,
            'event_type': event_type,
            'total_required': total_required,
            'total_uploaded': len(photos),
            'template': template,
            'item_progress': item_progress,
        })

def _batch_link_wo_photos(order_no, urls, file_size=0):
    """移动端已上传文件后，批量将URL关联到工单（模式3）"""
    with get_db() as db:
        wo = db.execute("SELECT id, status, images, event_type FROM work_orders WHERE order_no=?", (order_no,)).fetchone()
        if not wo:
            return jsonify({'error': '工单不存在'}), 404
        wo_id = wo['id']
        wo_status = wo['status']
        wo_images_str = wo['images']
        wo_event_type = wo['event_type'] or ''
    # 取模板第一项作为category（方便审核匹配）
    template = get_workorder_photo_template(wo_event_type)
    default_category = template[0]['item_name'] if template else '现场照片'
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    # 智能识别：按说明关键词匹配照片类型（接入审核链，工单照片不进独立影像审核队列）
    _b_desc = f'工单[{order_no}]处置照片'
    _b_match = match_photo_requirement('', None, '', _b_desc)
    _b_req_id = _b_match['requirement_id'] if _b_match else None
    _b_rec_cat = _b_match['recognized_category'] if _b_match else ''
    _b_match_status = _b_match['match_status'] if _b_match else 'manual'
    _b_match_conf = _b_match['match_confidence'] if _b_match else None
    _b_review_required = 0
    # 1. 插入 operation_attachments
    inserted = 0
    for url in urls:
        fname = os.path.basename(url) or f'photo_{uuid.uuid4()[:8]}.jpg'
        ext = os.path.splitext(fname)[1].lower() or '.jpg'
        try:
            with get_db() as db:
                db.execute("""INSERT INTO operation_attachments
                    (filename, stored_path, file_type, mime_type, file_size, description,
                     source_type, source_id, site_id, uploader_id, uploader_name,
                     gps_lat, gps_lng, taken_at, category,
             watermark_text, recognized_category, match_status, match_confidence,
             review_required, requirement_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (fname, url, 'image', ext, file_size,
                     f'工单[{order_no}]处置照片',
                     'workorder', wo_id,
                     0, None, '', None, None, now_str, default_category, '', _b_rec_cat, _b_match_status, _b_match_conf, _b_review_required, _b_req_id))
                db.commit()
                inserted += 1
        except Exception as e:
            continue
    # 2. 更新 work_orders.images（追加到JSON数组）
    existing = []
    if wo_images_str:
        try:
            existing = json.loads(wo_images_str)
        except (json.JSONDecodeError, TypeError):
            existing = []
    existing.extend(urls)
    images_json = json.dumps(existing)
    # 3. 如果状态是 dispatched（已派发），自动转为 in_progress（处置中）
    new_status = wo_status
    if wo_status == 'dispatched':
        new_status = 'in_progress'
    with get_db() as db:
        db.execute("UPDATE work_orders SET images=?, status=? WHERE order_no=?",
                   (images_json, new_status, order_no))
        db.commit()
    return {'success': True, 'count': inserted, 'status_changed': wo_status != new_status, 'images': images_json}

def _delete_wo_photo(order_no, delete_url):
    """从工单中删除指定URL的照片"""
    with get_db() as db:
        wo = db.execute("SELECT id, images FROM work_orders WHERE order_no=?", (order_no,)).fetchone()
        if not wo:
            return jsonify({'error': '工单不存在'}), 404
        wo_id = wo['id']
        wo_images_str = wo['images']
        if not wo_images_str:
            return jsonify({'error': '工单无照片'}), 404
        try:
            imgs = json.loads(wo_images_str)
        except (json.JSONDecodeError, TypeError):
            return jsonify({'error': '照片数据异常'}), 500
        if delete_url not in imgs:
            return jsonify({'error': '未找到该照片'}), 404
        imgs.remove(delete_url)
        # 更新 work_orders.images
        db.execute("UPDATE work_orders SET images=? WHERE order_no=?", (json.dumps(imgs), order_no))
        # 删除 operation_attachments 中的对应记录
        db.execute("DELETE FROM operation_attachments WHERE source_type='workorder' AND source_id=? AND stored_path=?", (wo_id, delete_url))
        db.commit()
    # 尝试删除物理文件
    fpath = os.path.join(UPLOAD_DIR, delete_url.lstrip('/'))
    if os.path.exists(fpath):
        try:
            os.remove(fpath)
        except:
            pass
    return jsonify({'success': True, 'images': imgs})

@app.route('/api/workorders/<order_no>/photos', methods=['POST'])
def api_workorder_upload_photo(order_no):
    """上传工单处置照片（关联到 operation_attachments）"""
    # 支持 multipart 和 base64 JSON 两种模式
    file = request.files.get('file')
    is_base64 = False
    if not file:
        data = request.get_json(silent=True) or {}
        # 模式4：删除照片（传 delete_url）
        delete_url = data.get('delete_url')
        if delete_url:
            return _delete_wo_photo(order_no, delete_url)
        # 模式3：移动端已上传文件，传URL列表批量关联
        photos_urls = data.get('photos')
        if isinstance(photos_urls, list) and len(photos_urls) > 0:
            res = _batch_link_wo_photos(order_no, photos_urls)
            return jsonify(res) if isinstance(res, dict) else res
        # 模式2：base64 JSON
        image_b64 = data.get('image', '')
        if image_b64:
            import base64
            file_data = base64.b64decode(image_b64.split(',')[-1])
            ext = '.jpg'
            is_base64 = True
            file = type('obj', (object,), {'filename': 'photo.jpg'})()
        else:
            return jsonify({'error': '请选择文件或提供图片数据'}), 400
    
    with get_db() as db:
        wo = db.execute("SELECT id FROM work_orders WHERE order_no=?", (order_no,)).fetchone()
        if not wo:
            return jsonify({'error': '工单不存在'}), 404
    
    ext = os.path.splitext(file.filename or '.jpg')[1].lower() or '.jpg'
    image_exts = {'.jpg', '.jpeg', '.png', '.webp', '.bmp'}
    if not is_base64:
        file_data = file.read()
    if len(file_data) > 20 * 1024 * 1024:
        return jsonify({'error': '文件大小超过20MB限制'}), 400
    
    # 压缩
    if ext in image_exts:
        try:
            from PIL import Image
            import io
            img = Image.open(io.BytesIO(file_data))
            max_side = 1920
            if max(img.size) > max_side:
                ratio = max_side / max(img.size)
                new_size = (int(img.size[0] * ratio), int(img.size[1] * ratio))
                img = img.resize(new_size, Image.LANCZOS)
            if ext in ('.jpg', '.jpeg') or ext == '.png':
                if img.mode in ('RGBA', 'P'):
                    img = img.convert('RGB')
            buf = io.BytesIO()
            if ext == '.png':
                img.save(buf, format='PNG', optimize=True)
            else:
                img.save(buf, format='JPEG', quality=70, optimize=True)
            file_data = buf.getvalue()
        except: pass
    
    fname = str(uuid.uuid4())[:12] + ext
    now = datetime.now()
    subdir = now.strftime('attachments/%Y/%m')
    stored_dir = os.path.join(UPLOAD_DIR, subdir)
    os.makedirs(stored_dir, exist_ok=True)
    stored_path = os.path.join(stored_dir, fname)
    with open(stored_path, 'wb') as f:
        f.write(file_data)
    url = f'/uploads/{subdir}/{fname}'
    
    # 支持 form 和 JSON 两种模式
    if is_base64:
        data = request.get_json(silent=True) or {}
        category = data.get('item_name', '')
        site_id = data.get('site_id', type=int) or 0
        uploader_id = data.get('uploader_id', type=int)
        uploader_name = data.get('uploader_name', '')
        gps_lat = data.get('gps_lat', type=float)
        gps_lng = data.get('gps_lng', type=float)
        taken_at = data.get('taken_at', '')
        description = data.get('description', f'工单[{order_no}]处置照片')
    else:
        category = request.form.get('item_name', '')
        site_id = request.form.get('site_id', type=int) or 0
        uploader_id = request.form.get('uploader_id', type=int)
        uploader_name = request.form.get('uploader_name', '')
        gps_lat = request.form.get('gps_lat', type=float)
        gps_lng = request.form.get('gps_lng', type=float)
        taken_at = request.form.get('taken_at', '')
        description = request.form.get('description', f'工单[{order_no}]处置照片')

    # 智能识别：按文件名/说明关键词匹配照片类型配置（水印/场景自动归类，接入审核链）
    _wm = (data.get('watermark_text', '') if is_base64 else request.form.get('watermark_text', '')) or ''
    _match = match_photo_requirement(_wm, site_id, file.filename or '', description)
    _req_id = _match['requirement_id'] if _match else None
    _rec_cat = _match['recognized_category'] if _match else ''
    _match_status = _match['match_status'] if _match else 'manual'
    _match_conf = _match['match_confidence'] if _match else None
    # 工单照片由工单审核流覆盖，不进独立影像审核队列（与 /audit/pending 排除 workorder 一致）
    _review_required = 0

    with get_db() as db:
        db.execute("""INSERT INTO operation_attachments
            (filename, stored_path, file_type, mime_type, file_size, description,
             source_type, source_id, site_id, uploader_id, uploader_name,
             gps_lat, gps_lng, taken_at, category,
             watermark_text, recognized_category, match_status, match_confidence,
             review_required, requirement_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (file.filename or fname, url,
             'image', ext, len(file_data),
             description,
             'workorder', wo['id'],
             site_id,
             uploader_id,
             uploader_name,
             gps_lat,
             gps_lng,
             taken_at or datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
             category,
             _wm, _rec_cat, _match_status, _match_conf,
             _review_required, _req_id))
        db.commit()
        aid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    
    return jsonify({'success': True, 'id': aid, 'url': url})

@app.route('/api/workorders/<order_no>/photo-progress', methods=['GET'])
def api_workorder_photo_progress(order_no):
    """工单拍照完成进度统计"""
    with get_db() as db:
        wo = db.execute("SELECT * FROM work_orders WHERE order_no=?", (order_no,)).fetchone()
        if not wo:
            return jsonify({'error': '工单不存在'}), 404
        wo_dict = dict(wo)
        event_type = wo_dict.get('event_type', '')
        template = get_workorder_photo_template(event_type)
        total_required = sum(t['max_photos'] for t in template)
        
        atts = db.execute("""
            SELECT category, COUNT(*) as cnt FROM operation_attachments
            WHERE source_type='workorder' AND source_id=? AND is_deleted=0
            GROUP BY category
        """, (wo['id'],)).fetchall()
        uploaded_map = {a['category']: a['cnt'] for a in atts}
        total_uploaded = sum(uploaded_map.values())
        
        items = []
        for t in template:
            act = uploaded_map.get(t['item_name'], 0)
            items.append({
                'item_name': t['item_name'],
                'required': t['max_photos'],
                'actual': act,
                'completed': act >= t['max_photos'],
                'need_review': t.get('need_review', 0),
            })
        
        return jsonify({
            'order_no': order_no,
            'total_required': total_required,
            'total_uploaded': total_uploaded,
            'completed': total_uploaded >= total_required,
            'progress_pct': round(total_uploaded / total_required * 100) if total_required > 0 else 0,
            'items': items,
        })

@app.route('/api/workorders/statistics')
@login_required
def workorder_statistics():
    site_ids = _filter_site_ids()
    with get_db() as db:
        if site_ids is not None:
            ph = ','.join('?' * len(site_ids))
            total = db.execute(f"SELECT COUNT(*) as c FROM work_orders WHERE site_id IN ({ph})", site_ids).fetchone()['c']
            by_status = {}
            for st in ['pending','accepted','dispatched','in_progress','reviewing','closed']:
                by_status[st] = db.execute(f"SELECT COUNT(*) as c FROM work_orders WHERE status=? AND site_id IN ({ph})", [st] + site_ids).fetchone()['c']
            today = datetime.now().strftime('%Y-%m-%d')
            today_new = db.execute(f"SELECT COUNT(*) as c FROM work_orders WHERE date(created_at)=? AND site_id IN ({ph})", [today] + site_ids).fetchone()['c']
            today_closed = db.execute(f"SELECT COUNT(*) as c FROM work_orders WHERE date(resolved_at)=? AND site_id IN ({ph})", [today] + site_ids).fetchone()['c']
        else:
            total = db.execute("SELECT COUNT(*) as c FROM work_orders").fetchone()['c']
            by_status = {}
            for st in ['pending','accepted','dispatched','in_progress','reviewing','closed']:
                by_status[st] = db.execute("SELECT COUNT(*) as c FROM work_orders WHERE status=?",(st,)).fetchone()['c']
            today = datetime.now().strftime('%Y-%m-%d')
            today_new = db.execute("SELECT COUNT(*) as c FROM work_orders WHERE date(created_at)=?",(today,)).fetchone()['c']
            today_closed = db.execute("SELECT COUNT(*) as c FROM work_orders WHERE date(resolved_at)=?",(today,)).fetchone()['c']
        return jsonify({'total':total, 'by_status':by_status, 'today_new':today_new, 'today_closed':today_closed})


@app.route('/api/workorders/<order_no>/related')
@login_required
def api_workorder_related(order_no):
    """获取工单关联的备件申请和设备回收记录"""
    with get_db() as db:
        parts = db.execute("""SELECT * FROM spare_part_requests WHERE work_order_no=? ORDER BY created_at DESC""",
                          (order_no,)).fetchall()
        recycles = db.execute("""SELECT * FROM device_recycle WHERE work_order_no=? ORDER BY created_at DESC""",
                             (order_no,)).fetchall()
    return jsonify({
        'parts': [dict(r) for r in parts],
        'recycles': [dict(r) for r in recycles],
    })


# --- Inspections ---
@app.route('/api/inspections')
@login_required
def get_inspections():
    site_ids = _filter_site_ids()
    freq = request.args.get('frequency', '')  # high/mid/low/annual
    with get_db() as db:
        q = """
            SELECT p.*,
                (SELECT COUNT(*) FROM inspection_tasks t WHERE t.plan_id=p.id) as total_items,
                (SELECT COUNT(*) FROM inspection_tasks t WHERE t.plan_id=p.id AND t.result IS NOT NULL) as completed_items
            FROM inspection_plans p
            WHERE 1=1
        """
        params = []
        if site_ids is not None:
            ph = ','.join('?' * len(site_ids))
            q += f" AND p.id IN (SELECT plan_id FROM plan_sites WHERE site_id IN ({ph}))"
            params.extend(site_ids)
        q += " ORDER BY p.created_at DESC"
        rows = db.execute(q, params).fetchall()
        plans = [dict(r) for r in rows]
        # 为每个计划加载关联站点列表
        for plan in plans:
            sites = db.execute("""
                SELECT s.id, s.name as site_name, s.code as site_code, s.type as site_type,
                    s.gps_lat as lat, s.gps_lng as lng, s.manager as assignee
                FROM plan_sites ps JOIN sites s ON ps.site_id=s.id
                WHERE ps.plan_id=?
            """, (plan['id'],)).fetchall()
            plan['sites'] = [dict(s) for s in sites]
            # 兼容旧字段：取第一个站点
            if sites:
                plan['site_id'] = sites[0]['id']
                plan['site_name'] = sites[0]['site_name']
                plan['site_code'] = sites[0]['site_code']
                plan['site_type'] = sites[0]['site_type']
                plan['lat'] = sites[0]['lat']
                plan['lng'] = sites[0]['lng']
                plan['assignee'] = sites[0]['assignee']
            else:
                plan['site_id'] = plan['site_name'] = plan['site_code'] = plan['site_type'] = None
                plan['lat'] = plan['lng'] = None
                plan['assignee'] = None
        # 按 site_type 统计分组
        site_type_map = {
            'station_yard': '站院', 'reservoir': '站院', 'sluice': '水文站',
            'dike': '水文站', 'pump': '水文站', 'water_supply': '水文站',
            'hydrology': '水文站', 'water_level': '水位站', 'rainfall': '雨量站',
            'groundwater': '地下水监测站', 'soil_moisture': '墒情站',
            'evaporation': '蒸发站',
        }
        site_cats = {}
        for p in plans:
            st = site_type_map.get(p.get('site_type',''), '其他')
            p['site_cat'] = st
            site_cats.setdefault(st, {'total':0,'pending':0,'in_progress':0,'completed':0})
            site_cats[st]['total'] += 1
            site_cats[st][p['status']] = site_cats[st].get(p['status'], 0) + 1
        return jsonify({'plans': plans, 'categories': site_cats, 'site_categories': site_cats})

@app.route('/api/inspections', methods=['POST'])
def create_inspection():
    data = request.json
    with get_db() as db:
        scheme_id = data.get('scheme_id')
        # 支持 site_ids 数组（多站点）和 site_id 单个站点（兼容旧版）
        site_ids = data.get('site_ids', [])
        site_id = data.get('site_id')
        if site_id and not site_ids:
            site_ids = [site_id]
        if not site_ids:
            return jsonify({'success': False, 'error': '请指定至少一个站点'}), 400
        first_site = site_ids[0]
        cursor = db.execute("""
            INSERT INTO inspection_plans (plan_name,site_id,type,start_date,end_date,period,description,category,scheme_id)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (data['plan_name'], first_site, data['type'], data['start_date'], data['end_date'], data.get('period','once'), data.get('description',''), data.get('category',''), scheme_id))
        plan_id = cursor.lastrowid
        # 写入 plan_sites（多站点关联）
        for sid in site_ids:
            db.execute("INSERT OR IGNORE INTO plan_sites (plan_id, site_id) VALUES (?,?)", (plan_id, sid))
        # 生成检查项：优先从scheme_id加载，否则用check_items
        check_items = data.get('check_items', [])
        if scheme_id:
            scheme_items = db.execute("SELECT check_item FROM inspection_scheme_items WHERE scheme_id=? ORDER BY sort_order",(scheme_id,)).fetchall()
            if scheme_items:
                check_items = [r['check_item'] for r in scheme_items]
        if not check_items:
            check_items = ['坝体外观检查','溢洪道检查','放水设施检查','监测设备检查','防汛物资检查','管理设施检查']
        for sid in site_ids:
            for item in check_items:
                db.execute(
                    "INSERT INTO inspection_tasks (plan_id,site_id,check_item) VALUES (?,?,?)",
                    (plan_id, sid, item)
                )
        db.commit()
        # 时间线记录
        operator = data.get('operator', '系统')
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('inspection', plan_id, 'created', operator, f'创建巡检计划-{data["plan_name"]}'))
        db.commit()
        # 通知站点负责人
        _notify_inspection_plan(plan_id, data['plan_name'], first_site, '已创建')
        return jsonify({'success': True, 'plan_id': plan_id})

@app.route('/api/inspections/<int:plan_id>', methods=['DELETE'])
@login_required
def delete_inspection(plan_id):
    """删除巡检计划及其检查项"""
    with get_db() as db:
        plan = db.execute("SELECT plan_name, site_id FROM inspection_plans WHERE id=?", (plan_id,)).fetchone()
        if not plan:
            return jsonify({'error': '计划不存在'}), 404
        db.execute("DELETE FROM plan_sites WHERE plan_id=?", (plan_id,))
        db.execute("DELETE FROM inspection_tasks WHERE plan_id=?", (plan_id,))
        db.execute("DELETE FROM timeline_events WHERE source_type='inspection' AND source_id=?", (plan_id,))
        db.execute("DELETE FROM inspection_plans WHERE id=?", (plan_id,))
        db.commit()

@app.route('/api/inspections/<int:plan_id>/tasks')
def get_inspection_tasks(plan_id):
    with get_db() as db:
        rows = db.execute("SELECT * FROM inspection_tasks WHERE plan_id=? ORDER BY id", (plan_id,)).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route('/api/inspections/<int:plan_id>/attachments')
def get_inspection_attachments(plan_id):
    """返回巡检计划的附件列表（有照片的检查项）"""
    with get_db() as db:
        rows = db.execute(
            "SELECT id, check_item, photo, remark, check_time, result FROM inspection_tasks WHERE plan_id=? AND photo IS NOT NULL AND photo != '' ORDER BY check_time DESC",
            (plan_id,)
        ).fetchall()
        return jsonify([dict(r) for r in rows])
@app.route('/api/inspections/tasks/<int:task_id>', methods=['PUT'])
def update_inspection_task(task_id):
    data = request.json
    with get_db() as db:
        # 构造动态更新字段
        updates = ["result=?", "photo=?", "gps_lat=?", "gps_lng=?", "check_time=?", "remark=?"]
        params = [data.get('result'), data.get('photo'), data.get('gps_lat'), data.get('gps_lng'),
                  data.get('check_time'), data.get('remark')]
        # 新增字段
        if 'photo_urls' in data:
            updates.append("photo_urls=?")
            params.append(data['photo_urls'])
        if 'calibrator' in data:
            updates.append("calibrator=?")
            params.append(data['calibrator'])
        if 'calibration_values' in data:
            updates.append("calibration_values=?")
            params.append(data['calibration_values'])
        params.append(task_id)
        db.execute(f"UPDATE inspection_tasks SET {','.join(updates)} WHERE id=?", params)

        # === 数据自洽性修复：巡检异常结果触发告警 ===
        if data.get('result') == 'abnormal':
            task = db.execute("SELECT site_id, check_item FROM inspection_tasks WHERE id=?", (task_id,)).fetchone()
            if task:
                check_item = data.get('remark', '') or task['check_item']
                create_alert_internal(db, task['site_id'], 'inspection', 0, 'yellow',
                    f'巡检异常：{task["check_item"]} - {check_item}')

        # 更新计划状态
        task = db.execute("SELECT plan_id FROM inspection_tasks WHERE id=?", (task_id,)).fetchone()
        if task:
            incomplete = db.execute(
                "SELECT COUNT(*) as c FROM inspection_tasks WHERE plan_id=? AND result IS NULL",
                (task['plan_id'],)
            ).fetchone()['c']
            if incomplete == 0:
                db.execute("UPDATE inspection_plans SET status='completed' WHERE id=?", (task['plan_id'],))
                plan = db.execute("SELECT plan_name, site_id FROM inspection_plans WHERE id=?", (task['plan_id'],)).fetchone()
                if plan:
                    db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                               ('inspection', task['plan_id'], 'completed', '系统', f'巡检计划完成-{plan["plan_name"]}'))
                    _notify_inspection_plan(task['plan_id'], plan['plan_name'], plan['site_id'], '已完成')
        db.commit()
        return jsonify({'success': True})

@app.route('/api/inspections/statistics')
@login_required
def inspection_statistics():
    site_ids = _filter_site_ids()
    with get_db() as db:
        if site_ids is not None:
            ph = ','.join('?' * len(site_ids))
            total_plans = db.execute(f"SELECT COUNT(DISTINCT plan_id) as c FROM plan_sites WHERE site_id IN ({ph})", site_ids).fetchone()['c']
            done = db.execute(f"SELECT COUNT(DISTINCT p.id) as c FROM inspection_plans p JOIN plan_sites ps ON p.id=ps.plan_id WHERE p.status='completed' AND ps.site_id IN ({ph})", site_ids).fetchone()['c']
            total_tasks = db.execute(f"SELECT COUNT(*) as c FROM inspection_tasks WHERE site_id IN ({ph})", site_ids).fetchone()['c']
            done_tasks = db.execute(f"SELECT COUNT(*) as c FROM inspection_tasks WHERE result IS NOT NULL AND site_id IN ({ph})", site_ids).fetchone()['c']
            abnormal = db.execute(f"SELECT COUNT(*) as c FROM inspection_tasks WHERE result='abnormal' AND site_id IN ({ph})", site_ids).fetchone()['c']
        else:
            total_plans = db.execute("SELECT COUNT(*) as c FROM inspection_plans").fetchone()['c']
            done = db.execute("SELECT COUNT(*) as c FROM inspection_plans WHERE status='completed'").fetchone()['c']
            total_tasks = db.execute("SELECT COUNT(*) as c FROM inspection_tasks").fetchone()['c']
            done_tasks = db.execute("SELECT COUNT(*) as c FROM inspection_tasks WHERE result IS NOT NULL").fetchone()['c']
            abnormal = db.execute("SELECT COUNT(*) as c FROM inspection_tasks WHERE result='abnormal'").fetchone()['c']
        return jsonify({
            'total_plans':total_plans, 'completed_plans':done,
            'total_tasks':total_tasks, 'completed_tasks':done_tasks,
            'abnormal_count':abnormal
        })

DEFAULT_CHECK_ITEMS = {
    '站院环境': ['水位井/站院/大门口全面打扫','设备表面及窗台擦拭','墙面天花板检查(无污迹/蜘蛛网)','草地灌木修剪维护','巡测站站房全面打扫','观测场草地维护(草高<20cm)'],
    '断面环境': ['测流断面上下游各5米清理杂草杂木','缆道铁塔四周清理','基本水尺断面上下游各10米清理','水尺码头/停船码头清理淤泥杂草','比降断面水尺道路清理','洪水退水时及时清理'],
    '水位观测': ['基本水尺读数观测记录','遥测水位及时间校对','人工与遥测水位偏差检测','水尺清洗检查','水位设备运行检查','填记水位巡查表并拍照存档'],
    '雨量监测': ['遥测雨量器现场运行维护','数据采集终端内部状态检查','供电设备及布线检查','雨量筒外观及水平检查','注水试验(季度≥12.5mm误差≤±4%)','特大暴雨后及时检查'],
    '蒸发监测': ['自动蒸发设备遥测终端巡检','蒸发器换水保持清洁','渗漏检查(半年/关闭阀门/邻站对比)','自动蒸发系统注水实验(汛前)','水圈清洁及环境维护'],
    '墒情监测': ['机箱内部清洁','周边杂草清理','无积水检查','数据校测记录','辅助站取土烘干法检验(干旱触发)'],
    '设施设备': ['水尺清洗检查','爬梯/护栏牢固度全面检查','设施设备外观检查','异常维修与拍照存档','上报中心站网监测科'],
    '缆道系统': ['行主索/循环索检查维护','拉线/卡头检查(异常通知甲方)','工作索毛刺断骨拍照留底','锚碇位移/土壤裂纹检查','导向轮/游轮/行车架运转检查','绞车运转检查','钢丝绳夹头/生锈/排水检查'],
    '安全防护': ['测验设施设备安全环境检查','灭火器压力及有效期检查','安全器材完好性检查','站房结构安全及电气线路检查','填写安全检查记录表','安全隐患及时告知中心'],
    '发电机': ['发电机维护保养(汛前/汛后:更换机油/线路/备足燃料)','机油液位检查','线路及各部件检查','发电运行≥30分钟并记录','燃料及机油储备检查'],
    '自定义': []
}

DAY_ITEMS = ['水位观测']
WEEK_ITEMS = ['站院环境','水位观测']
MONTH_ITEMS = ['站院环境','水位观测','雨量监测','蒸发监测','设施设备','安全防护','发电机']
QUARTER_ITEMS = ['雨量监测','墒情监测']
HALF_YEAR_ITEMS = ['蒸发监测']
YEAR_ITEMS = ['断面环境','蒸发监测','发电机']

# DOCX 巡查对象分类（按一）— 用于巡检计划分类显示
DOCX_CATEGORIES = [
    ('站院环境', '站院/观测场清洁、草地修剪维护'),
    ('断面环境', '测流断面及水尺码头清理'),
    ('水位观测', '水尺读数、遥测校对、水位设备检查'),
    ('雨量监测', '雨量器巡检、注水试验'),
    ('蒸发监测', '蒸发设备巡检、换水、渗漏检查'),
    ('墒情监测', '墒情站巡查、数据校测'),
    ('设施设备', '水尺、爬梯、护栏等设施检查'),
    ('缆道系统', '主索、绞车、锚碇等缆道检查'),
    ('安全防护', '灭火器、电气线路安全检查'),
    ('发电机', '发电机保养、运行检查'),
]

@app.route('/api/schemes/template')
def download_scheme_template():
    """下载巡检方案导入模板（CSV格式）"""
    import csv, io
    output = io.StringIO()
    output.write('\ufeff')
    w = csv.writer(output)
    w.writerow(['站点名称','分类','检查项','每日','每周','每月','每季度','每半年','每年'])
    all_cats = [c for c in DEFAULT_CHECK_ITEMS if DEFAULT_CHECK_ITEMS[c]]
    for cat in all_cats:
        for item in DEFAULT_CHECK_ITEMS[cat]:
            w.writerow([
                '', cat, item,
                '✓' if cat in DAY_ITEMS else '',
                '✓' if cat in WEEK_ITEMS else '',
                '✓' if cat in MONTH_ITEMS else '',
                '✓' if cat in QUARTER_ITEMS else '',
                '✓' if cat in HALF_YEAR_ITEMS else '',
                '✓' if cat in YEAR_ITEMS else '',
            ])
    data = output.getvalue().encode('utf-8-sig')
    output.close()
    from flask import Response
    from urllib.parse import quote
    cd = f"attachment; filename=\"inspection_template.csv\"; filename*=UTF-8''{quote('巡检方案导入模板.csv')}"
    return Response(data, mimetype='text/csv; charset=utf-8',
                    headers={'Content-Disposition': cd})

@app.route('/api/schemes/import', methods=['POST'])
def import_schemes():
    """导入站点级方案表：每行=站点+检查项，匹配站点名后写入对应方案"""
    try: import openpyxl
    except: return jsonify({'error':'openpyxl未安装'}),500
    file = request.files.get('file')
    if not file: return jsonify({'error':'请上传文件'}),400
    wb = openpyxl.load_workbook(file); ws = wb.active
    def _yes(v): return str(v).strip() in ('✓','√','Y','y','1','是','yes')
    created = 0
    with get_db() as db:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row)<3: continue
            site_name = str(row[0]).strip() if row[0] else ''
            cat = str(row[1]).strip() if row[1] else ''
            item = str(row[2]).strip() if row[2] else ''
            day_flag = row[3] if len(row)>3 and row[3] else ''
            week_flag = row[4] if len(row)>4 and row[4] else ''
            month_flag = row[5] if len(row)>5 and row[5] else ''
            if not site_name or not item: continue
            site = db.execute("SELECT id, name FROM sites WHERE name=? OR code=?",(site_name,site_name)).fetchone()
            if not site:
                site = db.execute("SELECT id, name FROM sites WHERE name LIKE ?",('%'+site_name+'%',)).fetchone()
            if not site: continue
            sid, sname = site[0], site[1]
            for period, flag, label in [('daily',day_flag,'日巡检方案'),('weekly',week_flag,'周巡检方案'),('monthly',month_flag,'月巡检方案')]:
                if not _yes(flag): continue
                db.execute("INSERT OR IGNORE INTO inspection_schemes (site_id,period,name) VALUES (?,?,?)",(sid,period,f'{sname}-{label}'))
                scheme = db.execute("SELECT id FROM inspection_schemes WHERE site_id=? AND period=?",(sid,period)).fetchone()
                if not scheme: continue
                sc_id = scheme['id']
                # Check if item already exists
                existing = db.execute("SELECT id FROM inspection_scheme_items WHERE scheme_id=? AND check_item=?",(sc_id,item)).fetchone()
                if not existing:
                    next_order = db.execute("SELECT COALESCE(MAX(sort_order),-1)+1 as n FROM inspection_scheme_items WHERE scheme_id=?",(sc_id,)).fetchone()['n']
                    db.execute("INSERT INTO inspection_scheme_items (scheme_id,category,check_item,sort_order) VALUES (?,?,?,?)",(sc_id,cat,item,next_order))
                    created += 1
        db.commit()
    return jsonify({'success':True,'created':created,"warn":"仅记录新增项，已有项未覆盖"})


UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'frontend', 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """上传图片/附件，返回可访问的URL。支持 multipart/form-data，字段名 file。
    图片自动压缩（最大边1920px，质量0.7），单文件上限5MB。"""
    file = request.files.get('file')
    if not file:
        return jsonify({'error': '请选择文件'}), 400

    ext = os.path.splitext(file.filename or '.jpg')[1].lower() or '.jpg'
    image_exts = {'.jpg', '.jpeg', '.png', '.webp', '.bmp'}

    # 读取文件内容，检查大小
    file_data = file.read()
    if len(file_data) > 5 * 1024 * 1024:
        return jsonify({'error': '文件大小超过5MB限制'}), 400

    # 图片压缩处理
    if ext in image_exts:
        try:
            from PIL import Image
            import io
            img = Image.open(io.BytesIO(file_data))
            # 限制最大边为1920px
            max_side = 1920
            if max(img.size) > max_side:
                ratio = max_side / max(img.size)
                new_size = (int(img.size[0] * ratio), int(img.size[1] * ratio))
                img = img.resize(new_size, Image.LANCZOS)
            # 转换为RGB（如果是PNG带透明通道）
            if ext in ('.jpg', '.jpeg') or ext == '.png':
                if img.mode in ('RGBA', 'P'):
                    img = img.convert('RGB')
            # 保存为压缩JPEG
            buf = io.BytesIO()
            save_ext = '.jpg'
            if ext == '.png':
                img.save(buf, format='PNG', optimize=True)
            else:
                img.save(buf, format='JPEG', quality=70, optimize=True)
            file_data = buf.getvalue()
        except ImportError:
            pass  # Pillow未安装，跳过压缩
        except Exception:
            pass  # 压缩失败，使用原文件

    fname = str(uuid.uuid4())[:8] + (ext if ext in image_exts else '.jpg')
    # 按日期分目录：YYYY/MM/fname
    now = datetime.now()
    subdir = now.strftime('%Y/%m')
    stored_dir = os.path.join(UPLOAD_DIR, subdir)
    os.makedirs(stored_dir, exist_ok=True)
    stored_path = os.path.join(stored_dir, fname)
    with open(stored_path, 'wb') as f:
        f.write(file_data)
    url = f'/uploads/{subdir}/{fname}'
    # 注册到files表
    import hashlib
    md5_val = hashlib.md5(file_data).hexdigest()
    source_type = request.form.get('source_type', '')
    source_id_str = request.form.get('source_id', '')
    try: source_id_val = int(source_id_str) if source_id_str else None
    except: source_id_val = None
    with get_db() as db:
        # 去重检查
        existing = db.execute("SELECT id, stored_path FROM files WHERE md5_hash=? AND is_deleted=0", (md5_val,)).fetchone()
        if existing:
            # 已有相同文件，删除刚存的，复用记录
            if os.path.exists(stored_path):
                os.remove(stored_path)
            url = existing['stored_path']
        else:
            db.execute("""
                INSERT INTO files (filename, stored_path, file_type, mime_type, file_size, md5_hash, source_type, source_id)
                VALUES (?,?,?,?,?,?,?,?)
            """, (file.filename or fname, url, 'image' if ext in image_exts else 'file',
                  ext, len(file_data), md5_val, source_type, source_id_val))
            db.commit()
    return jsonify({'success': True, 'url': url})


# ============================================================
# 运维影像附件管理 API
# ============================================================

@app.route('/api/attachments', methods=['GET'])
def list_attachments():
    """影像档案列表，支持多维筛选"""
    site_id = request.args.get('site_id', type=int)
    source_type = request.args.get('source_type')
    source_id = request.args.get('source_id', type=int)
    category = request.args.get('category')
    uploader = request.args.get('uploader')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    keyword = request.args.get('keyword')
    archived = request.args.get('archived')  # '0' / '1' / None(全部)
    review_status = request.args.get('review_status')
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 30, type=int)
    offset = (page - 1) * limit

    with get_db() as db:
        where = ["oa.is_deleted=0"]
        params = []
        if site_id:
            where.append("oa.site_id=?"); params.append(site_id)
        if source_type:
            where.append("oa.source_type=?"); params.append(source_type)
        if source_id:
            where.append("oa.source_id=?"); params.append(source_id)
        if category:
            where.append("oa.category=?"); params.append(category)
        if uploader:
            where.append("oa.uploader_name LIKE ?"); params.append(f'%{uploader}%')
        if date_from:
            where.append("oa.taken_at>=?"); params.append(date_from)
        if date_to:
            where.append("oa.taken_at<=?"); params.append(date_to + ' 23:59:59')
        if keyword:
            where.append("(oa.description LIKE ? OR oa.filename LIKE ?)")
            params.extend([f'%{keyword}%', f'%{keyword}%'])
        if archived in ('0', '1'):
            where.append("oa.archived=?")
            params.append(int(archived))
        if review_status:
            where.append("oa.review_status=?")
            params.append(review_status)

        # 范围隔离：操作员仅见本人站点影像。站点为 0/NULL 视为未绑定资料（多为测试/种子数据），
        # 不展示给受限角色，避免跨站点照片泄露；仅管理员/经理（scope=None）可见全部。
        scope = _filter_site_ids()
        if scope is not None:
            where.append("oa.site_id IN (%s)" % ','.join('?' * len(scope)))
            params.extend(scope)

        where_sql = ' AND '.join(where)
        total = db.execute(f"SELECT COUNT(*) FROM operation_attachments oa WHERE {where_sql}", params).fetchone()[0]
        rows = db.execute(f"""
            SELECT oa.*, s.name as site_name, u.real_name AS uploader_real_name
            FROM operation_attachments oa
            LEFT JOIN sites s ON oa.site_id=s.id
            LEFT JOIN users u ON oa.uploader_id=u.id
            WHERE {where_sql}
            ORDER BY oa.created_at DESC LIMIT ? OFFSET ?
        """, params + [limit, offset]).fetchall()

        result = []
        for r in rows:
            d = dict(r)
            if d.get('uploader_real_name'):
                d['uploader_name'] = d['uploader_real_name']
            result.append(d)
        return jsonify({
            'total': total,
            'page': page,
            'limit': limit,
            'items': result
        })


@app.route('/api/attachments', methods=['POST'])
def create_attachment():
    """手动创建附件记录（配合已有图片路径使用）"""
    data = request.json
    required = ['filename', 'stored_path']
    for f in required:
        if f not in data:
            return jsonify({'error': f'缺少必填字段: {f}'}), 400
    with get_db() as db:
        db.execute("""INSERT INTO operation_attachments
            (filename, stored_path, thumbnail_path, file_type, mime_type, file_size,
             description, source_type, source_id, site_id, uploader_id, uploader_name,
             gps_lat, gps_lng, taken_at, category)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (data['filename'], data['stored_path'], data.get('thumbnail_path',''),
             data.get('file_type','image'), data.get('mime_type',''), data.get('file_size',0),
             data.get('description',''), data.get('source_type',''), data.get('source_id',0),
             data.get('site_id'), data.get('uploader_id'), data.get('uploader_name',''),
             data.get('gps_lat'), data.get('gps_lng'), data.get('taken_at'),
             data.get('category','')))
        db.commit()
        aid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    return jsonify({'success': True, 'id': aid})


@app.route('/api/attachments/<int:aid>')
def get_attachment(aid):
    """获取单个附件详情"""
    with get_db() as db:
        row = db.execute("""
            SELECT oa.*, s.name as site_name
            FROM operation_attachments oa
            LEFT JOIN sites s ON oa.site_id=s.id
            WHERE oa.id=? AND oa.is_deleted=0
        """, (aid,)).fetchone()
        if not row:
            return jsonify({'error': '附件不存在'}), 404
        return jsonify(dict(row))


@app.route('/api/attachments/<int:aid>', methods=['PUT'])
def update_attachment(aid):
    """更新附件元信息（描述、分类等）"""
    data = request.json
    allowed = ['description', 'category', 'source_type', 'source_id', 'site_id']
    with get_db() as db:
        existing = db.execute("SELECT id FROM operation_attachments WHERE id=? AND is_deleted=0", (aid,)).fetchone()
        if not existing:
            return jsonify({'error': '附件不存在'}), 404
        updates = []
        params = []
        for key in allowed:
            if key in data:
                updates.append(f"{key}=?")
                params.append(data[key])
        if updates:
            params.append(aid)
            db.execute(f"UPDATE operation_attachments SET {','.join(updates)} WHERE id=?", params)
            db.commit()
    return jsonify({'success': True})


@app.route('/api/attachments/<int:aid>', methods=['DELETE'])
def delete_attachment(aid):
    """软删除附件"""
    with get_db() as db:
        db.execute("UPDATE operation_attachments SET is_deleted=1 WHERE id=?", (aid,))
        db.commit()
    return jsonify({'success': True})


@app.route('/api/attachments/<int:aid>/archive', methods=['POST'])
def archive_attachment(aid):
    """归档影像资料：标记为已归档，便于长期留存与独立检索"""
    data = request.get_json(silent=True) or {}
    reason = data.get('archive_reason', '')
    archived_by = data.get('archived_by') or (
        g.current_user.get('id') if hasattr(g, 'current_user') else None)
    with get_db() as db:
        att = db.execute(
            "SELECT id FROM operation_attachments WHERE id=? AND is_deleted=0", (aid,)
        ).fetchone()
        if not att:
            return jsonify({'error': '附件不存在'}), 404
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        db.execute(
            """UPDATE operation_attachments
               SET archived=1, archived_at=?, archived_by=?, archive_reason=?
               WHERE id=?""",
            (now, archived_by, reason, aid))
        db.commit()
    return jsonify({'success': True})


@app.route('/api/attachments/<int:aid>/unarchive', methods=['POST'])
def unarchive_attachment(aid):
    """取消归档"""
    with get_db() as db:
        att = db.execute(
            "SELECT id FROM operation_attachments WHERE id=? AND is_deleted=0", (aid,)
        ).fetchone()
        if not att:
            return jsonify({'error': '附件不存在'}), 404
        db.execute(
            """UPDATE operation_attachments
               SET archived=0, archived_at=NULL, archived_by=NULL, archive_reason=''
               WHERE id=?""",
            (aid,))
        db.commit()
    return jsonify({'success': True})


@app.route('/api/attachments/stats')
def attachment_stats():
    """影像档案统计：总数、按类型统计、按月统计"""
    with get_db() as db:
        total = db.execute("SELECT COUNT(*) FROM operation_attachments WHERE is_deleted=0").fetchone()[0]
        archived_count = db.execute("SELECT COUNT(*) FROM operation_attachments WHERE is_deleted=0 AND archived=1").fetchone()[0]
        review_pending = db.execute(
            "SELECT COUNT(*) FROM operation_attachments WHERE is_deleted=0 AND review_required=1 AND review_status='pending'"
        ).fetchone()[0]
        by_category = {}
        for row in db.execute("SELECT category, COUNT(*) as cnt FROM operation_attachments WHERE is_deleted=0 GROUP BY category"):
            by_category[row['category']] = row['cnt']
        by_source = {}
        for row in db.execute("SELECT source_type, COUNT(*) as cnt FROM operation_attachments WHERE is_deleted=0 GROUP BY source_type"):
            by_source[row['source_type']] = row['cnt']
        by_month = []
        for row in db.execute("SELECT substr(created_at,1,7) as ym, COUNT(*) as cnt FROM operation_attachments WHERE is_deleted=0 GROUP BY ym ORDER BY ym DESC LIMIT 12"):
            by_month.append({'month': row['ym'], 'count': row['cnt']})
        return jsonify({
            'total': total,
            'archived': archived_count,
            'review_pending': review_pending,
            'by_category': by_category,
            'by_source': by_source,
            'by_month': by_month,
        })


def match_photo_requirement(watermark_text='', site_id=None, filename='', description=''):
    """按水印文字/文件名关键词匹配照片类型配置（v1 轻量方案，无需 OCR）。
    返回 dict 或 None。v2 可替换为多模态模型识别，仅改此函数即可。
    匹配逻辑：把水印文字+文件名+说明拼成一个字符串，逐条比对 photo_requirements.watermark_keyword
    （支持用 | 分隔多关键词）。命中即返回该配置，自动带出是否需要审核。"""
    text = f'{watermark_text or ""} {(filename or "")} {(description or "")}'.lower()
    if not text.strip():
        return None
    with get_db() as db:
        rows = db.execute(
            "SELECT id, item_name, review_required, watermark_keyword, category "
            "FROM photo_requirements WHERE watermark_keyword IS NOT NULL AND watermark_keyword != ''"
        ).fetchall()
        for r in rows:
            kw = (r['watermark_keyword'] or '').strip().lower()
            if not kw:
                continue
            hit = any(k.strip() and k.strip() in text for k in kw.split('|'))
            if hit:
                return {
                    'requirement_id': r['id'],
                    'recognized_category': r['item_name'],
                    'review_required': r['review_required'] or 0,
                    'category': r['category'] or '',
                    'match_status': 'auto',
                    'match_confidence': 0.8,
                }
    return None


@app.route('/api/upload/attachment', methods=['POST'])
def upload_attachment():
    """上传影像附件（带完善元信息）"""
    file = request.files.get('file')
    if not file:
        return jsonify({'error': '请选择文件'}), 400

    ext = os.path.splitext(file.filename or '.jpg')[1].lower() or '.jpg'
    image_exts = {'.jpg', '.jpeg', '.png', '.webp', '.bmp', '.gif'}
    if ext not in image_exts:
        return jsonify({'error': '仅支持图片格式'}), 400
    file_data = file.read()
    if len(file_data) > 20 * 1024 * 1024:
        return jsonify({'error': '文件大小超过20MB限制'}), 400

    # 生成唯一文件名
    fname = str(uuid.uuid4())[:12] + ext
    now = datetime.now()
    subdir = now.strftime('attachments/%Y/%m')
    stored_dir = os.path.join(UPLOAD_DIR, subdir)
    os.makedirs(stored_dir, exist_ok=True)
    stored_path = os.path.join(stored_dir, fname)
    with open(stored_path, 'wb') as f:
        f.write(file_data)
    url = f'/uploads/{subdir}/{fname}'

    # 水印解析 + 自动归类（v1：关键词匹配，无需 OCR）
    watermark_text = request.form.get('watermark_text', '') or ''
    match = match_photo_requirement(
        watermark_text,
        request.form.get('site_id', type=int),
        file.filename or '',
        request.form.get('description', ''))
    req_id = match['requirement_id'] if match else None
    rec_cat = match['recognized_category'] if match else ''
    match_status = match['match_status'] if match else 'manual'
    match_conf = match['match_confidence'] if match else None
    review_required = match['review_required'] if match else 0
    auto_category = match['category'] if match else request.form.get('category', '')

    # 流程外资料（试剂配置/车辆里程加油/养护记录）仅归档，不强制进审核队列
    # 即便水印关键词命中配置，也不置 review_required，避免与在流程照片审核链混淆
    _src_type = request.form.get('source_type', '') or ''
    if _src_type in ('reagent', 'vehicle', 'maintenance'):
        review_required = 0
        req_id = None
        rec_cat = ''
        match_status = 'manual'
        match_conf = None
        auto_category = request.form.get('category', '') or ''

    # 写入数据库
    # 清洗乱码：替换字符 U+FFFD 或控制字符直接置空，避免污染档案展示
    def _clean_text(s):
        if not s:
            return ''
        return ''.join(c for c in s if c not in {'\x00', '\x01', '\x02', '\x03', '\x04', '\x05', '\x06', '\x07', '\x08', '\x0b', '\x0c', '\x0e', '\x0f', '\x10', '\x11', '\x12', '\x13', '\x14', '\x15', '\x16', '\x17', '\x18', '\x19', '\x1a', '\x1b', '\x1c', '\x1d', '\x1e', '\x1f'} and ord(c) != 0xFFFD)

    description = _clean_text(request.form.get('description', ''))
    watermark_text = _clean_text(watermark_text)

    with get_db() as db:
        db.execute("""INSERT INTO operation_attachments
            (filename, stored_path, file_type, mime_type, file_size, description,
             source_type, source_id, site_id, uploader_id, uploader_name,
             gps_lat, gps_lng, taken_at, category,
             watermark_text, recognized_category, match_status, match_confidence,
             review_required, requirement_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (file.filename or fname, url,
             'video' if ext in ('.mp4','.mov') else 'image',
             ext, len(file_data),
             description,
             request.form.get('source_type',''),
             request.form.get('source_id', 0, type=int),
             request.form.get('site_id', type=int),
             request.form.get('uploader_id', type=int),
             request.form.get('uploader_name',''),
             request.form.get('gps_lat', type=float),
             request.form.get('gps_lng', type=float),
             request.form.get('taken_at') or now.strftime('%Y-%m-%d %H:%M:%S'),
             auto_category,
             watermark_text, rec_cat, match_status, match_conf,
             review_required, req_id))
        db.commit()
        aid = db.execute("SELECT last_insert_rowid()").fetchone()[0]

    return jsonify({'success': True, 'id': aid, 'url': url,
                   'match': match, 'review_required': review_required})


@app.route('/api/sites/<int:site_id>/schemes')
def get_site_schemes(site_id):
    with get_db() as db:
        schemes = db.execute("SELECT s.*,(SELECT COUNT(*) FROM inspection_scheme_items i WHERE i.scheme_id=s.id) as item_count FROM inspection_schemes s WHERE s.site_id=? ORDER BY CASE s.period WHEN 'daily' THEN 1 WHEN 'weekly' THEN 2 ELSE 3 END",(site_id,)).fetchall()
        return jsonify([dict(r) for r in schemes])

@app.route('/api/schemes/<int:scheme_id>')
def get_scheme_detail(scheme_id):
    with get_db() as db:
        scheme = db.execute("SELECT * FROM inspection_schemes WHERE id=?",(scheme_id,)).fetchone()
        if not scheme: return jsonify({'error':'方案不存在'}),404
        items = db.execute("SELECT * FROM inspection_scheme_items WHERE scheme_id=? ORDER BY sort_order",(scheme_id,)).fetchall()
        result = dict(scheme); result['items']=[dict(r) for r in items]
        return jsonify(result)

@app.route('/api/schemes/<int:scheme_id>', methods=['PUT'])
def update_scheme(scheme_id):
    data = request.json
    with get_db() as db:
        scheme = db.execute("SELECT * FROM inspection_schemes WHERE id=?",(scheme_id,)).fetchone()
        if not scheme: return jsonify({'error':'方案不存在'}),404
        if 'name' in data: db.execute("UPDATE inspection_schemes SET name=?,updated_at=datetime('now','localtime') WHERE id=?",(data['name'],scheme_id))
        if 'items' in data:
            db.execute("DELETE FROM inspection_scheme_items WHERE scheme_id=?",(scheme_id,))
            for idx,item in enumerate(data['items']):
                db.execute("INSERT INTO inspection_scheme_items (scheme_id,category,check_item,sort_order,is_required) VALUES (?,?,?,?,?)",(scheme_id,item.get('category',''),item.get('check_item',''),idx,item.get('is_required',1)))
        db.commit()
        return jsonify({'success':True})

@app.route('/api/schemes/<int:scheme_id>/items', methods=['POST'])
def add_scheme_item(scheme_id):
    data = request.json
    item_name = data.get('check_item','').strip()
    if not item_name: return jsonify({'error':'检查项不能为空'}),400
    with get_db() as db:
        max_order = db.execute("SELECT COALESCE(MAX(sort_order),-1)+1 as n FROM inspection_scheme_items WHERE scheme_id=?",(scheme_id,)).fetchone()['n']
        db.execute("INSERT INTO inspection_scheme_items (scheme_id,category,check_item,sort_order) VALUES (?,?,?,?)",(scheme_id,data.get('category','自定义'),item_name,max_order))
        db.execute("UPDATE inspection_schemes SET updated_at=datetime('now','localtime') WHERE id=?",(scheme_id,))
        db.commit()
        return jsonify({'success':True})

@app.route('/api/schemes/items/<int:item_id>', methods=['DELETE'])
def delete_scheme_item_ep(item_id):
    with get_db() as db:
        db.execute("DELETE FROM inspection_scheme_items WHERE id=?",(item_id,))
        db.commit()
        return jsonify({'success':True})

@app.route('/api/inspections/auto-generate', methods=['POST'])
def auto_generate_inspections():
    """按频次分层的智能排程引擎（替代旧的日期轮询方案）
    
    请求参数（可选）：
    - user_id: 指定某人的组（默认全部分配）
    - period: daily/weekly/monthly (默认monthly)
    - start_date: 起始日期（默认今天）
    - end_date: 截止日期（默认+30天）
    """
    data = request.get_json(silent=True) or {}
    period = data.get('period', 'monthly')
    start_str = data.get('start_date', datetime.now().strftime('%Y-%m-%d'))
    user_id = data.get('user_id')
    force = data.get('force', False)  # 是否覆盖已存在的计划
    
    start = datetime.strptime(start_str, '%Y-%m-%d')
    if period == 'daily':
        end = start + timedelta(days=1)
    elif period == 'weekly':
        end = start + timedelta(days=7)
    else:
        end = start + timedelta(days=30)
    end_str = end.strftime('%Y-%m-%d')
    
    # 频次映射：period -> 应包含的frequency_level
    freq_map = {
        'high': ['high'],
        'mid': ['high', 'mid'],
        'low': ['high', 'mid', 'low'],
        'annual': ['high', 'mid', 'low', 'annual'],
    }
    applicable = freq_map.get(period, ['high', 'mid'])
    
    with get_db() as db:
        rows = db.execute("""
            SELECT si.*, s.name as scheme_name, s.site_id, s.period as speriod
            FROM inspection_scheme_items si
            JOIN inspection_schemes s ON si.scheme_id=s.id
            WHERE s.status='active'
            AND (si.frequency_level IS NULL OR si.frequency_level IN ({vals}))
            ORDER BY s.site_id, si.sort_order
        """.format(vals=','.join('?' * len(applicable))), applicable).fetchall()
        
        if not rows:
            # 降级：从已有inspection_plans读取站点列表作为参考
            plan_sites = db.execute("SELECT DISTINCT site_id FROM inspection_plans").fetchall()
            if not plan_sites:
                plan_sites = db.execute("SELECT id as site_id FROM sites WHERE id IN (1,5,108,193)").fetchall()
            for ps in plan_sites:
                # 先确保该站点有活跃方案
                scheme = db.execute("SELECT id FROM inspection_schemes WHERE site_id=? AND status='active' LIMIT 1", (ps['site_id'],)).fetchone()
                if not scheme:
                    db.execute("INSERT INTO inspection_schemes (site_id,period,name) VALUES (?,?,?)",
                               (ps['site_id'], period, f"站{ps['site_id']}巡检方案"))
                    scheme_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
                else:
                    scheme_id = scheme['id']
                # 检查是否有带频次的方案项
                existing_items = db.execute("SELECT id FROM inspection_scheme_items WHERE scheme_id=? AND frequency_level IS NOT NULL LIMIT 1", (scheme_id,)).fetchone()
                if not existing_items:
                    # 创建默认方案项
                    db.execute("DELETE FROM inspection_scheme_items WHERE scheme_id=?", (scheme_id,))
                    default_items = [
                        ('水位观测', 'high', 1), ('设备状态确认', 'high', 2), ('传感器外观清洁', 'high', 3),
                        ('数据通讯检查', 'high', 4), ('电池电压检查', 'mid', 5), ('太阳能板检查', 'mid', 6),
                        ('机箱密封性检查', 'mid', 7), ('站院环境维护', 'mid', 8), ('翻斗雨量计校准', 'low', 9),
                        ('水位计精度校验', 'low', 10), ('全面校准试验', 'annual', 11),
                    ]
                    for item_name, freq, order in default_items:
                        if freq in applicable:
                            db.execute("INSERT INTO inspection_scheme_items (scheme_id,category,check_item,frequency_level,sort_order) VALUES (?,'常规检查',?,?,?)",
                                       (scheme_id, item_name, freq, order))
                rows.extend(db.execute("""
                    SELECT si.*, s.name as scheme_name, s.site_id, s.period as speriod
                    FROM inspection_scheme_items si
                    JOIN inspection_schemes s ON si.scheme_id=s.id
                    WHERE s.site_id=? AND s.status='active'
                """, (ps['site_id'],)).fetchall() or [])
        
        if not rows:
            return jsonify({'success': False, 'error': '没有活跃的巡检方案项，请先在方案中配置检查项', 'generated': 0})
        
        # 按site_id分组
        site_groups = {}
        for r in rows:
            sid = r['site_id']
            site_groups.setdefault(sid, []).append(r)
        
        # 获取所有运维人员及其站点分配
        operators = db.execute("""
            SELECT u.id, u.real_name FROM users u WHERE u.role='operator' ORDER BY u.id
        """).fetchall()
        
        if user_id:
            operators = [op for op in operators if op['id'] == user_id]
        
        generated = 0
        for op in operators:
            user_sites = db.execute("SELECT site_id FROM user_sites WHERE user_id=?", (op['id'],)).fetchall()
            # 按站点打包：同一操作员的所有站点合并为一个计划
            op_site_ids = [us['site_id'] for us in user_sites]
            all_items = []
            check_items_set = set()
            for sid in op_site_ids:
                items = site_groups.get(sid, [])
                if not items:
                    continue
                all_items.append((sid, items))
                for item in items:
                    check_items_set.add(item['check_item'])
            if not all_items:
                continue
            # 检查是否已存在该操作员该时段的计划
            if not force:
                exist = db.execute(
                    "SELECT p.id FROM inspection_plans p JOIN plan_sites ps ON p.id=ps.plan_id WHERE ps.site_id IN ({}) AND p.start_date=? AND p.status='pending' LIMIT 1".format(
                        ','.join('?' * len(op_site_ids))
                    ), op_site_ids + [start_str]
                ).fetchone()
                if exist:
                    continue
            plan_name = f"{period}巡检-{op['real_name']}"
            first_sid = op_site_ids[0] if op_site_ids else 0
            db.execute("""
                INSERT INTO inspection_plans (plan_name,site_id,type,start_date,end_date,status)
                VALUES (?,?,?,?,?,?)
            """, (plan_name, first_sid, period, start_str, end_str, 'pending'))
            plan_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            # 写入所有站点到 plan_sites，并生成每个站点的检查项
            for sid, items in all_items:
                db.execute("INSERT OR IGNORE INTO plan_sites (plan_id, site_id) VALUES (?,?)", (plan_id, sid))
                for item in items:
                    db.execute("""
                        INSERT INTO inspection_tasks (plan_id,site_id,check_item)
                        VALUES (?,?,?)
                    """, (plan_id, sid, item['check_item']))
                generated += 1
        
        db.commit()
        msg = f"已生成 {generated} 个巡检计划"
        if generated == 0:
            msg = "该时段计划已存在，无需重复生成"
        return jsonify({'success': True, 'generated': generated, 'message': msg})


# ===================== 移动巡检方案新增 API =====================

@app.route('/api/inspections/skip', methods=['POST'])
@login_required
def skip_inspection_item():
    """跳过某项检查（记录跳过原因）"""
    data = request.get_json(silent=True) or {}
    with get_db() as db:
        db.execute("""
            INSERT INTO inspection_skip_logs (plan_id,task_id,site_id,check_item,reason,skip_type)
            VALUES (?,?,?,?,?,?)
        """, (data.get('plan_id'), data.get('task_id'), data.get('site_id'),
              data.get('check_item',''), data.get('reason',''), data.get('skip_type','user')))
        # 更新跳过计数
        exist = db.execute(
            "SELECT id, skip_count FROM inspection_skip_logs WHERE plan_id=? AND check_item=? ORDER BY id DESC LIMIT 1",
            (data['plan_id'], data['check_item'])
        ).fetchone()
        if exist:
            db.execute("UPDATE inspection_skip_logs SET skip_count=skip_count+1 WHERE id=?", (exist['id'],))
        db.commit()
        return jsonify({'success': True, 'message': '已记录跳过'})

@app.route('/api/inspections/skip/history')
@login_required
def get_skip_history():
    """查看跳过记录"""
    site_id = request.args.get('site_id', type=int)
    plan_id = request.args.get('plan_id', type=int)
    with get_db() as db:
        q = "SELECT * FROM inspection_skip_logs WHERE 1=1"
        params = []
        if site_id:
            q += " AND site_id=?"
            params.append(site_id)
        if plan_id:
            q += " AND plan_id=?"
            params.append(plan_id)
        q += " ORDER BY created_at DESC LIMIT 50"
        rows = db.execute(q, params).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route('/api/calibration-templates')
@login_required
def get_calibration_templates():
    """获取校准模板列表"""
    device_type = request.args.get('device_type', '')
    with get_db() as db:
        q = "SELECT * FROM calibration_templates WHERE 1=1"
        params = []
        if device_type:
            q += " AND device_type=?"
            params.append(device_type)
        q += " ORDER BY sort_order, category"
        rows = db.execute(q, params).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route('/api/calibration-templates', methods=['POST'])
@login_required
def create_calibration_template():
    """创建校准模板"""
    data = request.get_json(silent=True) or {}
    with get_db() as db:
        db.execute("""
            INSERT INTO calibration_templates (device_type,template_name,fields,calculations,thresholds,category,sort_order)
            VALUES (?,?,?,?,?,?,?)
        """, (data['device_type'], data['template_name'],
              data.get('fields','[]'), data.get('calculations','[]'),
              data.get('thresholds','[]'), data.get('category',''), data.get('sort_order',0)))
        db.commit()
        return jsonify({'success': True, 'id': db.execute("SELECT last_insert_rowid()").fetchone()[0]})

@app.route('/api/inspections/photo-types', methods=['GET', 'POST'])
@login_required
def manage_photo_types():
    """管理照片类型配置"""
    with get_db() as db:
        if request.method == 'POST':
            data = request.get_json(silent=True) or {}
            db.execute("""
                INSERT INTO inspection_photo_types (plan_id,site_type,photo_type,label,min_count,sort_order)
                VALUES (?,?,?,?,?,?)
            """, (data.get('plan_id'), data.get('site_type',''), data['photo_type'],
                  data['label'], data.get('min_count',1), data.get('sort_order',0)))
            db.commit()
            return jsonify({'success': True})
        else:
            plan_id = request.args.get('plan_id', type=int)
            site_type = request.args.get('site_type', '')
            q = "SELECT * FROM inspection_photo_types WHERE 1=1"
            params = []
            if plan_id:
                q += " AND plan_id=?"
                params.append(plan_id)
            if site_type:
                q += " AND site_type=?"
                params.append(site_type)
            q += " ORDER BY sort_order"
            rows = db.execute(q, params).fetchall()
            return jsonify([dict(r) for r in rows])


# ===================== 通知系统 API =====================

@app.route('/api/notifications')
@login_required
def get_notifications():
    """获取当前用户的通知列表"""
    user = g.current_user
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 50, type=int)
    offset = (page - 1) * limit
    with get_db() as db:
        rows = db.execute(
            "SELECT * FROM notifications WHERE user_id=? ORDER BY created_at DESC LIMIT ? OFFSET ?",
            (user['id'], limit, offset)
        ).fetchall()
        unread = db.execute(
            "SELECT COUNT(*) FROM notifications WHERE user_id=? AND is_read=0",
            (user['id'],)
        ).fetchone()[0]
        return jsonify({'notifications': [dict(r) for r in rows], 'unread_count': unread})

@app.route('/api/notifications/unread-count')
@login_required
def unread_notification_count():
    """获取未读通知数量"""
    user = g.current_user
    with get_db() as db:
        cnt = db.execute(
            "SELECT COUNT(*) FROM notifications WHERE user_id=? AND is_read=0",
            (user['id'],)
        ).fetchone()[0]
        return jsonify({'count': cnt})

@app.route('/api/notifications/<int:nid>/read', methods=['PUT'])
@login_required
def mark_notification_read(nid):
    """标记单条通知为已读"""
    user = g.current_user
    with get_db() as db:
        db.execute(
            "UPDATE notifications SET is_read=1 WHERE id=? AND user_id=?",
            (nid, user['id'])
        )
        db.commit()
        return jsonify({'success': True})

@app.route('/api/notifications/read-all', methods=['PUT'])
@login_required
def mark_all_notifications_read():
    """标记所有通知为已读"""
    user = g.current_user
    with get_db() as db:
        db.execute(
            "UPDATE notifications SET is_read=1 WHERE user_id=? AND is_read=0",
            (user['id'],)
        )
        db.commit()
        return jsonify({'success': True})


# --- Workorder management ---
@app.route('/api/workorders/<order_no>', methods=['DELETE'])
def delete_workorder(order_no):
    """删除工单（支持待受理、已受理、处置中、审核中或已完成的工单）"""
    with get_db() as db:
        cur = db.execute('SELECT status FROM work_orders WHERE order_no=?', (order_no,)).fetchone()
        if not cur:
            return jsonify({'error': 'not found'}), 404
        if cur['status'] not in ('pending', 'accepted', 'in_progress', 'reviewing', 'closed'):
            return jsonify({'error': '当前状态不允许删除'}), 400
        db.execute('DELETE FROM work_orders WHERE order_no=?', (order_no,))
        db.execute("DELETE FROM timeline_events WHERE source_type='workorder' AND source_id=?", (order_no,))
        db.commit()
        return jsonify({'success': True, 'message': '工单已删除'})
# --- Maintenance Templates ---
@app.route('/api/maintenance/templates')
def get_maintenance_templates():
    """返回所有运维模板"""
    with get_db() as db:
        rows = db.execute("SELECT * FROM maintenance_templates ORDER BY sort_order").fetchall()
        result = []
        for r in rows:
            d = dict(r)
            if d['check_items']:
                try:
                    d['check_items'] = json.loads(d['check_items'])
                except:
                    d['check_items'] = []
            result.append(d)
        return jsonify(result)


@app.route('/api/maintenance/templates', methods=['POST'])
@login_required
def create_maintenance_template():
    """新建运维模板"""
    data = request.get_json(force=True)
    title = (data.get('title') or '').strip()
    if not title:
        return jsonify({'error': '模板名称不能为空'}), 400
    category = (data.get('category') or '').strip()
    sub_category = (data.get('sub_category') or '').strip()
    frequency = data.get('frequency') or 'monthly'
    description = (data.get('description') or '').strip()
    standard = (data.get('standard') or '').strip()
    check_items = data.get('check_items')
    estimated_hours = data.get('estimated_hours')
    photo_required = 1 if data.get('photo_required') else 0

    if isinstance(check_items, list):
        check_items = json.dumps(check_items, ensure_ascii=False)

    with get_db() as db:
        cur = db.execute(
            """INSERT INTO maintenance_templates
               (title, category, sub_category, frequency, description, standard, check_items, estimated_hours, photo_required)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (title, category, sub_category, frequency, description, standard, check_items, estimated_hours, photo_required)
        )
        db.commit()
    return jsonify({'success': True, 'id': cur.lastrowid, 'message': '模板创建成功'})


@app.route('/api/maintenance/templates/<int:tid>', methods=['PUT'])
@login_required
def update_maintenance_template(tid):
    """编辑运维模板"""
    data = request.get_json(force=True)
    with get_db() as db:
        existing = db.execute("SELECT id FROM maintenance_templates WHERE id=?", (tid,)).fetchone()
        if not existing:
            return jsonify({'error': '模板不存在'}), 404

        fields = []
        values = []
        for col in ['title', 'category', 'sub_category', 'frequency', 'description', 'standard', 'estimated_hours', 'photo_required']:
            if col in data:
                if col == 'photo_required':
                    fields.append(f"{col}=?")
                    values.append(1 if data[col] else 0)
                else:
                    fields.append(f"{col}=?")
                    values.append(data[col])
        if 'check_items' in data:
            fields.append("check_items=?")
            ci = data['check_items']
            values.append(json.dumps(ci, ensure_ascii=False) if isinstance(ci, list) else ci)

        if not fields:
            return jsonify({'error': '没有可更新的字段'}), 400
        values.append(tid)
        db.execute(f"UPDATE maintenance_templates SET {', '.join(fields)} WHERE id=?", values)
        db.commit()
    return jsonify({'success': True, 'message': '模板已更新'})


@app.route('/api/maintenance/templates/<int:tid>', methods=['DELETE'])
@login_required
def delete_maintenance_template(tid):
    """删除运维模板"""
    with get_db() as db:
        existing = db.execute("SELECT id FROM maintenance_templates WHERE id=?", (tid,)).fetchone()
        if not existing:
            return jsonify({'error': '模板不存在'}), 404
        db.execute("DELETE FROM maintenance_templates WHERE id=?", (tid,))
        db.commit()
    return jsonify({'success': True, 'message': '模板已删除'})


# --- Maintenance Plans ---
@app.route('/api/maintenance/plans')
def get_maintenance_plans():
    with get_db() as db:
        status = request.args.get('status')
        category = request.args.get('category')
        q = "SELECT mp.*, s.name as site_name, s.code as site_code FROM maintenance_plans mp LEFT JOIN sites s ON mp.site_id=s.id"
        params = []
        conds = []
        if status:
            conds.append("mp.status=?")
            params.append(status)
        if category:
            conds.append("mp.category=?")
            params.append(category)
        if conds:
            q += " WHERE " + " AND ".join(conds)
        q += " ORDER BY mp.due_date ASC"
        rows = db.execute(q, params).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route('/api/maintenance/plans', methods=['POST'])
def create_maintenance_plan():
    data = request.json
    template_id = data.get('template_id')
    with get_db() as db:
        if template_id:
            # 从模板自动填充
            tpl = db.execute("SELECT * FROM maintenance_templates WHERE id=?", (template_id,)).fetchone()
            if tpl:
                site_name = db.execute("SELECT name FROM sites WHERE id=?", (data['site_id'],)).fetchone()
                site_label = site_name['name'] if site_name else ''
                plan_name = data.get('plan_name') or f"{tpl['title']}-{site_label}"
                category = tpl['category']
                frequency = tpl['frequency']
                sub_category = tpl['sub_category']
                remark = data.get('remark') or tpl['description']
                cur = db.execute(
                    "INSERT INTO maintenance_plans (site_id,plan_name,category,frequency,due_date,assignee,template_id,sub_category,remark) VALUES (?,?,?,?,?,?,?,?,?)",
                    (data['site_id'], plan_name, category, frequency, data.get('due_date'), data.get('assignee'), template_id, sub_category, remark)
                )
            else:
                return jsonify({'error': 'template not found'}), 404
        else:
            # 无模板的传统创建方式
            cur = db.execute(
                "INSERT INTO maintenance_plans (site_id,plan_name,category,frequency,due_date,assignee) VALUES (?,?,?,?,?,?)",
                (data['site_id'], data['plan_name'], data['category'], data.get('frequency','monthly'), data.get('due_date'), data.get('assignee'))
            )
        db.commit()

@app.route('/api/maintenance/plans/<int:plan_id>/complete', methods=['PUT'])
def complete_maintenance_plan(plan_id):
    data = request.get_json(silent=True) or {}
    check_results = data.get('check_results')
    with get_db() as db:
        if check_results:
            db.execute("UPDATE maintenance_plans SET status='completed', completed_at=datetime('now','localtime'), check_results=? WHERE id=?",
                       (json.dumps(check_results, ensure_ascii=False), plan_id))
        else:
            db.execute("UPDATE maintenance_plans SET status='completed', completed_at=datetime('now','localtime') WHERE id=?", (plan_id,))
        db.commit()
        # 记录时间线
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator) VALUES (?,?,?,?)",
                   ('maintenance', plan_id, 'completed', '系统'))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/maintenance/plans/<int:plan_id>/urge', methods=['POST'])
def urge_maintenance(plan_id):
    with get_db() as db:
        db.execute("UPDATE maintenance_plans SET urge_count=COALESCE(urge_count,0)+1, last_urged_at=datetime('now','localtime') WHERE id=?", (plan_id,))
        db.commit()
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator) VALUES (?,?,?,?)",
                   ('maintenance', plan_id, 'urged', '系统'))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/maintenance/stats')
def maintenance_stats():
    """运维统计：今日待办总数/各分类统计"""
    with get_db() as db:
        today = datetime.now().strftime('%Y-%m-%d')
        total_pending = db.execute("SELECT COUNT(*) as c FROM maintenance_plans WHERE status='pending'").fetchone()['c']
        overdue = db.execute("SELECT COUNT(*) as c FROM maintenance_plans WHERE status='pending' AND due_date < ?", (today,)).fetchone()['c']
        review_pending = db.execute("SELECT COUNT(*) as c FROM data_reviews WHERE status='smart_reviewed' AND smart_result='suspicious'").fetchone()['c']
        return jsonify({'total_pending': total_pending, 'overdue': overdue, 'review_pending': review_pending})

@app.route('/api/maintenance/plans/<int:plan_id>', methods=['PUT'])
def update_maintenance_plan(plan_id):
    """修改运维计划"""
    data = request.json
    with get_db() as db:
        cur = db.execute("SELECT * FROM maintenance_plans WHERE id=?", (plan_id,)).fetchone()
        if not cur:
            return jsonify({'error': 'not found'}), 404
        updates = []
        params = []
        for field in ['plan_name','category','frequency','due_date','site_id','assignee']:
            if field in data:
                updates.append(f"{field}=?")
                params.append(data[field])
        if not updates:
            return jsonify({'error': 'no fields to update'}), 400
        params.append(plan_id)
        db.execute(f"UPDATE maintenance_plans SET {','.join(updates)} WHERE id=?", params)
        db.commit()

@app.route('/api/maintenance/plans/<int:plan_id>', methods=['DELETE'])
def delete_maintenance_plan(plan_id):
    """删除运维计划"""
    with get_db() as db:
        db.execute("DELETE FROM maintenance_plans WHERE id=?", (plan_id,))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/maintenance/plans/<int:plan_id>/review', methods=['PUT'])
def review_maintenance_plan(plan_id):
    """审核运维计划"""
    data = request.json
    review_result = data.get('review_result', 'pending')
    review_comment = data.get('review_comment', '')
    operator = data.get('operator', '系统')
    with get_db() as db:
        db.execute("UPDATE maintenance_plans SET review_status=?, review_comment=? WHERE id=?",
                   (review_result, review_comment, plan_id))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('maintenance', plan_id, 'reviewed', operator, f'审核结果:{review_result} 意见:{review_comment}'))
        db.commit()
        return jsonify({'success': True})

# ===================== Inspection V2 API =====================

# --- 方案模板 CRUD ---

@app.route('/api/inspection-v2/templates')
def v2_get_templates():
    """获取方案模板列表"""
    category = request.args.get('category', '')
    with get_db() as db:
        q = "SELECT t.*, (SELECT COUNT(*) FROM inspection_template_items WHERE template_id=t.id) as item_count FROM inspection_templates t WHERE 1=1"
        params = []
        if category:
            q += " AND t.category=?"
            params.append(category)
        q += " ORDER BY t.sort_order, t.id"
        rows = db.execute(q, params).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route('/api/inspection-v2/templates', methods=['POST'])
def v2_create_template():
    """创建方案模板"""
    data = request.get_json(silent=True) or {}
    name = data.get('template_name', '').strip()
    category = data.get('category', '').strip()
    frequency = data.get('frequency', 'monthly')
    desc = data.get('description', '')
    items = data.get('items', [])
    if not name or not category:
        return jsonify({'error': '模板名称和分类不能为空'}), 400
    with get_db() as db:
        db.execute("INSERT INTO inspection_templates (template_name,category,frequency,description) VALUES (?,?,?,?)",
                   (name, category, frequency, desc))
        tid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        for i, item in enumerate(items):
            db.execute("""
                INSERT INTO inspection_template_items (template_id,item_name,category,frequency_level,photo_required,sort_order)
                VALUES (?,?,?,?,?,?)
            """, (tid, item.get('item_name',''), item.get('category',''), item.get('frequency_level','mid'),
                  1 if item.get('photo_required') else 0, i+1))
        db.commit()
        return jsonify({'id': tid, 'success': True})

@app.route('/api/inspection-v2/templates/<int:tid>', methods=['PUT'])
def v2_update_template(tid):
    """更新方案模板"""
    data = request.get_json(silent=True) or {}
    with get_db() as db:
        tpl = db.execute("SELECT id FROM inspection_templates WHERE id=?", (tid,)).fetchone()
        if not tpl:
            return jsonify({'error': '模板不存在'}), 404
        fields = []
        params = []
        for key in ('template_name', 'category', 'frequency', 'description', 'status', 'sort_order'):
            if key in data:
                fields.append(f"{key}=?")
                params.append(data[key])
        if fields:
            params.append(tid)
            db.execute(f"UPDATE inspection_templates SET {','.join(fields)} WHERE id=?", params)
        # 如果提供了items，全量替换
        if 'items' in data:
            db.execute("DELETE FROM inspection_template_items WHERE template_id=?", (tid,))
            for i, item in enumerate(data['items']):
                db.execute("""
                    INSERT INTO inspection_template_items (template_id,item_name,category,frequency_level,photo_required,sort_order)
                    VALUES (?,?,?,?,?,?)
                """, (tid, item.get('item_name',''), item.get('category',''), item.get('frequency_level','mid'),
                      1 if item.get('photo_required') else 0, i+1))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/inspection-v2/templates/<int:tid>', methods=['DELETE'])
def v2_delete_template(tid):
    """删除方案模板"""
    with get_db() as db:
        # 删除关联的配置、排程、检查项
        db.execute("DELETE FROM inspection_template_items WHERE template_id=?", (tid,))
        db.execute("DELETE FROM inspection_configs WHERE template_id=?", (tid,))
        db.execute("DELETE FROM inspection_schedules WHERE template_id=?", (tid,))
        db.execute("DELETE FROM inspection_templates WHERE id=?", (tid,))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/inspection-v2/templates/<int:tid>/items')
def v2_get_template_items(tid):
    """获取模板检查项列表"""
    with get_db() as db:
        rows = db.execute("SELECT * FROM inspection_template_items WHERE template_id=? ORDER BY sort_order", (tid,)).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route('/api/inspection-v2/templates/<int:tid>/items', methods=['POST'])
def v2_add_template_item(tid):
    """添加模板检查项"""
    data = request.get_json(silent=True) or {}
    item_name = data.get('item_name', '').strip()
    if not item_name:
        return jsonify({'error': '检查项名称不能为空'}), 400
    with get_db() as db:
        max_sort = db.execute("SELECT MAX(sort_order) FROM inspection_template_items WHERE template_id=?", (tid,)).fetchone()[0] or 0
        db.execute("""
            INSERT INTO inspection_template_items (template_id,item_name,category,frequency_level,photo_required,sort_order)
            VALUES (?,?,?,?,?,?)
        """, (tid, item_name, data.get('category',''), data.get('frequency_level','mid'),
              1 if data.get('photo_required') else 0, max_sort + 1))
        db.commit()
        return jsonify({'id': db.execute("SELECT last_insert_rowid()").fetchone()[0], 'success': True})

@app.route('/api/inspection-v2/templates/<int:tid>/items/<int:item_id>', methods=['PUT'])
def v2_update_template_item(tid, item_id):
    """更新模板检查项"""
    data = request.get_json(silent=True) or {}
    with get_db() as db:
        fields = []
        params = []
        for key in ('item_name', 'category', 'frequency_level', 'sort_order'):
            if key in data:
                fields.append(f"{key}=?")
                params.append(data[key])
        if 'photo_required' in data:
            fields.append("photo_required=?")
            params.append(1 if data['photo_required'] else 0)
        if fields:
            params.append(item_id)
            db.execute(f"UPDATE inspection_template_items SET {','.join(fields)} WHERE id=? AND template_id=?", params + [tid])
        db.commit()
        return jsonify({'success': True})

@app.route('/api/inspection-v2/templates/<int:tid>/items/<int:item_id>', methods=['DELETE'])
def v2_delete_template_item(tid, item_id):
    """删除模板检查项"""
    with get_db() as db:
        db.execute("DELETE FROM inspection_template_items WHERE id=? AND template_id=?", (item_id, tid))
        db.commit()
        return jsonify({'success': True})

# --- 巡检配置 CRUD + 匹配引擎 ---

@app.route('/api/inspection-v2/configs')
def v2_get_configs():
    """获取巡检配置列表"""
    site_type = request.args.get('site_type', '')
    with get_db() as db:
        q = """
            SELECT ic.*, it.template_name, it.category as tpl_category, it.frequency as tpl_frequency,
                   (SELECT COUNT(*) FROM inspection_template_items WHERE template_id=ic.template_id) as item_count
            FROM inspection_configs ic
            JOIN inspection_templates it ON ic.template_id = it.id
            WHERE 1=1
        """
        params = []
        if site_type:
            q += " AND ic.site_type=?"
            params.append(site_type)
        q += " ORDER BY ic.site_type, it.sort_order"
        rows = db.execute(q, params).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route('/api/inspection-v2/configs', methods=['POST'])
def v2_create_config():
    """创建巡检配置规则"""
    data = request.get_json(silent=True) or {}
    site_type = data.get('site_type', '').strip()
    template_id = data.get('template_id')
    if not site_type or not template_id:
        return jsonify({'error': '站点类型和模板不能为空'}), 400
    with get_db() as db:
        # 检查是否已存在相同配置
        existing = db.execute("SELECT id FROM inspection_configs WHERE site_type=? AND template_id=?",
                              (site_type, template_id)).fetchone()
        if existing:
            return jsonify({'error': '该站点类型已配置此模板'}), 409
        db.execute("""
            INSERT INTO inspection_configs (site_type,device_types,template_id,is_active,remark)
            VALUES (?,?,?,?,?)
        """, (site_type, _json.dumps(data.get('device_types', []), ensure_ascii=False),
              template_id, 1, data.get('remark', '')))
        db.commit()
        return jsonify({'id': db.execute("SELECT last_insert_rowid()").fetchone()[0], 'success': True})

@app.route('/api/inspection-v2/configs/<int:cid>', methods=['PUT'])
def v2_update_config(cid):
    """更新巡检配置"""
    data = request.get_json(silent=True) or {}
    with get_db() as db:
        fields = []
        params = []
        for key in ('site_type', 'is_active', 'remark'):
            if key in data:
                fields.append(f"{key}=?")
                params.append(data[key])
        if 'device_types' in data:
            fields.append("device_types=?")
            params.append(_json.dumps(data['device_types'], ensure_ascii=False))
        if 'template_id' in data:
            fields.append("template_id=?")
            params.append(data['template_id'])
        if fields:
            params.append(cid)
            db.execute(f"UPDATE inspection_configs SET {','.join(fields)} WHERE id=?", params)
        db.commit()
        return jsonify({'success': True})

@app.route('/api/inspection-v2/configs/<int:cid>', methods=['DELETE'])
def v2_delete_config(cid):
    """删除巡检配置"""
    with get_db() as db:
        db.execute("DELETE FROM inspection_configs WHERE id=?", (cid,))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/inspection-v2/configs/match')
def v2_match_configs():
    """匹配引擎：给定站点，返回适配的所有模板+检查项"""
    site_id = request.args.get('site_id', type=int)
    if not site_id:
        return jsonify({'error': '缺少site_id参数'}), 400
    with get_db() as db:
        site = db.execute("SELECT id, type, name FROM sites WHERE id=?", (site_id,)).fetchone()
        if not site:
            return jsonify({'error': '站点不存在'}), 404
        # 获取站点设备类型
        dev_types = [r['device_type'] for r in db.execute(
            "SELECT DISTINCT device_type FROM device_shadows WHERE site_id=?", (site_id,)).fetchall()]
        # 查找匹配的配置
        configs = db.execute("""
            SELECT ic.*, it.template_name, it.category, it.frequency, it.description
            FROM inspection_configs ic
            JOIN inspection_templates it ON ic.template_id = it.id
            WHERE ic.site_type=? AND ic.is_active=1
            ORDER BY it.sort_order
        """, (site['type'],)).fetchall()
        result = []
        for cfg in configs:
            # 设备类型过滤（如果配置指定了device_types）
            cfg_dev = cfg['device_types']
            if cfg_dev and cfg_dev != '[]' and cfg_dev != '':
                try:
                    required = _json.loads(cfg_dev)
                    if required and not any(d in dev_types for d in required):
                        continue
                except:
                    pass
            # 获取模板检查项
            items = db.execute("""
                SELECT * FROM inspection_template_items WHERE template_id=? ORDER BY sort_order
            """, (cfg['template_id'],)).fetchall()
            result.append({
                'config_id': cfg['id'],
                'template_id': cfg['template_id'],
                'template_name': cfg['template_name'],
                'category': cfg['category'],
                'frequency': cfg['frequency'],
                'description': cfg['description'],
                'items': [dict(it) for it in items],
            })
        return jsonify({
            'site_id': site_id,
            'site_name': site['name'],
            'site_type': site['type'],
            'device_types': dev_types,
            'matched_templates': result,
            'total_items': sum(len(t['items']) for t in result),
        })

# --- 排程管理 ---

@app.route('/api/inspection-v2/schedules')
def v2_get_schedules():
    """获取排程列表"""
    site_id = request.args.get('site_id', type=int)
    status = request.args.get('status', '')  # due/upcoming/overdue
    user_id = request.args.get('user_id', type=int)
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    with get_db() as db:
        today = datetime.now().strftime('%Y-%m-%d')
        q = """
            SELECT s.*, st.name as site_name, st.type as site_type,
                   it.template_name, iti.item_name, iti.category as item_category,
                   iti.frequency_level, iti.photo_required
            FROM inspection_schedules s
            JOIN sites st ON s.site_id = st.id
            JOIN inspection_templates it ON s.template_id = it.id
            JOIN inspection_template_items iti ON s.template_item_id = iti.id
            WHERE s.status='active'
        """
        params = []
        if site_id:
            q += " AND s.site_id=?"
            params.append(site_id)
        if user_id:
            q += " AND s.site_id IN (SELECT site_id FROM user_sites WHERE user_id=?)"
            params.append(user_id)
        if status == 'due':
            q += " AND s.next_due_date<=?"
            params.append(today)
        elif status == 'overdue':
            q += " AND s.next_due_date<?"
            params.append(today)
        elif status == 'upcoming':
            q += " AND s.next_due_date>? AND s.next_due_date<=date(?, '+7 days')"
            params.extend([today, today])
        if date_from:
            q += " AND s.next_due_date>=?"
            params.append(date_from)
        if date_to:
            q += " AND s.next_due_date<=?"
            params.append(date_to)
        q += " ORDER BY s.next_due_date, st.name"
        rows = db.execute(q, params).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route('/api/inspection-v2/schedules/init', methods=['POST'])
def v2_init_schedules():
    """重新初始化排程（根据当前配置）"""
    with get_db() as db:
        _init_v2_schedules(db)
        cnt = db.execute("SELECT COUNT(*) FROM inspection_schedules").fetchone()[0]
        return jsonify({'success': True, 'total_schedules': cnt})

@app.route('/api/inspection-v2/schedules/<int:sid>/complete', methods=['PUT'])
def v2_complete_schedule(sid):
    """标记排程完成并推进next_due_date"""
    data = request.get_json(silent=True) or {}
    freq_days = {'daily': 1, 'weekly': 7, 'monthly': 30, 'quarterly': 90, 'semi_annual': 180, 'annual': 365}
    now = datetime.now()
    with get_db() as db:
        sch = db.execute("SELECT * FROM inspection_schedules WHERE id=?", (sid,)).fetchone()
        if not sch:
            return jsonify({'error': '排程不存在'}), 404
        fd = freq_days.get(sch['frequency'], 30)
        new_due = (now + timedelta(days=fd)).strftime('%Y-%m-%d')
        db.execute("""
            UPDATE inspection_schedules
            SET last_completed_at=?, next_due_date=?, cycle_count=cycle_count+1
            WHERE id=?
        """, (now.strftime('%Y-%m-%d %H:%M:%S'), new_due, sid))
        db.commit()
        return jsonify({'success': True, 'next_due_date': new_due})

# --- 巡检计划 V2 ---

@app.route('/api/inspection-v2/plans')
def v2_get_plans():
    """获取巡检计划列表（按当前用户站点范围隔离）"""
    assignee_id = request.args.get('assignee_id', type=int)
    status = request.args.get('status', '')
    period = request.args.get('period', '')
    allowed = _filter_site_ids()
    with get_db() as db:
        q = """
            SELECT p.*,
                   v.plate_no as vehicle_name,
                   (SELECT COUNT(*) FROM insp_plan_items i WHERE i.plan_id=p.id
                        AND COALESCE(i.execution_status, 'active')='active') as total_items,
                   (SELECT COUNT(*) FROM insp_plan_items i WHERE i.plan_id=p.id AND i.result IS NOT NULL
                        AND COALESCE(i.execution_status, 'active')='active') as completed_items,
                   (SELECT COUNT(DISTINCT site_id) FROM insp_plan_items WHERE plan_id=p.id) as site_count
            FROM insp_plans p
            LEFT JOIN vehicles v ON p.vehicle_id = v.id
            WHERE 1=1
        """
        params = []
        if allowed is not None:
            ph = ','.join('?' * len(allowed))
            q += f" AND p.id IN (SELECT plan_id FROM insp_plan_items WHERE site_id IN ({ph}))"
            params.extend(allowed)
        if assignee_id:
            q += " AND p.assignee_id=?"
            params.append(assignee_id)
        if status:
            q += " AND p.status=?"
            params.append(status)
        if period:
            q += " AND p.period=?"
            params.append(period)
        q += " ORDER BY p.created_at DESC"
        rows = db.execute(q, params).fetchall()
        result = []
        for r in rows:
            d = dict(r)
            d['completion_rate'] = round(d['completed_items'] / d['total_items'] * 100, 1) if d['total_items'] > 0 else 0
            result.append(d)
        return jsonify(result)

@app.route('/api/inspection-v2/plans/generate', methods=['POST'])
def v2_generate_plans():
    """核心：根据排程生成巡检计划（按负责人打包）"""
    data = request.get_json(silent=True) or {}
    remind_days = data.get('remind_days', 1)
    today = datetime.now()
    today_str = today.strftime('%Y-%m-%d')
    cutoff = (today + timedelta(days=remind_days)).strftime('%Y-%m-%d')
    with get_db() as db:
        # 1. 获取所有到期/临近到期的排程
        schedules = db.execute("""
            SELECT s.*, st.name as site_name, st.type as site_type,
                   iti.item_name, iti.category as item_category
            FROM inspection_schedules s
            JOIN sites st ON s.site_id = st.id
            JOIN inspection_template_items iti ON s.template_item_id = iti.id
            WHERE s.status='active' AND s.next_due_date <= ?
            ORDER BY s.site_id, s.next_due_date
        """, (cutoff,)).fetchall()
        if not schedules:
            return jsonify({'success': True, 'plans_created': 0, 'message': '没有到期的检查项'})

        # 2. 获取负责人→站点映射
        operators = db.execute("SELECT id, real_name FROM users WHERE role IN ('operator','admin')").fetchall()
        user_site_map = {}  # user_id -> [site_ids]
        for op in operators:
            sites = db.execute("SELECT site_id FROM user_sites WHERE user_id=?", (op['id'],)).fetchall()
            user_site_map[op['id']] = {s['site_id'] for s in sites}
            user_site_map[op['real_name']] = {s['site_id'] for s in sites}

        # 3. 按站点分组排程
        site_schedules = {}
        for sch in schedules:
            sid = sch['site_id']
            if sid not in site_schedules:
                site_schedules[sid] = []
            site_schedules[sid].append(sch)

        # 4. 按负责人+频次拆分打包
        #    同一负责人不同频次的排程生成独立的计划，避免"每日计划含月检项"的混乱
        _freq_cn_map = {'daily':'日检','weekly':'周检','monthly':'月检','quarterly':'季检','semi_annual':'半年检','yearly':'年检'}
        plans_created = 0
        total_items = 0
        for op in operators:
            op_name = op['real_name']
            op_sites = user_site_map.get(op['id'], set())
            # 该负责人管辖的到期站点
            due_sites = {sid for sid in site_schedules if sid in op_sites}
            if not due_sites:
                continue

            # 按频次分组：frequency -> {site_id -> [schedules]}
            freq_groups = {}
            for sid in due_sites:
                for sch in site_schedules[sid]:
                    freq = sch['frequency'] or 'daily'
                    if freq not in freq_groups:
                        freq_groups[freq] = {}
                    if sid not in freq_groups[freq]:
                        freq_groups[freq][sid] = []
                    freq_groups[freq][sid].append(sch)

            # 每个频次生成一个独立计划
            for freq in sorted(freq_groups.keys()):
                freq_sites = freq_groups[freq]
                cn_label = _freq_cn_map.get(freq, freq)
                # dedup: 检查是否已有今日生成的同负责人+同频次计划
                existing = db.execute("""
                    SELECT id FROM insp_plans
                    WHERE assignee_id=? AND period=? AND generate_date=? AND status NOT IN ('completed','rejected')
                    LIMIT 1
                """, (op['id'], freq, today_str)).fetchone()
                if existing:
                    plan_id = existing['id']
                    # 追加新的检查项到已有计划
                    for sid in freq_sites:
                        for sch in freq_sites[sid]:
                            # 检查是否已存在相同schedule_id的项
                            dup = db.execute("""
                                SELECT id FROM insp_plan_items
                                WHERE plan_id=? AND schedule_id=? AND site_id=?
                                LIMIT 1
                            """, (plan_id, sch['id'], sid)).fetchone()
                            if dup:
                                continue
                            db.execute("""
                                INSERT INTO insp_plan_items (plan_id, site_id, schedule_id, template_id, item_name, category, frequency)
                                VALUES (?,?,?,?,?,?,?)
                            """, (plan_id, sid, sch['id'], sch['template_id'], sch['item_name'],
                                  sch['item_category'], sch['frequency']))
                            total_items += 1
                    continue

                plan_name = f"{op_name}·{cn_label}-{today.strftime('%Y%m%d')}"
                db.execute("""
                    INSERT INTO insp_plans (plan_name, assignee, assignee_id, period, generate_date, status)
                    VALUES (?,?,?,?,?,?)
                """, (plan_name, op_name, op['id'], freq, today_str, 'draft'))
                plan_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
                # 为该频次的每个到期站点创建plan_items
                for sid in freq_sites:
                    for sch in freq_sites[sid]:
                        db.execute("""
                            INSERT INTO insp_plan_items (plan_id, site_id, schedule_id, template_id, item_name, category, frequency)
                            VALUES (?,?,?,?,?,?,?)
                        """, (plan_id, sid, sch['id'], sch['template_id'], sch['item_name'],
                              sch['item_category'], sch['frequency']))
                        total_items += 1
                plans_created += 1

        db.commit()
        return jsonify({
            'success': True,
            'plans_created': plans_created,
            'total_items': total_items,
            'due_sites': len(site_schedules),
            'date': today_str,
        })

_freq_cn_map = {'daily': '日检', 'weekly': '周检', 'monthly': '月检',
                 'quarterly': '季检', 'semi_annual': '半年检', 'yearly': '年检'}

@app.route('/api/inspection-v2/plans/smart-preview', methods=['POST'])
@login_required
def v2_smart_preview():
    """智能生成预览：基于当前登录账号负责的站点，
    判断哪些需要巡检（排程到期/逾期结合），并推荐可用车辆。不直接创建。"""
    user = g.current_user
    data = request.get_json(silent=True) or {}
    remind_days = data.get('remind_days', 1)
    today = datetime.now()
    today_str = today.strftime('%Y-%m-%d')
    cutoff = (today + timedelta(days=remind_days)).strftime('%Y-%m-%d')
    site_ids = getattr(g, 'user_site_ids', None)  # 空=全部

    with get_db() as db:
        # 1. 到期/临近排程（含已逾期），按当前账号负责站点过滤
        q = """
            SELECT s.id AS schedule_id, s.site_id, s.frequency, s.next_due_date,
                   st.name AS site_name, st.type AS site_type,
                   iti.item_name, iti.category AS item_category
            FROM inspection_schedules s
            JOIN sites st ON s.site_id = st.id
            JOIN inspection_template_items iti ON s.template_item_id = iti.id
            WHERE s.status='active' AND s.next_due_date <= ?
        """
        params = [cutoff]
        if site_ids:
            placeholders = ','.join('?' for _ in site_ids)
            q += f" AND s.site_id IN ({placeholders})"
            params += list(site_ids)
        q += " ORDER BY s.site_id, s.next_due_date"
        schedules = db.execute(q, params).fetchall()

        # 2. 按站点分组，标记逾期
        site_map = {}
        for sch in schedules:
            sid = sch['site_id']
            if sid not in site_map:
                site_map[sid] = {
                    'site_id': sid, 'site_name': sch['site_name'],
                    'type': sch['site_type'], 'schedules': [], 'overdue': False,
                }
            is_over = sch['next_due_date'] < today_str
            site_map[sid]['schedules'].append({
                'schedule_id': sch['schedule_id'],
                'item_name': sch['item_name'],
                'category': sch['item_category'],
                'frequency': sch['frequency'] or 'daily',
                'next_due_date': sch['next_due_date'],
                'overdue': is_over,
            })
            if is_over:
                site_map[sid]['overdue'] = True

        # 3. 每站点近30天最后生成计划时间（结合逾期判定的辅助信息）
        for sid, info in site_map.items():
            last = db.execute("""
                SELECT MAX(p.generate_date) AS d FROM insp_plans p
                JOIN insp_plan_items i ON i.plan_id = p.id
                WHERE i.site_id=? AND p.status NOT IN ('completed','rejected')
            """, (sid,)).fetchone()
            info['last_plan_at'] = last['d'] if last else None

        # 4. 可用车辆：status='idle'，且未被 active/submitted 计划占用、未被 pending 用车申请占用
        occ = db.execute("""
            SELECT DISTINCT vehicle_id FROM insp_plans
            WHERE vehicle_id IS NOT NULL AND status IN ('active','submitted')
        """).fetchall()
        pend = db.execute("""
            SELECT DISTINCT vehicle_id FROM vehicle_applications
            WHERE status='pending' AND vehicle_id IS NOT NULL
        """).fetchall()
        exclude = {r['vehicle_id'] for r in occ} | {r['vehicle_id'] for r in pend}
        veh_rows = db.execute(
            "SELECT id, plate_no, model, status FROM vehicles WHERE status='idle'"
        ).fetchall()
        available_vehicles = [dict(v) for v in veh_rows if v['id'] not in exclude]

        # 5. 预打包建议（按频次分组）
        freq_groups = {}
        for sid, info in site_map.items():
            for sch in info['schedules']:
                f = sch['frequency'] or 'daily'
                freq_groups.setdefault(f, []).append({
                    'site_id': sid, 'site_name': info['site_name'],
                    'schedule_id': sch['schedule_id'],
                    'item_name': sch['item_name'], 'category': sch['category'],
                })
        suggested = []
        for f in sorted(freq_groups.keys()):
            sites_set = {}
            for it in freq_groups[f]:
                key = it['site_id']
                sites_set.setdefault(key, {
                    'site_id': key, 'site_name': it['site_name'], 'items': []
                })
                sites_set[key]['items'].append({
                    'schedule_id': it['schedule_id'],
                    'item_name': it['item_name'], 'category': it['category'],
                })
            suggested.append({
                'period': f,
                'period_label': _freq_cn_map.get(f, f),
                'sites': list(sites_set.values()),
            })

    return jsonify({
        'success': True,
        'user': {'id': user.get('id'), 'name': user.get('real_name', '')},
        'remind_days': remind_days,
        'due_sites': list(site_map.values()),
        'available_vehicles': available_vehicles,
        'suggested': suggested,
        'date': today_str,
    })

@app.route('/api/inspection-v2/plans/confirm', methods=['POST'])
@login_required
def v2_confirm_plan():
    """确认生成：接收前端调整后的 payload 创建 draft 计划。"""
    user = g.current_user
    data = request.get_json(silent=True) or {}
    plan_name = (data.get('plan_name') or '').strip()
    if user.get('role') == 'admin':
        assignee = (data.get('assignee') or user.get('real_name', '') or '').strip()
        assignee_id = data.get('assignee_id') or user.get('id')
    else:
        # 非管理员：负责人强制为自己，避免冒名
        assignee = user.get('real_name', '') or ''
        assignee_id = user.get('id')
    period = data.get('period') or 'weekly'
    vehicle_id = data.get('vehicle_id') or None
    site_items = data.get('site_items') or []  # [{site_id, items:[{schedule_id,item_name,category,frequency}]}]

    if not plan_name:
        return jsonify({'error': '计划名称必填'}), 400
    total = sum(len(s.get('items', [])) for s in site_items)
    if total == 0:
        return jsonify({'error': '请至少选择一个检查项'}), 400

    today_str = datetime.now().strftime('%Y-%m-%d')
    with get_db() as db:
        # 车辆冲突检测（与创建/更新一致）
        if vehicle_id:
            conflict = db.execute(
                "SELECT id FROM insp_plans WHERE vehicle_id=? AND status IN ('active','submitted') LIMIT 1",
                (vehicle_id,)).fetchone()
            if conflict:
                return jsonify({'error': '该车辆已被其他执行中计划占用'}), 400

        db.execute("""
            INSERT INTO insp_plans (plan_name, assignee, assignee_id, period, generate_date, status, vehicle_id)
            VALUES (?,?,?,?,?,?,?)
        """, (plan_name, assignee, assignee_id, period, today_str, 'draft', vehicle_id))
        plan_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]

        for s in site_items:
            sid = s.get('site_id')
            for it in s.get('items', []):
                db.execute("""
                    INSERT INTO insp_plan_items (plan_id, site_id, schedule_id, item_name, category, frequency)
                    VALUES (?,?,?,?,?,?)
                """, (plan_id, sid, it.get('schedule_id'), it.get('item_name'),
                       it.get('category'), it.get('frequency', period)))
        db.commit()

    return jsonify({'success': True, 'plan_id': plan_id, 'total_items': total})

@app.route('/api/inspection-v2/favorites', methods=['GET'])
@login_required
def v2_list_favorites():
    user = g.current_user
    with get_db() as db:
        rows = db.execute(
            "SELECT * FROM plan_favorites WHERE user_id=? ORDER BY created_at DESC",
            (user.get('id'),)).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/inspection-v2/favorites', methods=['POST'])
@login_required
def v2_add_favorite():
    user = g.current_user
    data = request.get_json(silent=True) or {}
    plan_id = data.get('plan_id')
    name = (data.get('name') or '').strip()
    if not plan_id:
        return jsonify({'error': 'plan_id 必填'}), 400
    with get_db() as db:
        plan = db.execute("SELECT * FROM insp_plans WHERE id=?", (plan_id,)).fetchone()
        if not plan:
            return jsonify({'error': '计划不存在'}), 404
        items = db.execute("""
            SELECT i.site_id, s.name AS site_name, i.item_name, i.category, i.frequency, i.schedule_id
            FROM insp_plan_items i LEFT JOIN sites s ON s.id=i.site_id
            WHERE i.plan_id=?
        """, (plan_id,)).fetchall()
        site_groups = {}
        for it in items:
            sid = it['site_id']
            site_groups.setdefault(sid, {'site_id': sid, 'site_name': it['site_name'], 'items': []})
            site_groups[sid]['items'].append({
                'schedule_id': it['schedule_id'], 'item_name': it['item_name'],
                'category': it['category'], 'frequency': it['frequency'],
            })
        snapshot = {
            'plan_name': plan['plan_name'],
            'period': plan['period'],
            'vehicle_id': plan['vehicle_id'],
            'site_groups': list(site_groups.values()),
        }
        if not name:
            name = plan['plan_name']
        db.execute(
            "INSERT INTO plan_favorites (user_id, plan_id, name, snapshot) VALUES (?,?,?,?)",
            (user.get('id'), plan_id, name, json.dumps(snapshot, ensure_ascii=False)))
        fid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        db.commit()
    return jsonify({'success': True, 'favorite_id': fid})

@app.route('/api/inspection-v2/favorites/<int:fid>', methods=['DELETE'])
@login_required
def v2_delete_favorite(fid):
    user = g.current_user
    with get_db() as db:
        db.execute("DELETE FROM plan_favorites WHERE id=? AND user_id=?",
                   (fid, user.get('id')))
        db.commit()
    return jsonify({'success': True})

@app.route('/api/inspection-v2/favorites/<int:fid>/apply', methods=['POST'])
@login_required
def v2_apply_favorite(fid):
    """从收藏复用：基于快照生成新草稿计划。"""
    user = g.current_user
    with get_db() as db:
        fav = db.execute("SELECT * FROM plan_favorites WHERE id=? AND user_id=?",
                         (fid, user.get('id'))).fetchone()
        if not fav:
            return jsonify({'error': '收藏不存在'}), 404
        snap = json.loads(fav['snapshot']) if fav['snapshot'] else {}
        today_str = datetime.now().strftime('%Y-%m-%d')
        base_name = snap.get('plan_name') or '收藏计划'
        plan_name = f"{base_name}-复用{today_str}"
        db.execute("""
            INSERT INTO insp_plans (plan_name, assignee, assignee_id, period, generate_date, status, vehicle_id)
            VALUES (?,?,?,?,?,?,?)
        """, (plan_name, user.get('real_name', ''), user.get('id'),
               snap.get('period', 'weekly'), today_str, 'draft', snap.get('vehicle_id')))
        plan_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        total = 0
        for s in snap.get('site_groups', []):
            for it in s.get('items', []):
                db.execute("""
                    INSERT INTO insp_plan_items (plan_id, site_id, schedule_id, item_name, category, frequency)
                    VALUES (?,?,?,?,?,?)
                """, (plan_id, s.get('site_id'), it.get('schedule_id'),
                       it.get('item_name'), it.get('category'),
                       it.get('frequency', snap.get('period'))))
                total += 1
        db.commit()
    return jsonify({'success': True, 'plan_id': plan_id, 'total_items': total})

@app.route('/api/inspection-v2/plans/<int:plan_id>')
def v2_get_plan_detail(plan_id):
    """获取巡检计划详情"""
    with get_db() as db:
        plan = db.execute("""
            SELECT p.*, v.plate_no as vehicle_name
            FROM insp_plans p
            LEFT JOIN vehicles v ON p.vehicle_id = v.id
            WHERE p.id=?
        """, (plan_id,)).fetchone()
        if not plan:
            return jsonify({'error': '计划不存在'}), 404
        items = db.execute("""
            SELECT pi.*, s.name as site_name, s.code as site_code
            FROM insp_plan_items pi
            JOIN sites s ON pi.site_id = s.id
            WHERE pi.plan_id=? AND COALESCE(pi.execution_status, 'active')='active'
            ORDER BY pi.site_id, pi.id
        """, (plan_id,)).fetchall()
        result = dict(plan)
        result['items'] = [dict(it) for it in items]
        result['total_items'] = len(items)
        result['completed_items'] = sum(1 for it in items if it['result'] is not None)
        result['completion_rate'] = round(result['completed_items'] / result['total_items'] * 100, 1) if result['total_items'] > 0 else 0
        # 各站点近30天告警数
        site_ids = list(set(it['site_id'] for it in items))
        alert_counts = {}
        if site_ids:
            placeholders = ','.join('?' for _ in site_ids)
            alert_rows = db.execute(f"""
                SELECT site_id, COUNT(*) as cnt
                FROM alerts
                WHERE site_id IN ({placeholders}) AND created_at >= datetime('now','-30 days')
                GROUP BY site_id
            """, site_ids).fetchall()
            for ar in alert_rows:
                alert_counts[ar['site_id']] = ar['cnt']
        # 按站点分组
        site_groups = {}
        for it in items:
            sn = it['site_name']
            if sn not in site_groups:
                site_groups[sn] = {'site_id': it['site_id'], 'site_name': sn, 'items': [], 'alert_count_30d': alert_counts.get(it['site_id'], 0)}
            site_groups[sn]['items'].append(dict(it))
        result['site_groups'] = list(site_groups.values())
        return jsonify(result)

@app.route('/api/inspection-v2/plans/<int:plan_id>', methods=['PUT'])
def v2_update_plan(plan_id):
    """更新巡检计划（添加/修改检查项）"""
    data = request.get_json(silent=True) or {}
    with get_db() as db:
        plan = db.execute("SELECT id, status FROM insp_plans WHERE id=?", (plan_id,)).fetchone()
        if not plan:
            return jsonify({'error': '计划不存在'}), 404
        # 更新计划字段
        for key in ('status', 'period', 'vehicle_id'):
            if key in data:
                if key == 'vehicle_id' and data[key]:
                    conflict = db.execute("SELECT id,plan_name FROM insp_plans WHERE vehicle_id=? AND status IN ('active','submitted') AND id!=?",
                                          (data[key], plan_id)).fetchall()
                    if conflict:
                        names = '；'.join(f"#{c['id']} {c['plan_name']}" for c in conflict)
                        db.close()
                        return jsonify({'error': f'该车辆已被计划 {names} 占用，无法重复分配'}), 400
                db.execute(f"UPDATE insp_plans SET {key}=? WHERE id=?", (data[key], plan_id))
        # 添加检查项
        if 'add_items' in data:
            for item in data['add_items']:
                db.execute("""
                    INSERT INTO insp_plan_items (plan_id,site_id,schedule_id,template_id,item_name,category,frequency)
                    VALUES (?,?,?,?,?,?,?)
                """, (plan_id, item.get('site_id'), item.get('schedule_id'), item.get('template_id'),
                      item.get('item_name',''), item.get('category',''), item.get('frequency','')))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/inspection-v2/plans/<int:plan_id>', methods=['DELETE'])
def v2_delete_plan(plan_id):
    """删除巡检计划"""
    with get_db() as db:
        plan = db.execute("SELECT id, status FROM insp_plans WHERE id=?", (plan_id,)).fetchone()
        if not plan:
            return jsonify({'error': '计划不存在'}), 404
        if plan['status'] == 'completed':
            return jsonify({'error': '已完成的计划不可删除'}), 400
        db.execute("DELETE FROM insp_plan_items WHERE plan_id=?", (plan_id,))
        db.execute("DELETE FROM insp_plans WHERE id=?", (plan_id,))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/inspection-v2/plans/<int:plan_id>/submit', methods=['POST'])
def v2_submit_plan(plan_id):
    """提交巡检计划审批：draft→submitted"""
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as db:
        plan = db.execute("SELECT id, status FROM insp_plans WHERE id=?", (plan_id,)).fetchone()
        if not plan:
            return jsonify({'error': '计划不存在'}), 404
        if plan['status'] != 'draft':
            return jsonify({'error': '仅草稿状态可提交'}), 400
        db.execute("UPDATE insp_plans SET status='submitted', submitted_at=? WHERE id=?", (now, plan_id))
        db.commit()
        return jsonify({'success': True, 'status': 'submitted'})

@app.route('/api/inspection-v2/plans/<int:plan_id>/complete', methods=['POST'])
@login_required
def v2_complete_plan(plan_id):
    """完成巡检计划执行：active→completed，不触发逐项审核"""
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as db:
        plan = db.execute("SELECT id, status FROM insp_plans WHERE id=?", (plan_id,)).fetchone()
        if not plan:
            return jsonify({'error': '计划不存在'}), 404
        if plan['status'] != 'active':
            return jsonify({'error': '仅执行中状态可完成'}), 400
        pending = db.execute("""SELECT COUNT(*) FROM insp_plan_items
            WHERE plan_id=? AND result IS NULL
              AND COALESCE(execution_status, 'active')='active'""", (plan_id,)).fetchone()[0]
        if pending > 0:
            return jsonify({'error': f'还有 {pending} 个检查项未执行，无法完成执行'}), 400
        db.execute("UPDATE insp_plans SET status='completed', completion_rate=100 WHERE id=?", (plan_id,))
        db.commit()
        return jsonify({'success': True, 'status': 'completed'})


@app.route('/api/inspection-v2/plans/<int:plan_id>/approve', methods=['POST'])
@login_required
def v2_approve_plan(plan_id):
    """巡检计划审批：submitted→active/rejected→draft"""
    g_ = require_admin()
    if g_:
        return g_
    data = request.get_json(silent=True) or {}
    action = data.get('action', '')
    approver_id = data.get('approver_id', 0)
    comment = data.get('comment', '')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as db:
        plan = db.execute("SELECT id, status FROM insp_plans WHERE id=?", (plan_id,)).fetchone()
        if not plan:
            return jsonify({'error': '计划不存在'}), 404
        # 站点归属校验：非本人站点的计划不可审批
        allowed = _filter_site_ids()
        if allowed is not None:
            plan_sites = [r['site_id'] for r in db.execute(
                "SELECT DISTINCT site_id FROM insp_plan_items WHERE plan_id=?", (plan_id,)).fetchall()]
            if not (set(plan_sites) & set(allowed)):
                return jsonify({'error': '无权限审批非本人负责站点的计划'}), 403
        if action == 'approve':
            if plan['status'] != 'submitted':
                return jsonify({'error': '仅已提交状态可批准'}), 400
            db.execute("UPDATE insp_plans SET status='active', approver_id=?, approve_comment=?, submitted_at=? WHERE id=?",
                       (approver_id, comment, now, plan_id))
        elif action == 'reject':
            # 看板中的车辆冲突/站点重叠计划多为执行中(active)，管理员可直接驳回回草稿
            if plan['status'] not in ('submitted', 'active'):
                return jsonify({'error': '仅已提交/执行中状态可驳回'}), 400
            reason = (data.get('reason') or data.get('comment') or '').strip()
            if not reason:
                return jsonify({'error': '驳回必须填写原因'}), 400
            db.execute("UPDATE insp_plans SET status='draft', approver_id=?, reject_reason=?, submitted_at=? WHERE id=?",
                       (approver_id, reason, now, plan_id))
            # 通知负责人：驳回后计划退回草稿，需负责人修改后重新提交
            p = db.execute("SELECT plan_name, assignee_id, assignee FROM insp_plans WHERE id=?", (plan_id,)).fetchone()
            if p and p['assignee_id']:
                _create_notification(
                    p['assignee_id'], 'inspection', plan_id,
                    f'巡检计划被驳回',
                    f'「{p["plan_name"]}」被驳回，原因：{reason}。请修改后重新提交。',
                    db=db)
        else:
            return jsonify({'error': '无效操作，请指定action为approve或reject'}), 400
        db.commit()
        return jsonify({'success': True, 'status': action})

@app.route('/api/inspection-v2/dashboard', methods=['GET'])
def v2_inspection_dashboard():
    """巡检态势看板：该检未检、车辆冲突、站点重叠、计划汇总（按站点范围隔离）"""
    today = datetime.now().strftime('%Y-%m-%d')
    allowed = _filter_site_ids()
    site_clause = ''
    sp = []
    if allowed is not None:
        site_clause = f' AND site_id IN ({",".join("?" * len(allowed))})'
        sp = list(allowed)
    with get_db() as db:
        # 到期排程（应检）：活跃排程 next_due_date <= 今日
        due = db.execute(f"""
            SELECT s.site_id, COUNT(*) cnt,
                   SUM(CASE WHEN s.next_due_date <= ? THEN 1 ELSE 0 END) overdue
            FROM inspection_schedules s
            WHERE s.status='active'{site_clause}
            GROUP BY s.site_id
        """, (today,) + tuple(sp)).fetchall()
        # 已覆盖站点：有 执行中/待审核/草稿 计划
        covered = db.execute(f"""
            SELECT DISTINCT site_id FROM insp_plan_items
            WHERE plan_id IN (SELECT id FROM insp_plans WHERE status IN ('active','submitted','draft')){site_clause}
        """, sp).fetchall()
        covered_ids = {r['site_id'] for r in covered}
        names = {r['id']: r['name'] for r in db.execute("SELECT id,name FROM sites")}
        # 该检未检：应检站点且未被覆盖
        due_sites = []
        for r in due:
            sid = r['site_id']
            if sid not in covered_ids:
                due_sites.append({
                    'site_id': sid,
                    'site_name': names.get(sid, f'站点{sid}'),
                    'schedule_count': r['cnt'],
                    'overdue': r['overdue'],
                })
        due_sites.sort(key=lambda x: (-x['overdue'], x['site_id']))
        # 车辆冲突：同一车辆被 >=2 个执行中/待审核计划占用
        vc = db.execute("""
            SELECT vehicle_id, COUNT(*) c FROM insp_plans
            WHERE status IN ('active','submitted') AND vehicle_id IS NOT NULL AND vehicle_id > 0
            GROUP BY vehicle_id HAVING COUNT(*) > 1
        """).fetchall()
        vehicle_conflicts = []
        for r in vc:
            plans = db.execute(
                "SELECT id,status,plan_name,assignee_id,assignee FROM insp_plans WHERE vehicle_id=? AND status IN ('active','submitted')",
                (r['vehicle_id'],)).fetchall()
            v = db.execute("SELECT plate_no FROM vehicles WHERE id=?", (r['vehicle_id'],)).fetchone()
            vehicle_conflicts.append({
                'vehicle_id': r['vehicle_id'],
                'plate_no': v['plate_no'] if v else '',
                'plans': [dict(p) for p in plans],
            })
        # 站点重叠：收窄为「同一站点 + 同一周期 + 执行中」被 >=2 个计划覆盖
        # （不同周期并行属正常多频次覆盖，不计入冲突噪音）
        so = db.execute("""
            SELECT i.site_id, p.period, COUNT(DISTINCT p.id) c
            FROM insp_plan_items i JOIN insp_plans p ON p.id=i.plan_id
            WHERE p.status='active'
            GROUP BY i.site_id, p.period
            HAVING COUNT(DISTINCT p.id) > 1
        """).fetchall()
        site_overlaps = []
        for r in so:
            plans = db.execute("""
                SELECT DISTINCT p.id, p.status, p.plan_name, p.assignee_id, p.assignee
                FROM insp_plan_items i
                JOIN insp_plans p ON p.id=i.plan_id
                WHERE i.site_id=? AND p.period=? AND p.status='active'
            """, (r['site_id'], r['period'])).fetchall()
            site_overlaps.append({
                'site_id': r['site_id'],
                'site_name': names.get(r['site_id'], f'站点{r["site_id"]}'),
                'period': r['period'],
                'plans': [dict(p) for p in plans],
            })
        # 可用车辆：status='idle' 且未被 active/submitted 计划占用、未被 pending 用车申请占用
        veh_all = db.execute("SELECT id FROM vehicles WHERE status='idle'").fetchall()
        occ = db.execute("""
            SELECT DISTINCT vehicle_id FROM insp_plans
            WHERE vehicle_id IS NOT NULL AND vehicle_id > 0 AND status IN ('active','submitted')
        """).fetchall()
        pend = db.execute("""
            SELECT DISTINCT vehicle_id FROM vehicle_applications
            WHERE status='pending' AND vehicle_id IS NOT NULL
        """).fetchall()
        exclude = {r['vehicle_id'] for r in occ} | {r['vehicle_id'] for r in pend}
        veh_total = db.execute("SELECT COUNT(*) c FROM vehicles").fetchone()['c']
        veh_available = len([v for v in veh_all if v['id'] not in exclude])
        # 计划状态汇总
        sm = db.execute("SELECT status, COUNT(*) c FROM insp_plans GROUP BY status").fetchall()
        summary = {r['status']: r['c'] for r in sm}
        summary['due_total'] = len(due_sites)
        summary['vehicle_conflict'] = len(vehicle_conflicts)
        summary['site_overlap'] = len(site_overlaps)
        summary['vehicle_total'] = veh_total
        summary['vehicle_available'] = veh_available
        return jsonify({
            'due_sites': due_sites,
            'vehicle_conflicts': vehicle_conflicts,
            'site_overlaps': site_overlaps,
            'summary': summary,
        })


@app.route('/api/inspection-v2/plans/manual', methods=['POST'])
@login_required
def v2_create_plan_manual():
    """手动创建巡检计划：选择站点+车辆+名称
    非管理员：负责人强制为当前用户，站点限定在其负责站点内。"""
    user = g.current_user
    data = request.get_json(silent=True) or {}
    plan_name = data.get('plan_name', '').strip()
    raw_site_ids = data.get('site_ids', [])
    site_ids = list(raw_site_ids)
    vehicle_id = data.get('vehicle_id')
    period = data.get('period', 'weekly')
    allowed = getattr(g, 'user_site_ids', None)  # None = 全部（管理员）
    if user.get('role') == 'admin':
        assignee = (data.get('assignee') or user.get('real_name', '')).strip()
        assignee_id = data.get('assignee_id') or user.get('id')
    else:
        # 非管理员：负责人必须是自己，站点必须是自己负责的
        assignee = user.get('real_name', '')
        assignee_id = user.get('id')
        if allowed is not None:
            site_ids = [s for s in site_ids if s in allowed]
    if not plan_name:
        return jsonify({'error': '计划名称不能为空'}), 400
    if not site_ids:
        return jsonify({'error': '请选择至少一个自己负责的站点'}), 400
    now = datetime.now()
    today = now.strftime('%Y-%m-%d')
    with get_db() as db:
        # 车辆冲突校验：该车辆已被其他执行中/待审核计划占用则拦截
        if vehicle_id:
            conflict = db.execute("""
                SELECT id, plan_name, status FROM insp_plans
                WHERE vehicle_id=? AND status IN ('active','submitted')
            """, (vehicle_id,)).fetchall()
            if conflict:
                names = '；'.join(f"#{c['id']} {c['plan_name']}({c['status']})" for c in conflict)
                return jsonify({'error': f'该车辆已被以下计划占用，无法重复分配：{names}'}), 400
        db.execute("""
            INSERT INTO insp_plans (plan_name, assignee, assignee_id, period, generate_date, status, vehicle_id)
            VALUES (?,?,?,?,?, 'draft', ?)
        """, (plan_name, assignee, assignee_id, period, today, vehicle_id))
        plan_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        for sid in site_ids:
            site = db.execute("SELECT name FROM sites WHERE id=?", (sid,)).fetchone()
            # 查找该站点的排程项
            scheds = db.execute("""
                SELECT s.id, s.template_id, s.template_item_id, iti.item_name, iti.category
                FROM inspection_schedules s
                JOIN inspection_template_items iti ON s.template_item_id = iti.id
                WHERE s.site_id=? AND s.status='active'
            """, (sid,)).fetchall()
            if scheds:
                for sch in scheds:
                    db.execute("""
                        INSERT INTO insp_plan_items (plan_id, site_id, schedule_id, template_id, item_name, category, frequency)
                        VALUES (?,?,?,?,?,?,?)
                    """, (plan_id, sid, sch['id'], sch['template_id'], sch['item_name'],
                          sch['category'], period))
            else:
                # 无排程：创建占位项
                db.execute("""
                    INSERT INTO insp_plan_items (plan_id, site_id, item_name, category, frequency)
                    VALUES (?,?,?,?,?)
                """, (plan_id, sid, f"{site['name'] if site else '站点'+str(sid)}巡检", '', period))
        db.commit()
        return jsonify({'id': plan_id, 'success': True, 'plan_name': plan_name}), 201

@app.route('/api/inspection-v2/plans/<int:plan_id>/sites', methods=['PUT'])
def v2_update_plan_sites(plan_id):
    """更新计划关联站点"""
    data = request.get_json(silent=True) or {}
    site_ids = data.get('site_ids', [])
    if not isinstance(site_ids, list):
        return jsonify({'error': 'site_ids 必须是数组'}), 400
    with get_db() as db:
        plan = db.execute("SELECT * FROM insp_plans WHERE id=?", (plan_id,)).fetchone()
        if not plan:
            return jsonify({'error': '计划不存在'}), 404
        # 获取当前计划中的站点
        current = db.execute("SELECT DISTINCT site_id FROM insp_plan_items WHERE plan_id=?", (plan_id,)).fetchall()
        current_ids = {r['site_id'] for r in current}
        new_ids = set(site_ids)
        # 要删除的站点：移除其所有检查项
        to_remove = current_ids - new_ids
        for sid in to_remove:
            db.execute("DELETE FROM insp_plan_items WHERE plan_id=? AND site_id=?", (plan_id, sid))
        # 要新增的站点：从模板/排程中获取检查项并添加
        to_add = new_ids - current_ids
        for sid in to_add:
            # 查找匹配的排程项
            scheds = db.execute("""
                SELECT s.id, s.template_id, s.template_item_id, iti.item_name, iti.category
                FROM inspection_schedules s
                JOIN inspection_template_items iti ON s.template_item_id = iti.id
                WHERE s.site_id=? AND s.status='active'
            """, (sid,)).fetchall()
            if scheds:
                for sch in scheds:
                    db.execute("""
                        INSERT INTO insp_plan_items (plan_id, site_id, schedule_id, template_id, item_name, category, frequency)
                        VALUES (?,?,?,?,?,?,?)
                    """, (plan_id, sid, sch['id'], sch['template_id'], sch['item_name'],
                          sch['category'], plan['period'] or 'weekly'))
            else:
                # 无排程：插入一条占位项
                site = db.execute("SELECT name FROM sites WHERE id=?", (sid,)).fetchone()
                db.execute("""
                    INSERT INTO insp_plan_items (plan_id, site_id, item_name)
                    VALUES (?,?,?)
                """, (plan_id, sid, f"{site['name'] if site else '站点'+str(sid)}巡检"))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/inspection-v2/plans/<int:plan_id>/items/<int:item_id>', methods=['DELETE'])
def v2_delete_plan_item(plan_id, item_id):
    """删除单条检查项"""
    with get_db() as db:
        item = db.execute("SELECT id, result FROM insp_plan_items WHERE id=? AND plan_id=?", (item_id, plan_id)).fetchone()
        if not item:
            return jsonify({'error': '检查项不存在'}), 404
        if item['result']:
            return jsonify({'error': '已有结果的检查项不能删除'}), 400
        db.execute("DELETE FROM insp_plan_items WHERE id=? AND plan_id=?", (item_id, plan_id))
        # 检查计划是否全部完成
        total = db.execute("""SELECT COUNT(*) FROM insp_plan_items
            WHERE plan_id=? AND COALESCE(execution_status, 'active')='active'""", (plan_id,)).fetchone()[0]
        if total == 0:
            db.execute("UPDATE insp_plans SET status='draft', completion_rate=0 WHERE id=?", (plan_id,))
        else:
            done = db.execute("""SELECT COUNT(*) FROM insp_plan_items
                WHERE plan_id=? AND result IS NOT NULL
                  AND COALESCE(execution_status, 'active')='active'""", (plan_id,)).fetchone()[0]
            rate = round(done / total * 100, 1) if total > 0 else 0
            db.execute("UPDATE insp_plans SET completion_rate=? WHERE id=?", (rate, plan_id))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/inspection-v2/plans/<int:plan_id>/items/<int:item_id>', methods=['PUT'])
def v2_update_plan_item(plan_id, item_id):
    """提交单个检查项结果"""
    data = request.get_json(silent=True) or {}
    now = datetime.now()
    with get_db() as db:
        plan = db.execute("SELECT id, status FROM insp_plans WHERE id=?", (plan_id,)).fetchone()
        if not plan:
            return jsonify({'error': '计划不存在'}), 404
        if plan['status'] != 'active':
            return jsonify({'error': '仅执行中状态可提交检查项结果'}), 400
        item = db.execute("SELECT * FROM insp_plan_items WHERE id=? AND plan_id=?", (item_id, plan_id)).fetchone()
        if not item:
            return jsonify({'error': '检查项不存在'}), 404
        # 结果枚举归一化（兼容历史中文值）
        result_raw = data.get('result', 'normal')
        result_norm = {'正常': 'normal', '正常完成': 'normal', '异常': 'abnormal'}.get(result_raw, result_raw)
        # 拍照校验：要求拍照但未满足，不可判为正常
        if result_norm == 'normal':
            req_photos = item['required_photos'] or 0
            act_photos = data.get('actual_photos', item['actual_photos'] or 0) or 0
            if req_photos > 0 and act_photos < req_photos:
                return jsonify({'error': f'该项需上传 {req_photos} 张照片，当前仅 {act_photos} 张，不可判为正常'}), 400
        # 更新检查项结果
        fields = ["result=?", "check_time=?"]
        params = [result_norm, now.strftime('%Y-%m-%d %H:%M:%S')]
        for key in ('photo_urls', 'gps_lat', 'gps_lng', 'remark', 'calibrator', 'calibration_values', 'actual_photos', 'location_address', 'part_consumed'):
            if key in data:
                fields.append(f"{key}=?")
                params.append(data[key])

        # 根据模板判断是否需要审核
        if item['schedule_id']:
            sch = db.execute("""
                SELECT t.id as template_id FROM inspection_schedules s
                JOIN inspection_templates t ON s.template_id = t.id
                WHERE s.id = ?
            """, (item['schedule_id'],)).fetchone()
            if sch:
                # 查找模板项中与当前 item_name 匹配的记录
                tpl_item = db.execute("""
                    SELECT need_review FROM inspection_template_items
                    WHERE template_id=? AND item_name=?
                """, (sch['template_id'], item['item_name'])).fetchone()
                if tpl_item and tpl_item['need_review'] == 1:
                    # 需要审核
                    fields.append("review_status=?")
                    params.append(1)  # pending review
                else:
                    # 自动通过
                    fields.append("review_status=?")
                    params.append(2)  # approved
            else:
                # schedule not found, auto approve
                fields.append("review_status=?")
                params.append(2)
        else:
            # no schedule, auto approve
            fields.append("review_status=?")
            params.append(2)

        fields.append("completed_at=?")
        params.append(now.strftime('%Y-%m-%d %H:%M:%S'))
        params.extend([item_id, plan_id])
        db.execute(f"UPDATE insp_plan_items SET {','.join(fields)} WHERE id=? AND plan_id=?", params)
        # 如果异常，创建告警
        if data.get('result') == 'abnormal':
            create_alert_internal(db, item['site_id'], 'inspection', 0, 'yellow',
                f'巡检异常：{item["item_name"]}' + (f' - {data.get("remark","")}' if data.get("remark") else ''))
        # 如果关联了schedule_id，推进排程
        if item['schedule_id'] and data.get('result'):
            freq_days_map = {'daily': 1, 'weekly': 7, 'monthly': 30, 'quarterly': 90, 'semi_annual': 180, 'annual': 365}
            sch = db.execute("SELECT frequency FROM inspection_schedules WHERE id=?", (item['schedule_id'],)).fetchone()
            if sch:
                fd = freq_days_map.get(sch['frequency'], 30)
                new_due = (now + timedelta(days=fd)).strftime('%Y-%m-%d')
                db.execute("""
                    UPDATE inspection_schedules SET last_completed_at=?, next_due_date=?, cycle_count=cycle_count+1
                    WHERE id=?
                """, (now.strftime('%Y-%m-%d %H:%M:%S'), new_due, item['schedule_id']))
        # 如果需更换且有备件消耗，扣减库存
        part_consumed = data.get('part_consumed', '')
        if data.get('result') == '需更换' and part_consumed:
            inv = db.execute(
                "SELECT id, quantity FROM spare_parts_inventory WHERE part_code=? LIMIT 1",
                (part_consumed,)).fetchone()
            if inv:
                new_qty = max(0, inv['quantity'] - 1)
                db.execute("UPDATE spare_parts_inventory SET quantity=?, updated_at=datetime('now','localtime') WHERE id=?",
                           (new_qty, inv['id']))
        # 更新完成率
        total = db.execute("""SELECT COUNT(*) FROM insp_plan_items
            WHERE plan_id=? AND COALESCE(execution_status, 'active')='active'""", (plan_id,)).fetchone()[0]
        done = db.execute("""SELECT COUNT(*) FROM insp_plan_items
            WHERE plan_id=? AND result IS NOT NULL
              AND COALESCE(execution_status, 'active')='active'""", (plan_id,)).fetchone()[0]
        if total > 0:
            rate = round(done / total * 100, 1)
            db.execute("UPDATE insp_plans SET completion_rate=? WHERE id=?", (rate, plan_id))
        # 全部检查项已提交结果 → 自动完成计划（替代手动"完成执行"按钮，避免冗余操作）
        if plan['status'] == 'active' and total > 0 and done == total:
            db.execute("UPDATE insp_plans SET status='completed', completion_rate=100 WHERE id=?", (plan_id,))
        # 同步巡检现场照片到 operation_attachments（统一归口：影像/归档/审核可见，内联 photo_urls 保留兼容）
        if 'photo_urls' in data:
            import json as _json
            try:
                new_list = _json.loads(data['photo_urls']) if isinstance(data['photo_urls'], str) else (data['photo_urls'] or [])
            except Exception:
                new_list = []
            old_list = []
            try:
                old_list = _json.loads(item['photo_urls']) if item['photo_urls'] else []
            except Exception:
                old_list = []
            for u in (new_list or []):
                if u and u not in old_list:
                    # 智能识别：按文件名/说明关键词匹配照片类型配置（水印/场景自动归类，接入审核链）
                    _desc = f"巡检照片-{item['item_name']}"
                    _match = match_photo_requirement('', item['site_id'], str(u), _desc)
                    _req_id = _match['requirement_id'] if _match else None
                    _rec_cat = _match['recognized_category'] if _match else ''
                    _match_status = _match['match_status'] if _match else 'manual'
                    _match_conf = _match['match_confidence'] if _match else None
                    _review_required = _match['review_required'] if _match else 0
                    db.execute("""INSERT INTO operation_attachments
                        (filename, stored_path, file_type, mime_type, file_size, description,
                         source_type, source_id, site_id, category, created_at,
                         watermark_text, recognized_category, match_status, match_confidence,
                         review_required, requirement_id)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (os.path.basename(str(u)), str(u), 'image', '', 0,
                         _desc, 'inspection', item_id, item['site_id'],
                         '巡检照片', now.strftime('%Y-%m-%d %H:%M:%S'),
                         '', _rec_cat, _match_status, _match_conf,
                         _review_required, _req_id))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/inspection-v2/plans/<int:plan_id>/items/<int:item_id>/anomaly', methods=['POST'])
def v2_report_anomaly(plan_id, item_id):
    """巡检中异常上报：更新检查项状态 + 创建告警 + 自动生成工单"""
    data = request.get_json(silent=True) or {}
    report_type = data.get('report_type', 'equipment')
    description = data.get('description', '')
    photo_urls = data.get('photo_urls', [])
    reporter_id = data.get('reporter_id', 0)
    import json as _json
    photos_json = _json.dumps(photo_urls, ensure_ascii=False) if photo_urls else '[]'
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as db:
        item = db.execute("SELECT * FROM insp_plan_items WHERE id=? AND plan_id=?", (item_id, plan_id)).fetchone()
        if not item:
            return jsonify({'error': '检查项不存在'}), 404
        # 1. 更新检查项状态
        db.execute("UPDATE insp_plan_items SET result='anomaly_reported', remark=?, photo_urls=?, check_time=? WHERE id=? AND plan_id=?",
                   (description, photos_json, now, item_id, plan_id))
        # 2. 创建告警
        report_type_label = {'equipment': '设备异常', 'environment': '环境异常', 'sensory': '感官异常', 'operation': '操作异常'}
        event_type = report_type_label.get(report_type, report_type)
        alert_msg = f'【巡检异常】{event_type}: {description[:200]}'
        # 创建工单
        order_no = 'IX' + now[:10].replace('-','') + str(int(time.time()*1000))[-6:]
        db.execute(
            '''INSERT INTO work_orders (order_no, site_id, source, event_type, level, title, description, status)
               VALUES (?,?, 'inspection', ?, 'normal', ?, ?, 'pending')''',
            (order_no, item['site_id'], event_type, f'【巡检异常】{event_type}', description))
        # 告警级别映射
        level_map = {'equipment': 'orange', 'environment': 'orange', 'sensory': 'yellow', 'operation': 'red'}
        alert_level = level_map.get(report_type, 'orange')
        existing_alert = db.execute(
            "SELECT id FROM alerts WHERE site_id=? AND message=? AND created_at > datetime('now','-30 seconds')",
            (item['site_id'], alert_msg)).fetchone()
        if not existing_alert:
            db.execute(
                '''INSERT INTO alerts (site_id, metric, value, level, message, status, related_order_no, flow_type)
                   VALUES (?,?,?,?,?,'pending',?,'manual')''',
                (item['site_id'], 'inspection_anomaly', 0, alert_level, alert_msg, order_no))
        alert_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        db.commit()
        return jsonify({'success': True, 'alert_id': alert_id, 'order_no': order_no}), 201

@app.route('/api/inspection-v2/plans/<int:plan_id>/parts-request', methods=['POST'])
def v2_parts_request(plan_id):
    """巡检备件预申报"""
    data = request.get_json(silent=True) or {}
    items_data = data.get('items', [])
    requester_id = data.get('requester_id', 0)
    if not items_data:
        return jsonify({'error': '备件列表不能为空'}), 400
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as db:
        # 确保表存在
        db.execute("""
            CREATE TABLE IF NOT EXISTS parts_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                plan_id INTEGER NOT NULL,
                requester_id INTEGER DEFAULT 0,
                status TEXT DEFAULT 'pending',
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (plan_id) REFERENCES insp_plans(id)
            )
        """)
        db.execute("""
            CREATE TABLE IF NOT EXISTS parts_request_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_id INTEGER NOT NULL,
                part_sku TEXT NOT NULL,
                quantity INTEGER NOT NULL DEFAULT 1,
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (request_id) REFERENCES parts_requests(id)
            )
        """)
        # 创建总申请记录
        cur = db.execute(
            "INSERT INTO parts_requests (plan_id, requester_id) VALUES (?,?)",
            (plan_id, requester_id))
        request_id = cur.lastrowid
        # 创建明细
        for item in items_data:
            db.execute(
                "INSERT INTO parts_request_items (request_id, part_sku, quantity) VALUES (?,?,?)",
                (request_id, item['part_sku'], item['quantity']))
        db.commit()
        return jsonify({'success': True, 'request_id': request_id, 'id': request_id}), 201

@app.route('/api/inspection-v2/parts-request/<int:rid>/approve', methods=['PUT'])
def v2_parts_request_approve(rid):
    """审批通过巡检备件预申报（聚合到 /audit 待办）"""
    g_ = require_approver()
    if g_:
        return g_
    data = request.get_json(silent=True) or {}
    comment = (data.get('comment') or '').strip()
    with get_db() as db:
        # 确保审批字段存在（旧表兼容）
        try:
            db.execute("ALTER TABLE parts_requests ADD COLUMN approver_id INTEGER")
            db.execute("ALTER TABLE parts_requests ADD COLUMN approve_comment TEXT")
        except Exception:
            pass
        req = db.execute('SELECT * FROM parts_requests WHERE id=?', (rid,)).fetchone()
        if not req:
            return jsonify({'error': '申请不存在'}), 404
        if req['status'] != 'pending':
            return jsonify({'error': '该申请已处理'}), 400
        db.execute(
            "UPDATE parts_requests SET status='approved', approver_id=?, approve_comment=? WHERE id=?",
            (data.get('approver_id', 1), comment, rid))
        # === 扣除库存 + 写流水（与 spare_part_requests 审批保持相同模式）===
        items = db.execute(
            "SELECT * FROM parts_request_items WHERE request_id=?", (rid,)).fetchall()
        for item in items:
            sku = item['part_sku']
            qty = item['quantity']
            # 先按 part_code 精确匹配，再按 part_name 模糊匹配
            inv = db.execute(
                """SELECT * FROM spare_parts_inventory WHERE part_code=? OR part_name LIKE ?
                   ORDER BY quantity DESC LIMIT 1""",
                (sku, f'%{sku}%')).fetchone()
            if inv:
                new_qty = max(0, inv['quantity'] - qty)
                db.execute("UPDATE spare_parts_inventory SET quantity=?, updated_at=datetime('now','localtime') WHERE id=?",
                           (new_qty, inv['id']))
                db.execute("""INSERT INTO inventory_logs (part_id, type, quantity, ref_type, ref_id, operator, remark)
                    VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (inv['id'], 'out', qty, 'parts_request', rid,
                     g.current_user['username'] or 'admin',
                     f"巡检备件审批 #{rid}"))
        db.commit()
    return jsonify({'success': True, 'message': '已批准'})

@app.route('/api/inspection-v2/parts-request/<int:rid>/reject', methods=['PUT'])
def v2_parts_request_reject(rid):
    """驳回巡检备件预申报（需填写原因）"""
    g_ = require_approver()
    if g_:
        return g_
    data = request.get_json(silent=True) or {}
    comment = (data.get('comment') or '').strip()
    if not comment:
        return jsonify({'error': '驳回需填写原因'}), 400
    with get_db() as db:
        try:
            db.execute("ALTER TABLE parts_requests ADD COLUMN approver_id INTEGER")
            db.execute("ALTER TABLE parts_requests ADD COLUMN approve_comment TEXT")
        except Exception:
            pass
        req = db.execute('SELECT * FROM parts_requests WHERE id=?', (rid,)).fetchone()
        if not req:
            return jsonify({'error': '申请不存在'}), 404
        if req['status'] != 'pending':
            return jsonify({'error': '该申请已处理'}), 400
        db.execute(
            "UPDATE parts_requests SET status='rejected', approver_id=?, approve_comment=? WHERE id=?",
            (data.get('approver_id', 1), comment, rid))
        db.commit()
    return jsonify({'success': True, 'message': '已驳回'})

@app.route('/api/inspection-v2/plans/<int:plan_id>/items/<int:item_id>/checkin', methods=['POST'])
def v2_item_checkin(plan_id, item_id):
    """巡检签到打卡"""
    data = request.get_json(silent=True) or {}
    gps_lat = data.get('gps_lat')
    gps_lng = data.get('gps_lng')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as db:
        item = db.execute("SELECT id FROM insp_plan_items WHERE id=? AND plan_id=?", (item_id, plan_id)).fetchone()
        if not item:
            return jsonify({'error': '检查项不存在'}), 404
        db.execute("UPDATE insp_plan_items SET check_in_time=?, gps_lat=?, gps_lng=? WHERE id=? AND plan_id=?",
                   (now, gps_lat, gps_lng, item_id, plan_id))
        db.commit()
        return jsonify({'success': True, 'check_in_time': now})

@app.route('/api/inspection-v2/plans/<int:plan_id>/items/<int:item_id>/checkout', methods=['POST'])
def v2_item_checkout(plan_id, item_id):
    """巡检签退打卡"""
    data = request.get_json(silent=True) or {}
    gps_lat = data.get('gps_lat')
    gps_lng = data.get('gps_lng')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as db:
        item = db.execute("SELECT id, gps_lat FROM insp_plan_items WHERE id=? AND plan_id=?", (item_id, plan_id)).fetchone()
        if not item:
            return jsonify({'error': '检查项不存在'}), 404
        # 只更新gps如果签到未记录gps
        if not item['gps_lat'] and gps_lat is not None:
            db.execute("UPDATE insp_plan_items SET check_out_time=?, gps_lat=?, gps_lng=? WHERE id=? AND plan_id=?",
                       (now, gps_lat, gps_lng, item_id, plan_id))
        else:
            db.execute("UPDATE insp_plan_items SET check_out_time=? WHERE id=? AND plan_id=?",
                       (now, item_id, plan_id))
        db.commit()
        return jsonify({'success': True, 'check_out_time': now})

@app.route('/api/inspection-v2/items/pending', methods=['GET'])
def v2_pending_review_items():
    """获取待审核的巡检检查项列表（按站点范围隔离）"""
    denied = require_reviewer()
    if denied:
        return denied
    allowed = _filter_site_ids()
    site_clause = ''
    sp = []
    if allowed is not None:
        site_clause = f' AND i.site_id IN ({",".join("?" * len(allowed))})'
        sp = list(allowed)
    with get_db() as db:
        items = db.execute(f"""
            SELECT i.id, i.item_name, i.site_id, s.name as site_name,
                   i.plan_id, p.plan_name, i.actual_photos, i.required_photos,
                   i.remark, i.check_time, i.photo_urls, i.calibration_values
            FROM insp_plan_items i
            JOIN sites s ON i.site_id = s.id
            JOIN insp_plans p ON i.plan_id = p.id
            WHERE i.review_status = 1{site_clause}
            ORDER BY i.check_time DESC
        """, sp).fetchall()
        return jsonify([dict(r) for r in items])

@app.route('/api/inspection-v2/items/<int:item_id>/review', methods=['PUT'])
def v2_review_item(item_id):
    """审核通过或驳回检查项"""
    denied = require_reviewer()
    if denied:
        return denied
    data = request.get_json(silent=True) or {}
    action = data.get('action', 'approve')  # 'approve' or 'reject'
    comment = data.get('comment', '')
    reviewer_id = g.current_user.get('id', 0) if hasattr(g, 'current_user') else 0

    with get_db() as db:
        item = db.execute("SELECT * FROM insp_plan_items WHERE id=?", (item_id,)).fetchone()
        if not item:
            return jsonify({'error': '检查项不存在'}), 404
        if action == 'approve':
            db.execute("""
                UPDATE insp_plan_items SET review_status=2, review_comment=?, reviewer_id=?, review_time=?
                WHERE id=?
            """, (comment, reviewer_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), item_id))
            msg = '审核通过'
        else:
            db.execute("""
                UPDATE insp_plan_items SET review_status=3, review_comment=?, reviewer_id=?, review_time=?
                WHERE id=?
            """, (comment, reviewer_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), item_id))
            msg = '审核驳回'
        db.commit()
        return jsonify({'success': True, 'message': msg})

@app.route('/api/inspection-v2/items/batch-review', methods=['POST'])
def v2_batch_review_items():
    """异常驱动批量审核：一次调用完成"通过N项 + 驳回M项"。
    body: { approve_ids: [id,...], reject_items: [{id, reason},...] }
    设计意图：审核者只标记异常项，其余一键通过，减少逐条点击。"""
    denied = require_reviewer()
    if denied:
        return denied
    data = request.get_json(silent=True) or {}
    approve_ids = data.get('approve_ids') or []
    reject_items = data.get('reject_items') or []
    if not approve_ids and not reject_items:
        return jsonify({'error': '无审核内容'}), 400
    reviewer_id = g.current_user.get('id', 0) if hasattr(g, 'current_user') else 0
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    approved_count, rejected_count = 0, 0
    with get_db() as db:
        # 批量通过
        if approve_ids:
            ph = ','.join('?' * len(approve_ids))
            db.execute(f"""
                UPDATE insp_plan_items SET review_status=2, reviewer_id=?, review_time=?, review_comment='批量通过'
                WHERE id IN ({ph}) AND review_status=1
            """, [reviewer_id, now_str] + list(approve_ids))
            approved_count = db.execute("SELECT changes()").fetchone()[0]
        # 逐项驳回（各有原因）
        for ri in reject_items:
            rid = ri.get('id')
            reason = ri.get('reason') or '照片不合格'
            if not rid:
                continue
            db.execute("""
                UPDATE insp_plan_items SET review_status=3, reviewer_id=?, review_time=?, review_comment=?
                WHERE id=? AND review_status=1
            """, (reviewer_id, now_str, reason, rid))
            rejected_count += db.execute("SELECT changes()").fetchone()[0]
            # 通知被驳回的运维人员
            item_row = db.execute("SELECT ipi.id, ipi.item_name, ip.user_id FROM insp_plan_items ipi JOIN insp_plans ip ON ipi.plan_id=ip.id WHERE ipi.id=?", (rid,)).fetchone()
            if item_row and item_row['user_id']:
                _create_notification(
                    user_id=item_row['user_id'],
                    title='巡检审核驳回',
                    content=f'检查项"{item_row["item_name"]}"被驳回：{reason}',
                    ntype='review_reject', db=db)
        db.commit()
    return jsonify({'success': True, 'approved': approved_count, 'rejected': rejected_count,
                    'message': f'通过 {approved_count} 项，驳回 {rejected_count} 项'})

@app.route('/api/inspection-v2/plans/<int:plan_id>/stats')
def v2_plan_stats(plan_id):
    """计划统计（按站点范围隔离）"""
    allowed = _filter_site_ids()
    with get_db() as db:
        plan = db.execute("SELECT * FROM insp_plans WHERE id=?", (plan_id,)).fetchone()
        if not plan:
            return jsonify({'error': '计划不存在'}), 404
        if allowed is not None:
            plan_sites = [r['site_id'] for r in db.execute(
                "SELECT DISTINCT site_id FROM insp_plan_items WHERE plan_id=?", (plan_id,)).fetchall()]
            if not (set(plan_sites) & set(allowed)):
                return jsonify({'error': '无权限访问该计划'}), 403
        total = db.execute("SELECT COUNT(*) FROM insp_plan_items WHERE plan_id=? AND COALESCE(execution_status, 'active')='active'", (plan_id,)).fetchone()[0]
        done = db.execute("SELECT COUNT(*) FROM insp_plan_items WHERE plan_id=? AND result IS NOT NULL AND COALESCE(execution_status, 'active')='active'", (plan_id,)).fetchone()[0]
        normal = db.execute("SELECT COUNT(*) FROM insp_plan_items WHERE plan_id=? AND result='normal' AND COALESCE(execution_status, 'active')='active'", (plan_id,)).fetchone()[0]
        abnormal = db.execute("SELECT COUNT(*) FROM insp_plan_items WHERE plan_id=? AND result='abnormal' AND COALESCE(execution_status, 'active')='active'", (plan_id,)).fetchone()[0]
        sites = db.execute("SELECT COUNT(DISTINCT site_id) FROM insp_plan_items WHERE plan_id=? AND COALESCE(execution_status, 'active')='active'", (plan_id,)).fetchone()[0]
        return jsonify({
            'total_items': total, 'completed_items': done,
            'normal_count': normal, 'abnormal_count': abnormal,
            'site_count': sites,
            'completion_rate': round(done / total * 100, 1) if total > 0 else 0,
        })

@app.route('/api/inspection-v2/stats')
def v2_overall_stats():
    """总体统计（按站点范围隔离）"""
    today = datetime.now().strftime('%Y-%m-%d')
    allowed = _filter_site_ids()
    sc = ''
    sp = []
    if allowed is not None:
        sc = f' AND site_id IN ({",".join("?" * len(allowed))})'
        sp = list(allowed)
    plan_filter = ''
    pf = []
    if allowed is not None:
        ph = ','.join('?' * len(allowed))
        plan_filter = f" AND id IN (SELECT plan_id FROM insp_plan_items WHERE site_id IN ({ph}))"
        pf = list(allowed)
    with get_db() as db:
        total_schedules = db.execute(f"SELECT COUNT(*) FROM inspection_schedules WHERE status='active'{sc}", sp).fetchone()[0]
        due_items = db.execute(f"SELECT COUNT(*) FROM inspection_schedules WHERE status='active' AND next_due_date<=?{sc}", (today,) + tuple(sp)).fetchone()[0]
        overdue_items = db.execute(f"SELECT COUNT(*) FROM inspection_schedules WHERE status='active' AND next_due_date<?{sc}", (today,) + tuple(sp)).fetchone()[0]
        upcoming_items = db.execute(f"""
            SELECT COUNT(*) FROM inspection_schedules
            WHERE status='active' AND next_due_date>? AND next_due_date<=date(?, '+7 days'){sc}
        """, (today, today) + tuple(sp)).fetchone()[0]
        total_plans = db.execute(f"SELECT COUNT(*) FROM insp_plans{plan_filter}", pf).fetchone()[0]
        active_plans = db.execute(f"SELECT COUNT(*) FROM insp_plans WHERE status='active'{plan_filter}", pf).fetchone()[0]
        completed_plans = db.execute(f"SELECT COUNT(*) FROM insp_plans WHERE status='completed'{plan_filter}", pf).fetchone()[0]
        total_templates = db.execute("SELECT COUNT(*) FROM inspection_templates WHERE status='active'").fetchone()[0]
        total_configs = db.execute("SELECT COUNT(*) FROM inspection_configs WHERE is_active=1").fetchone()[0]
        return jsonify({
            'total_schedules': total_schedules,
            'due_items': due_items,
            'overdue_items': overdue_items,
            'upcoming_items': upcoming_items,
            'total_plans': total_plans,
            'active_plans': active_plans,
            'completed_plans': completed_plans,
            'total_templates': total_templates,
            'total_configs': total_configs,
        })

# ===== 统一待办审核（巡检质控审核 + 工单影像审核） =====
def require_admin():
    """纵深防御：系统级管理动作仅管理员可执行；非管理员直接 403。
    用于补足前端角色门禁（路由守卫 + 菜单隐藏），即使绕过前端直调接口也调不通。"""
    u = g.get('current_user')
    if not u or u.get('role') != 'admin':
        return jsonify({'success': False, 'error': '需要管理员权限'}), 403
    return None

def require_approver():
    """审核/审批门禁：管理员(admin)与主管(manager)均可；其余角色 403。
    对应产品定位——网页端与移动端的审核/审批由管理者与审批者执行。"""
    u = g.get('current_user')
    if not u or u.get('role') not in ('admin', 'manager'):
        return jsonify({'success': False, 'error': '需要管理员或主管权限'}), 403
    return None

def _resolve_period(period, custom_start=None, custom_end=None):
    """统一考核周期解析。返回 (start, end, label, days)。
    支持 month(默认自然月) / 7d / 30d / quarter(本季度) / year(本年度) / custom。"""
    now = datetime.now()
    if period == '7d':
        start_dt = now - timedelta(days=6)
        start = start_dt.strftime('%Y-%m-%d 00:00:00'); end = now.strftime('%Y-%m-%d 23:59:59')
        label = '近7天'
    elif period == '30d':
        start_dt = now - timedelta(days=29)
        start = start_dt.strftime('%Y-%m-%d 00:00:00'); end = now.strftime('%Y-%m-%d 23:59:59')
        label = '近30天'
    elif period == 'quarter':
        q = (now.month - 1) // 3           # 0..3
        qm = q * 3 + 1                      # 季度首月
        start = now.strftime(f'%Y-{qm:02d}-01 00:00:00'); end = now.strftime('%Y-%m-%d 23:59:59')
        label = f'{now.year}年第{q+1}季度'
    elif period == 'year':
        start = now.strftime('%Y-01-01 00:00:00'); end = now.strftime('%Y-%m-%d 23:59:59')
        label = f'{now.year}年度'
    elif period == 'custom' and custom_start and custom_end:
        start = custom_start + ' 00:00:00'; end = custom_end + ' 23:59:59'
        label = custom_start + '~' + custom_end
    else:  # month
        start = now.strftime('%Y-%m-01 00:00:00'); end = now.strftime('%Y-%m-%d 23:59:59')
        label = now.strftime('%Y-%m')
    d0 = datetime.strptime(start[:10], '%Y-%m-%d'); d1 = datetime.strptime(end[:10], '%Y-%m-%d')
    return start, end, label, (d1 - d0).days + 1


@app.route('/api/evaluation/personnel')
@login_required
def evaluation_personnel():
    """人员评估（运维口径，多维指标）：按人员管理(users.real_name)归一聚合。
    维度：工单处理量/闭环数/闭环率、平均响应时长(h)、平均处理时长(天)、
          SLA超时数、SLA达标率、巡检执行数、巡检审核数。
    按考核周期(period)过滤 work_orders.created_at。管理员看全部，操作员仅看本人。"""
    period = request.args.get('period', 'month')
    denied = require_reviewer()
    if denied:
        return denied
    start, end, period_label, days = _resolve_period(
        period, request.args.get('start'), request.args.get('end'))
    now_s = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    with get_db() as db:
        rows = db.execute("""
            SELECT u.id, u.real_name, u.role, u.phone,
                   -- 工单处理量（周期内按创建时间）
                   (SELECT COUNT(*) FROM work_orders w
                      WHERE w.assignee=u.real_name AND w.created_at BETWEEN ? AND ?) AS wo_total,
                   (SELECT COUNT(*) FROM work_orders w
                      WHERE w.assignee=u.real_name AND w.status='closed'
                        AND w.created_at BETWEEN ? AND ?) AS wo_closed,
                   -- 平均响应时长(小时)：到场签到 - 创建
                   (SELECT AVG((julianday(w.check_in_time)-julianday(w.created_at))*24.0)
                      FROM work_orders w
                      WHERE w.assignee=u.real_name AND w.check_in_time IS NOT NULL
                        AND w.created_at BETWEEN ? AND ?) AS response_hours,
                   -- 平均处理时长(天)：关单 - 创建
                   (SELECT AVG(julianday(w.resolved_at)-julianday(w.created_at))
                      FROM work_orders w
                      WHERE w.assignee=u.real_name AND w.status='closed' AND w.resolved_at IS NOT NULL
                        AND w.created_at BETWEEN ? AND ?) AS wo_avg_days,
                   -- SLA超时数：已关单且关单晚于截止 或 未关单且已过截止
                   (SELECT COUNT(*) FROM work_orders w
                      WHERE w.assignee=u.real_name AND w.sla_deadline IS NOT NULL
                        AND w.created_at BETWEEN ? AND ?
                        AND ((w.resolved_at IS NOT NULL AND w.resolved_at > w.sla_deadline)
                             OR (w.status!='closed' AND ? > w.sla_deadline))) AS sla_breach,
                   -- 巡检执行数：本人负责计划下已完成检查项
                   (SELECT COUNT(*) FROM insp_plan_items i
                      JOIN insp_plans p ON i.plan_id=p.id
                      WHERE p.assignee_id=u.id AND i.completed_at BETWEEN ? AND ?) AS insp_done,
                   -- 巡检审核数
                   (SELECT COUNT(*) FROM insp_plan_items i
                      WHERE i.reviewer_id=u.id AND i.review_time BETWEEN ? AND ?) AS insp_reviewed
            FROM users u
            WHERE u.role IN ('operator','inspector','manager')
            ORDER BY CASE u.role WHEN 'manager' THEN 0 WHEN 'operator' THEN 1 ELSE 2 END, u.real_name
        """, (start, end, start, end, start, end, start, end, start, end, now_s,
              start, end, start, end)).fetchall()

    result = []
    for r in rows:
        d = dict(r)
        total = d.get('wo_total') or 0
        closed = d.get('wo_closed') or 0
        breach = d.get('sla_breach') or 0
        d['wo_closed_rate'] = round(closed / total * 100, 1) if total else 0.0
        d['on_time_rate'] = round((total - breach) / total * 100, 1) if total else 0.0
        d['response_hours'] = round(d['response_hours'], 1) if d.get('response_hours') is not None else None
        d['wo_avg_days'] = round(d['wo_avg_days'], 1) if d.get('wo_avg_days') is not None else None
        d['sla_breach'] = breach
        result.append(d)

    # 人均概览
    n = len(result) or 1
    overview = {
        'staff_count': len(result),
        'wo_total': sum(d['wo_total'] for d in result),
        'wo_closed': sum(d['wo_closed'] for d in result),
        'sla_breach': sum(d['sla_breach'] for d in result),
        'insp_done': sum(d['insp_done'] for d in result),
    }
    overview['closed_rate'] = round(overview['wo_closed'] / overview['wo_total'] * 100, 1) if overview['wo_total'] else 0.0
    overview['on_time_rate'] = round((overview['wo_total'] - overview['sla_breach']) / overview['wo_total'] * 100, 1) if overview['wo_total'] else 0.0
    _resp = [d['response_hours'] for d in result if d['response_hours'] is not None]
    overview['avg_response_hours'] = round(sum(_resp) / len(_resp), 1) if _resp else None

    return jsonify({
        'period': period, 'period_label': period_label, 'days': days,
        'overview': overview, 'list': result,
    })


# =============================================================================
# Excel 导出（xlsx）+ 季度/年度运维报告
# =============================================================================
def _xlsx_send(wb, filename):
    """将 openpyxl Workbook 作为 xlsx 文件下载返回。"""
    if not _HAS_OPENPYXL:
        return jsonify({'success': False, 'error': '缺少 openpyxl，无法导出'}), 500
    tmp_path = os.path.join(tempfile.gettempdir(), filename)
    wb.save(tmp_path)
    return send_file(tmp_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


def _set_header_style(cell):
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='2B6CFF')
    cell.alignment = Alignment(horizontal='center', vertical='center')


def _build_work_orders_export(db, start, end, site_id=None, status=None, assignee=None):
    """构造工单导出数据（周期内创建）。"""
    params = [start, end]
    where = "WHERE w.created_at BETWEEN ? AND ?"
    if site_id:
        where += " AND w.site_id=?"; params.append(int(site_id))
    if status:
        where += " AND w.status=?"; params.append(status)
    if assignee:
        where += " AND w.assignee=?"; params.append(assignee)
    rows = db.execute(f"""
        SELECT w.order_no, s.name as site_name, w.title, w.event_type, w.level,
               w.assignee, w.status, w.created_at, w.check_in_time, w.resolved_at,
               w.sla_deadline, w.description
        FROM work_orders w LEFT JOIN sites s ON w.site_id=s.id
        {where} ORDER BY w.created_at DESC
    """, params).fetchall()
    data = [dict(r) for r in rows]
    status_cn = {'pending':'待处理','accepted':'已接受','dispatched':'已派发','in_progress':'处理中',
                 'reviewing':'审核中','acceptance':'验收中','resolved':'已解决','closed':'已关闭'}
    level_cn = {'normal':'普通','urgent':'紧急','critical':'严重'}
    for d in data:
        d['status_cn'] = status_cn.get(d['status'], d['status'])
        d['level_cn'] = level_cn.get(d['level'], d['level'])
    return data


@app.route('/api/export/work-orders')
@login_required
def export_work_orders():
    """导出工单列表为 xlsx。参数：period(default month)/start/end/site_id/status/assignee。"""
    if not _HAS_OPENPYXL:
        return jsonify({'success': False, 'error': '缺少 openpyxl'}), 500
    period = request.args.get('period', 'month')
    start, end, label, _ = _resolve_period(period, request.args.get('start'), request.args.get('end'))
    site_id = request.args.get('site_id')
    status = request.args.get('status')
    assignee = request.args.get('assignee')

    with get_db() as db:
        data = _build_work_orders_export(db, start, end, site_id, status, assignee)

    wb = Workbook()
    ws = wb.active
    ws.title = '工单明细'
    headers = ['工单编号','站点','标题','事件类型','等级','负责人','状态','创建时间','到场时间','关单时间','SLA截止','描述']
    keys = ['order_no','site_name','title','event_type','level_cn','assignee','status_cn','created_at','check_in_time','resolved_at','sla_deadline','description']
    ws.append(headers)
    for c in ws[1]: _set_header_style(c)
    for d in data:
        ws.append([d.get(k, '') or '' for k in keys])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(36, max(10, len(str(col[0].value)) + 4))
    filename = f"工单明细_{label}.xlsx"
    return _xlsx_send(wb, filename)


@app.route('/api/export/inspection-records')
@login_required
def export_inspection_records():
    """导出巡检记录为 xlsx。参数：period(default month)/start/end。"""
    if not _HAS_OPENPYXL:
        return jsonify({'success': False, 'error': '缺少 openpyxl'}), 500
    period = request.args.get('period', 'month')
    start, end, label, _ = _resolve_period(period, request.args.get('start'), request.args.get('end'))
    with get_db() as db:
        rows = db.execute("""
            SELECT p.plan_name, s.name as site_name, i.item_name, i.category, i.frequency,
                   i.result, i.check_time, i.completed_at, i.review_status, i.actual_photos,
                   i.remark
            FROM insp_plan_items i
            JOIN insp_plans p ON i.plan_id=p.id
            JOIN sites s ON i.site_id=s.id
            WHERE (i.completed_at BETWEEN ? AND ?) OR (i.check_time BETWEEN ? AND ?)
            ORDER BY i.completed_at DESC, i.check_time DESC
        """, (start, end, start, end)).fetchall()
        data = [dict(r) for r in rows]
    review_cn = {1:'待审核',2:'已通过',3:'已驳回'}
    result_cn = {'normal':'正常','abnormal':'异常'}
    for d in data:
        d['review_status_cn'] = review_cn.get(d.get('review_status'), d.get('review_status') or '—')
        d['result_cn'] = result_cn.get(d.get('result'), d.get('result') or '—')

    wb = Workbook(); ws = wb.active; ws.title = '巡检记录'
    headers = ['巡检计划','站点','检查项','分类','频次','结果','检查时间','完成时间','审核状态','拍照数','备注']
    keys = ['plan_name','site_name','item_name','category','frequency','result_cn','check_time','completed_at','review_status_cn','actual_photos','remark']
    ws.append(headers)
    for c in ws[1]: _set_header_style(c)
    for d in data: ws.append([d.get(k, '') or '' for k in keys])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(36, max(10, len(str(col[0].value)) + 4))
    filename = f"巡检记录_{label}.xlsx"
    return _xlsx_send(wb, filename)


@app.route('/api/export/evaluation')
@login_required
def export_evaluation():
    denied = require_reviewer()
    if denied: return denied
    """导出人员评估报表为 xlsx。参数：period(default month)。"""
    if not _HAS_OPENPYXL:
        return jsonify({'success': False, 'error': '缺少 openpyxl'}), 500
    period = request.args.get('period', 'month')
    start, end, label, _ = _resolve_period(period, request.args.get('start'), request.args.get('end'))
    now_s = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    role_order = {'manager':0,'operator':1,'inspector':2}
    with get_db() as db:
        rows = db.execute("""
            SELECT u.id, u.real_name, u.role, u.phone,
                   (SELECT COUNT(*) FROM work_orders w WHERE w.assignee=u.real_name AND w.created_at BETWEEN ? AND ?) AS wo_total,
                   (SELECT COUNT(*) FROM work_orders w WHERE w.assignee=u.real_name AND w.status='closed' AND w.created_at BETWEEN ? AND ?) AS wo_closed,
                   (SELECT AVG((julianday(w.check_in_time)-julianday(w.created_at))*24.0)
                      FROM work_orders w WHERE w.assignee=u.real_name AND w.check_in_time IS NOT NULL AND w.created_at BETWEEN ? AND ?) AS response_hours,
                   (SELECT AVG(julianday(w.resolved_at)-julianday(w.created_at))
                      FROM work_orders w WHERE w.assignee=u.real_name AND w.status='closed' AND w.resolved_at IS NOT NULL AND w.created_at BETWEEN ? AND ?) AS wo_avg_days,
                   (SELECT COUNT(*) FROM work_orders w WHERE w.assignee=u.real_name AND w.sla_deadline IS NOT NULL AND w.created_at BETWEEN ? AND ?
                      AND ((w.resolved_at IS NOT NULL AND w.resolved_at > w.sla_deadline) OR (w.status!='closed' AND ? > w.sla_deadline))) AS sla_breach,
                   (SELECT COUNT(*) FROM insp_plan_items i JOIN insp_plans p ON i.plan_id=p.id
                      WHERE p.assignee_id=u.id AND i.completed_at IS NOT NULL) AS insp_done,
                   (SELECT COUNT(*) FROM insp_plan_items i WHERE i.reviewer_id=u.id) AS insp_reviewed
            FROM users u WHERE u.role IN ('operator','inspector','manager')
            ORDER BY CASE u.role WHEN 'manager' THEN 0 WHEN 'operator' THEN 1 ELSE 2 END, u.real_name
        """, (start, end, start, end, start, end, start, end, start, end, now_s)).fetchall()
        data = []
        for r in rows:
            d = dict(r)
            total = d.get('wo_total') or 0
            closed = d.get('wo_closed') or 0
            breach = d.get('sla_breach') or 0
            d['wo_closed_rate'] = round(closed / total * 100, 1) if total else 0.0
            d['on_time_rate'] = round((total - breach) / total * 100, 1) if total else 0.0
            data.append(d)

    role_cn = {'manager':'主管','operator':'运维员','inspector':'审核员'}
    wb = Workbook(); ws = wb.active; ws.title = '人员评估'
    headers = ['姓名','角色','工单处理','闭环数','闭环率(%)','平均响应(h)','平均处理(天)','SLA超时','SLA达标率(%)','巡检执行','巡检审核']
    ws.append(headers)
    for c in ws[1]: _set_header_style(c)
    for d in data:
        ws.append([
            d['real_name'], role_cn.get(d['role'], d['role']), d['wo_total'], d['wo_closed'],
            d['wo_closed_rate'], d['response_hours'] if d['response_hours'] is not None else '—',
            d['wo_avg_days'] if d['wo_avg_days'] is not None else '—',
            d['sla_breach'], d['on_time_rate'], d['insp_done'], d['insp_reviewed']
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(20, max(10, len(str(col[0].value)) + 4))
    filename = f"人员评估_{label}.xlsx"
    return _xlsx_send(wb, filename)


@app.route('/api/export/ops-report')
@login_required
def export_ops_report():
    denied = require_reviewer()
    if denied: return denied
    """导出季度/年度运维报告（多 sheet）：概览、人员绩效、工单明细、站点健康度、告警统计。
    参数：period(default quarter) / year；也支持 custom start/end。"""
    if not _HAS_OPENPYXL:
        return jsonify({'success': False, 'error': '缺少 openpyxl'}), 500
    period = request.args.get('period', 'quarter')
    start, end, label, days = _resolve_period(period, request.args.get('start'), request.args.get('end'))
    now_s = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    wb = Workbook()
    # 移除默认 sheet，后续按需创建
    wb.remove(wb.active)

    with get_db() as db:
        # ---- Sheet 1: 概览 ----
        ws1 = wb.create_sheet('概览')
        site_total = db.execute("SELECT COUNT(*) FROM sites").fetchone()[0]
        sites_online = db.execute("SELECT COUNT(*) FROM sites WHERE status='online'").fetchone()[0]
        alerts_total = db.execute("SELECT COUNT(*) FROM alerts WHERE created_at BETWEEN ? AND ?", (start, end)).fetchone()[0]
        alerts_pending = db.execute("SELECT COUNT(*) FROM alerts WHERE status='pending' AND created_at BETWEEN ? AND ?", (start, end)).fetchone()[0]
        wo_total = db.execute("SELECT COUNT(*) FROM work_orders WHERE created_at BETWEEN ? AND ?", (start, end)).fetchone()[0]
        wo_closed = db.execute("SELECT COUNT(*) FROM work_orders WHERE status='closed' AND created_at BETWEEN ? AND ?", (start, end)).fetchone()[0]
        insp_done = db.execute("SELECT COUNT(*) FROM insp_plan_items WHERE completed_at BETWEEN ? AND ?", (start, end)).fetchone()[0]
        overview_rows = [
            ['运维报告', label],
            ['统计周期天数', days],
            ['站点总数', site_total],
            ['在线站点', sites_online],
            ['告警总数', alerts_total],
            ['待处理告警', alerts_pending],
            ['工单总数', wo_total],
            ['已闭环工单', wo_closed],
            ['闭环率(%)', round(wo_closed / wo_total * 100, 1) if wo_total else 0],
            ['巡检完成项', insp_done],
        ]
        for r in overview_rows:
            ws1.append(r)
        ws1.column_dimensions['A'].width = 20
        ws1.column_dimensions['B'].width = 20

        # ---- Sheet 2: 人员绩效 ----
        ws2 = wb.create_sheet('人员绩效')
        headers2 = ['姓名','角色','工单处理','闭环数','闭环率(%)','平均响应(h)','平均处理(天)','SLA超时','SLA达标率(%)','巡检执行','巡检审核']
        ws2.append(headers2)
        for c in ws2[1]: _set_header_style(c)
        rows = db.execute("""
            SELECT u.real_name, u.role,
                   (SELECT COUNT(*) FROM work_orders w WHERE w.assignee=u.real_name AND w.created_at BETWEEN ? AND ?) AS wo_total,
                   (SELECT COUNT(*) FROM work_orders w WHERE w.assignee=u.real_name AND w.status='closed' AND w.created_at BETWEEN ? AND ?) AS wo_closed,
                   (SELECT AVG((julianday(w.check_in_time)-julianday(w.created_at))*24.0)
                      FROM work_orders w WHERE w.assignee=u.real_name AND w.check_in_time IS NOT NULL AND w.created_at BETWEEN ? AND ?) AS response_hours,
                   (SELECT AVG(julianday(w.resolved_at)-julianday(w.created_at))
                      FROM work_orders w WHERE w.assignee=u.real_name AND w.status='closed' AND w.resolved_at IS NOT NULL AND w.created_at BETWEEN ? AND ?) AS wo_avg_days,
                   (SELECT COUNT(*) FROM work_orders w WHERE w.assignee=u.real_name AND w.sla_deadline IS NOT NULL AND w.created_at BETWEEN ? AND ?
                      AND ((w.resolved_at IS NOT NULL AND w.resolved_at > w.sla_deadline) OR (w.status!='closed' AND ? > w.sla_deadline))) AS sla_breach,
                   (SELECT COUNT(*) FROM insp_plan_items i JOIN insp_plans p ON i.plan_id=p.id
                      WHERE p.assignee_id=u.id AND i.completed_at IS NOT NULL) AS insp_done,
                   (SELECT COUNT(*) FROM insp_plan_items i WHERE i.reviewer_id=u.id) AS insp_reviewed
            FROM users u WHERE u.role IN ('operator','inspector','manager')
            ORDER BY CASE u.role WHEN 'manager' THEN 0 WHEN 'operator' THEN 1 ELSE 2 END, u.real_name
        """, (start, end, start, end, start, end, start, end, start, end, now_s)).fetchall()
        role_cn = {'manager':'主管','operator':'运维员','inspector':'审核员'}
        for r in rows:
            d = dict(r); total = d['wo_total'] or 0; closed = d['wo_closed'] or 0; breach = d['sla_breach'] or 0
            ws2.append([
                d['real_name'], role_cn.get(d['role'], d['role']), total, closed,
                round(closed / total * 100, 1) if total else 0,
                round(d['response_hours'], 1) if d['response_hours'] is not None else '—',
                round(d['wo_avg_days'], 1) if d['wo_avg_days'] is not None else '—',
                breach,
                round((total - breach) / total * 100, 1) if total else 0,
                d['insp_done'] or 0, d['insp_reviewed'] or 0
            ])
        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = min(20, max(10, len(str(col[0].value)) + 4))

        # ---- Sheet 3: 工单明细 ----
        ws3 = wb.create_sheet('工单明细')
        wo_data = _build_work_orders_export(db, start, end)
        headers3 = ['工单编号','站点','标题','事件类型','等级','负责人','状态','创建时间','到场时间','关单时间','SLA截止']
        ws3.append(headers3)
        for c in ws3[1]: _set_header_style(c)
        for d in wo_data:
            ws3.append([d.get(k, '') or '' for k in ['order_no','site_name','title','event_type','level_cn','assignee','status_cn','created_at','check_in_time','resolved_at','sla_deadline']])
        for col in ws3.columns:
            ws3.column_dimensions[col[0].column_letter].width = min(36, max(10, len(str(col[0].value)) + 4))

        # ---- Sheet 4: 站点健康度（按阈值实时判定异常/超限） ----
        ws4 = wb.create_sheet('站点健康度')
        ws4.append(['站点','负责人','应报数','异常数','正常数','缺失','超限','健康度(%)'])
        for c in ws4[1]: _set_header_style(c)
        thresholds_map = {t['metric']: dict(t) for t in db.execute("SELECT * FROM param_thresholds").fetchall()}
        site_rows = db.execute("SELECT id, name, manager FROM sites ORDER BY id").fetchall()
        for s in site_rows:
            records = db.execute("""
                SELECT metric, value FROM sensor_data
                WHERE site_id=? AND recorded_at BETWEEN ? AND ?
            """, (s['id'], start, end)).fetchall()
            expected = len(records)
            missing = sum(1 for r in records if r['value'] is None or r['value'] == '')
            over_limit = 0
            for r in records:
                v = r['value']
                if v is None or v == '':
                    continue
                try:
                    v = float(v)
                except (TypeError, ValueError):
                    continue
                t = thresholds_map.get(r['metric'])
                if not t:
                    continue
                if (t.get('critical_low') is not None and v <= float(t['critical_low'])) or \
                   (t.get('critical_high') is not None and v >= float(t['critical_high'])) or \
                   (t.get('low') is not None and v < float(t['low'])) or \
                   (t.get('high') is not None and v > float(t['high'])):
                    over_limit += 1
            abnormal = over_limit + missing
            normal = expected - abnormal
            rate = round(normal / expected * 100, 1) if expected else 100.0
            ws4.append([s['name'], s['manager'] or '—', expected, abnormal, normal, missing, over_limit, rate])
        for col in ws4.columns:
            ws4.column_dimensions[col[0].column_letter].width = min(24, max(10, len(str(col[0].value)) + 4))

        # ---- Sheet 5: 告警统计 ----
        ws5 = wb.create_sheet('告警统计')
        ws5.append(['指标','等级','状态','数量'])
        for c in ws5[1]: _set_header_style(c)
        alert_rows = db.execute("""
            SELECT metric, level, status, COUNT(*) as cnt
            FROM alerts WHERE created_at BETWEEN ? AND ?
            GROUP BY metric, level, status
            ORDER BY cnt DESC
        """, (start, end)).fetchall()
        metric_cn = {'codmn':'高锰酸盐指数','ammonia':'氨氮','total_phosphorus':'总磷','total_nitrogen':'总氮',
                     'water_temp':'水温','dissolved_oxygen':'溶解氧','ph':'pH','turbidity':'浊度',
                     'conductivity':'电导率','device_status':'设备状态','data_gap':'数据缺失','data_freeze':'数据冻结','data_spike':'数据突变'}
        level_cn = {'red':'红色','orange':'橙色','yellow':'黄色','blue':'蓝色'}
        status_cn = {'pending':'待处理','acknowledged':'已确认','resolved':'已解决'}
        for r in alert_rows:
            d = dict(r)
            ws5.append([metric_cn.get(d['metric'], d['metric']), level_cn.get(d['level'], d['level']),
                        status_cn.get(d['status'], d['status']), d['cnt']])
        for col in ws5.columns:
            ws5.column_dimensions[col[0].column_letter].width = min(24, max(10, len(str(col[0].value)) + 4))

    filename = f"运维报告_{label}.xlsx"
    return _xlsx_send(wb, filename)


@app.route('/api/audit/pending')
def audit_pending():
    """返回所有待审核项（巡检检查项 + 工单照片项），统一列表"""
    g_ = require_reviewer()
    if g_:
        return g_
    result = []
    allowed = _filter_site_ids()  # 非管理员/有站点绑定 → 仅本人站点待办
    with get_db() as db:
        # 1. 巡检待审项
        insp_items = db.execute("""
            SELECT i.id, i.item_name as title, i.site_id, s.name as site_name,
                   i.plan_id, p.plan_name as source_name, i.actual_photos, i.required_photos,
                   i.remark, i.check_time as submit_time, i.photo_urls
            FROM insp_plan_items i
            JOIN sites s ON i.site_id = s.id
            JOIN insp_plans p ON i.plan_id = p.id
            WHERE i.review_status = 1
            ORDER BY i.check_time DESC
        """).fetchall()
        for r in insp_items:
            if allowed is not None and r['site_id'] not in allowed:
                continue
            rd = dict(r)
            rd['source_type'] = 'inspection'
            rd['source_label'] = '巡检质控'
            rd['id'] = f'insp_{rd["id"]}'
            result.append(rd)

        # 2. 工单待审项：两种类型
        # 2a. 有未审核照片的工单（照片审核）
        # 2b. 状态为reviewing/pending但无照片的工单（工单审核）
        wos = db.execute("""
            SELECT w.id, w.order_no, w.site_id, s.name as site_name, w.event_type, w.title, w.description, w.status, w.images, w.assignee
            FROM work_orders w
            LEFT JOIN sites s ON w.site_id = s.id
            WHERE w.status != 'closed'
            ORDER BY w.created_at DESC
        """).fetchall()
        for wo in wos:
            wo_dict = dict(wo)
            if allowed is not None and wo_dict.get('site_id') not in allowed:
                continue
            # 取该工单下所有未审核的照片（含ID和URL）
            photo_rows = db.execute("""
                SELECT id, filename, stored_path, description, category
                FROM operation_attachments
                WHERE source_type='workorder' AND source_id=?
                  AND file_type='image' AND is_deleted=0
                  AND (review_status IS NULL OR review_status='pending')
                ORDER BY created_at DESC
            """, (wo_dict['id'],)).fetchall()
            total_photos = len(photo_rows)
            
            # 2a: 有未审照片 → 照片审核项
            if total_photos > 0:
                result.append({
                    'source_type': 'workorder_photo',
                    'source_label': '工单影像',
                    'id': f'wo_pic_{wo_dict["id"]}',
                    'title': f'处置照片审核({total_photos}张)',
                    'site_id': wo_dict.get('site_id'),
                    'site_name': wo_dict.get('site_name') or '',
                    'source_name': wo_dict.get('order_no', ''),
                    'source_title': wo_dict.get('title', ''),
                    'actual_photos': total_photos,
                    'required_photos': 1,
                    'remark': wo_dict.get('description', '') or '',
                    'submit_time': wo_dict.get('created_at', ''),
                    'photo_urls': wo_dict.get('images') or '',
                    'order_no': wo_dict.get('order_no', ''),
                    'attachment_ids': [p['id'] for p in photo_rows],
                    'attachment_details': [dict(p) for p in photo_rows],
                    'status': wo_dict.get('status', ''),
                })
            
            # 2b: 无照片且状态为 reviewing（待审核，需审核人处理）→ 工单状态审核项；pending 为受理人环节，不纳入审核
            elif wo_dict['status'] == 'reviewing':
                status_label = '工单办结审核'
                result.append({
                    'source_type': 'workorder_status',
                    'source_label': status_label,
                    'id': f'wo_stat_{wo_dict["id"]}',
                    'title': f'{status_label}: {wo_dict.get("title", "")}',
                    'site_id': wo_dict.get('site_id'),
                    'site_name': wo_dict.get('site_name') or '',
                    'source_name': wo_dict.get('order_no', ''),
                    'source_title': wo_dict.get('title', ''),
                    'actual_photos': 0,
                    'required_photos': 0,
                    'remark': wo_dict.get('description', '') or '',
                    'submit_time': wo_dict.get('created_at', ''),
                    'photo_urls': '',
                    'order_no': wo_dict.get('order_no', ''),
                    'attachment_ids': [],
                    'attachment_details': [],
                    'status': wo_dict['status'],
                    'assignee': wo_dict.get('assignee', ''),
                })
        
        # 3. 巡检备件预申报待审项（parts_requests, status='pending'）
        prs = db.execute("""
            SELECT pr.id, pr.plan_id, pr.requester_id, pr.created_at as submit_time,
                   p.plan_name, u.real_name as requester_name
            FROM parts_requests pr
            LEFT JOIN insp_plans p ON pr.plan_id = p.id
            LEFT JOIN users u ON pr.requester_id = u.id
            WHERE pr.status = 'pending'
            ORDER BY pr.created_at DESC
        """).fetchall()
        for pr in prs:
            pd = dict(pr)
            site_row = db.execute("""
                SELECT s.id as site_id, s.name FROM insp_plan_items i JOIN sites s ON i.site_id = s.id
                WHERE i.plan_id = ? LIMIT 1
            """, (pd['plan_id'],)).fetchone()
            pd['site_name'] = site_row['name'] if site_row else ''
            if allowed is not None and (site_row['site_id'] if site_row else None) not in allowed:
                continue
            items = db.execute("""
                SELECT pri.part_sku, pri.quantity,
                       spi.part_name, spi.manufacturer, spi.model
                FROM parts_request_items pri
                LEFT JOIN spare_parts_inventory spi ON pri.part_sku = spi.part_code
                WHERE pri.request_id=?
            """, (pd['id'],)).fetchall()
            pd['parts_detail'] = [dict(x) for x in items]
            pd['source_type'] = 'parts_request'
            pd['source_label'] = '备件预申报'
            pd['id'] = f'pr_{pd["id"]}'
            pd['title'] = f'巡检备件预申报（{len(items)}项）'
            pd['source_name'] = pd['plan_name'] or f'计划#{pd["plan_id"]}'
            pd['requester_name'] = pd.get('requester_name') or ''
            pd['actual_photos'] = 0
            pd['required_photos'] = 0
            pd['remark'] = ''
            pd['submit_time'] = pd['submit_time'] or ''
            result.append(pd)

        # 3b. 移动端/网页端备件申请待审项（spare_part_requests, status='pending'）
        # 与巡检备件预申报（parts_requests）是两张表，这里单独补读，使工单场景的备件申请进入审核程序
        sprs = db.execute("""
            SELECT sr.id, sr.request_no, sr.site_id, s.name as site_name,
                   sr.applicant, sr.part_name, sr.quantity, sr.reason,
                   sr.work_order_no, sr.created_at as submit_time
            FROM spare_part_requests sr
            LEFT JOIN sites s ON sr.site_id = s.id
            WHERE sr.status = 'pending'
            ORDER BY sr.created_at DESC
        """).fetchall()
        for spr in sprs:
            sd = dict(spr)
            if allowed is not None and sd.get('site_id') not in allowed:
                continue
            sd['source_type'] = 'spare_part_request'
            sd['source_label'] = '备件申请'
            sd['id'] = f'spr_{sd["id"]}'
            sd['title'] = f'备件申请：{sd.get("part_name", "")} ×{sd.get("quantity", 1)}'
            sd['source_name'] = sd.get('request_no', '')
            sd['site_name'] = sd.get('site_name') or ''
            sd['actual_photos'] = 0
            sd['required_photos'] = 0
            sd['remark'] = sd.get('reason') or ''
            sd['submit_time'] = sd.get('submit_time') or ''
            sd['requester_name'] = sd.get('applicant') or ''
            result.append(sd)

        # 4. 用车申请待审项（vehicle_applications, status='pending'）
        vas = db.execute("""
            SELECT va.id, va.vehicle_id, va.applicant_id, va.start_at, va.end_at,
                   va.site_id, va.destination, va.reason, va.created_at as submit_time,
                   v.plate_no, v.model, u.real_name as applicant_name
            FROM vehicle_applications va
            LEFT JOIN vehicles v ON va.vehicle_id = v.id
            LEFT JOIN users u ON va.applicant_id = u.id
            WHERE va.status = 'pending'
            ORDER BY va.created_at DESC
        """).fetchall()
        for va in vas:
            vd = dict(va)
            vd['source_type'] = 'vehicle_application'
            vd['source_label'] = '用车审批'
            vd['id'] = f'va_{vd["id"]}'
            vd['title'] = f'用车申请：{vd.get("plate_no") or ""}'
            vd['source_name'] = vd.get('plate_no') or ''
            vd['site_name'] = ''
            vd['site_id'] = vd.get('site_id')
            vd['actual_photos'] = 0
            vd['required_photos'] = 0
            vd['remark'] = vd.get('reason') or ''
            vd['submit_time'] = vd['submit_time'] or ''
            vd['applicant_name'] = vd.get('applicant_name') or ''
            vd['start_at'] = vd.get('start_at') or ''
            vd['end_at'] = vd.get('end_at') or ''
            vd['destination'] = vd.get('destination') or ''
            result.append(vd)

        # 5. 影像资料审核项：需审核的照片（operation_attachments, review_required=1, review_status=pending，排除工单照片避免与 2a 重复）
        photo_reviews = db.execute("""
            SELECT a.id as aid, a.description, a.watermark_text, a.recognized_category,
                   a.site_id, s.name as site_name, a.source_type, a.source_id,
                   a.created_at as submit_time, a.stored_path, a.filename, a.taken_at
            FROM operation_attachments a
            LEFT JOIN sites s ON a.site_id = s.id
            WHERE a.is_deleted=0 AND a.review_required=1 AND a.review_status='pending'
              AND a.source_type != 'workorder'
            ORDER BY a.created_at DESC
        """).fetchall()
        for pr in photo_reviews:
            if allowed is not None and pr['site_id'] not in allowed:
                continue
            pd = dict(pr)
            raw_id = pd['aid']
            pd['source_type'] = 'photo_review'
            pd['source_label'] = '影像审核'
            pd['id'] = f'photo_{raw_id}'
            pd['title'] = f"影像审核：{pd.get('recognized_category') or pd.get('description') or pd.get('filename') or '照片'}"
            pd['site_name'] = pd.get('site_name') or ''
            pd['actual_photos'] = 1
            pd['required_photos'] = 1
            pd['remark'] = pd.get('watermark_text') or ''
            pd['submit_time'] = pd.get('submit_time') or ''
            pd['attachment_ids'] = [raw_id]
            pd['attachment_details'] = [{
                'id': raw_id,
                'filename': pd.get('filename'),
                'stored_path': pd.get('stored_path'),
                'description': pd.get('description'),
            }]
            result.append(pd)

        # 补充工单的站点名称（历史兼容分支）
        site_names = {}
        for item in result:
            if item['source_type'] == 'workorder' and not item['site_name'] and item.get('site_id'):
                if item['site_id'] not in site_names:
                    s = db.execute("SELECT name FROM sites WHERE id=?", (item['site_id'],)).fetchone()
                    site_names[item['site_id']] = s['name'] if s else ''
                item['site_name'] = site_names[item['site_id']]

    # 按提交时间降序（巡检在前，工单在后）
    result.sort(key=lambda x: x.get('submit_time') or '', reverse=True)
    return jsonify(result)


# ---------- 操作附件审核（照片驳回/通过）----------
@app.route('/api/operation-attachments/review', methods=['POST'])
def api_operation_attachments_review():
    """审核操作附件（照片），支持批量通过/驳回"""
    g_ = require_reviewer()
    if g_:
        return g_
    data = request.get_json() or {}
    attachment_ids = data.get('attachment_ids', [])
    action = data.get('action', 'approve')  # approve | reject
    reviewer_id = data.get('reviewer_id', 1)
    reject_reason = data.get('reject_reason', '')

    if not attachment_ids:
        return jsonify({'error': '缺少 attachment_ids'}), 400

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    new_status = 'approved' if action == 'approve' else 'rejected'
    with get_db() as db:
        for aid in attachment_ids:
            att = db.execute('SELECT * FROM operation_attachments WHERE id=?', (aid,)).fetchone()
            if not att:
                continue
            db.execute(
                '''UPDATE operation_attachments
                   SET review_status=?, reviewer_id=?, reviewed_at=?, reject_reason=?
                   WHERE id=?''',
                (new_status, reviewer_id, now, reject_reason if action == 'reject' else None, aid))
            if action == 'reject' and att['uploader_id']:
                name = att['description'] or att['filename'] or f'照片#{aid}'
                db.execute(
                    'INSERT INTO notifications (user_id, source_type, source_id, title, content) VALUES (?,?,?,?,?)',
                    (att['uploader_id'], 'photo_review', aid,
                     f'照片被驳回',
                     f'「{name}」被驳回，原因：{reject_reason or "未达标"}。请重新拍摄上传。'))
        db.commit()
    return jsonify({'ok': True, 'count': len(attachment_ids), 'status': new_status})


@app.route('/api/audit/stats')
def audit_stats():
    """待审核统计：返回巡检+工单+备件预申报+用车申请的待审总数"""
    g_ = require_reviewer()
    if g_:
        return g_
    total = 0
    insp_pending = 0
    wo_pending = 0
    parts_pending = 0
    vehicle_pending = 0
    with get_db() as db:
        insp_pending = db.execute("SELECT COUNT(*) as c FROM insp_plan_items WHERE review_status=1").fetchone()['c']
        # 工单待审项：与 /audit/pending 同口径（有未审照片的工单 + 状态为 reviewing/pending 的工单），保证徽标=列表行数
        wo_pending = 0
        for wo in db.execute(
            "SELECT w.id, w.status FROM work_orders w WHERE w.status != 'closed'"
        ).fetchall():
            photo_rows = db.execute(
                "SELECT COUNT(*) c FROM operation_attachments "
                "WHERE source_type='workorder' AND source_id=? AND file_type='image' AND is_deleted=0 "
                "AND (review_status IS NULL OR review_status='pending')",
                (wo['id'],)
            ).fetchone()['c']
            if photo_rows > 0 or wo['status'] == 'reviewing':
                wo_pending += 1
        # 审核员只处理材料与影像；资源审批属于管理者决策，不能作为审核员待办。
        if g.current_user['role'] in ('admin', 'manager'):
            parts_pending = db.execute("SELECT COUNT(*) as c FROM parts_requests WHERE status='pending'").fetchone()['c']
            vehicle_pending = db.execute("SELECT COUNT(*) as c FROM vehicle_applications WHERE status='pending'").fetchone()['c']
        photo_pending = db.execute(
            "SELECT COUNT(*) as c FROM operation_attachments "
            "WHERE is_deleted=0 AND review_required=1 AND review_status='pending' AND source_type!='workorder'"
        ).fetchone()['c']
        total = insp_pending + wo_pending + parts_pending + vehicle_pending + photo_pending
    return jsonify({
        'total': total,
        'inspection_pending': insp_pending,
        'workorder_pending': wo_pending,
        'parts_pending': parts_pending,
        'vehicle_pending': vehicle_pending,
        'photo_pending': photo_pending,
    })

# --- 提醒配置 ---

@app.route('/api/inspection-v2/reminders')
def v2_get_reminders():
    """获取提醒配置"""
    with get_db() as db:
        rows = db.execute("SELECT * FROM inspection_reminders").fetchall()
        return jsonify([dict(r) for r in rows] if rows else [{'remind_days_before': 1, 'remind_method': 'notification',
                 'overdue_escalation': 0, 'escalation_days': 3, 'is_active': 1}])

@app.route('/api/inspection-v2/reminders', methods=['PUT'])
def v2_update_reminders():
    """更新提醒配置"""
    data = request.get_json(silent=True) or {}
    with get_db() as db:
        existing = db.execute("SELECT id FROM inspection_reminders LIMIT 1").fetchone()
        if existing:
            fields = []
            params = []
            for key in ('remind_days_before', 'remind_method', 'overdue_escalation', 'escalation_days', 'is_active'):
                if key in data:
                    fields.append(f"{key}=?")
                    params.append(data[key])
            if fields:
                params.append(existing['id'])
                db.execute(f"UPDATE inspection_reminders SET {','.join(fields)} WHERE id=?", params)
        else:
            db.execute("""
                INSERT INTO inspection_reminders (remind_days_before,remind_method,overdue_escalation,escalation_days,is_active)
                VALUES (?,?,?,?,?)
            """, (data.get('remind_days_before', 1), data.get('remind_method', 'notification'),
                  data.get('overdue_escalation', 0), data.get('escalation_days', 3), 1))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/inspection-v2/reminders/pending')
def v2_pending_reminders():
    """获取待提醒列表"""
    today = datetime.now()
    with get_db() as db:
        cfg = db.execute("SELECT * FROM inspection_reminders WHERE is_active=1 LIMIT 1").fetchone()
        if not cfg:
            return jsonify([])
        remind_days = cfg['remind_days_before']
        cutoff = (today + timedelta(days=remind_days)).strftime('%Y-%m-%d')
        today_str = today.strftime('%Y-%m-%d')
        # 到期+逾期的排程
        rows = db.execute("""
            SELECT s.*, st.name as site_name, iti.item_name,
                   CASE WHEN s.next_due_date < ? THEN 'overdue' ELSE 'due' END as urgency
            FROM inspection_schedules s
            JOIN sites st ON s.site_id = st.id
            JOIN inspection_template_items iti ON s.template_item_id = iti.id
            WHERE s.status='active' AND s.next_due_date <= ?
            ORDER BY s.next_due_date
        """, (today_str, cutoff)).fetchall()
        return jsonify([dict(r) for r in rows])

# --- Water Level Checks (Phase A-3) ---
@app.route('/api/water-level/checks')
def get_water_level_checks():
    """水位差值校验列表"""
    site_id = request.args.get('site_id', '', type=int)
    limit = request.args.get('limit', 50, type=int)
    with get_db() as db:
        q = "SELECT w.*, s.name as site_name, s.code as site_code FROM water_level_checks w LEFT JOIN sites s ON w.site_id=s.id WHERE 1=1"
        params = []
        if site_id:
            q += " AND w.site_id=?"
            params.append(site_id)
        q += " ORDER BY w.created_at DESC LIMIT ?"
        params.append(limit)
        return jsonify([dict(r) for r in db.execute(q, params).fetchall()])

@app.route('/api/water-level/checks', methods=['POST'])
def create_water_level_check():
    """手动录入水位校验"""
    data = request.json
    site_id = data.get('site_id')
    manual_level = data.get('manual_level')
    telemetry_level = data.get('telemetry_level')
    operator = data.get('operator', '系统')
    diff = round(abs(manual_level - telemetry_level), 3)
    status = 'abnormal' if diff > 0.02 else 'normal'
    adjust_action = data.get('adjust_action', '')
    with get_db() as db:
        db.execute("""
            INSERT INTO water_level_checks (site_id,manual_level,telemetry_level,diff,status,adjust_action,operator)
            VALUES (?,?,?,?,?,?,?)
        """, (site_id, manual_level, telemetry_level, diff, status, adjust_action, operator))
        wlc_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        # 如果差值超标，生成告警
        if diff > 0.02:
            site = db.execute("SELECT name, code FROM sites WHERE id=?", (site_id,)).fetchone()
            site_name = site['name'] if site else f'站点{site_id}'
            level = 'red' if diff > 0.05 else 'orange'
            db.execute("""
                INSERT INTO alerts (site_id,level,metric,value,message,status)
                VALUES (?,?,?,?,?,?)
            """, (site_id, level, 'water_level_diff', diff,
                  f'{site_name}水位校验差值{diff}m超过阈值', 'pending'))
            # 时间线
            db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                       ('water_level', wlc_id, 'alert_generated', operator, f'水位差值{diff}m触发告警'))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('water_level', wlc_id, 'checked', operator, f'录入校验-人工{manual_level}m/遥测{telemetry_level}m/差值{diff}m'))
        db.commit()
        return jsonify({'success': True, 'id': wlc_id, 'diff': diff, 'status': status})

@app.route('/api/water-level/checks/auto', methods=['POST'])
def auto_water_level_check():
    """自动生成水位校验记录（模拟手工录入与遥测数据的比较）"""
    with get_db() as db:
        # 选取水位站/水文站
        sites = db.execute(
            "SELECT id, name FROM sites WHERE type IN ('water_level','hydrology') ORDER BY RANDOM() LIMIT 5"
        ).fetchall()
        results = []
        for site in sites:
            manual = round(random.uniform(17.0, 24.0), 2)
            telemetry = round(manual + random.uniform(-0.03, 0.03), 2)
            diff = round(abs(manual - telemetry), 3)
            status = 'abnormal' if diff > 0.02 else 'normal'
            db.execute("""
                INSERT INTO water_level_checks (site_id,manual_level,telemetry_level,diff,status,operator)
                VALUES (?,?,?,?,?,?)
            """, (site['id'], manual, telemetry, diff, status, '自动'))
            wlc_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            if diff > 0.02:
                level = 'red' if diff > 0.05 else 'orange'
                db.execute("""
                    INSERT INTO alerts (site_id,level,metric,value,message,status)
                    VALUES (?,?,?,?,?,?)
                """, (site['id'], level, 'water_level_diff', diff,
                      f'{site["name"]}自动校验差值{diff}m超过阈值', 'pending'))
                db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                           ('water_level', wlc_id, 'alert_generated', '自动', f'自动校验差值{diff}m触发告警'))
            db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                       ('water_level', wlc_id, 'auto_checked', '自动', f'自动校验-人工{manual}m/遥测{telemetry}m/差值{diff}m'))
            results.append({'site_id': site['id'], 'site_name': site['name'], 'diff': diff, 'status': status})
        db.commit()
        return jsonify({'success': True, 'count': len(results), 'results': results})

# --- Data Arrival ---
@app.route('/api/data/arrival')
def get_data_arrival():
    """当日到报率数据（按站点范围隔离）"""
    date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    allowed = _filter_site_ids()
    sc = f' AND da.site_id IN ({",".join("?" * len(allowed))})' if allowed is not None else ''
    scp = list(allowed) if allowed is not None else []
    with get_db() as db:
        rows = db.execute(
            f"SELECT da.*, s.name as site_name, s.code as site_code, s.type as site_type FROM data_arrival da LEFT JOIN sites s ON da.site_id=s.id WHERE da.date=?{sc} ORDER BY da.arrival_rate ASC",
            (date,) + tuple(scp)
        ).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route('/api/data/arrival/summary')
def data_arrival_summary():
    """到报率汇总：按项目分类（按站点范围隔离）"""
    date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    allowed = _filter_site_ids()
    sc = f' AND da.site_id IN ({",".join("?" * len(allowed))})' if allowed is not None else ''
    scp = list(allowed) if allowed is not None else []
    ssc = f' AND s.id IN ({",".join("?" * len(allowed))})' if allowed is not None else ''
    with get_db() as db:
        # 检查是否有今天的数据，没有则尝试从sensor_data实时计算
        has_data = db.execute(f"SELECT COUNT(*) as c FROM data_arrival da WHERE date=?{sc}", (date,) + tuple(scp)).fetchone()['c']
        if has_data == 0:
            # 回退：从sensor_data统计今天的到报情况
            rows = db.execute(f"""
                SELECT s.type as metric, 
                    COUNT(DISTINCT sd.site_id) as site_count,
                    ROUND(AVG(CASE WHEN sd.id IS NOT NULL THEN 100.0 ELSE 0 END),1) as avg_rate,
                    0 as below_threshold
                FROM sites s
                LEFT JOIN sensor_data sd ON s.id = sd.site_id AND sd.recorded_at >= ?
                {ssc}
                GROUP BY s.type
            """, (date + ' 00:00:00',) + tuple(scp)).fetchall()
            # 如果回退也无数据，返回空
            if not rows or all((r['avg_rate'] or 0) == 0 for r in rows):
                return jsonify({'total_avg': 0, 'by_metric': []})
            total_avg = round(sum(r['avg_rate'] or 0 for r in rows) / max(len(rows), 1), 1)
            return jsonify({
                'total_avg': total_avg,
                'by_metric': [dict(r) for r in rows]
            })
        total = db.execute(f"SELECT AVG(arrival_rate) as avg FROM data_arrival da WHERE date=?{sc}", (date,) + tuple(scp)).fetchone()
        rows = db.execute(f"""
            SELECT da.metric,
                COUNT(da.site_id) as site_count,
                ROUND(AVG(da.arrival_rate),1) as avg_rate,
                0 as below_threshold
            FROM data_arrival da
            WHERE da.date=?{sc}
            GROUP BY da.metric
        """, (date,) + tuple(scp)).fetchall()
        return jsonify({
            'total_avg': round(total['avg'],1) if total and total['avg'] else 0,
            'by_metric': [dict(r) for r in rows]
        })

@app.route('/api/analysis/trends')
@login_required
def analysis_trends():
    """统计分析页趋势数据（按站点范围隔离）：
       近7日到报率趋势、近12月巡检完成率趋势、设备在线/离线/维护统计。"""
    site_ids = _filter_site_ids()
    sc = f" AND da.site_id IN ({','.join('?' * len(site_ids))})" if site_ids is not None else ''
    scp = list(site_ids) if site_ids is not None else []
    with get_db() as db:
        # 近7日到报率（按天聚合，跨站点跨指标平均；无数据记为 None 以呈现断点而非假数据）
        today = datetime.now().date()
        arrival = []
        for i in range(6, -1, -1):
            d = (today - timedelta(days=i)).strftime('%Y-%m-%d')
            row = db.execute(
                f"SELECT AVG(arrival_rate) as avg FROM data_arrival da WHERE da.date=?{sc}",
                [d] + scp
            ).fetchone()
            rate = round(row['avg'], 1) if row and row['avg'] is not None else None
            arrival.append({'date': d, 'rate': rate})

        # 近12月巡检完成率（按月聚合）
        now = datetime.now()
        months = []
        for i in range(11, -1, -1):
            y = now.year + (now.month - 1 - i) // 12
            m = (now.month - 1 - i) % 12 + 1
            months.append(f"{y:04d}-{m:02d}")
        task_sc = f" AND site_id IN ({','.join('?' * len(site_ids))})" if site_ids is not None else ''
        task_scp = list(site_ids) if site_ids is not None else []
        rows = db.execute(
            f"""SELECT substr(created_at,1,7) as ym,
                       COUNT(*) as total,
                       SUM(CASE WHEN result IS NOT NULL THEN 1 ELSE 0 END) as done
                FROM inspection_tasks
                WHERE created_at >= date('now','-12 months'){task_sc}
                GROUP BY ym ORDER BY ym""",
            task_scp
        ).fetchall()
        done_map = {r['ym']: (r['done'] or 0) for r in rows}
        total_map = {r['ym']: (r['total'] or 0) for r in rows}
        inspection = []
        for ym in months:
            total = total_map.get(ym, 0)
            rate = round(done_map.get(ym, 0) / total * 100, 1) if total > 0 else None
            inspection.append({'month': ym, 'rate': rate, 'total': total,
                               'completed': done_map.get(ym, 0)})

        # 设备在线/离线/维护统计（device_shadows 按站点隔离）
        dev_sc = f" AND ds.site_id IN ({','.join('?' * len(site_ids))})" if site_ids is not None else ''
        dev_scp = list(site_ids) if site_ids is not None else []
        dev_total = db.execute(f"SELECT COUNT(*) as c FROM device_shadows ds WHERE 1=1{dev_sc}", dev_scp).fetchone()['c']
        dev_online = db.execute(f"SELECT COUNT(*) as c FROM device_shadows ds WHERE status='online'{dev_sc}", dev_scp).fetchone()['c']
        dev_offline = db.execute(f"SELECT COUNT(*) as c FROM device_shadows ds WHERE status='offline'{dev_sc}", dev_scp).fetchone()['c']
        dev_maint = db.execute(f"SELECT COUNT(*) as c FROM device_shadows ds WHERE status='maintenance'{dev_sc}", dev_scp).fetchone()['c']
        devices = {
            'total': dev_total,
            'online': dev_online,
            'offline': dev_offline,
            'fault': dev_maint,
        }
    return jsonify({
        'arrival': arrival,
        'inspection': inspection,
        'devices': devices,
    })

# ============================================================
# 数据健康度分析（动态基准学习版）
# ============================================================

# 站点类型中文映射
_STATION_TYPE_CN = {
    'rainfall': '雨量站',
    'water_level': '水位站',
    'hydrology': '水文站',
    'soil_moisture': '墒情站',
    'evaporation': '蒸发站',
    'groundwater': '地下水站',
    'station_yard': '站院',
    'water_quality': '水质监测站',
}

# 各类型站点标准日频次（用于历史数据不足时的回退）
_STANDARD_DAILY_COUNT = {
    ('rainfall', 'precipitation'): 288,       # 5min
    ('rainfall', 'cumulative_rainfall'): 288,
    ('water_level', 'water_level'): 96,       # 15min
    ('water_level', 'flow'): 96,
    ('hydrology', 'water_level'): 144,        # 10min
    ('hydrology', 'velocity'): 144,
    ('hydrology', 'flow'): 144,
    ('hydrology', 'precipitation'): 144,
    ('soil_moisture', 'soil_moisture'): 48,   # 30min
    ('soil_moisture', 'soil_temperature'): 48,
    ('evaporation', 'evaporation'): 24,       # 60min
    ('evaporation', 'temperature'): 24,
    ('evaporation', 'wind_speed'): 24,
    ('groundwater', 'groundwater_level'): 24,   # 60min
    ('groundwater', 'water_quality'): 24,
    # 水质监测站（water_quality）：真实数据约 96 条/天/组合（每参数，15min 级）
    ('water_quality', 'ph'): 96,
    ('water_quality', 'ammonia'): 96,
    ('water_quality', 'cod'): 96,
    ('water_quality', 'dissolved_oxygen'): 96,
    ('water_quality', 'total_nitrogen'): 96,
    ('water_quality', 'total_phosphorus'): 96,
    ('water_quality', 'turbidity'): 96,
    ('water_quality', 'water_temp'): 96,
    ('groundwater', 'water_level'): 24,          # 兼容旧数据
    ('station_yard', 'water_level'): 24,
    ('station_yard', 'temperature'): 288,   # 站院环境温度，5min
    ('station_yard', 'noise'): 288,         # 站院噪声，5min
}

_HIGH_THRESHOLD_TYPES = {'rainfall', 'water_level', 'hydrology'}
_HIGH_RATE = 98.0
_NORMAL_RATE = 95.0


def _p90(values):
    """计算 P90（第90百分位数）"""
    if not values:
        return 0
    sorted_v = sorted(values)
    idx = int(len(sorted_v) * 0.9)
    return sorted_v[min(idx, len(sorted_v) - 1)]


@app.route('/api/data/health')
@login_required
def data_health():
    """数据健康度分析（考核口径：正常数据量 / 应报数据总量）

    公式：健康度 = (应报总数 - 异常数) / 应报总数 x 100%
    异常 = 缺报(应报-实到) + 超限(逐条比对 param_thresholds 正常区间)
    应报 = 系统标准日频次(站点类型x参数) x 考核周期天数

    返回：
        period / period_label / days
        total:        {expected, abnormal, normal, rate, missing, over_limit}
        by_manager:   [{manager, site_count, expected, abnormal, normal, rate}]  (按负责人考核)
        by_site:      [{site_id, site_name, manager, expected, abnormal, normal, missing, over_limit, rate, metrics}]
        by_metric:     [{metric, label, expected, abnormal, normal, rate}]
    """
    period = request.args.get('period', 'month')   # month / 30d / 7d / custom
    custom_start = request.args.get('start')
    custom_end = request.args.get('end')

    now = datetime.now()
    if period == '7d':
        start_dt = now - timedelta(days=6)
        start = start_dt.strftime('%Y-%m-%d 00:00:00')
        end = now.strftime('%Y-%m-%d 23:59:59')
        period_label = '近7天'
    elif period == '30d':
        start_dt = now - timedelta(days=29)
        start = start_dt.strftime('%Y-%m-%d 00:00:00')
        end = now.strftime('%Y-%m-%d 23:59:59')
        period_label = '近30天'
    elif period == 'custom' and custom_start and custom_end:
        start = custom_start + ' 00:00:00'
        end = custom_end + ' 23:59:59'
        period_label = custom_start + '~' + custom_end
    else:  # month（默认自然月）
        start = now.strftime('%Y-%m-01 00:00:00')
        end = now.strftime('%Y-%m-%d 23:59:59')
        period_label = now.strftime('%Y-%m')

    d0 = datetime.strptime(start[:10], '%Y-%m-%d')
    d1 = datetime.strptime(end[:10], '%Y-%m-%d')
    days = (d1 - d0).days + 1
    freshness_reference = min(datetime.strptime(end, '%Y-%m-%d %H:%M:%S'), datetime.now())

    # 范围隔离：操作员仅本人负责站点；管理员返回 None=全部
    site_ids = _filter_site_ids()
    _qm = (','.join('?' * len(site_ids))) if site_ids else ''
    _site_where = f" WHERE id IN ({_qm})" if site_ids else ''
    _site_params = list(site_ids) if site_ids else []
    _sd_where = f" AND site_id IN ({_qm})" if site_ids else ''
    _sd_params = list(site_ids) if site_ids else []

    with get_db() as db:
        sites = db.execute(
            "SELECT id, name, code, type, manager FROM sites" + _site_where,
            _site_params).fetchall()
        th = {r['metric']: dict(r) for r in db.execute("SELECT * FROM param_thresholds").fetchall()}

        actual_rows = db.execute(
            "SELECT site_id, metric, COUNT(*) as actual FROM sensor_data "
            "WHERE recorded_at >= ? AND recorded_at <= ?" + _sd_where + " GROUP BY site_id, metric",
            (start, end) + tuple(_sd_params)).fetchall()
        actual_map = {(r['site_id'], r['metric']): r['actual'] for r in actual_rows}

        over_rows = db.execute(
            "SELECT s.site_id, s.metric, COUNT(*) as over FROM sensor_data s "
            "JOIN param_thresholds t ON t.metric = s.metric "
            "WHERE s.recorded_at >= ? AND s.recorded_at <= ?" + _sd_where.replace('site_id', 's.site_id') + " "
            "AND (s.value < t.low OR s.value > t.high) GROUP BY s.site_id, s.metric",
            (start, end) + tuple(_sd_params)).fetchall()
        over_map = {(r['site_id'], r['metric']): r['over'] for r in over_rows}
        latest_rows = db.execute(
            "SELECT site_id, metric, MAX(recorded_at) as latest FROM sensor_data "
            "WHERE recorded_at <= ?" + _sd_where + " GROUP BY site_id, metric",
            (end,) + tuple(_sd_params)).fetchall()
        latest_map = {(r['site_id'], r['metric']): r['latest'] for r in latest_rows}

        by_site = []
        manager_agg = {}
        metric_agg = {}
        total_expected = 0
        total_abnormal = 0
        total_missing = 0
        total_over = 0
        total_actual = 0
        total_metric_count = 0
        total_timely_metrics = 0

        for site in sites:
            sid = site['id']
            stype = site['type']
            mgr = site['manager'] or '未分配'
            site_metrics = set()
            for (s, m) in actual_map:
                if s == sid:
                    site_metrics.add(m)
            for m in th:
                site_metrics.add(m)

            exp = 0
            abn = 0
            miss = 0
            over = 0
            actual_total = 0
            timely_metrics = 0
            metric_detail = []
            for metric in sorted(site_metrics):
                day_freq = _STANDARD_DAILY_COUNT.get((stype, metric), 12)
                expected = day_freq * days
                actual = actual_map.get((sid, metric), 0)
                m_missing = max(0, expected - actual)
                m_over = over_map.get((sid, metric), 0)
                m_abn = min(expected, m_missing + m_over)
                latest_at = latest_map.get((sid, metric))
                freshness_hours = max(2.0, 48.0 / max(day_freq, 1))
                timely = False
                if latest_at:
                    try:
                        timely = (freshness_reference
                                  - datetime.strptime(latest_at, '%Y-%m-%d %H:%M:%S')).total_seconds() <= freshness_hours * 3600
                    except Exception:
                        timely = False
                exp += expected
                abn += m_abn
                miss += m_missing
                over += m_over
                actual_total += actual
                timely_metrics += 1 if timely else 0
                metric_detail.append({
                    'metric': metric,
                    'label': th.get(metric, {}).get('label', metric),
                    'expected': expected, 'actual': actual,
                    'missing': m_missing, 'over_limit': m_over, 'abnormal': m_abn,
                    'latest_at': latest_at, 'timely': timely,
                })
            normal = max(0, exp - abn)
            rate = round(normal / exp * 100, 1) if exp > 0 else 100.0
            completeness_rate = round((exp - miss) / exp * 100, 1) if exp > 0 else 100.0
            validity_rate = round(max(0, actual_total - over) / actual_total * 100, 1) if actual_total > 0 else 100.0
            timeliness_rate = round(timely_metrics / len(metric_detail) * 100, 1) if metric_detail else 100.0
            by_site.append({
                'site_id': sid, 'site_name': site['name'], 'site_code': site['code'],
                'manager': mgr, 'expected': exp, 'abnormal': abn, 'normal': normal,
                'actual': actual_total, 'missing': miss, 'over_limit': over, 'rate': rate,
                'completeness_rate': completeness_rate, 'validity_rate': validity_rate,
                'timeliness_rate': timeliness_rate, 'metrics': metric_detail,
            })
            total_expected += exp
            total_abnormal += abn
            total_missing += miss
            total_over += over
            total_actual += actual_total
            total_metric_count += len(metric_detail)
            total_timely_metrics += timely_metrics
            if mgr not in manager_agg:
                manager_agg[mgr] = {
                    'expected': 0, 'actual': 0, 'missing': 0, 'over_limit': 0,
                    'timely_metrics': 0, 'metric_count': 0, 'site_count': 0,
                }
            manager_agg[mgr]['expected'] += exp
            manager_agg[mgr]['actual'] += actual_total
            manager_agg[mgr]['missing'] += miss
            manager_agg[mgr]['over_limit'] += over
            manager_agg[mgr]['timely_metrics'] += timely_metrics
            manager_agg[mgr]['metric_count'] += len(metric_detail)
            manager_agg[mgr]['site_count'] += 1

        for site in sites:
            for m in th:
                key = (site['id'], m)
                day_freq = _STANDARD_DAILY_COUNT.get((site['type'], m), 12)
                expected = day_freq * days
                actual = actual_map.get(key, 0)
                m_abn = max(0, expected - actual) + over_map.get(key, 0)
                if m not in metric_agg:
                    metric_agg[m] = {'expected': 0, 'abnormal': 0}
                metric_agg[m]['expected'] += expected
                metric_agg[m]['abnormal'] += m_abn

    by_manager = []
    for mgr, a in sorted(manager_agg.items(), key=lambda x: x[0]):
        e = a['expected']; actual = a['actual']; missing = a['missing']; over = a['over_limit']
        abnormal = min(e, missing + over)
        by_manager.append({
            'manager': mgr, 'site_count': a['site_count'],
            'expected': e, 'actual': actual, 'abnormal': abnormal, 'normal': max(0, e - abnormal),
            'missing': missing, 'over_limit': over,
            'completeness_rate': round((e - missing) / e * 100, 1) if e > 0 else 100.0,
            'validity_rate': round(max(0, actual - over) / actual * 100, 1) if actual > 0 else 100.0,
            'timeliness_rate': round(a['timely_metrics'] / a['metric_count'] * 100, 1) if a['metric_count'] else 100.0,
        })

    by_metric = []
    for m, a in sorted(metric_agg.items(), key=lambda x: x[0]):
        e = a['expected']; ab = a['abnormal']
        by_metric.append({
            'metric': m, 'label': th.get(m, {}).get('label', m),
            'expected': e, 'abnormal': ab, 'normal': e - ab,
            'rate': round((e - ab) / e * 100, 1) if e > 0 else 100.0,
        })

    total_normal = total_expected - total_abnormal
    completeness_rate = round((total_expected - total_missing) / total_expected * 100, 1) if total_expected > 0 else 100.0
    validity_rate = round(max(0, total_actual - total_over) / total_actual * 100, 1) if total_actual > 0 else 100.0
    timeliness_rate = round(total_timely_metrics / total_metric_count * 100, 1) if total_metric_count else 100.0
    return jsonify({
        'period': period, 'period_label': period_label, 'days': days,
        'total': {
            'expected': total_expected, 'abnormal': total_abnormal,
            'normal': total_normal,
            'rate': round(total_normal / total_expected * 100, 1) if total_expected > 0 else 100.0,
            'missing': total_missing, 'over_limit': total_over,
            'actual': total_actual,
            'completeness_rate': completeness_rate,
            'validity_rate': validity_rate,
            'timeliness_rate': timeliness_rate,
        },
        'by_manager': by_manager,
        'by_site': by_site,
        'by_metric': by_metric,
    })


@app.route('/api/hotline/events')
def get_hotline_events():
    limit = request.args.get('limit', 50, type=int)
    with get_db() as db:
        rows = db.execute("SELECT * FROM hotline_events ORDER BY created_at DESC LIMIT ?", (limit,)).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route('/api/hotline/events', methods=['POST'])
def create_hotline_event():
    data = request.json
    with get_db() as db:
        db.execute("""
            INSERT INTO hotline_events (caller_name,caller_phone,event_type,description,location,operator)
            VALUES (?,?,?,?,?,?)
        """, (data.get('caller_name',''),data.get('caller_phone',''),
              data.get('event_type',''),data.get('description',''),
              data.get('location',''),data.get('operator','')))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/hotline/events/<int:event_id>/convert', methods=['POST'])
def convert_hotline_to_order(event_id):
    """热线事件转工单"""
    data = request.json
    with get_db() as db:
        event = db.execute("SELECT * FROM hotline_events WHERE id=?", (event_id,)).fetchone()
        if not event:
            return jsonify({'error': 'not found'}), 404
        now = datetime.now()
        order_no = f"WO-{now.strftime('%Y%m%d')}-{random.randint(100,999)}"
        leve = data.get('level','normal')
        sla_hours = {'normal': 72, 'urgent': 24, 'critical': 2}.get(leve, 72)
        sla_deadline = (now + timedelta(hours=sla_hours)).strftime('%Y-%m-%d %H:%M')

        # === 数据自洽性修复：尝试根据位置匹配站点 ===
        site_id = data.get('site_id')
        if not site_id and event.get('location'):
            loc = event['location']
            site_row = db.execute(
                "SELECT id FROM sites WHERE name LIKE ? OR district LIKE ? LIMIT 1",
                (f'%{loc}%', f'%{loc}%')
            ).fetchone()
            if site_row:
                site_id = site_row['id']

        db.execute("""
            INSERT INTO work_orders (order_no,site_id,source,event_type,level,title,description,assignee,status,sla_deadline)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (order_no, site_id, 'hotline', event['event_type'], leve,
              f"[热线]{event['event_type']}", event['description'],
              data.get('assignee',''), 'pending', sla_deadline))
        # 更新热线事件：设置状态、关联工单号、站点ID
        db.execute("UPDATE hotline_events SET status='dispatched', related_order_no=?, site_id=? WHERE id=?",
                   (order_no, site_id, event_id))
        db.commit()
        return jsonify({'success': True, 'order_no': order_no})

# --- Dashboard ---
@app.route('/api/dashboard/summary')
@login_required
def dashboard_summary():
    """全系统统一数据源——返回所有面板需要的计数和状态"""
    site_ids = _filter_site_ids()
    site_filter = ''
    sf_params = []
    if site_ids is not None:
        placeholders = ','.join('?' * len(site_ids))
        site_filter = f' AND site_id IN ({placeholders})'
        sf_params = list(site_ids)
    with get_db() as db:
        # 告警按状态计数
        pending = db.execute(f"SELECT COUNT(*) FROM alerts WHERE status='pending'{site_filter}", sf_params).fetchone()[0]
        acknowledged = db.execute(f"SELECT COUNT(*) FROM alerts WHERE status='acknowledged'{site_filter}", sf_params).fetchone()[0]
        resolved = db.execute(f"SELECT COUNT(*) FROM alerts WHERE status='resolved'{site_filter}", sf_params).fetchone()[0]
        total_alerts = pending + acknowledged + resolved

        # 告警按级别计数（仅活跃告警）
        alert_by_level = {}
        for lv in ['red','orange','yellow']:
            alert_by_level[lv] = db.execute(f"SELECT COUNT(*) FROM alerts WHERE level=? AND status IN ('pending','acknowledged'){site_filter}", [lv] + sf_params).fetchone()[0]

        # 告警按类型计数（仅活跃）
        alert_by_type = {
            '数据质量': db.execute(f"SELECT COUNT(*) FROM alerts WHERE metric IN ('data_gap','data_freeze','data_spike') AND status IN ('pending','acknowledged'){site_filter}", sf_params).fetchone()[0],
            '设备状态': db.execute(f"SELECT COUNT(*) FROM alerts WHERE metric='device_status' AND status IN ('pending','acknowledged'){site_filter}", sf_params).fetchone()[0],
        }
        alert_by_type['运维时效'] = total_alerts - alert_by_type['数据质量'] - alert_by_type['设备状态'] - resolved

        # 今日新增告警
        today_new = db.execute(f"SELECT COUNT(*) FROM alerts WHERE date(created_at)=date('now','localtime'){site_filter}", sf_params).fetchone()[0]

        # 告警站点数（有活跃告警的站点）
        alert_sites = db.execute(f"SELECT COUNT(DISTINCT site_id) FROM alerts WHERE status IN ('pending','acknowledged'){site_filter}", sf_params).fetchone()[0]

        # 站点状态
        if site_ids is not None:
            sp = list(site_ids)
            sites_online = db.execute(f"SELECT COUNT(*) FROM sites WHERE status='online' AND id IN ({','.join('?'*len(site_ids))})", sp).fetchone()[0]
            sites_offline = db.execute(f"SELECT COUNT(*) FROM sites WHERE status='offline' AND id IN ({','.join('?'*len(site_ids))})", sp).fetchone()[0]
        else:
            sites_online = db.execute("SELECT COUNT(*) FROM sites WHERE status='online'").fetchone()[0]
            sites_offline = db.execute("SELECT COUNT(*) FROM sites WHERE status='offline'").fetchone()[0]
        sites_with_alerts = alert_sites

        # 工单状态分布
        wo_by_status = {}
        for st in ['pending','accepted','dispatched','in_progress','reviewing','acceptance','closed']:
            wo_by_status[st] = db.execute(f"SELECT COUNT(*) FROM work_orders WHERE status=?{site_filter}", [st] + sf_params).fetchone()[0]

        # 今日工单
        today_wo = db.execute(f"SELECT COUNT(*) FROM work_orders WHERE date(created_at)=date('now','localtime'){site_filter}", sf_params).fetchone()[0]
        today_closed = db.execute(f"SELECT COUNT(*) FROM work_orders WHERE date(resolved_at)=date('now','localtime'){site_filter}", sf_params).fetchone()[0]

        # 数据到达
        arrival_row = db.execute("SELECT AVG(arrival_rate) FROM data_arrival WHERE date=(SELECT MAX(date) FROM data_arrival)").fetchone()
        arrival = arrival_row[0] if arrival_row and arrival_row[0] is not None else 0

        # 巡检
        insp_total = db.execute("SELECT COUNT(*) FROM inspection_tasks").fetchone()[0]
        insp_done = db.execute("SELECT COUNT(*) FROM inspection_tasks WHERE result IS NOT NULL").fetchone()[0]

        # 按metric分类的告警详情（预警中心分类卡片用）
        alerts_detail = list(db.execute(f"""
            SELECT metric, level, status, COUNT(*) as cnt
            FROM alerts WHERE status IN ('pending','acknowledged'){site_filter}
            GROUP BY metric, level, status
        """, sf_params))
        _METRIC_CN = {'codmn':'高锰酸盐指数','ammonia':'氨氮','total_phosphorus':'总磷','total_nitrogen':'总氮',
            'water_temp':'水温','dissolved_oxygen':'溶解氧','ph':'pH','turbidity':'浊度',
            'conductivity':'电导率','device_status':'设备状态','data_gap':'数据缺失','data_freeze':'数据冻结','data_spike':'数据突变'}
        alerts_detail_cn = []
        for r in alerts_detail:
            d = dict(r)
            d['metric_cn'] = _METRIC_CN.get(d.get('metric',''), d.get('metric',''))
            alerts_detail_cn.append(d)

        # 最新告警TOP5
        latest_alerts = db.execute(f"""
            SELECT a.*, s.name as site_name FROM alerts a LEFT JOIN sites s ON a.site_id=s.id
            WHERE a.status='pending'{site_filter} ORDER BY CASE level WHEN 'red' THEN 1 WHEN 'orange' THEN 2 ELSE 3 END, a.created_at DESC LIMIT 5
        """, sf_params).fetchall()

        # 待处理工单TOP5
        pending_orders = db.execute(f"""
            SELECT w.*, s.name as site_name FROM work_orders w LEFT JOIN sites s ON w.site_id=s.id
            WHERE w.status NOT IN ('closed'){site_filter} ORDER BY w.created_at DESC LIMIT 5
        """, sf_params).fetchall()

        _dash_level_cn = {'red':'红色','orange':'橙色','yellow':'黄色','blue':'蓝色'}
        _la = []
        for a in latest_alerts:
            d = dict(a)
            _m = d.get('metric','')
            d['metric_cn'] = _METRIC_CN.get(_m, _m)
            d['level_cn'] = _dash_level_cn.get(d.get('level',''), d.get('level',''))
            _la.append(d)

        _wo_status_cn2 = {'pending':'待处理','accepted':'已接受','dispatched':'已派发','in_progress':'处理中','reviewing':'审核中','resolved':'已解决','closed':'已关闭'}
        _wo_level_cn2 = {'normal':'普通','urgent':'紧急','critical':'严重'}
        _po = []
        for o in pending_orders:
            d = dict(o)
            d['status_cn'] = _wo_status_cn2.get(d.get('status',''), d.get('status',''))
            d['level_cn'] = _wo_level_cn2.get(d.get('level',''), d.get('level',''))
            _po.append(d)

        return jsonify({
            'alerts': {
                'total': total_alerts,
                'pending': pending,
                'acknowledged': acknowledged,
                'resolved': resolved,
                'by_level': alert_by_level,
                'by_type': alert_by_type,
                'today_new': today_new,
                'alert_sites': alert_sites,
                'detail': alerts_detail_cn,
            },
            'sites': {
                'total': sites_online + sites_offline,
                'online': sites_online,
                'offline': sites_offline,
                'with_alerts': sites_with_alerts,
            },
            'workorders': {
                'total': sum(wo_by_status.values()),
                'by_status': wo_by_status,
                'today_new': today_wo,
                'today_closed': today_closed,
            },
            'inspections': {
                'total': insp_total,
                'completed': insp_done,
            },
            'arrival_rate': round(arrival, 1),
            # 兼容旧版字段
            'overview': {
                'total_sites': sites_online + sites_offline,
                'online_sites': sites_online,
                'device_total': 0,
                'device_online': 0,
                'active_alerts': pending,
                'open_orders': sum(wo_by_status.values()) - wo_by_status.get('closed', 0),
                'today_orders': today_wo,
            },
            'latest_alerts': _la,
            'pending_orders': _po,
        })

# --- 实时天气获取 ---
# 默认坐标：南昌（可修改为项目所在地）
WEATHER_LAT = 28.68
WEATHER_LON = 115.89

def fetch_real_weather():
    """从 Open-Meteo 免费接口获取实时天气，写入数据库"""
    url = (
        f"https://api.open-meteo.com/v1/forecast?"
        f"latitude={WEATHER_LAT}&longitude={WEATHER_LON}"
        f"&current=temperature_2m,relative_humidity_2m,precipitation,pressure_msl,"
        f"wind_speed_10m,wind_direction_10m,weather_code"
        f"&timezone=auto"
    )
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'WaterOps/1.0'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = _json.loads(resp.read().decode())
    except Exception as e:
        print(f"[Weather] 获取天气失败: {e}")
        return False

    current = data.get('current', {})
    if not current:
        return False

    # WMO 天气代码 → 中文天气类型
    wmo_map = {
        0:'晴',1:'晴',2:'多云',3:'阴',45:'雾',48:'雾',
        51:'小雨',53:'小雨',55:'中雨',56:'冻雨',57:'冻雨',
        61:'小雨',63:'中雨',65:'大雨',66:'冻雨',67:'冻雨',
        71:'小雪',73:'中雪',75:'大雪',77:'雪粒',
        80:'阵雨',81:'中雨',82:'大雨',85:'阵雪',86:'阵雪',
        95:'雷雨',96:'雷雨',99:'雷雨'
    }
    wcode = current.get('weather_code', 0)
    weather_type = wmo_map.get(wcode, '多云')

    # 温度(°C), 湿度(%), 降水量(mm), 气压(hPa), 风速(km/h)
    temp = current.get('temperature_2m', 25)
    humidity = current.get('relative_humidity_2m', 60)
    precip = current.get('precipitation', 0)
    pressure = current.get('pressure_msl', 1013)
    wind_speed = current.get('wind_speed_10m', 5)
    wind_dir_deg = current.get('wind_direction_10m', 0)

    # 风向角度 → 中文
    dirs = ['北','东北','东','东南','南','西南','西','西北']
    wind_dir = dirs[round(wind_dir_deg / 45) % 8]

    # 生成预警信息
    warnings = []
    if precip > 10: warnings.append('暴雨')
    if wind_speed > 40: warnings.append('大风')
    if temp > 35: warnings.append('高温')

    with get_db() as db:
        # 清理旧数据，只保留最新一条
        db.execute("DELETE FROM weather_data WHERE id NOT IN (SELECT id FROM weather_data ORDER BY id DESC LIMIT 1)")
        db.execute(
            """INSERT INTO weather_data (temperature,humidity,wind_speed,wind_direction,
               precipitation,pressure,weather_type,warning_info,recorded_at)
               VALUES (?,?,?,?,?,?,?,?,datetime('now','localtime'))""",
            (temp, humidity, wind_speed, wind_dir, precip, pressure, weather_type,
             ','.join(warnings) if warnings else '')
        )
        db.commit()
    print(f"[Weather] 已更新: {weather_type} {temp}°C 降水{precip}mm 风速{wind_speed}km/h")
    return True

# --- 天气数据 ---
@app.route('/api/weather')
def get_weather():
    """返回当前天气数据、未来24小时逐时预报、天气预警"""
    with get_db() as db:
        # 获取最新天气记录
        current = db.execute(
            "SELECT * FROM weather_data ORDER BY recorded_at DESC LIMIT 1"
        ).fetchone()

        # 无数据或数据超过5分钟 → 刷新（确保移动端/PC端看到一致的最新数据）
        need_fetch = False
        if not current:
            need_fetch = True
        else:
            try:
                from datetime import datetime as _dt
                last = _dt.strptime(current['recorded_at'], '%Y-%m-%d %H:%M:%S')
                if (_dt.now() - last).total_seconds() > 300:  # 5分钟
                    need_fetch = True
            except:
                need_fetch = True

        if need_fetch:
            fetch_real_weather()
            # 重新读取
            current = db.execute(
                "SELECT * FROM weather_data ORDER BY recorded_at DESC LIMIT 1"
            ).fetchone()

        if not current:
            return jsonify({'error': '暂无天气数据，请稍后重试'}), 503

        # 构建当前天气
        now = datetime.now()
        result = {
            'current': {
                'temperature': current['temperature'],
                'humidity': current['humidity'],
                'wind_speed': current['wind_speed'],
                'wind_direction': current['wind_direction'],
                'precipitation': current['precipitation'],
                'pressure': current['pressure'],
                'weather_type': current['weather_type'],
                'recorded_at': current['recorded_at'],
            },
            'hourly_forecast': [],  # 未来24小时逐小时预报
            'warnings': [],  # 天气预警
        }

        # 解析预警信息
        if current['warning_info']:
            for w in current['warning_info'].split(','):
                if '暴雨' in w:
                    result['warnings'].append({'type': '暴雨', 'level': '黄色', 'message': '预计未来6小时有暴雨，请加强防范'})
                elif '大风' in w:
                    result['warnings'].append({'type': '大风', 'level': '蓝色', 'message': '风力已达6级以上，请加固设施'})
                elif '高温' in w:
                    result['warnings'].append({'type': '高温', 'level': '橙色', 'message': '最高气温超过35℃，请做好防暑降温'})

        # 生成未来24小时逐小时预报
        directions = ['北', '东北', '东', '东南', '南', '西南', '西', '西北']
        weather_types = ['晴', '多云', '阴', '小雨', '中雨', '大雨']
        for i in range(24):
            forecast_time = now + timedelta(hours=i+1)
            # 基于当前值做小幅随机波动
            temp_forecast = round(current['temperature'] + random.uniform(-3, 3), 1)
            hum_forecast = round(current['humidity'] + random.uniform(-10, 10), 1)
            hum_forecast = max(20, min(100, hum_forecast))
            wind_forecast = round(current['wind_speed'] + random.uniform(-2, 2), 1)
            wind_forecast = max(0, wind_forecast)
            precip_forecast = round(current['precipitation'] * random.uniform(0.5, 1.5), 1)

            result['hourly_forecast'].append({
                'time': forecast_time.strftime('%H:%M'),
                'datetime': forecast_time.strftime('%Y-%m-%d %H:%M'),
                'temperature': temp_forecast,
                'humidity': hum_forecast,
                'wind_speed': wind_forecast,
                'wind_direction': random.choice(directions),
                'precipitation': precip_forecast,
                'weather': random.choices(weather_types,
                    weights=[0.35, 0.25, 0.15, 0.1, 0.1, 0.05])[0],
            })

        return jsonify(result)

# --- 降雨预报 ---
@app.route('/api/rainfall/forecast')
def rainfall_forecast():
    """降雨预报数据：当前降雨 + 未来48小时逐小时预报 + 数据来源"""
    with get_db() as db:
        # 当前实时降雨（取最新天气记录的降水量）
        current = db.execute("SELECT precipitation, weather_type, temperature FROM weather_data ORDER BY id DESC LIMIT 1").fetchone()
        now = datetime.now()
        # 模拟48小时逐小时降雨预报
        hours = []
        base_precip = current['precipitation'] if current else 5.0
        base_weather = current['weather_type'] if current else '多云'
        rainy_weights = [0.6, 0.2, 0.1, 0.05, 0.05]
        rain_types = ['无雨', '小雨', '中雨', '大雨', '暴雨']
        for i in range(48):
            t = now + timedelta(hours=i)
            # 不同时段降雨概率不同
            hour_of_day = t.hour
            if 2 <= hour_of_day <= 6:
                prob = 0.15  # 凌晨降雨概率低
            elif 14 <= hour_of_day <= 17:
                prob = 0.40  # 午后对流高
            else:
                prob = 0.25
            is_rain = random.random() < prob
            if is_rain:
                p = round(random.uniform(0.5, max(2, base_precip * 1.5)), 1)
                wt = random.choices(rain_types[1:], weights=[0.5, 0.3, 0.15, 0.05])[0]
            else:
                p = 0
                wt = '无雨'
            hours.append({
                'time': t.strftime('%m-%d %H:00'),
                'precipitation': p,
                'rain_type': wt,
                'probability': round(prob * 100),
            })
        sources = ['自动监测站', '气象局', '雷达估测']
        return jsonify({
            'current_rainfall': round(base_precip, 1) if base_precip else 0,
            'current_weather': base_weather,
            'forecast': hours,
            'sources': sources,
        })

# --- 水质监测 ---
@app.route('/api/water-quality')
def water_quality():
    """水质监测数据：返回各供水站/水库的水质指标及7日均值对比"""
    site_id = request.args.get('site_id', type=int)
    with get_db() as db:
        # 查询所有水源相关站点（水库 + 供水站）
        q = "SELECT * FROM sites WHERE type IN ('reservoir','water_supply')"
        params = []
        if site_id:
            q += " AND id=?"
            params.append(site_id)
        sites = db.execute(q, params).fetchall()

        if not sites:
            return jsonify({'error': '没有找到水源相关站点'}), 404

        # 水质指标定义
        water_metrics = [
            ('turbidity', 'NTU', '浊度'),
            ('ph', '', 'pH值'),
            ('chlorine', 'mg/L', '余氯'),
            ('ammonia', 'mg/L', '氨氮'),
            ('cod', 'mg/L', 'COD'),
        ]
        seven_days_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d %H:%M:%S')

        result = []
        for site in sites:
            site_data = {
                'site_id': site['id'],
                'site_code': site['code'],
                'site_name': site['name'],
                'site_type': site['type'],
                'metrics': {},
            }
            for metric, unit, label in water_metrics:
                # 最新值 + 阈值
                latest = db.execute(
                    "SELECT value, threshold_high, threshold_critical, recorded_at FROM sensor_data WHERE site_id=? AND metric=? ORDER BY recorded_at DESC LIMIT 1",
                    (site['id'], metric)
                ).fetchone()

                # 7天均值
                avg_row = db.execute(
                    "SELECT AVG(value) as avg_val FROM sensor_data WHERE site_id=? AND metric=? AND recorded_at > ?",
                    (site['id'], metric, seven_days_ago)
                ).fetchone()

                site_data['metrics'][metric] = {
                    'label': label,
                    'unit': unit,
                    'current': round(latest['value'], 3) if latest and latest['value'] is not None else None,
                    'current_time': latest['recorded_at'] if latest else None,
                    'avg_7d': round(avg_row['avg_val'], 3) if avg_row and avg_row['avg_val'] is not None else None,
                    'thresh_high': round(latest['threshold_high'], 3) if latest and latest['threshold_high'] is not None else None,
                    'thresh_critical': round(latest['threshold_critical'], 3) if latest and latest['threshold_critical'] is not None else None,
                }

            result.append(site_data)

        return jsonify(result)

# --- 设备状态监控 (新增) ---
@app.route('/api/devices/status')
@login_required
def device_status():
    """设备心跳状态汇总：在线/离线统计、各类型统计、离线设备明细"""
    with get_db() as db:
        # 设备总数与状态统计
        total = db.execute("SELECT COUNT(*) as c FROM device_shadows").fetchone()['c']
        online = db.execute("SELECT COUNT(*) as c FROM device_shadows WHERE status='online'").fetchone()['c']
        offline = db.execute("SELECT COUNT(*) as c FROM device_shadows WHERE status='offline'").fetchone()['c']

        # 按设备类型统计
        by_type = db.execute("""
            SELECT device_type,
                COUNT(*) as total,
                SUM(CASE WHEN status='online' THEN 1 ELSE 0 END) as online,
                SUM(CASE WHEN status='offline' THEN 1 ELSE 0 END) as offline
            FROM device_shadows GROUP BY device_type ORDER BY device_type
        """).fetchall()

        # 离线设备明细
        offline_devices = db.execute("""
            SELECT d.*, s.name as site_name, s.code as site_code
            FROM device_shadows d LEFT JOIN sites s ON d.site_id=s.id
            WHERE d.status='offline'
            ORDER BY d.last_data_time DESC
        """).fetchall()

        # 各站点设备状态
        site_devices = db.execute("""
            SELECT s.id as site_id, s.code, s.name,
                COUNT(d.id) as total_devices,
                SUM(CASE WHEN d.status='online' THEN 1 ELSE 0 END) as online_devices,
                SUM(CASE WHEN d.status='offline' THEN 1 ELSE 0 END) as offline_devices
            FROM sites s LEFT JOIN device_shadows d ON s.id=d.site_id
            GROUP BY s.id ORDER BY s.id
        """).fetchall()

        return jsonify({
            'summary': {
                'total': total,
                'online': online,
                'offline': offline,
                'online_rate': round(online / total * 100, 1) if total > 0 else 0,
            },
            'by_type': [dict(r) for r in by_type],
            'offline_devices': [dict(d) for d in offline_devices],
            'site_devices': [dict(s) for s in site_devices],
        })

# --- 数据质量报告 (新增) ---
@app.route('/api/data-quality')
def data_quality():
    """数据质量报告：今日到达率/完整率/及时率、异常站点、24小时趋势"""
    with get_db() as db:
        today = datetime.now().strftime('%Y-%m-%d')
        now = datetime.now()
        twenty_four_hours_ago = (now - timedelta(hours=24)).strftime('%Y-%m-%d %H:%M:%S')

        # 获取所有站点及其类型，确定每个站点期望的指标列表
        sites = db.execute("SELECT * FROM sites").fetchall()
        expected_metrics = {
            'rainfall': ['rainfall', 'precipitation', 'cumulative_rainfall'],
            'water_level': ['water_level', 'flow', 'velocity'],
            'hydrology': ['water_level', 'rainfall', 'flow', 'velocity', 'sediment'],
            'soil_moisture': ['soil_moisture', 'soil_temperature'],
            'evaporation': ['evaporation', 'temperature', 'wind_speed'],
        }

        # 计算今日数据到达率和完整率
        total_expected = 0
        total_received = 0
        anomaly_sites = []  # 异常站点列表

        for site in sites:
            site_metrics = expected_metrics.get(site['type'], [])
            if not site_metrics:
                continue

            # 期望数据点：按每小时12条(每5分钟一条)估算，匹配后端的回填频率
            hours_elapsed = max(1, (now.hour * 60 + now.minute) / 60)
            expected_per_metric = max(1, int(hours_elapsed * 12))  # 每小时12条(每5分钟一次)
            expected_today = len(site_metrics) * expected_per_metric
            total_expected += expected_today

            # 实际收到数据条数
            received = db.execute(
                "SELECT COUNT(*) as c FROM sensor_data WHERE site_id=? AND recorded_at LIKE ?",
                (site['id'], today + '%')
            ).fetchone()['c']
            total_received += received

            # 计算该站点数据完整率
            completeness = round(received / expected_today * 100, 1) if expected_today > 0 else 0

            # 标注异常站点（数据到达率<50%或数据为0的站点）
            if completeness < 50 and expected_today > 5:
                anomaly_sites.append({
                    'site_id': site['id'],
                    'site_code': site['code'],
                    'site_name': site['name'],
                    'site_type': site['type'],
                    'expected': expected_today,
                    'received': received,
                    'completeness': completeness,
                    'reason': '数据到达率低于50%' if received < expected_today / 2 else '数据采集异常',
                })

        # 总体指标
        arrival_rate = round(min(100, total_received / total_expected * 100), 1) if total_expected > 0 else 0
        completeness_rate = arrival_rate  # 在模拟场景下，到达率≈完整率

        # 数据及时率 (最近1小时内是否有数据视为及时)
        one_hour_ago = (now - timedelta(hours=1)).strftime('%Y-%m-%d %H:%M:%S')
        timely_sites = db.execute("""
            SELECT COUNT(DISTINCT site_id) as c FROM sensor_data WHERE recorded_at > ?
        """, (one_hour_ago,)).fetchone()['c']
        total_sites = len(sites) if sites else 1
        timeliness_rate = round(timely_sites / total_sites * 100, 1)

        # 最近24小时数据质量趋势 (按小时分组)
        hourly_trend = db.execute("""
            SELECT strftime('%Y-%m-%d %H:00', recorded_at) as hour,
                COUNT(*) as data_count
            FROM sensor_data
            WHERE recorded_at > ?
            GROUP BY strftime('%Y-%m-%d %H:00', recorded_at)
            ORDER BY hour ASC
        """, (twenty_four_hours_ago,)).fetchall()

        # 每个小时期望数据量（所有站点 × 12条/小时×指标数）
        expected_per_hour = sum(len(expected_metrics.get(s['type'], [])) for s in sites) * 12

        trend = []
        for row in hourly_trend:
            trend.append({
                'hour': row['hour'],
                'count': row['data_count'],
                'rate': round(min(100, row['data_count'] / expected_per_hour * 100), 1) if expected_per_hour > 0 else 0,
            })

        return jsonify({
            'today': {
                'arrival_rate': arrival_rate,         # 数据到达率(%)
                'completeness_rate': completeness_rate, # 数据完整率(%)
                'timeliness_rate': timeliness_rate,     # 数据及时率(%)
                'total_expected': total_expected,
                'total_received': total_received,
                'active_sites': total_sites,
                'timely_sites': timely_sites,
            },
            'anomaly_sites': anomaly_sites,
            'hourly_trend': trend,
        })

# ===================== 数据回填（生成历史数据用于图表展示） =====================

def backfill_history(hours=72):
    """回填历史监测数据，让图表有历史数据可展示"""
    with get_db() as db:
        count = db.execute("SELECT COUNT(*) as c FROM sensor_data").fetchone()['c']
        if count > 10000:
            print(f"[Backfill] 已有 {count} 条数据，跳过回填")
            return
        sites = db.execute("SELECT * FROM sites").fetchall()
        now = datetime.now()
        print(f"[Backfill] 开始回填 {hours} 小时历史数据...")
        for h in range(hours * 12, 0, -1):  # 每5分钟一条，共hours小时
            ts = (now - timedelta(minutes=5 * h)).strftime('%Y-%m-%d %H:%M:%S')
            for site in sites:
                sid = site['id']
                stype = site['type']
                if stype == 'reservoir':
                    wl = round(random.uniform(46.0, 52.5), 2)
                    sp = round(random.uniform(0.05, 0.6), 3)
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'water_level',wl,'m',ts))
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'seepage',sp,'L/s',ts))
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'inflow',round(random.uniform(10,50),1),'m³/s',ts))
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'outflow',round(random.uniform(5,40),1),'m³/s',ts))
                elif stype == 'sluice':
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'gate_opening',round(random.uniform(20,80),1),'%',ts))
                elif stype == 'dike':
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'displacement',round(random.uniform(0,12),1),'mm',ts))
                elif stype == 'pump':
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'flow',round(random.uniform(0.5,3.5),1),'m³/s',ts))
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'vibration',round(random.uniform(2,9),1),'mm/s',ts))
                elif stype == 'water_supply':
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'turbidity',round(random.uniform(0.05,0.6),2),'NTU',ts))
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'ph',round(random.uniform(6.8,7.8),1),'',ts))
                elif stype == 'irrigation':
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'water_level',round(random.uniform(2.0,4.5),1),'m',ts))
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'soil_moisture',round(random.uniform(30,90),1),'%',ts))
        db.commit()
        total = db.execute("SELECT COUNT(*) as c FROM sensor_data").fetchone()['c']
        print(f"[Backfill] 完成！共 {total} 条历史数据")


# ===================== 认证API端点 =====================

@app.route('/api/auth/login', methods=['POST'])
def api_login():
    data = request.get_json(silent=True) or {}
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    if not username or not password:
        return jsonify({'error': '请输入用户名和密码'}), 400
    with get_db() as db:
        user = db.execute("SELECT id, username, password_hash, role, real_name, phone, status FROM users WHERE username=? AND status='active'",
                          (username,)).fetchone()
    if not user or user['password_hash'] != _hash_pw(password):
        return jsonify({'error': '用户名或密码错误'}), 401
    token = secrets.token_urlsafe(32)
    _tokens[token] = {
        'id': user['id'],
        'username': user['username'],
        'role': user['role'],
        'real_name': user['real_name'],
        'phone': user['phone'] or '',
    }
    # 获取此用户可管理的站点列表
    with get_db() as db:
        site_rows = db.execute("SELECT s.id, s.name, s.code, s.type FROM sites s JOIN user_sites us ON s.id=us.site_id WHERE us.user_id=?", (user['id'],)).fetchall()
    sites = [{'id': r['id'], 'name': r['name'], 'code': r['code'], 'type': r['type']} for r in site_rows]
    return jsonify({
        'success': True,
        'token': token,
        'user': {
            'id': user['id'],
            'username': user['username'],
            'role': user['role'],
            'real_name': user['real_name'],
            'phone': user['phone'] or '',
        },
        'sites_count': len(sites),
        'sites': sites,
    })

@app.route('/api/auth/me')
@login_required
def api_me():
    return jsonify({
        'success': True,
        'user': g.current_user,
        'site_ids': g.user_site_ids,
    })


# ===================== 用户管理API（管理员） =====================

@app.route('/api/users')
@login_required
def api_users():
    if g.current_user['role'] != 'admin':
        return jsonify({'error': '无权限'}), 403
    with get_db() as db:
        rows = db.execute("SELECT id, username, role, real_name, phone, status, created_at FROM users ORDER BY id").fetchall()
    users = []
    for r in rows:
        with get_db() as db2:
            cnt = db2.execute("SELECT COUNT(*) as c FROM user_sites WHERE user_id=?", (r['id'],)).fetchone()['c']
        users.append({
            'id': r['id'], 'username': r['username'], 'role': r['role'],
            'real_name': r['real_name'], 'phone': r['phone'] or '',
            'status': r['status'], 'sites_count': cnt,
            'created_at': r['created_at'],
        })
    return jsonify(users)

@app.route('/api/assignees')
@login_required
def api_assignees():
    """返回可用负责人名单（用于B级预警复核转工单下拉选择）"""
    with get_db() as db:
        rows = db.execute("SELECT id, real_name, role FROM users ORDER BY real_name").fetchall()
    return jsonify([{'id': r['id'], 'name': r['real_name'], 'role': r['role']} for r in rows])

@app.route('/api/users/<int:uid>/sites', methods=['GET'])
@login_required
def api_user_sites(uid):
    if g.current_user['role'] != 'admin':
        return jsonify({'error': '无权限'}), 403
    with get_db() as db:
        sids = [r['site_id'] for r in db.execute("SELECT site_id FROM user_sites WHERE user_id=?", (uid,)).fetchall()]
    return jsonify({'site_ids': sids})

@app.route('/api/users/<int:uid>/sites', methods=['PUT'])
@login_required
def api_update_user_sites(uid):
    if g.current_user['role'] != 'admin':
        return jsonify({'error': '无权限'}), 403
    data = request.get_json(silent=True) or {}
    site_ids = data.get('site_ids', [])
    if not isinstance(site_ids, list):
        return jsonify({'error': 'site_ids格式错误'}), 400
    with get_db() as db:
        db.execute("DELETE FROM user_sites WHERE user_id=?", (uid,))
        for sid in site_ids:
            db.execute("INSERT OR IGNORE INTO user_sites (user_id,site_id) VALUES (?,?)", (uid, sid))
        db.commit()
    # 失效该用户的站点缓存（按 token 反查）
    for t, u in list(_tokens.items()):
        if u.get('id') == uid:
            _site_ids_cache.pop(t, None)
    return jsonify({'success': True, 'count': len(site_ids)})

@app.route('/api/users/<int:uid>/reset-password', methods=['PUT'])
@login_required
def api_reset_password(uid):
    if g.current_user['role'] != 'admin':
        return jsonify({'error': '无权限'}), 403
    data = request.get_json(silent=True) or {}
    new_pw = data.get('new_password', 'yw123456')
    with get_db() as db:
        db.execute("UPDATE users SET password_hash=? WHERE id=?", (_hash_pw(new_pw), uid))
        db.commit()
    return jsonify({'success': True})

@app.route('/api/users/<int:uid>/status', methods=['PUT'])
@login_required
def api_user_status(uid):
    if g.current_user['role'] != 'admin':
        return jsonify({'error': '无权限'}), 403
    data = request.get_json(silent=True) or {}
    new_status = data.get('status', 'active')
    with get_db() as db:
        db.execute("UPDATE users SET status=? WHERE id=?", (new_status, uid))
        db.commit()
    return jsonify({'success': True})


# ===================== 设备管理 API =====================

@app.route('/api/devices')
@login_required
def api_devices_list():
    """设备台账列表，支持按站点/类型筛选"""
    site_id = request.args.get('site_id', '').strip()
    device_type = request.args.get('type', '').strip()
    search = request.args.get('search', '').strip()
    with get_db() as db:
        sql = """SELECT d.id, d.device_code, d.device_name, d.device_type, d.device_model, d.manufacturer, d.install_date,
                        d.last_data_time,
                        s.name as site_name, s.code as site_code, s.id as site_id
                 FROM device_shadows d LEFT JOIN sites s ON d.site_id=s.id WHERE 1=1"""
        params = []
        if site_id:
            sql += " AND d.site_id=?"
            params.append(site_id)
        if device_type:
            sql += " AND d.device_type=?"
            params.append(device_type)
        if search:
            sql += " AND (d.device_name LIKE ? OR d.device_code LIKE ? OR s.name LIKE ?)"
            like = f"%{search}%"
            params.extend([like, like, like])
        sql += " ORDER BY d.id"
        rows = db.execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/devices/<int:device_id>')
@login_required
def api_device_detail(device_id):
    """设备详情 + 维护记录"""
    with get_db() as db:
        dev = db.execute("""SELECT d.*, s.name as site_name, s.code as site_code,
                                   s.district, s.manager
                            FROM device_shadows d LEFT JOIN sites s ON d.site_id=s.id
                            WHERE d.id=?""", (device_id,)).fetchone()
        if not dev:
            return jsonify({'error': '设备不存在'}), 404
        logs = db.execute("""SELECT * FROM inventory_logs
                             WHERE ref_type='maintenance' AND ref_id=?
                             ORDER BY created_at DESC LIMIT 20""", (device_id,)).fetchall()
        # 增加设备操作日志（CRUD）
        op_logs = db.execute("""SELECT * FROM operation_logs
                                WHERE target_type='device' AND target_id=?
                                ORDER BY created_at DESC LIMIT 20""", (device_id,)).fetchall()
    return jsonify({'device': dict(dev),
                    'logs': [dict(l) for l in logs],
                    'operation_logs': [dict(o) for o in op_logs]})


@app.route('/api/devices', methods=['POST'])
@login_required
def api_device_create():
    """注册新设备"""
    g_ = require_admin()
    if g_:
        return g_
    data = request.get_json(force=True)
    device_code = (data.get('device_code') or '').strip()
    device_name = (data.get('device_name') or '').strip()
    device_type = (data.get('device_type') or '').strip()
    site_id = data.get('site_id')

    if not device_code or not device_name:
        return jsonify({'error': '设备编码和名称不能为空'}), 400

    with get_db() as db:
        existing = db.execute("SELECT id FROM device_shadows WHERE device_code=?", (device_code,)).fetchone()
        if existing:
            return jsonify({'error': '设备编码已存在'}), 409
        cur = db.execute(
            """INSERT INTO device_shadows (device_code, device_name, device_type, site_id)
               VALUES (?, ?, ?, ?)""",
            (device_code, device_name, device_type, site_id)
        )
        new_id = cur.lastrowid
        db.execute("""INSERT INTO operation_logs (module, action, target_type, target_id, operator, operator_id, details)
            VALUES (?, ?, ?, ?, ?, ?, ?)""",
            ('device', 'create', 'device', new_id,
             g.current_user['username'], g.current_user['id'],
             f"注册设备「{device_name}」({device_code})"))
        db.commit()
    return jsonify({'success': True, 'id': new_id, 'message': '设备注册成功'})


@app.route('/api/devices/<int:device_id>', methods=['PUT'])
@login_required
def api_device_update(device_id):
    """编辑设备信息"""
    g_ = require_admin()
    if g_:
        return g_
    data = request.get_json(force=True)
    with get_db() as db:
        dev = db.execute("SELECT id FROM device_shadows WHERE id=?", (device_id,)).fetchone()
        if not dev:
            return jsonify({'error': '设备不存在'}), 404

        fields = []
        values = []
        for col in ['device_code', 'device_name', 'device_type', 'site_id']:
            if col in data:
                fields.append(f"{col}=?")
                values.append(data[col])
        if not fields:
            return jsonify({'error': '没有可更新的字段'}), 400

        # 如果修改了编码，检查唯一性
        if 'device_code' in data:
            dup = db.execute("SELECT id FROM device_shadows WHERE device_code=? AND id!=?",
                             (data['device_code'], device_id)).fetchone()
            if dup:
                return jsonify({'error': '设备编码已存在'}), 409

        values.append(device_id)
        db.execute(f"UPDATE device_shadows SET {', '.join(fields)} WHERE id=?", values)
        # 记录操作日志
        changed = [f"{k}→{data[k]}" for k in data if k in ['device_code','device_name','device_type','site_id']]
        db.execute("""INSERT INTO operation_logs (module, action, target_type, target_id, operator, operator_id, details)
            VALUES (?, ?, ?, ?, ?, ?, ?)""",
            ('device', 'update', 'device', device_id,
             g.current_user['username'], g.current_user['id'],
             '设备更新: ' + '; '.join(changed) if changed else '设备信息更新'))
        db.commit()
    return jsonify({'success': True, 'message': '设备信息已更新'})


@app.route('/api/devices/<int:device_id>', methods=['DELETE'])
@login_required
def api_device_delete(device_id):
    """删除设备（直接从台账移除。如需回收流程请走 /api/device-recycle POST）"""
    g_ = require_admin()
    if g_:
        return g_
    try:
        with get_db() as db:
            dev = db.execute("SELECT * FROM device_shadows WHERE id=?", (device_id,)).fetchone()
            if not dev:
                return jsonify({'error': '设备不存在'}), 404
            db.execute("DELETE FROM device_shadows WHERE id=?", (device_id,))
            db.execute("""INSERT INTO operation_logs (module, action, target_type, target_id, operator, operator_id, details)
                VALUES (?, ?, ?, ?, ?, ?, ?)""",
                ('device', 'delete', 'device', device_id,
                 g.current_user['username'], g.current_user['id'],
                 f"删除设备「{dev['device_name']}」({dev['device_code']})"))
            db.commit()
        return jsonify({'success': True, 'message': '设备已删除'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# --- 操作日志 ---

@app.route('/api/operation-logs')
@login_required
def api_operation_logs():
    """操作日志列表（设备管理全部操作记录）"""
    module = request.args.get('module', '').strip()
    target_type = request.args.get('target_type', '').strip()
    target_id = request.args.get('target_id', type=int)
    limit = min(int(request.args.get('limit', 50)), 200)
    with get_db() as db:
        sql = """SELECT ol.*, u.real_name AS operator_name
                 FROM operation_logs ol
                 LEFT JOIN users u ON ol.operator_id = u.id
                 WHERE 1=1"""
        params = []
        if module:
            sql += " AND ol.module=?"
            params.append(module)
        if target_type:
            sql += " AND ol.target_type=?"
            params.append(target_type)
        if target_id:
            sql += " AND ol.target_id=?"
            params.append(target_id)
        sql += " ORDER BY ol.created_at DESC LIMIT ?"
        params.append(limit)
        rows = db.execute(sql, params).fetchall()
        result = []
        for r in rows:
            d = dict(r)
            if d.get('operator_name'):
                d['operator'] = d['operator_name']
            result.append(d)
    return jsonify(result)


# --- 设备回收 ---

@app.route('/api/device-recycle')
@login_required
def api_device_recycle_list():
    """设备回收记录列表，支持按设备/站点搜索"""
    search = request.args.get('search', '').strip()
    status = request.args.get('status', '').strip()
    with get_db() as db:
        sql = "SELECT * FROM device_recycle WHERE 1=1"
        params = []
        if search:
            sql += " AND (device_code LIKE ? OR device_name LIKE ? OR site_name LIKE ? OR destination LIKE ?)"
            like = f"%{search}%"
            params.extend([like, like, like, like])
        if status:
            sql += " AND status=?"
            params.append(status)
        sql += " ORDER BY created_at DESC"
        rows = db.execute(sql, params).fetchall()
        records = [dict(r) for r in rows]
        return jsonify(records)

@app.route('/api/device-recycle', methods=['POST'])
@login_required
def api_device_recycle_create():
    """登记设备回收"""
    g_ = require_admin()
    if g_:
        return g_
    data = request.get_json(silent=True) or {}
    with get_db() as db:
        # 验证设备是否存在
        device_id = data.get('device_id')
        device = db.execute("SELECT * FROM device_shadows WHERE id=?", (device_id,)).fetchone()
        if not device:
            return jsonify({'error': '设备不存在'}), 404
        # 插入回收记录
        site = db.execute("SELECT name FROM sites WHERE id=?", (device['site_id'],)).fetchone()
        db.execute("""
            INSERT INTO device_recycle (device_id, device_code, device_name, device_type,
                site_id, site_name, recycle_date, reason, destination, operator, remark, status)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (device_id, device['device_code'], device['device_name'], device['device_type'],
              device['site_id'], site['name'] if site else '',
              data.get('recycle_date', ''), data.get('reason', ''),
              data.get('destination', ''), data.get('operator', ''),
              data.get('remark', ''), data.get('status', 'recycled')))
        # 同时将设备状态设为 offline（已回收）
        db.execute("UPDATE device_shadows SET status='offline' WHERE id=?", (device_id,))
        # 记录时间线事件
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('device', device_id, 'recycled', data.get('operator', '系统'),
                    f'设备回收-{device["device_name"]}({device["device_code"]})->{data.get("destination","")}'))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/device-recycle/<int:rec_id>', methods=['PUT'])
@login_required
def api_device_recycle_update(rec_id):
    """更新回收记录（如去向变更）"""
    data = request.get_json(silent=True) or {}
    with get_db() as db:
        fields = []
        params = []
        for f in ['destination', 'remark', 'status', 'reason', 'operator']:
            if f in data:
                fields.append(f"{f}=?")
                params.append(data[f])
        if fields:
            params.append(rec_id)
            db.execute(f"UPDATE device_recycle SET {','.join(fields)} WHERE id=?", params)
            db.commit()
        return jsonify({'success': True})


# --- 备件库存 ---

@app.route('/api/parts/inventory')
@login_required
def api_parts_inventory():
    """备件库存列表"""
    category = request.args.get('category', '').strip()
    low = request.args.get('low', '').strip()
    search = request.args.get('search', '').strip()
    with get_db() as db:
        sql = """SELECT p.*, s.name as site_name
                 FROM spare_parts_inventory p LEFT JOIN sites s ON p.site_id=s.id WHERE 1=1"""
        params = []
        if category:
            sql += " AND p.category=?"
            params.append(category)
        if low == '1':
            sql += " AND p.quantity <= p.min_quantity"
        if search:
            sql += " AND (p.part_name LIKE ? OR p.part_code LIKE ?)"
            like = f"%{search}%"
            params.extend([like, like])
        sql += " ORDER BY p.quantity ASC"
        rows = db.execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/parts/inventory', methods=['POST'])
@login_required
def api_parts_inventory_add():
    """新增备件或入库"""
    g_ = require_admin()
    if g_:
        return g_
    data = request.get_json(silent=True) or {}
    part_code = data.get('part_code', '').strip()
    part_name = data.get('part_name', '').strip()
    category = data.get('category', '其他').strip()
    unit = data.get('unit', '个').strip()
    quantity = int(data.get('quantity', 1))
    min_quantity = int(data.get('min_quantity', 5))
    site_id = data.get('site_id')
    remark = data.get('remark', '').strip()
    if not part_name:
        return jsonify({'error': '备件名称不能为空'}), 400
    if not part_code:
        import uuid
        part_code = f"BJ-{uuid.uuid4().hex[:6].upper()}"
    with get_db() as db:
        cur = db.execute("""INSERT INTO spare_parts_inventory
            (part_code,part_name,category,unit,quantity,min_quantity,site_id,remark)
            VALUES (?,?,?,?,?,?,?,?)""",
            (part_code, part_name, category, unit, quantity, min_quantity, site_id, remark))
        pid = cur.lastrowid
        db.execute("""INSERT INTO inventory_logs (part_id,type,quantity,ref_type,operator,remark)
            VALUES (?,'in',?,'purchase',?,?)""",
            (pid, quantity, g.current_user['username'] or 'admin', f'入库: {part_name}'))
        db.commit()
    return jsonify({'success': True, 'id': pid, 'part_code': part_code})


@app.route('/api/parts/inventory/<int:pid>', methods=['PUT'])
@login_required
def api_parts_inventory_update(pid):
    """更新备件信息或手动出库"""
    g_ = require_admin()
    if g_:
        return g_
    data = request.get_json(silent=True) or {}
    with get_db() as db:
        part = db.execute("SELECT * FROM spare_parts_inventory WHERE id=?", (pid,)).fetchone()
        if not part:
            return jsonify({'error': '备件不存在'}), 404
        # 出库操作
        if 'out_qty' in data:
            qty = int(data['out_qty'])
            if qty <= 0:
                return jsonify({'error': '出库数量需大于0'}), 400
            if part['quantity'] - qty < 0:
                return jsonify({'error': '库存不足'}), 400
            db.execute("UPDATE spare_parts_inventory SET quantity=quantity-?, updated_at=datetime('now','localtime') WHERE id=?", (qty, pid))
            remark_out = data.get('remark', '').strip() or '手动出库'
            db.execute("""INSERT INTO inventory_logs (part_id,type,quantity,ref_type,operator,remark)
                VALUES (?,'out',?,'adjust',?,?)""",
                (pid, qty, g.current_user['username'] or 'admin', remark_out))
        # 入库操作
        if 'in_qty' in data:
            qty = int(data['in_qty'])
            if qty <= 0:
                return jsonify({'error': '入库数量需大于0'}), 400
            db.execute("UPDATE spare_parts_inventory SET quantity=quantity+?, updated_at=datetime('now','localtime') WHERE id=?", (qty, pid))
            remark_in = data.get('remark', '').strip() or '手动入库'
            db.execute("""INSERT INTO inventory_logs (part_id,type,quantity,ref_type,operator,remark)
                VALUES (?,'in',?,'purchase',?,?)""",
                (pid, qty, g.current_user['username'] or 'admin', remark_in))
        # 更新基本信息
        for field in ['part_name', 'category', 'unit', 'min_quantity', 'remark']:
            if field in data:
                db.execute(f"UPDATE spare_parts_inventory SET {field}=? WHERE id=?", (data[field], pid))
        db.commit()
    return jsonify({'success': True})


@app.route('/api/parts/inventory/<int:pid>/logs')
@login_required
def api_parts_inventory_logs(pid):
    """备件库存变更流水"""
    with get_db() as db:
        rows = db.execute("""SELECT il.*, u.real_name AS operator_name
                             FROM inventory_logs il
                             LEFT JOIN users u ON il.operator = u.username
                             WHERE il.part_id=? ORDER BY il.created_at DESC LIMIT 50""", (pid,)).fetchall()
        result = []
        for r in rows:
            d = dict(r)
            if d.get('operator_name'):
                d['operator'] = d['operator_name']
            result.append(d)
    return jsonify(result)


@app.route('/api/parts/inventory/<int:pid>/stock', methods=['POST'])
@login_required
def api_parts_inventory_stock(pid):
    """备件入库/出库操作"""
    g_ = require_admin()
    if g_:
        return g_
    try:
        data = request.get_json(silent=True) or {}
        stock_type = data.get('type', '')  # 'in' or 'out'
        quantity = int(data.get('quantity', 0))
        reason = data.get('reason', '').strip()
        operator = data.get('operator', '').strip() or g.current_user.get('username', 'unknown')
        work_order_no = data.get('work_order_no', '').strip()
        if stock_type not in ('in', 'out'):
            return jsonify({'error': '操作类型必须为 in 或 out'}), 400
        if quantity <= 0:
            return jsonify({'error': '数量必须大于0'}), 400
        with get_db() as db:
            # 兼容旧表：加 work_order_no 列
            try:
                db.execute("ALTER TABLE inventory_logs ADD COLUMN work_order_no TEXT DEFAULT ''")
            except Exception:
                pass
            part = db.execute("SELECT * FROM spare_parts_inventory WHERE id=?", (pid,)).fetchone()
            if not part:
                return jsonify({'error': '备件不存在'}), 404
            if stock_type == 'out' and part['quantity'] < quantity:
                return jsonify({'error': f"库存不足，当前库存 {part['quantity']}"}), 400
            if stock_type == 'in':
                db.execute("UPDATE spare_parts_inventory SET quantity=quantity+?, updated_at=datetime('now','localtime') WHERE id=?", (quantity, pid))
            else:
                db.execute("UPDATE spare_parts_inventory SET quantity=quantity-?, updated_at=datetime('now','localtime') WHERE id=?", (quantity, pid))
            db.execute("""INSERT INTO inventory_logs (part_id,type,quantity,ref_type,operator,remark,work_order_no)
                VALUES (?,?,?, 'stock', ?, ?, ?)""",
                (pid, stock_type, quantity, operator, reason, work_order_no))
            db.commit()
        return jsonify({'success': True, 'message': f"{'入库' if stock_type == 'in' else '出库'}成功"})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/parts/recovery', methods=['POST'])
def api_parts_recovery():
    """旧件回收：更换备件后旧件退回公司库存。
    与采购入库区分——inventory_logs 记 type='in' 且 ref_type='recovery'，
    便于后续维修/报废/采购决策追溯（备件全生命周期 P2）。
    一线运维即可提交（非管理员专属），故不加 require_admin。"""
    u = g.current_user
    data = request.get_json(silent=True) or {}
    part_id = data.get('part_id')
    part_code = (data.get('part_code') or '').strip()
    try:
        quantity = int(data.get('quantity', 1))
    except (TypeError, ValueError):
        quantity = 0
    if quantity <= 0:
        return jsonify({'error': '回收数量必须大于0'}), 400
    work_order_no = (data.get('work_order_no') or '').strip()
    remark = (data.get('remark') or '').strip() or '旧件回收'
    operator = (data.get('operator') or '').strip() or u.get('real_name') or u.get('username', 'unknown')
    with get_db() as db:
        # 兼容旧表：补 work_order_no 列
        try:
            db.execute("ALTER TABLE inventory_logs ADD COLUMN work_order_no TEXT DEFAULT ''")
        except Exception:
            pass
        part = None
        if part_id:
            part = db.execute("SELECT * FROM spare_parts_inventory WHERE id=?", (part_id,)).fetchone()
        elif part_code:
            part = db.execute("SELECT * FROM spare_parts_inventory WHERE part_code=?", (part_code,)).fetchone()
        if not part:
            return jsonify({'error': '备件不存在（请提供 part_id 或 part_code）'}), 404
        db.execute("UPDATE spare_parts_inventory SET quantity=quantity+?, updated_at=datetime('now','localtime') WHERE id=?",
                   (quantity, part['id']))
        db.execute("""INSERT INTO inventory_logs (part_id, type, quantity, ref_type, operator, remark, work_order_no)
            VALUES (?, 'in', ?, 'recovery', ?, ?, ?)""",
            (part['id'], quantity, operator, remark, work_order_no))
        db.commit()
        return jsonify({'success': True, 'part_id': part['id'], 'part_name': part['part_name'],
                        'recovered': quantity, 'message': f"已回收 {part['part_name']} ×{quantity}"})


# --- 备件仪表盘 ---

@app.route('/api/parts/dashboard')
@login_required
def api_parts_dashboard():
    """备件仪表盘：总数、低库存数、最新动态"""
    with get_db() as db:
        total = db.execute("SELECT COUNT(*) as cnt FROM spare_parts_inventory").fetchone()['cnt']
        low = db.execute("SELECT COUNT(*) as cnt FROM spare_parts_inventory WHERE quantity < min_quantity").fetchone()['cnt']
        device_count = db.execute("SELECT COUNT(*) as cnt FROM device_shadows").fetchone()['cnt']
        latest_ops = db.execute("""SELECT l.*, p.part_name, p.part_code
            FROM inventory_logs l JOIN spare_parts_inventory p ON l.part_id=p.id
            ORDER BY l.created_at DESC LIMIT 10""").fetchall()
    return jsonify({
        'total_parts': total,
        'low_stock': low,
        'device_count': device_count,
        'latest_operations': [dict(r) for r in latest_ops],
    })


# --- 备件申请 ---

@app.route('/api/parts/requests', methods=['GET'])
@login_required
def api_parts_requests_list():
    """备件申请列表（Web端：全部；移动端按申请人过滤）"""
    status_f = request.args.get('status', '').strip()
    applicant = request.args.get('applicant', '').strip()
    with get_db() as db:
        sql = """SELECT r.*, s.name as site_name
                 FROM spare_part_requests r LEFT JOIN sites s ON r.site_id=s.id WHERE 1=1"""
        params = []
        if status_f:
            sql += " AND r.status=?"
            params.append(status_f)
        if applicant:
            sql += " AND r.applicant=?"
            params.append(applicant)
        sql += " ORDER BY r.created_at DESC"
        rows = db.execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/parts/requests', methods=['POST'])
@login_required
def api_parts_requests_create():
    """创建备件申请（移动端调用）"""
    data = request.get_json(silent=True) or {}
    site_id = data.get('site_id')
    raw_part_name = (data.get('part_name') or '').strip()
    quantity = int(data.get('quantity', 1) or 1)
    reason = (data.get('reason') or '').strip()
    work_order_no = (data.get('work_order_no') or '').strip()
    spare_part_id = data.get('spare_part_id') or None
    applicant = g.current_user['username'] or 'unknown'
    from datetime import datetime
    today = datetime.now().strftime('%Y%m%d')
    with get_db() as db:
        part_name = raw_part_name
        # 关联库存时若未手填名称，自动带出库存名称
        if spare_part_id and not part_name:
            inv = db.execute("SELECT part_name FROM spare_parts_inventory WHERE id=?", (spare_part_id,)).fetchone()
            if inv:
                part_name = inv['part_name']
        if not site_id or not part_name:
            return jsonify({'error': '站点和备件名称不能为空'}), 400
        count = db.execute("SELECT COUNT(*) as c FROM spare_part_requests WHERE request_no LIKE ?", (f"BJ-{today}%",)).fetchone()['c']
        request_no = f"BJ-{today}-{count+1:03d}"
        db.execute("""INSERT INTO spare_part_requests
            (request_no,site_id,applicant,part_name,quantity,reason,work_order_no,spare_part_id)
            VALUES (?,?,?,?,?,?,?,?)""",
            (request_no, site_id, applicant, part_name, quantity, reason, work_order_no, spare_part_id))
        db.commit()
    return jsonify({'success': True, 'request_no': request_no})


@app.route('/api/parts/requests/mine')
@login_required
def api_parts_requests_mine():
    """我的备件申请记录（移动端）"""
    applicant = g.current_user['username'] or 'unknown'
    with get_db() as db:
        rows = db.execute("""SELECT r.*, s.name as site_name
            FROM spare_part_requests r LEFT JOIN sites s ON r.site_id=s.id
            WHERE r.applicant=? ORDER BY r.created_at DESC""", (applicant,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/parts/requests/<int:rid>/approve', methods=['PUT'])
@login_required
def api_parts_request_approve(rid):
    """审批通过：更新状态 + 扣减库存"""
    if g.current_user['role'] != 'admin':
        return jsonify({'error': '仅管理员可审批'}), 403
    data = request.get_json(silent=True) or {}
    comment = data.get('comment', '审批通过')
    with get_db() as db:
        req = db.execute("SELECT * FROM spare_part_requests WHERE id=?", (rid,)).fetchone()
        if not req:
            return jsonify({'error': '申请不存在'}), 404
        if req['status'] != 'pending':
            return jsonify({'error': '该申请已处理'}), 400
        # 更新申请状态
        db.execute("""UPDATE spare_part_requests SET status='approved', approver=?,
            approval_comment=?, updated_at=datetime('now','localtime') WHERE id=?""",
            (g.current_user['username'] or 'admin', comment, rid))
        # 尝试扣减库存：查找匹配的备件（按名称模糊匹配）
        inv = db.execute("""SELECT * FROM spare_parts_inventory
            WHERE part_name LIKE ? ORDER BY quantity DESC LIMIT 1""",
            (f"%{req['part_name']}%",)).fetchone()
        if inv:
            new_qty = max(0, inv['quantity'] - req['quantity'])
            db.execute("UPDATE spare_parts_inventory SET quantity=?, updated_at=datetime('now','localtime') WHERE id=?", (new_qty, inv['id']))
            db.execute("""INSERT INTO inventory_logs (part_id,type,quantity,ref_type,ref_id,operator,remark,work_order_no)
                VALUES (?,?,?,'request',?,?,?,?)""",
                (inv['id'], 'out', req['quantity'], rid,
                 g.current_user['username'] or 'admin',
                 f"备件申请 #{req['request_no']}",
                 req['work_order_no'] if req['work_order_no'] else ''))
        db.commit()
    return jsonify({'success': True, 'message': '已批准，库存已扣减'})


@app.route('/api/parts/requests/<int:rid>/reject', methods=['PUT'])
@login_required
def api_parts_request_reject(rid):
    """驳回申请"""
    if g.current_user['role'] != 'admin':
        return jsonify({'error': '仅管理员可审批'}), 403
    data = request.get_json(silent=True) or {}
    comment = data.get('comment', '驳回')
    with get_db() as db:
        req = db.execute("SELECT * FROM spare_part_requests WHERE id=?", (rid,)).fetchone()
        if not req:
            return jsonify({'error': '申请不存在'}), 404
        if req['status'] != 'pending':
            return jsonify({'error': '该申请已处理'}), 400
        db.execute("""UPDATE spare_part_requests SET status='rejected', approver=?,
            approval_comment=?, updated_at=datetime('now','localtime') WHERE id=?""",
            (g.current_user['username'] or 'admin', comment, rid))
        db.commit()
    return jsonify({'success': True, 'message': '已驳回'})


# ===================== 站点过滤辅助函数 =====================

def _filter_by_user(where_clause='', table_prefix=''):
    """为API查询注入站点过滤条件。返回 (where_extra, params)
    管理员不限制，操作员限制为分配的站点。
    在路由函数中使用：在原始WHERE后加上此函数的返回。
    """
    site_ids = getattr(g, 'user_site_ids', None)
    if site_ids is None:
        return '', []
    prefix = table_prefix + '.' if table_prefix else ''
    site_condition = f"{prefix}site_id IN ({','.join('?' * len(site_ids))})" if site_ids else '1=0'
    extra = f" AND {site_condition}" if where_clause else f" WHERE {site_condition}"
    return extra, site_ids


# ===================== 前端静态文件服务 =====================
FRONTEND_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'frontend')
V2_DIR = os.path.join(FRONTEND_DIR, 'v2')

@app.route('/')
def index_html():
    return send_from_directory(V2_DIR, 'index.html')

# ===================== Site Data Import & Data Sources =====================

@app.route('/api/sites/import', methods=['POST'])
@login_required
def import_sites():
    """批量导入站点（CSV文件）"""
    if 'file' not in request.files:
        return jsonify({'error': '请上传CSV文件'}), 400
    f = request.files['file']
    if not f.filename.endswith('.csv'):
        return jsonify({'error': '仅支持CSV格式文件'}), 400
    import csv as csv_mod, io
    try:
        content = f.read().decode('utf-8-sig')
        reader = csv_mod.DictReader(io.StringIO(content))
        success, failed, errors = 0, 0, []
        with get_db() as db:
            for i, row in enumerate(reader, 2):
                code = (row.get('code') or row.get('编码') or row.get('站点编码') or '').strip()
                name = (row.get('name') or row.get('名称') or row.get('站点名称') or '').strip()
                stype = (row.get('type') or row.get('类型') or row.get('站点类型') or '').strip()
                if not code or not name or not stype:
                    failed += 1
                    errors.append(f'第{i}行: 缺少必填字段(code/name/type)')
                    continue
                try:
                    lat = float(row.get('lat') or row.get('纬度') or 0)
                    lng = float(row.get('lng') or row.get('经度') or 0)
                except (ValueError, TypeError):
                    lat, lng = 0, 0
                district = (row.get('district') or row.get('区域') or '').strip()
                river = (row.get('river') or row.get('河流') or '').strip()
                manager = (row.get('manager') or row.get('负责人') or '').strip()
                phone = (row.get('phone') or row.get('电话') or '').strip()
                try:
                    db.execute(
                        "INSERT INTO sites (code,name,type,lat,lng,district,river,manager,phone) VALUES (?,?,?,?,?,?,?,?,?)",
                        (code, name, stype, lat, lng, district, river, manager, phone)
                    )
                    success += 1
                except Exception as e:
                    failed += 1
                    errors.append(f'第{i}行({code}): {str(e)[:60]}')
            db.commit()
        return jsonify({
            'success': True,
            'imported': success,
            'failed': failed,
            'errors': errors[:10],
        })
    except Exception as e:
        return jsonify({'error': f'解析文件失败: {str(e)[:100]}'}), 400

@app.route('/api/sites/data-sources', methods=['GET'])
@login_required
def list_data_sources():
    """获取数据源列表"""
    with get_db() as db:
        rows = db.execute("SELECT * FROM data_sources ORDER BY created_at DESC").fetchall()
        return jsonify([dict(r) for r in rows])

@app.route('/api/sites/data-sources', methods=['POST'])
@login_required
def create_data_source():
    """新增数据源配置"""
    data = request.get_json(silent=True) or {}
    name = data.get('name', '').strip()
    url = data.get('url', '').strip()
    if not name or not url:
        return jsonify({'error': '名称和URL不能为空'}), 400
    with get_db() as db:
        db.execute(
            "INSERT INTO data_sources (name,source_type,protocol,url,auth_type,auth_config,sync_interval,remark) VALUES (?,?,?,?,?,?,?,?)",
            (name, data.get('source_type', 'api'), data.get('protocol', 'HTTP'), url,
             data.get('auth_type', 'none'), json.dumps(data.get('auth_config', {})),
             data.get('sync_interval', 60), data.get('remark', ''))
        )
        db.commit()
    return jsonify({'success': True})

@app.route('/api/sites/data-sources/<int:ds_id>', methods=['DELETE'])
@login_required
def delete_data_source(ds_id):
    """删除数据源"""
    with get_db() as db:
        db.execute("DELETE FROM data_sources WHERE id=?", (ds_id,))
        db.commit()
    return jsonify({'success': True})

@app.route('/api/sites/data-sources/<int:ds_id>/test', methods=['POST'])
@login_required
def test_data_source(ds_id):
    """测试数据源连通性（模拟）"""
    with get_db() as db:
        ds = db.execute("SELECT * FROM data_sources WHERE id=?", (ds_id,)).fetchone()
        if not ds:
            return jsonify({'error': '数据源不存在'}), 404
    # 模拟测试：返回成功
    import random
    latency = random.randint(50, 300)
    return jsonify({
        'success': True,
        'latency_ms': latency,
        'message': f'连接成功，响应时间 {latency}ms',
    })

@app.route('/api/sites/template', methods=['GET'])
def download_site_template():
    """下载站点导入CSV模板"""
    import csv as csv_mod, io
    output = io.StringIO()
    writer = csv_mod.writer(output)
    writer.writerow(['code', 'name', 'type', 'lat', 'lng', 'district', 'river', 'manager', 'phone'])
    writer.writerow(['GST001', '示例雨量站', 'rainfall', '28.68', '115.89', '南昌市', '赣江', '张工', '13800138000'])
    from flask import Response
    return Response(
        '\ufeff' + output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=site_import_template.csv'}
    )

@app.route('/uploads/<path:filename>')
def serve_uploads(filename):
    # 使用与上传写入相同的 UPLOAD_DIR，确保读写一致
    return send_from_directory(UPLOAD_DIR, filename)

@app.route('/<path:filename>')
def serve_frontend(filename):
    # SPA 路由回退：非 API 路径且非实际文件 → 返回 React 入口
    # 让 React Router 处理客户端路由
    v2_path = os.path.join(V2_DIR, filename)
    frontend_path = os.path.join(FRONTEND_DIR, filename)
    if os.path.exists(v2_path):
        return send_from_directory(V2_DIR, filename)
    if os.path.exists(frontend_path):
        return send_from_directory(FRONTEND_DIR, filename)
    # SPA 回退：所有未匹配的路径都返回 v2/index.html
    return send_from_directory(V2_DIR, 'index.html')

# ===================== Startup =====================

def fix_site_river():
    """为水位站/水文站设置正确的河流字段，避免水位基值漂移"""
    with get_db() as conn:
        conn.execute("UPDATE sites SET river='赣江' WHERE type IN ('water_level','hydrology') AND (river IS NULL OR river='')")
        conn.execute("UPDATE sites SET river='鄱阳湖' WHERE type='groundwater' AND (river IS NULL OR river='')")
        conn.commit()
        updated = conn.total_changes
        print(f"[Fix] 已更新 {updated} 个站点的河流字段")


# =============================================================================
# 移动端专用接口 —— 一线执行助手
# =============================================================================

# 频次中文映射（移动端统一使用此映射，不用 high/mid/low）
_FREQ_CN = {
    'daily': '每日', 'weekly': '每周', 'monthly': '每月',
    'quarterly': '每季', 'semi_annual': '每半年', 'annual': '每年',
    'high': '每日', 'mid': '每月', 'low': '每季', 'annual': '每年',
}

# 指标中文映射（移动端告警/巡检统一使用，供 mobile_my_today 等模块级 handler 引用）
METRIC_CN = {
    'codmn': '高锰酸盐指数', 'ammonia': '氨氮', 'total_phosphorus': '总磷', 'total_nitrogen': '总氮',
    'water_temp': '水温', 'dissolved_oxygen': '溶解氧', 'ph': 'pH', 'turbidity': '浊度',
    'conductivity': '电导率', 'temperature': '气温',
    'data_spike': '数据突变', 'data_freeze': '数据冻结', 'data_gap': '数据缺失',
    'device_status': '设备状态',
}


@app.route('/api/mobile/bind-openid', methods=['POST'])
@login_required
def mobile_bind_openid():
    """绑定微信 openid：前端 wx.login 拿到 code 后调用，服务端用 code 换 openid 并落库。
    失败不阻断登录流程（如未配置 AppSecret），前端静默处理即可。"""
    data = request.get_json(silent=True) or {}
    code = (data.get('code') or '').strip()
    if not code:
        return jsonify({'success': False, 'error': '缺少 code'}), 400
    openid = _wx_code2openid(code)
    if not openid:
        # 换 openid 失败（未配置密钥 / 微信接口异常）不阻断，前端照常进入首页
        return jsonify({'success': True, 'bound': False, 'warn': 'openid 换取失败'})
    try:
        with get_db() as db:
            db.execute("UPDATE users SET openid=? WHERE id=?", (openid, g.current_user['id']))
            db.commit()
        return jsonify({'success': True, 'bound': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/mobile/my-today')
@login_required
def mobile_my_today():
    """移动端首页聚合接口：一次请求返回当前用户今日所有任务数据。
    包含：今日作业包、巡检任务（按站点分组）、待处理工单、未处理告警。
    """
    user = g.current_user
    today = datetime.now().strftime('%Y-%m-%d')

    with get_db() as db:
        # ---- 1. 巡检任务：与网页端一致，按站点范围隔离（_filter_site_ids）----
        allowed = _filter_site_ids()  # None=全部（管理员/无绑定），否则为可见站点列表

        # 未完成检查项（待执行），按站点范围隔离
        insp_q = """SELECT pi.*, s.name as site_name, s.code as site_code,
                           s.gps_lat as lat, s.gps_lng as lng, s.type as site_type
                    FROM insp_plan_items pi
                    JOIN insp_plans ip ON ip.id = pi.plan_id
                    JOIN sites s ON s.id = pi.site_id
                    WHERE pi.result IS NULL AND date(ip.generate_date)=?"""
        insp_params = [today]
        if allowed is not None:
            ph = ','.join('?' * len(allowed))
            insp_q += f" AND pi.site_id IN ({ph})"
            insp_params = list(allowed)
        insp_q += " ORDER BY pi.category, pi.item_name"
        items = db.execute(insp_q, insp_params).fetchall()

        # 按站点分组
        site_tasks = {}
        for item in items:
            sid = item['site_id']
            if sid not in site_tasks:
                site_tasks[sid] = {
                    'site_id': sid,
                    'site_name': item['site_name'],
                    'site_code': item['site_code'],
                    'lat': item['lat'],
                    'lng': item['lng'],
                    'site_type': item['site_type'],
                    'items': [],
                    'categories': {},
                }
            freq_cn = _FREQ_CN.get(item['frequency'] or '', item['frequency'] or '')
            item_dict = {
                'item_id': item['id'],
                'plan_id': item['plan_id'],
                'item_name': item['item_name'],
                'category': item['category'] or '其他',
                'frequency': item['frequency'] or '',
                'frequency_cn': freq_cn,
                'result': item['result'],
                'calibrator': item['calibrator'],
                'calibration_values': item['calibration_values'],
                'photo_urls': item['photo_urls'],
                'remark': item['remark'],
            }
            site_tasks[sid]['items'].append(item_dict)
            cat = item['category'] or '其他'
            if cat not in site_tasks[sid]['categories']:
                site_tasks[sid]['categories'][cat] = []
            site_tasks[sid]['categories'][cat].append(item_dict)

        # 统计（含已完成），同按站点范围
        stat_q = """SELECT pi.result FROM insp_plan_items pi
                    JOIN insp_plans ip ON ip.id=pi.plan_id
                    WHERE date(ip.generate_date)=?"""
        stat_params = [today]
        if allowed is not None:
            ph = ','.join('?' * len(allowed))
            stat_q += f" AND pi.site_id IN ({ph})"
            stat_params = list(allowed)
        total_items = 0
        completed_items = 0
        abnormal_items = 0
        for it in db.execute(stat_q, stat_params).fetchall():
            total_items += 1
            if it['result'] is not None:
                completed_items += 1
                if it['result'] == 'abnormal':
                    abnormal_items += 1

        # 构建站点列表
        sites_list = []
        _TM_CN = {'water_quality':'水质自动站','manual_station':'水质手动站','drinking_source':'饮用水源站','cross_boundary':'跨界断面站','groundwater':'地下水站'}
        for sid, st in site_tasks.items():
            cats_summary = []
            for cat_name, cat_items in st['categories'].items():
                cats_summary.append({'category': cat_name, 'pending': len(cat_items)})
            sites_list.append({
                'site_id': sid,
                'site_name': st['site_name'],
                'site_code': st['site_code'],
                'lat': st['lat'],
                'lng': st['lng'],
                'site_type': st['site_type'],
                'site_type_cn': _TM_CN.get(st['site_type'], st['site_type']),
                'pending_items': len(st['items']),
                'categories': cats_summary,
            })
        sites_list.sort(key=lambda x: x['site_name'])

        # ---- 2. 待处理工单：与网页端一致，按站点范围隔离（_filter_site_ids）----
        wo_q = """SELECT order_no, title, status, source, level, site_id,
                         (SELECT name FROM sites WHERE id=work_orders.site_id) as site_name,
                         created_at, sla_deadline
                  FROM work_orders WHERE 1=1"""
        wo_params = []
        if allowed is not None:
            ph = ','.join('?' * len(allowed))
            wo_q += f" AND site_id IN ({ph})"
            wo_params.extend(allowed)
        wo_q += """ AND status NOT IN ('closed')
                  ORDER BY
                    CASE level WHEN 'critical' THEN 1 WHEN 'urgent' THEN 2 ELSE 3 END,
                    created_at DESC
                  LIMIT 20"""
        workorders = db.execute(wo_q, wo_params).fetchall()

        _wo_status_cn = {'pending':'待处理','accepted':'已接受','dispatched':'已派发','in_progress':'处理中','reviewing':'审核中','resolved':'已解决','closed':'已关闭'}
        _wo_level_cn = {'normal':'普通','urgent':'紧急','critical':'严重'}
        _wo_source_cn = {'auto':'自动派发','auto_created':'自动派发','patrol':'巡检生成','manual':'手动创建','report':'上报工单'}

        wo_list = []
        for wo in workorders:
            wo_list.append({
                'order_no': wo['order_no'],
                'title': wo['title'],
                'status': wo['status'],
                'status_cn': _wo_status_cn.get(wo['status'], wo['status']),
                'source': wo['source'],
                'source_cn': _wo_source_cn.get(wo['source'], wo['source']),
                'level': wo['level'],
                'level_cn': _wo_level_cn.get(wo['level'], wo['level']),
                'site_id': wo['site_id'],
                'site_name': wo['site_name'],
                'created_at': wo['created_at'],
                'sla_deadline': wo['sla_deadline'],
            })

        # ---- 3. 未处理告警：与网页端一致，按站点范围隔离（不再依赖巡检计划站点）----
        if allowed is not None:
            ph = ','.join('?' * len(allowed))
            alerts = db.execute(
                f"""SELECT id, site_id, metric, level, message, status, created_at,
                           (SELECT name FROM sites WHERE id=alerts.site_id) as site_name
                    FROM alerts
                    WHERE status='pending' AND site_id IN ({ph})
                    ORDER BY
                      CASE level WHEN 'red' THEN 1 WHEN 'orange' THEN 2 WHEN 'yellow' THEN 3 ELSE 4 END,
                      created_at DESC
                    LIMIT 15""",
                allowed
            ).fetchall()
        else:
            alerts = db.execute(
                """SELECT id, site_id, metric, level, message, status, created_at,
                          (SELECT name FROM sites WHERE id=alerts.site_id) as site_name
                   FROM alerts
                   WHERE status='pending'
                   ORDER BY
                     CASE level WHEN 'red' THEN 1 WHEN 'orange' THEN 2 WHEN 'yellow' THEN 3 ELSE 4 END,
                     created_at DESC
                   LIMIT 15"""
            ).fetchall()

        _alert_level_cn = {'red':'红色','orange':'橙色','yellow':'黄色','blue':'蓝色'}
        alert_list = []
        for a in alerts:
            _m = a['metric'] or ''
            alert_list.append({
                'id': a['id'],
                'site_id': a['site_id'],
                'site_name': a['site_name'],
                'metric': _m,
                'metric_cn': METRIC_CN.get(_m, _m),
                'level': a['level'],
                'level_cn': _alert_level_cn.get(a['level'], a['level']),
                'message': a['message'],
                'status': a['status'],
                'created_at': a['created_at'],
            })

        # ---- 4. 今日作业包：由已批准计划派生，不增加一线手工维护 ----
        schedule_q = """SELECT * FROM plan_schedules
                        WHERE status='approved' AND period_start<=? AND period_end>=?"""
        schedule_params = [today, today]
        if user['role'] not in ('admin', 'manager', 'reviewer'):
            schedule_q += " AND user_id=?"
            schedule_params.append(user['id'])
        schedules = db.execute(schedule_q, schedule_params).fetchall()
        package_sites = set()
        package_vehicle_ids = set()
        package_parts = []
        package_work_order_ids = set()
        package_schedule_ids = []
        for schedule in schedules:
            try:
                plan_data = json.loads(schedule['plan_data'] or '{}')
                vehicle_days = json.loads(schedule['vehicle_days'] or '{}')
                spare_parts = json.loads(schedule['spare_parts'] or '[]')
                work_order_ids = json.loads(schedule['work_order_ids'] or '[]')
            except Exception:
                continue
            day_plan = plan_data.get(today) or {}
            day_sites = day_plan.get('sites') or []
            if not day_sites:
                continue
            package_schedule_ids.append(schedule['id'])
            package_sites.update(int(sid) for sid in day_sites)
            if vehicle_days.get(today):
                package_vehicle_ids.add(int(vehicle_days[today]))
            package_parts.extend(p for p in spare_parts if isinstance(p, dict))
            package_work_order_ids.update(int(wid) for wid in work_order_ids if str(wid).isdigit())

        package_vehicles = []
        if package_vehicle_ids:
            ph = ','.join('?' * len(package_vehicle_ids))
            package_vehicles = [dict(r) for r in db.execute(
                f"SELECT id, plate_number, name, status FROM vehicles WHERE id IN ({ph})",
                list(package_vehicle_ids)).fetchall()]
        package_orders = []
        if package_work_order_ids:
            ph = ','.join('?' * len(package_work_order_ids))
            package_orders = [dict(r) for r in db.execute(
                f"""SELECT id, order_no, site_id, title, level, status
                    FROM work_orders WHERE id IN ({ph}) AND status!='closed'""",
                list(package_work_order_ids)).fetchall()]
        package_site_rows = []
        if package_sites:
            ph = ','.join('?' * len(package_sites))
            package_site_rows = [dict(r) for r in db.execute(
                f"SELECT id, name, code FROM sites WHERE id IN ({ph}) ORDER BY name",
                list(package_sites)).fetchall()]

        work_package = {
            'has_plan': bool(package_schedule_ids),
            'schedule_ids': package_schedule_ids,
            'sites': package_site_rows,
            'vehicles': package_vehicles,
            'spare_parts': package_parts,
            'workorders': package_orders,
            'readiness': {
                'vehicle_confirmed': bool(package_vehicles),
                'parts_count': sum(int(p.get('quantity') or 1) for p in package_parts),
                'linked_workorders': len(package_orders),
            },
        }

        # ---- 汇总 ----
        return jsonify({
            'summary': {
                'total_sites': len(sites_list),
                'total_items': total_items,
                'completed_items': completed_items,
                'pending_items': total_items - completed_items,
                'abnormal_items': abnormal_items,
                'pending_workorders': len(wo_list),
                'pending_alerts': len(alert_list),
                'date': today,
            },
            'sites': sites_list,
            'workorders': wo_list,
            'alerts': alert_list,
            'work_package': work_package,
        })


@app.route('/api/mobile/site-tasks/<int:site_id>')
@login_required
def mobile_site_tasks(site_id):
    """移动端站点任务详情：返回该站点所有待检检查项（含已完成），按类别分组。"""
    user = g.current_user

    with get_db() as db:
        site = db.execute("SELECT * FROM sites WHERE id=?", (site_id,)).fetchone()
        if not site:
            return jsonify({'error': '站点不存在'}), 404

        # 获取该站点的所有检查项（来自活跃计划）
        plans = db.execute(
            "SELECT id FROM insp_plans WHERE status IN ('active','draft')"
        ).fetchall()
        plan_ids = [p['id'] for p in plans]

        if not plan_ids:
            return jsonify({
                'site': {'id': site['id'], 'name': site['name'], 'code': site['code'],
                         'lat': site['gps_lat'], 'lng': site['gps_lng'], 'type': site['type'],
                         'type_cn': {'water_quality':'水质自动站','manual_station':'水质手动站','drinking_source':'饮用水源站','cross_boundary':'跨界断面站','groundwater':'地下水站'}.get(site['type'], site['type'])},
                'categories': [],
                'total': 0, 'completed': 0,
            })

        placeholders = ','.join(['?'] * len(plan_ids))
        items = db.execute(
            f"""SELECT pi.* FROM insp_plan_items pi
                WHERE pi.site_id=? AND pi.plan_id IN ({placeholders})
                ORDER BY pi.category, pi.item_name""",
            [site_id] + plan_ids
        ).fetchall()

        # 按类别分组
        categories = {}
        total = 0
        completed = 0
        for item in items:
            total += 1
            if item['result'] is not None:
                completed += 1
            cat = item['category'] or '其他'
            if cat not in categories:
                categories[cat] = {'category': cat, 'items': [], 'total': 0, 'completed': 0}
            categories[cat]['total'] += 1
            if item['result'] is not None:
                categories[cat]['completed'] += 1
            freq_cn = _FREQ_CN.get(item['frequency'] or '', item['frequency'] or '')
            categories[cat]['items'].append({
                'item_id': item['id'],
                'plan_id': item['plan_id'],
                'item_name': item['item_name'],
                'frequency': item['frequency'] or '',
                'frequency_cn': freq_cn,
                'result': item['result'],
                'remark': item['remark'],
                'check_time': item['check_time'],
                'calibrator': item['calibrator'],
                'calibration_values': item['calibration_values'],
                'photo_urls': item['photo_urls'],
                'required_photos': item['required_photos'] if 'required_photos' in item.keys() else 0,
                'actual_photos': item['actual_photos'] if 'actual_photos' in item.keys() else 0,
            })

        _TM_CN = {'water_quality':'水质自动站','manual_station':'水质手动站','drinking_source':'饮用水源站','cross_boundary':'跨界断面站','groundwater':'地下水站'}
        return jsonify({
            'site': {'id': site['id'], 'name': site['name'], 'code': site['code'],
                     'lat': site['gps_lat'], 'lng': site['gps_lng'], 'type': site['type'],
                     'type_cn': _TM_CN.get(site['type'], site['type'])},
            'categories': list(categories.values()),
            'total': total,
            'completed': completed,
        })


def _mobile_execution_categories(db, plan_id, site_id):
    """将一个已批准执行包中的单站检查项整形成移动端可直接渲染的数据。"""
    items = db.execute("""SELECT * FROM insp_plan_items
        WHERE plan_id=? AND site_id=? AND COALESCE(execution_status, 'active')='active'
        ORDER BY category, item_name""", (plan_id, site_id)).fetchall()
    categories, total, completed = {}, 0, 0
    for item in items:
        total += 1
        completed += item['result'] is not None
        category = item['category'] or '其他'
        categories.setdefault(category, {'category': category, 'items': [], 'total': 0, 'completed': 0})
        categories[category]['total'] += 1
        categories[category]['completed'] += item['result'] is not None
        categories[category]['items'].append({
            'item_id': item['id'], 'plan_id': plan_id, 'item_name': item['item_name'],
            'frequency': item['frequency'] or '', 'frequency_cn': _FREQ_CN.get(item['frequency'] or '', item['frequency'] or ''),
            'result': item['result'], 'remark': item['remark'], 'check_time': item['check_time'],
            'calibrator': item['calibrator'], 'calibration_values': item['calibration_values'],
            'photo_urls': item['photo_urls'], 'required_photos': item['required_photos'] if 'required_photos' in item.keys() else 0,
            'actual_photos': item['actual_photos'] if 'actual_photos' in item.keys() else 0,
        })
    return list(categories.values()), total, completed


@app.route('/api/mobile/today-execution')
@login_required
def mobile_today_execution():
    """移动端唯一巡检入口：当前用户今日、已审批排程生成的执行包。"""
    user = g.current_user
    today = request.args.get('date') or datetime.now().strftime('%Y-%m-%d')
    with get_db() as db:
        rows = db.execute("""SELECT ip.id, ip.plan_name, ip.generate_date, ip.status, ip.completion_rate,
                ip.plan_schedule_id, ps.schedule_type, ps.vehicle_days, ps.spare_parts, ps.work_order_ids,
                ps.version, ps.remarks
            FROM insp_plans ip JOIN plan_schedules ps ON ps.id=ip.plan_schedule_id
            WHERE ip.assignee_id=? AND ip.generate_date=? AND ps.status='approved'
              AND ip.status IN ('active','completed')
            ORDER BY ip.id""", (user['id'], today)).fetchall()
        packages = []
        for row in rows:
            p = dict(row)
            try: vehicle_days = json.loads(p['vehicle_days'] or '{}')
            except Exception: vehicle_days = {}
            try: parts = json.loads(p['spare_parts'] or '[]')
            except Exception: parts = []
            try: order_ids = json.loads(p['work_order_ids'] or '[]')
            except Exception: order_ids = []
            vehicle_id = vehicle_days.get(today)
            vehicle = db.execute("SELECT id, plate_no FROM vehicles WHERE id=?", (vehicle_id,)).fetchone() if vehicle_id else None
            sites = db.execute("""SELECT pi.site_id, s.name, s.type, s.code,
                    COUNT(*) AS total, SUM(CASE WHEN pi.result IS NOT NULL THEN 1 ELSE 0 END) AS completed,
                    SUM(CASE WHEN pi.result='abnormal' THEN 1 ELSE 0 END) AS abnormal
                FROM insp_plan_items pi JOIN sites s ON s.id=pi.site_id
                WHERE pi.plan_id=? AND COALESCE(pi.execution_status,'active')='active'
                GROUP BY pi.site_id ORDER BY MIN(pi.id)""", (p['id'],)).fetchall()
            packages.append({
                'plan_id': p['id'], 'schedule_id': p['plan_schedule_id'], 'schedule_type': p['schedule_type'],
                'version': p['version'], 'status': p['status'], 'progress': p['completion_rate'],
                'vehicle': dict(vehicle) if vehicle else None, 'spare_parts': parts, 'work_order_ids': order_ids,
                'remarks': p['remarks'], 'sites': [dict(s) for s in sites],
            })
        return jsonify({'date': today, 'packages': packages})


@app.route('/api/mobile/execution-plans/<int:plan_id>/sites/<int:site_id>')
@login_required
def mobile_execution_site_tasks(plan_id, site_id):
    """只允许从今日已批准执行包进入单站检查，不再暴露按任意站点取任务的入口。"""
    user = g.current_user
    with get_db() as db:
        plan = db.execute("""SELECT ip.*, ps.status AS schedule_status FROM insp_plans ip
            JOIN plan_schedules ps ON ps.id=ip.plan_schedule_id
            WHERE ip.id=? AND ip.assignee_id=? AND ps.status='approved' AND ip.status IN ('active','completed')""",
            (plan_id, user['id'])).fetchone()
        if not plan:
            return jsonify({'error': '该执行任务不存在、未批准或不属于当前用户'}), 404
        site = db.execute("SELECT * FROM sites WHERE id=?", (site_id,)).fetchone()
        if not site:
            return jsonify({'error': '站点不存在'}), 404
        categories, total, completed = _mobile_execution_categories(db, plan_id, site_id)
        if not total:
            return jsonify({'error': '该站点不在当前执行包中'}), 404
        return jsonify({'site': {'id': site['id'], 'name': site['name'], 'code': site['code'],
                         'lat': site['gps_lat'], 'lng': site['gps_lng'], 'type': site['type'],
                         'type_cn': {'water_quality':'水质自动站','manual_station':'水质手动站','drinking_source':'饮用水源站','cross_boundary':'跨界断面站','groundwater':'地下水站'}.get(site['type'], site['type'])},
                        'plan_id': plan_id, 'categories': categories, 'total': total, 'completed': completed})


@app.route('/api/sites/<int:site_id>/calibrate', methods=['PUT'])
@login_required
def calibrate_site_location(site_id):
    """站点位置校准：一线人员到场后校准站点经纬度。"""
    data = request.get_json(silent=True) or {}
    new_lat = data.get('lat')
    new_lng = data.get('lng')
    if new_lat is None or new_lng is None:
        return jsonify({'error': '请提供经纬度'}), 400

    with get_db() as db:
        site = db.execute("SELECT * FROM sites WHERE id=?", (site_id,)).fetchone()
        if not site:
            return jsonify({'error': '站点不存在'}), 404

        old_lat = site['gps_lat']
        old_lng = site['gps_lng']

        # 计算偏移距离（简化公式，单位：米）
        import math
        dlat = math.radians(new_lat - old_lat) if old_lat else 0
        dlng = math.radians(new_lng - old_lng) if old_lng else 0
        a = math.sin(dlat/2)**2 + math.cos(math.radians(old_lat or 0)) * math.cos(math.radians(new_lat)) * math.sin(dlng/2)**2
        distance_m = 6371000 * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a)) if old_lat and old_lng else 0

        # 更新站点坐标
        db.execute("UPDATE sites SET gps_lat=?, gps_lng=? WHERE id=?", (new_lat, new_lng, site_id))

        # 记录校准日志
        user = g.current_user
        db.execute(
            "INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
            ('site', site_id, 'calibrated', user.get('username', ''),
             f'位置校准: ({old_lat},{old_lng}) → ({new_lat},{new_lng}), 偏移{distance_m:.1f}m')
        )
        db.commit()

        return jsonify({
            'success': True,
            'old_lat': old_lat, 'old_lng': old_lng,
            'new_lat': new_lat, 'new_lng': new_lng,
            'distance_m': round(distance_m, 1),
        })


def _mobile_idempotency_get(db, key, endpoint):
    if not key:
        return None
    db.execute("""CREATE TABLE IF NOT EXISTS mobile_idempotency (
        idempotency_key TEXT PRIMARY KEY,
        endpoint TEXT NOT NULL,
        response_json TEXT NOT NULL,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    )""")
    row = db.execute(
        "SELECT response_json FROM mobile_idempotency WHERE idempotency_key=? AND endpoint=?",
        (key, endpoint)).fetchone()
    if not row:
        return None
    try:
        return json.loads(row['response_json'])
    except Exception:
        return None


def _mobile_idempotency_store(db, key, endpoint, response):
    if not key:
        return
    db.execute("""INSERT OR REPLACE INTO mobile_idempotency
        (idempotency_key, endpoint, response_json) VALUES (?,?,?)""",
        (key, endpoint, json.dumps(response, ensure_ascii=False)))


@app.route('/api/mobile/submit-item', methods=['POST'])
@login_required
def mobile_submit_item():
    """移动端提交检查项结果（统一入口，支持普通项和校准项）。"""
    data = request.get_json(silent=True) or {}
    item_id = data.get('item_id')
    plan_id = data.get('plan_id')
    result = data.get('result')  # 'normal' or 'abnormal'
    idempotency_key = data.get('_idempotency_key')

    if not item_id or not result:
        return jsonify({'error': '缺少必要参数'}), 400

    with get_db() as db:
        cached = _mobile_idempotency_get(db, idempotency_key, 'submit-item')
        if cached is not None:
            return jsonify(cached)
        item = db.execute("SELECT * FROM insp_plan_items WHERE id=?", (item_id,)).fetchone()
        if not item:
            return jsonify({'error': '检查项不存在'}), 404
        if (item['execution_status'] if 'execution_status' in item.keys() else 'active') == 'cancelled':
            return jsonify({'error': '该检查项已被计划变更取消'}), 400
        execution = db.execute("""SELECT ip.assignee_id, ip.status, ps.status AS schedule_status
            FROM insp_plans ip LEFT JOIN plan_schedules ps ON ps.id=ip.plan_schedule_id WHERE ip.id=?""",
            (item['plan_id'],)).fetchone()
        if not execution or execution['assignee_id'] != g.current_user['id'] \
                or execution['status'] not in ('active', 'completed') \
                or (execution['schedule_status'] is not None and execution['schedule_status'] != 'approved'):
            return jsonify({'error': '该检查项不属于当前已批准执行任务'}), 403
        if item['result'] is not None:
            return jsonify({'error': '该检查项已完成', 'duplicate': True}), 400

        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        photo_error = validate_submission_photos(
            result, item['required_photos'] if 'required_photos' in item.keys() else 0,
            data.get('photo_urls'),
        )
        if photo_error:
            return jsonify({'error': photo_error}), 400

        updates = ["result=?", "check_time=?", "completed_at=?"]
        params = [result, now, now]

        if data.get('remark'):
            updates.append("remark=?")
            params.append(data['remark'])
        if data.get('gps_lat') is not None:
            updates.append("gps_lat=?")
            params.append(data['gps_lat'])
        if data.get('gps_lng') is not None:
            updates.append("gps_lng=?")
            params.append(data['gps_lng'])
        if data.get('photo_urls'):
            updates.append("photo_urls=?")
            params.append(data['photo_urls'])
        if data.get('calibrator'):
            updates.append("calibrator=?")
            params.append(data['calibrator'])
        if data.get('calibration_values'):
            updates.append("calibration_values=?")
            params.append(data['calibration_values'])

        params.append(item_id)
        db.execute(f"UPDATE insp_plan_items SET {','.join(updates)} WHERE id=?", params)

        order_no = None
        # 异常项自动触发告警并生成关联处置工单，避免一线人员再次切换页面补报。
        if result == 'abnormal':
            task = db.execute("SELECT site_id, item_name FROM insp_plan_items WHERE id=?", (item_id,)).fetchone()
            if task:
                remark_text = data.get('remark', '')
                msg = f'巡检异常：{task["item_name"]}'
                if remark_text:
                    msg += f' - {remark_text}'
                alert_id = create_alert_internal(db, task['site_id'], 'inspection', 0, 'yellow', msg)
                now_dt = datetime.now()
                order_no = f"IX{now_dt.strftime('%Y%m%d')}{item_id:06d}"
                assignee = _station_operator(task['site_id']) or g.current_user.get('real_name', '')
                db.execute("""INSERT OR IGNORE INTO work_orders
                    (order_no, site_id, source, event_type, level, title, description, assignee, status, sla_deadline, related_alert_id)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)""", (
                    order_no, task['site_id'], 'inspection', '巡检异常', 'normal',
                    f'【巡检异常】{task["item_name"]}', remark_text or msg, assignee, 'pending',
                    (now_dt + timedelta(hours=72)).strftime('%Y-%m-%d %H:%M:%S'), alert_id))
                if alert_id:
                    db.execute("UPDATE alerts SET related_order_no=? WHERE id=?", (order_no, alert_id))

        # 更新计划完成率和状态
        if plan_id:
            total = db.execute("""SELECT COUNT(*) as c FROM insp_plan_items
                WHERE plan_id=? AND COALESCE(execution_status, 'active')='active'""", (plan_id,)).fetchone()['c']
            done = db.execute("""SELECT COUNT(*) as c FROM insp_plan_items
                WHERE plan_id=? AND result IS NOT NULL
                  AND COALESCE(execution_status, 'active')='active'""", (plan_id,)).fetchone()['c']
            rate = round(done / total * 100, 1) if total > 0 else 0
            new_status = 'completed' if done == total else 'active'
            db.execute("UPDATE insp_plans SET completion_rate=?, status=? WHERE id=?",
                       (rate, new_status, plan_id))
            if done == total:
                plan = db.execute("SELECT plan_name FROM insp_plans WHERE id=?", (plan_id,)).fetchone()
                db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                           ('inspection', plan_id, 'completed', '系统', f'巡检计划完成-{plan["plan_name"] if plan else ""}'))

        response = {'success': True, 'result': result, 'order_no': order_no}
        _mobile_idempotency_store(db, idempotency_key, 'submit-item', response)
        db.commit()
        return jsonify(response)


@app.route('/api/mobile/check-in', methods=['POST'])
@login_required
def mobile_check_in():
    """移动端到站打卡：支持巡检站点打卡与工单到场签到（含GPS围栏校验）"""
    data = request.get_json(silent=True) or {}
    site_id = data.get('site_id')
    order_no = data.get('order_no')
    lat = data.get('lat')
    lng = data.get('lng')
    user = g.current_user
    idempotency_key = data.get('_idempotency_key')
    with get_db() as db:
        cached = _mobile_idempotency_get(db, idempotency_key, 'check-in')
        if cached is not None:
            return jsonify(cached)
        # 工单到场签到分支：校验距站点 ≤500m 后才记录
        if order_no:
            wo = db.execute("SELECT site_id, status FROM work_orders WHERE order_no=?", (order_no,)).fetchone()
            if not wo:
                return jsonify({'error': '工单不存在'}), 404
            site = db.execute("SELECT gps_lat, gps_lng FROM sites WHERE id=?", (wo['site_id'],)).fetchone()
            # GPS 围栏校验（站点坐标缺失则降级放行，不阻断既有流程）
            if site and site['gps_lat'] is not None and site['gps_lng'] is not None \
                    and lat is not None and lng is not None:
                import math
                s_lat = float(site['gps_lat']); s_lng = float(site['gps_lng'])
                dlat = math.radians(float(lat) - s_lat)
                dlng = math.radians(float(lng) - s_lng)
                a = math.sin(dlat / 2) ** 2 + math.cos(math.radians(s_lat)) * math.cos(math.radians(float(lat))) * math.sin(dlng / 2) ** 2
                distance_m = 6371000 * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
                if distance_m > 500:
                    return jsonify({'error': f'距站点约 {distance_m:.0f}m，超出 500m 到场范围，无法签到',
                                    'distance_m': round(distance_m)}), 400
            db.execute(
                "UPDATE work_orders SET check_in_lat=?, check_in_lng=?, check_in_time=datetime('now','localtime'), check_in_user=? WHERE order_no=?",
                (lat, lng, user.get('real_name') or user.get('username'), order_no))
            response = {'success': True, 'message': f'工单 {order_no} 已到场签到'}
            _mobile_idempotency_store(db, idempotency_key, 'check-in', response)
            db.commit()
            return jsonify(response)
        # 巡检站点打卡（原有逻辑）
        if not site_id:
            return jsonify({'error': '缺少站点ID'}), 400
        db.execute("""
            INSERT INTO inspection_checkins (site_id, site_name, user_id, user_name, check_time, lat, lng)
            VALUES (?,?,?,?,?,?,?)
        """, (site_id, data.get('site_name'), user['id'], user['real_name'],
              data.get('check_time'), lat, lng))
        response = {'success': True, 'message': f'已打卡站点 #{site_id}'}
        _mobile_idempotency_store(db, idempotency_key, 'check-in', response)
        db.commit()
        return jsonify(response)


@app.route('/api/mobile/upload-site-photo', methods=['POST'])
@login_required
def mobile_upload_site_photo():
    """移动端上传站点/巡检现场影像：保存文件并统一归口 operation_attachments，
    使照片在网页端「影像档案」立即可见（此前只存盘不落表，网页端看不到）。"""
    data = request.get_json(silent=True) or {}
    site_id = data.get('site_id')
    image = data.get('image', '')
    idempotency_key = data.get('_idempotency_key')
    if not site_id or not image:
        return jsonify({'error': '缺少站点ID或图片数据'}), 400
    try:
        import base64
        img_data = base64.b64decode(image.split(',')[-1])
        if len(img_data) < 200:
            return jsonify({'error': '图片数据异常，请重新拍摄或选择'}), 400
        with get_db() as db:
            cached = _mobile_idempotency_get(db, idempotency_key, 'upload-site-photo')
            if cached is not None:
                return jsonify(cached)
        photo_dir = os.path.join(UPLOAD_DIR, 'site_photos')
        os.makedirs(photo_dir, exist_ok=True)
        filename = f"site_{site_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.jpg"
        filepath = os.path.join(photo_dir, filename)
        with open(filepath, 'wb') as f:
            f.write(img_data)
        url = f'/uploads/site_photos/{filename}'
        # 统一归口：写入 operation_attachments，网页端影像档案可见
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        uploader = g.current_user
        with get_db() as db:
            db.execute("""INSERT INTO operation_attachments
                (filename, stored_path, file_type, mime_type, file_size, description,
                 source_type, source_id, site_id, uploader_id, uploader_name,
                 taken_at, created_at, category)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (filename, url, 'image', '.jpg', len(img_data),
                 f'站点[{site_id}]现场照片', 'site_photo', 0, site_id,
                 uploader['id'], uploader.get('real_name', ''),
                 now_str, now_str, '现场照片'))
            new_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]
            response = {'success': True, 'url': url, 'size': len(img_data), 'id': new_id}
            _mobile_idempotency_store(db, idempotency_key, 'upload-site-photo', response)
            db.commit()
        return jsonify(response)
    except Exception as e:
        return jsonify({'error': f'保存失败: {str(e)}'}), 500


@app.route('/api/mobile/workorder/<order_no>/image', methods=['POST'])
@login_required
def mobile_workorder_image(order_no):
    """移动端上传工单处置影像：保存文件并统一归口 operation_attachments（与网页端同一函数），
    使照片同时进入「影像档案」与「工单影像审核」队列。"""
    data = request.get_json(silent=True) or {}
    image = data.get('image', '')
    if not image:
        return jsonify({'error': '缺少图片数据'}), 400
    try:
        import base64
        img_data = base64.b64decode(image.split(',')[-1])
        # 拦截空图/占位图： base64 解码后不足 200 字节的基本是 1×1 占位或损坏图
        if len(img_data) < 200:
            return jsonify({'error': '图片数据异常，请重新拍摄或选择'}), 400
        photo_dir = os.path.join(UPLOAD_DIR, 'workorder_photos')
        os.makedirs(photo_dir, exist_ok=True)
        filename = f"wo_{order_no}_{datetime.now().strftime('%Y%m%d%H%M%S')}.jpg"
        filepath = os.path.join(photo_dir, filename)
        with open(filepath, 'wb') as f:
            f.write(img_data)
        url = f'/uploads/workorder_photos/{filename}'
        # 复用网页端同一归口：写 operation_attachments + work_orders.images
        res = _batch_link_wo_photos(order_no, [url], file_size=len(img_data))
        if isinstance(res, dict):
            return jsonify({'success': True, 'url': url, 'images': res.get('images')})
        return res  # 错误元组（工单不存在等）
    except Exception as e:
        return jsonify({'error': f'保存失败: {str(e)}'}), 500


@app.route('/api/mobile/delete-photo', methods=['POST'])
@login_required
def mobile_delete_photo():
    """删除检查项中的某张照片"""
    data = request.get_json(silent=True) or {}
    item_id = data.get('item_id')
    photo_index = data.get('photo_index')
    if item_id is None or photo_index is None:
        return jsonify({'error': '缺少参数'}), 400
    with get_db() as db:
        item = db.execute("SELECT photo_urls FROM insp_plan_items WHERE id=?", (item_id,)).fetchone()
        if not item or not item['photo_urls']:
            return jsonify({'error': '无照片可删除'}), 404
        try:
            urls = json.loads(item['photo_urls'])
            if not isinstance(urls, list) or photo_index < 0 or photo_index >= len(urls):
                return jsonify({'error': '照片索引无效'}), 400
            removed = urls.pop(photo_index)
            db.execute("UPDATE insp_plan_items SET photo_urls=? WHERE id=?", (json.dumps(urls), item_id))
            # 尝试删除物理文件
            if isinstance(removed, str) and removed.startswith('/uploads/'):
                filepath = os.path.join(os.path.dirname(DB_PATH), '..', removed.lstrip('/'))
                if os.path.exists(filepath):
                    try: os.remove(filepath)
                    except: pass
            db.commit()
            return jsonify({'success': True})
        except (json.JSONDecodeError, TypeError, ValueError):
            return jsonify({'error': '照片数据格式错误'}), 500


@app.route('/api/inspection-v2/photos')
@login_required
def v2_inspection_photos():
    """返回指定站点的巡检照片列表（供网页端展示，按站点范围隔离）"""
    allowed = _filter_site_ids()
    site_id = request.args.get('site_id', type=int)
    plan_id = request.args.get('plan_id', type=int)
    limit = min(request.args.get('limit', 50, type=int), 200)
    with get_db() as db:
        q = "SELECT pi.id, pi.item_name, pi.photo_urls, pi.result, pi.check_time, pi.site_id, s.name as site_name, p.plan_name FROM insp_plan_items pi JOIN sites s ON pi.site_id=s.id JOIN insp_plans p ON pi.plan_id=p.id WHERE pi.photo_urls IS NOT NULL AND pi.photo_urls != ''"
        params = []
        if site_id:
            if allowed is not None and site_id not in allowed:
                return jsonify([])
            q += " AND pi.site_id=?"
            params.append(site_id)
        elif allowed is not None:
            ph = ','.join('?' * len(allowed))
            q += f" AND pi.site_id IN ({ph})"
            params.extend(allowed)
        if plan_id:
            q += " AND pi.plan_id=?"
            params.append(plan_id)
        q += " ORDER BY pi.check_time DESC LIMIT ?"
        params.append(limit)
        rows = db.execute(q, params).fetchall()
        result = []
        for r in rows:
            item = dict(r)
            try:
                urls = json.loads(item['photo_urls'])
                item['photos'] = urls if isinstance(urls, list) else [item['photo_urls']]
            except:
                item['photos'] = [item['photo_urls']] if item['photo_urls'] else []
            result.append(item)
        return jsonify(result)


# ===== 统一文件管理API =====

@app.route('/api/files', methods=['GET'])
@login_required
def list_files():
    """按业务类型查询文件列表：?source_type=xxx&source_id=xxx"""
    source_type = request.args.get('source_type', '')
    source_id = request.args.get('source_id', type=int)
    limit = min(request.args.get('limit', 50, type=int), 200)
    with get_db() as db:
        q = "SELECT * FROM files WHERE is_deleted=0"
        params = []
        if source_type:
            q += " AND source_type=?"
            params.append(source_type)
        if source_id:
            q += " AND source_id=?"
            params.append(source_id)
        q += " ORDER BY created_at DESC LIMIT ?"
        params.append(limit)
        rows = db.execute(q, params).fetchall()
        return jsonify([dict(r) for r in rows])


@app.route('/api/files/<int:fid>', methods=['GET'])
@login_required
def get_file(fid):
    """获取文件元数据"""
    with get_db() as db:
        f = db.execute("SELECT * FROM files WHERE id=? AND is_deleted=0", (fid,)).fetchone()
        if not f:
            return jsonify({'error': '文件不存在'}), 404
        return jsonify(dict(f))


@app.route('/api/files/<int:fid>', methods=['DELETE'])
@login_required
def delete_file(fid):
    """软删除文件"""
    with get_db() as db:
        f = db.execute("SELECT * FROM files WHERE id=? AND is_deleted=0", (fid,)).fetchone()
        if not f:
            return jsonify({'error': '文件不存在'}), 404
        db.execute("UPDATE files SET is_deleted=1 WHERE id=?", (fid,))
        db.commit()
        return jsonify({'success': True})


# ===== 时序数据聚合与清理 =====

def _aggregate_hourly():
    """每小时时序聚合：sensor_data_raw → sensor_data_hourly"""
    with get_db() as db:
        last_hour = (datetime.now() - timedelta(hours=1)).strftime('%Y-%m-%d %H:00:00')
        rows = db.execute("""
            SELECT site_id, metric,
                   ROUND(AVG(value),2) as avg_val,
                   ROUND(MIN(value),2) as min_val,
                   ROUND(MAX(value),2) as max_val,
                   COUNT(*) as cnt
            FROM sensor_data_raw
            WHERE recorded_at >= ? AND recorded_at < ?
            GROUP BY site_id, metric
        """, (last_hour, (datetime.now()).strftime('%Y-%m-%d %H:00:00'))).fetchall()
        for r in rows:
            db.execute("""
                INSERT INTO sensor_data_hourly (site_id, metric, hour, avg_value, min_value, max_value, sample_count)
                VALUES (?,?,?,?,?,?,?)
                ON CONFLICT(site_id, metric, hour) DO UPDATE SET
                    avg_value=excluded.avg_value, min_value=excluded.min_value,
                    max_value=excluded.max_value, sample_count=excluded.sample_count
            """, (r['site_id'], r['metric'], last_hour, r['avg_val'], r['min_val'], r['max_val'], r['cnt']))
        db.commit()
        return len(rows)


def _aggregate_daily():
    """每日时序聚合：sensor_data_hourly → sensor_data_daily"""
    yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    with get_db() as db:
        rows = db.execute("""
            SELECT site_id, metric,
                   ROUND(AVG(avg_value),2) as avg_val,
                   ROUND(MIN(min_value),2) as min_val,
                   ROUND(MAX(max_value),2) as max_val,
                   SUM(sample_count) as cnt
            FROM sensor_data_hourly
            WHERE hour >= ? AND hour < ?
            GROUP BY site_id, metric
        """, (yesterday, (datetime.now()).strftime('%Y-%m-%d'))).fetchall()
        for r in rows:
            db.execute("""
                INSERT INTO sensor_data_daily (site_id, metric, date, avg_value, min_value, max_value, sample_count)
                VALUES (?,?,?,?,?,?,?)
                ON CONFLICT(site_id, metric, date) DO UPDATE SET
                    avg_value=excluded.avg_value, min_value=excluded.min_value,
                    max_value=excluded.max_value, sample_count=excluded.sample_count
            """, (r['site_id'], r['metric'], yesterday, r['avg_val'], r['min_val'], r['max_val'], r['cnt']))
        db.commit()
        return len(rows)


def _cleanup_raw_data(retain_days=7):
    """清理原始时序数据（保留retain_days天）"""
    cutoff = (datetime.now() - timedelta(days=retain_days)).strftime('%Y-%m-%d')
    with get_db() as db:
        db.execute("DELETE FROM sensor_data_raw WHERE recorded_at < ?", (cutoff,))
        db.commit()


@app.route('/api/data/aggregate', methods=['POST'])
@login_required
def trigger_aggregation():
    """手动触发聚合（供定时任务或管理员调用）"""
    action = request.args.get('action', 'hourly')
    try:
        if action == 'hourly':
            n = _aggregate_hourly()
        elif action == 'daily':
            n = _aggregate_daily()
        elif action == 'cleanup':
            _cleanup_raw_data()
            n = 0
        else:
            return jsonify({'error': '未知操作'}), 400
        return jsonify({'success': True, 'action': action, 'processed': n})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/data/realtime-v2')
@login_required
def get_realtime_data_v2():
    """查询实时数据（优先小时表，回退原始表）"""
    site_id = request.args.get('site_id', type=int)
    metric = request.args.get('metric', '')
    hours = min(request.args.get('hours', 24, type=int), 168)
    with get_db() as db:
        cutoff = (datetime.now() - timedelta(hours=hours)).strftime('%Y-%m-%d %H:%M:%S')
        q = "SELECT * FROM sensor_data_raw WHERE 1=1"
        params = []
        if site_id:
            q += " AND site_id=?"
            params.append(site_id)
        if metric:
            q += " AND metric=?"
            params.append(metric)
        q += " AND recorded_at>=? ORDER BY recorded_at ASC"
        params.append(cutoff)
        rows = db.execute(q, params).fetchall()
        if len(rows) < 10 and hours > 48:
            # 原始数据不足，回退到小时表
            q2 = "SELECT site_id, metric, hour as recorded_at, avg_value as value FROM sensor_data_hourly WHERE 1=1"
            if site_id:
                q2 += " AND site_id=?"
            if metric:
                q2 += " AND metric=?"
            q2 += " AND hour>=? ORDER BY hour ASC"
            rows = db.execute(q2, params + [cutoff]).fetchall()
        return jsonify([dict(r) for r in rows])


# ============================================================
# 水质智慧运维平台 v2 新增路由
# ============================================================

# ---------- 2.1 照片需求模板 ----------
@app.route('/api/photo-requirements', methods=['GET'])
def api_photo_requirements():
    """获取照片需求模板，可按site_type/period筛选"""
    site_type = request.args.get('site_type', 'water_quality')
    period = request.args.get('period')
    with get_db() as db:
        if period:
            rows = db.execute(
                'SELECT * FROM photo_requirements WHERE site_type=? AND period=? ORDER BY seq',
                (site_type, period)).fetchall()
        else:
            rows = db.execute(
                'SELECT * FROM photo_requirements WHERE site_type=? ORDER BY period, seq',
                (site_type,)).fetchall()
        return jsonify([dict(r) for r in rows])


# ---------- 2.2 巡检照片审核状态查询 ----------
@app.route('/api/inspection/photos/<int:site_id>', methods=['GET'])
def api_inspection_photos_site(site_id):
    """获取某站点某周期的照片上传与审核状态"""
    period = request.args.get('period', 'weekly')
    with get_db() as db:
        # 获取该站点类型对应的需求模板
        site = db.execute('SELECT type FROM sites WHERE id=?', (site_id,)).fetchone()
        site_type = site['type'] if site else 'water_quality'
        requirements = db.execute(
            'SELECT * FROM photo_requirements WHERE site_type=? AND period=? ORDER BY seq',
            (site_type, period)).fetchall()
        result = []
        for req in requirements:
            # 查已上传的照片
            photos = db.execute(
                '''SELECT id, filename, stored_path, review_status, reviewer_id, reviewed_at, reject_reason
                   FROM operation_attachments
                   WHERE site_id=? AND requirement_id=? AND is_deleted=0
                   ORDER BY created_at DESC''',
                (site_id, req['id'])).fetchall()
            item = dict(req)
            item['photos'] = [dict(p) for p in photos]
            item['uploaded'] = len(photos)
            item['required'] = req['photo_count']
            item['complete'] = len(photos) >= req['photo_count']
            item['pending_review'] = sum(1 for p in photos if p['review_status'] == 'pending')
            result.append(item)
        return jsonify(result)


# ---------- 2.3 上传巡检照片（关联需求项） ----------
@app.route('/api/inspection/photos/upload', methods=['POST'])
def api_inspection_photo_upload():
    """巡检照片上传，关联photo_requirements"""
    data = request.get_json() or {}
    site_id = data.get('site_id')
    requirement_id = data.get('requirement_id')
    stored_path = data.get('stored_path', '')
    filename = data.get('filename', '')
    uploader_id = data.get('uploader_id')
    uploader_name = data.get('uploader_name', '')
    gps_lat = data.get('gps_lat')
    gps_lng = data.get('gps_lng')
    category = data.get('category', '巡检照片')
    description = data.get('description', '')

    if not site_id or not requirement_id:
        return jsonify({'error': '缺少 site_id 或 requirement_id'}), 400

    with get_db() as db:
        # 查需求项是否需审核
        req = db.execute('SELECT review_required FROM photo_requirements WHERE id=?',
                         (requirement_id,)).fetchone()
        review_required = req['review_required'] if req else 0

        cur = db.execute(
            '''INSERT INTO operation_attachments
               (filename, stored_path, file_type, source_type, source_id, site_id,
                uploader_id, uploader_name, gps_lat, gps_lng, category, description,
                requirement_id, review_status)
               VALUES (?,?,'image','inspection',0,?,?,?,?,?,?,?,?,?)''',
            (filename, stored_path, site_id, uploader_id, uploader_name,
             gps_lat, gps_lng, category, description, requirement_id,
             'pending' if review_required else 'approved'))
        db.commit()
        new_id = cur.lastrowid
        row = db.execute('SELECT * FROM operation_attachments WHERE id=?', (new_id,)).fetchone()
        return jsonify(dict(row)), 201


# ---------- 2.4 批量审核照片 ----------
@app.route('/api/inspection/photos/batch-review', methods=['POST'])
def api_inspection_photos_batch_review():
    """批量审核照片"""
    data = request.get_json() or {}
    photo_ids = data.get('photo_ids', [])
    action = data.get('action', 'approve')  # 'approve' or 'reject'
    reviewer_id = data.get('reviewer_id')
    reject_reason = data.get('reject_reason', '')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if not photo_ids:
        return jsonify({'error': '缺少 photo_ids'}), 400

    with get_db() as db:
        new_status = 'approved' if action == 'approve' else 'rejected'
        for pid in photo_ids:
            # 查照片信息（用于通知和重传）
            photo = db.execute(
                'SELECT site_id, uploader_id, uploader_name, description FROM operation_attachments WHERE id=?',
                (pid,)).fetchone()
            db.execute(
                '''UPDATE operation_attachments
                   SET review_status=?, reviewer_id=?, reviewed_at=?, reject_reason=?
                   WHERE id=?''',
                (new_status, reviewer_id, now, reject_reason if action == 'reject' else None, pid))
            # 驳回时写通知，通知原上传人
            if action == 'reject' and photo:
                item_name = photo['description'] or f'照片#{pid}'
                db.execute(
                    'INSERT INTO notifications (user_id, source_type, source_id, title, content) VALUES (?,?,?,?,?)',
                    (photo['uploader_id'] or 1, 'photo_review', pid,
                     f'照片被驳回',
                     f'「{item_name}」被驳回，原因：{reject_reason or "未达标"}。请重新拍摄上传。'))
        db.commit()
        return jsonify({'ok': True, 'count': len(photo_ids), 'status': new_status})


# ---------- 2.5 单张照片审核 ----------
@app.route('/api/inspection/photos/<int:photo_id>/review', methods=['POST'])
def api_inspection_photo_review(photo_id):
    """审核单张照片"""
    data = request.get_json() or {}
    action = data.get('action', 'approve')
    reviewer_id = data.get('reviewer_id')
    reject_reason = data.get('reject_reason', '')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    new_status = 'approved' if action == 'approve' else 'rejected'
    with get_db() as db:
        db.execute(
            '''UPDATE operation_attachments
               SET review_status=?, reviewer_id=?, reviewed_at=?, reject_reason=?
               WHERE id=?''',
            (new_status, reviewer_id, now, reject_reason if action == 'reject' else None, photo_id))
        db.commit()
        row = db.execute('SELECT * FROM operation_attachments WHERE id=?', (photo_id,)).fetchone()
        if not row:
            return jsonify({'error': '照片不存在'}), 404
        return jsonify(dict(row))


# ---------- 2.6 缺项检查（巡检提交时校验） ----------
@app.route('/api/inspection/photos/<int:site_id>/check', methods=['GET'])
def api_inspection_photos_check(site_id):
    """检查某站点某周期照片是否传齐，返回缺项列表"""
    period = request.args.get('period', 'weekly')
    with get_db() as db:
        site = db.execute('SELECT type FROM sites WHERE id=?', (site_id,)).fetchone()
        site_type = site['type'] if site else 'water_quality'
        requirements = db.execute(
            'SELECT * FROM photo_requirements WHERE site_type=? AND period=? ORDER BY seq',
            (site_type, period)).fetchall()
        missing = []
        for req in requirements:
            count = db.execute(
                'SELECT COUNT(*) as cnt FROM operation_attachments WHERE site_id=? AND requirement_id=? AND is_deleted=0',
                (site_id, req['id'])).fetchone()['cnt']
            if count < req['photo_count']:
                missing.append({
                    'requirement_id': req['id'],
                    'item_name': req['item_name'],
                    'required': req['photo_count'],
                    'uploaded': count,
                    'missing': req['photo_count'] - count
                })
        return jsonify({'site_id': site_id, 'period': period, 'complete': len(missing) == 0, 'missing': missing})


# ---------- 试剂状态计算（纯手动，不预测）----------
def compute_reagent_status(inv):
    """根据库存行计算试剂状态与剩余可用天数。
    剩余使用时间完全由运维人员经验手动填写（expected_duration_days），系统不做预测。
    触发预警的两类：剩余天数 <= 临期阈值（warning_days）或 余量 <= 低余量阈值。
    """
    expected = inv.get('expected_duration_days')
    last = inv.get('last_replaced_at')
    warning = inv.get('warning_days') or 7
    qty = inv.get('current_qty')
    threshold = inv.get('low_stock_threshold')
    remaining = None
    expires_at = None
    flags = []
    if expected is not None and last:
        try:
            ld = datetime.strptime(last[:19], '%Y-%m-%d %H:%M:%S')
        except ValueError:
            try:
                ld = datetime.strptime(last[:10], '%Y-%m-%d')
            except ValueError:
                ld = None
        if ld:
            exp = ld + timedelta(days=expected)
            expires_at = exp.strftime('%Y-%m-%d')
            remaining = (exp.date() - datetime.now().date()).days
            if remaining <= 0:
                flags.append('已过期')
            elif remaining <= warning:
                flags.append('临期')
    if qty is not None and threshold is not None and qty <= threshold:
        flags.append('低余量')
    if not flags:
        status = '未设置' if (expected is None and last is None) else '正常'
    else:
        status = '已过期' if '已过期' in flags else ('临期' if '临期' in flags else '低余量')
    return {'remaining_days': remaining, 'expires_at': expires_at, 'status': status}


# ---------- 3.1 试剂主数据 ----------
@app.route('/api/reagents', methods=['GET'])
def api_reagents():
    """获取试剂主数据列表"""
    with get_db() as db:
        rows = db.execute('SELECT * FROM reagents ORDER BY name').fetchall()
        return jsonify([dict(r) for r in rows])


@app.route('/api/reagents', methods=['POST'])
@login_required
def api_reagents_create():
    """新增试剂主数据（重名校验）"""
    g_ = require_admin()
    if g_:
        return g_
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': '请填写试剂名称'}), 400
    manufacturer = (data.get('manufacturer') or '').strip()
    spec = (data.get('spec') or '').strip()
    unit = (data.get('unit') or '瓶').strip()
    try:
        shelf_life_days = int(data.get('shelf_life_days', 365) or 365)
    except (ValueError, TypeError):
        shelf_life_days = 365
    with get_db() as db:
        dup = db.execute('SELECT id FROM reagents WHERE name=?', (name,)).fetchone()
        if dup:
            return jsonify({'error': f'已存在名为「{name}」的试剂，请勿重复添加'}), 400
        cur = db.execute(
            'INSERT INTO reagents (name, manufacturer, spec, unit, shelf_life_days) VALUES (?,?,?,?,?)',
            (name, manufacturer, spec, unit, shelf_life_days))
        db.commit()
        row = db.execute('SELECT * FROM reagents WHERE id=?', (cur.lastrowid,)).fetchone()
        return jsonify(dict(row)), 201


@app.route('/api/reagents/<int:rid>', methods=['PUT'])
def api_reagents_update(rid):
    """编辑试剂主数据（重名校验）"""
    g_ = require_admin()
    if g_:
        return g_
    data = request.get_json() or {}
    with get_db() as db:
        row = db.execute('SELECT * FROM reagents WHERE id=?', (rid,)).fetchone()
        if not row:
            return jsonify({'error': '试剂不存在'}), 404
        name = (data.get('name', row['name']) or '').strip()
        if not name:
            return jsonify({'error': '请填写试剂名称'}), 400
        dup = db.execute('SELECT id FROM reagents WHERE name=? AND id<>?', (name, rid)).fetchone()
        if dup:
            return jsonify({'error': f'已存在名为「{name}」的试剂，请勿重复'}), 400
        manufacturer = (data.get('manufacturer', row['manufacturer'] or '') or '').strip()
        spec = (data.get('spec', row['spec'] or '') or '').strip()
        unit = (data.get('unit', row['unit'] or '瓶') or '').strip()
        try:
            shelf_life_days = int(data.get('shelf_life_days', row['shelf_life_days']) or 365)
        except (ValueError, TypeError):
            shelf_life_days = 365
        db.execute(
            'UPDATE reagents SET name=?, manufacturer=?, spec=?, unit=?, shelf_life_days=? WHERE id=?',
            (name, manufacturer, spec, unit, shelf_life_days, rid))
        db.commit()
        nr = db.execute('SELECT * FROM reagents WHERE id=?', (rid,)).fetchone()
        return jsonify(dict(nr))


@app.route('/api/reagents/<int:rid>', methods=['DELETE'])
def api_reagents_delete(rid):
    """删除试剂主数据（被引用时拦截，避免产生孤儿记录）"""
    g_ = require_admin()
    if g_:
        return g_
    with get_db() as db:
        row = db.execute('SELECT * FROM reagents WHERE id=?', (rid,)).fetchone()
        if not row:
            return jsonify({'error': '试剂不存在'}), 404
        inv = db.execute('SELECT COUNT(*) FROM reagent_inventory WHERE reagent_id=?', (rid,)).fetchone()[0]
        al = db.execute('SELECT COUNT(*) FROM reagent_alerts WHERE reagent_id=?', (rid,)).fetchone()[0]
        us = db.execute('SELECT COUNT(*) FROM reagent_usage WHERE reagent_id=?', (rid,)).fetchone()[0]
        if inv or al or us:
            parts = []
            if inv: parts.append(f'{inv} 条站点库存')
            if al: parts.append(f'{al} 条告警')
            if us: parts.append(f'{us} 条用量')
            return jsonify({'error': f'该试剂已被{"、".join(parts)}引用，无法删除；请先清理相关记录后再试。'}), 400
        db.execute('DELETE FROM reagents WHERE id=?', (rid,))
        db.commit()
        return jsonify({'ok': True, 'id': rid})


# ---------- 3.2 站点试剂库存 ----------
@app.route('/api/reagent-inventory/<int:site_id>', methods=['GET'])
def api_reagent_inventory(site_id):
    """获取站点试剂库存（含计算的剩余可用天数与状态）"""
    with get_db() as db:
        rows = db.execute(
            '''SELECT ri.*, r.name as reagent_name, r.manufacturer, r.unit, r.shelf_life_days
               FROM reagent_inventory ri
               JOIN reagents r ON ri.reagent_id = r.id
               WHERE ri.site_id=? ORDER BY r.name''',
            (site_id,)).fetchall()
        out = []
        for r in rows:
            d = dict(r)
            d.update(compute_reagent_status(d))
            out.append(d)
        return jsonify(out)


# ---------- 3.2.1 新增站点试剂库存 ----------
@app.route('/api/reagent-inventory', methods=['POST'])
def api_reagent_inventory_create():
    """从试剂主数据目录中为本站点新增一条试剂库存记录（受站点范围约束）"""
    data = request.get_json() or {}
    site_id = data['site_id']
    allowed = _filter_site_ids()
    if allowed is not None and (not site_id or int(site_id) not in allowed):
        return jsonify({'ok': False, 'error': '无权限为非本人站点添加试剂库存'}), 403
    reagent_id = data['reagent_id']
    current_qty = data.get('current_qty')
    last_replaced_at = data.get('last_replaced_at')
    expected_duration_days = data.get('expected_duration_days')
    with get_db() as db:
        exist = db.execute(
            'SELECT id FROM reagent_inventory WHERE site_id=? AND reagent_id=?',
            (site_id, reagent_id)).fetchone()
        if exist:
            return jsonify({'ok': False, 'error': '该站点已存在此试剂库存，请勿重复添加'}), 400
        db.execute(
            'INSERT INTO reagent_inventory (site_id, reagent_id, current_qty, last_replaced_at, expected_duration_days) VALUES (?,?,?,?,?)',
            (site_id, reagent_id, current_qty, last_replaced_at, expected_duration_days))
        db.commit()
        row = db.execute('SELECT * FROM reagent_inventory WHERE site_id=? AND reagent_id=?', (site_id, reagent_id)).fetchone()
        return jsonify(dict(row)), 201


# ---------- 3.2.2 删除站点试剂库存 ----------
@app.route('/api/reagent-inventory/<int:site_id>/<int:reagent_id>', methods=['DELETE'])
def api_reagent_inventory_delete(site_id, reagent_id):
    """删除某站点的某条试剂库存记录"""
    with get_db() as db:
        db.execute('DELETE FROM reagent_inventory WHERE site_id=? AND reagent_id=?', (site_id, reagent_id))
        db.commit()
        return jsonify({'ok': True})


# ---------- 3.3 记录试剂用量 ----------
@app.route('/api/reagent-inventory/usage', methods=['POST'])
def api_reagent_usage():
    """记录试剂用量，自动更新库存"""
    data = request.get_json() or {}
    site_id = data['site_id']
    reagent_id = data['reagent_id']
    used_qty = data['used_qty']
    expected_duration_days = data.get('expected_duration_days')
    operator_id = data.get('operator_id')
    remark = data.get('remark', '')

    with get_db() as db:
        # 插入用量记录
        db.execute(
            'INSERT INTO reagent_usage (site_id, reagent_id, used_qty, expected_duration_days, operator_id, remark) VALUES (?,?,?,?,?,?)',
            (site_id, reagent_id, used_qty, expected_duration_days, operator_id, remark))
        # 更新库存
        db.execute(
            'UPDATE reagent_inventory SET current_qty=current_qty-?, updated_at=datetime("now","localtime") WHERE site_id=? AND reagent_id=?',
            (used_qty, site_id, reagent_id))
        db.commit()
        # 检查是否低于阈值
        inv = db.execute(
            'SELECT current_qty, low_stock_threshold FROM reagent_inventory WHERE site_id=? AND reagent_id=?',
            (site_id, reagent_id)).fetchone()
        if inv and inv['current_qty'] <= inv['low_stock_threshold']:
            db.execute(
                "INSERT INTO reagent_alerts (site_id, reagent_id, alert_type, current_qty, threshold_qty) VALUES (?,?,'low_stock',?,?)",
                (site_id, reagent_id, inv['current_qty'], inv['low_stock_threshold']))
            db.commit()
        return jsonify({'ok': True, 'remaining_qty': inv['current_qty'] if inv else 0})


# ---------- 3.4 记录试剂更换 ----------
@app.route('/api/reagent-inventory/replacement', methods=['POST'])
def api_reagent_replacement():
    """记录试剂更换，更新库存+写reagent_records"""
    data = request.get_json() or {}
    site_id = data['site_id']
    reagent_id = data['reagent_id']
    old_qty = data.get('old_qty', 0)
    new_qty = data.get('new_qty')
    old_batch_no = data.get('old_batch_no', '')
    new_batch_no = data.get('new_batch_no', '')
    operator = data.get('operator', '')
    operator_id = data.get('operator_id')
    expected_duration_days = data.get('expected_duration_days')
    replaced_at = data.get('replaced_at') or datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    with get_db() as db:
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        # 写历史表（原有表）
        reagent_name = db.execute('SELECT name FROM reagents WHERE id=?', (reagent_id,)).fetchone()
        db.execute(
            'INSERT INTO reagent_records (site_id, reagent_name, reagent_type, usage_date, replacement_date, operator, notes, old_batch_no, new_batch_no, old_qty, new_qty) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
            (site_id, reagent_name['name'] if reagent_name else '', '', replaced_at, replaced_at, operator, '更换试剂', old_batch_no, new_batch_no, old_qty, new_qty))
        # 更新库存（last_replaced_at 用本次更换日期；expected_duration_days 由运维人员手动填写）
        # 更换后进入"待质控"状态：必须跑标样验证通过才算更换完成
        inv = db.execute(
            'SELECT * FROM reagent_inventory WHERE site_id=? AND reagent_id=?',
            (site_id, reagent_id)).fetchone()
        if inv:
            db.execute(
                "UPDATE reagent_inventory SET current_qty=?, last_replaced_at=?, expected_duration_days=?, updated_at=?, qc_status='pending' WHERE site_id=? AND reagent_id=?",
                (new_qty, replaced_at, expected_duration_days, now, site_id, reagent_id))
        else:
            db.execute(
                "INSERT INTO reagent_inventory (site_id, reagent_id, current_qty, last_replaced_at, expected_duration_days, qc_status) VALUES (?,?,?,?,?,'pending')",
                (site_id, reagent_id, new_qty, replaced_at, expected_duration_days))
        db.commit()
        return jsonify({'ok': True, 'qc_status': 'pending'})


# ---------- 3.4.2 试剂质控（更换后跑标样验证） ----------
@app.route('/api/reagent-qc', methods=['POST'])
def api_reagent_qc_submit():
    """记录试剂质控结果（更换后跑标样）：
    passed=1 → qc_status='passed'（更换完成）；passed=0 → qc_status='failed' + 记录处置动作（校准/报修）。"""
    u = g.current_user
    data = request.get_json(silent=True) or {}
    site_id = data.get('site_id')
    reagent_id = data.get('reagent_id')
    if not site_id or not reagent_id:
        return jsonify({'error': '缺少 site_id 或 reagent_id'}), 400
    passed = 1 if data.get('passed') in (1, True, '1', 'true') else 0
    standard_value = data.get('standard_value')
    measured_value = data.get('measured_value')
    deviation = None
    if standard_value not in (None, '') and measured_value not in (None, ''):
        try:
            deviation = round(float(measured_value) - float(standard_value), 4)
        except (TypeError, ValueError):
            deviation = None
    fail_action = (data.get('fail_action') or '').strip()  # calibrate=校准 / repair=报修
    if not passed and fail_action not in ('calibrate', 'repair'):
        return jsonify({'error': '质控不通过时必须选择处置动作（calibrate 校准 / repair 报修）'}), 400
    operator = data.get('operator') or u.get('real_name') or u.get('username', '')
    qc_time = data.get('qc_time') or datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    remark = data.get('remark', '')
    with get_db() as db:
        inv = db.execute('SELECT * FROM reagent_inventory WHERE site_id=? AND reagent_id=?',
                         (site_id, reagent_id)).fetchone()
        if not inv:
            return jsonify({'error': '该站点无此试剂库存记录'}), 404
        db.execute("""
            INSERT INTO reagent_qc_records
                (site_id, reagent_id, standard_value, measured_value, deviation, passed,
                 fail_action, operator, operator_id, qc_time, remark)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (site_id, reagent_id, standard_value, measured_value, deviation, passed,
              fail_action if not passed else '', operator, u.get('id'), qc_time, remark))
        new_status = 'passed' if passed else 'failed'
        db.execute("UPDATE reagent_inventory SET qc_status=?, updated_at=? WHERE site_id=? AND reagent_id=?",
                   (new_status, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), site_id, reagent_id))
        if not passed:
            # 质控不通过 → 通知管理者跟进（校准或报修）
            reagent_name = db.execute('SELECT name FROM reagents WHERE id=?', (reagent_id,)).fetchone()
            site_name = db.execute('SELECT name FROM sites WHERE id=?', (site_id,)).fetchone()
            action_cn = '校准' if fail_action == 'calibrate' else '报修'
            approvers = db.execute("SELECT id FROM users WHERE role IN ('admin','manager')").fetchall()
            for ap in approvers:
                _create_notification(ap['id'], 'reagent_qc', site_id, '试剂质控不通过',
                                     f'{site_name["name"] if site_name else site_id} 的 '
                                     f'{reagent_name["name"] if reagent_name else reagent_id} '
                                     f'标样质控不通过（偏差 {deviation}），需{action_cn}。', db=db)
        db.commit()
        return jsonify({'ok': True, 'qc_status': new_status, 'deviation': deviation})


@app.route('/api/reagent-qc/pending', methods=['GET'])
def api_reagent_qc_pending():
    """待质控清单：更换后尚未跑标样验证的试剂（qc_status='pending'），供移动端/PC 工作台提示。"""
    with get_db() as db:
        rows = db.execute("""
            SELECT ri.site_id, ri.reagent_id, ri.current_qty, ri.last_replaced_at, ri.qc_status,
                   s.name AS site_name, r.name AS reagent_name, r.spec, r.unit
            FROM reagent_inventory ri
            LEFT JOIN sites s ON ri.site_id = s.id
            LEFT JOIN reagents r ON ri.reagent_id = r.id
            WHERE ri.qc_status = 'pending'
            ORDER BY ri.last_replaced_at DESC
        """).fetchall()
        return jsonify([dict(x) for x in rows])


# ---------- 3.5 试剂告警 ----------
@app.route('/api/reagent-alerts', methods=['GET'])
def api_reagent_alerts():
    """获取未处理的试剂告警（按站点范围隔离）"""
    allowed = _filter_site_ids()
    site_id = request.args.get('site_id')
    site_id_int = int(site_id) if site_id and site_id.isdigit() else None
    with get_db() as db:
        if site_id_int is not None:
            if allowed is not None and site_id_int not in allowed:
                return jsonify([])
            rows = db.execute(
                '''SELECT ra.*, r.name as reagent_name
                   FROM reagent_alerts ra
                   JOIN reagents r ON ra.reagent_id = r.id
                   WHERE ra.handled=0 AND ra.site_id=? ORDER BY ra.alert_at DESC''',
                (site_id_int,)).fetchall()
        elif allowed is not None:
            ph = ','.join('?' * len(allowed))
            rows = db.execute(
                f'''SELECT ra.*, r.name as reagent_name
                   FROM reagent_alerts ra
                   JOIN reagents r ON ra.reagent_id = r.id
                   WHERE ra.handled=0 AND ra.site_id IN ({ph}) ORDER BY ra.alert_at DESC''',
                list(allowed)).fetchall()
        else:
            rows = db.execute(
                '''SELECT ra.*, r.name as reagent_name
                   FROM reagent_alerts ra
                   JOIN reagents r ON ra.reagent_id = r.id
                   WHERE ra.handled=0 ORDER BY ra.alert_at DESC''').fetchall()
        return jsonify([dict(r) for r in rows])


# ---------- 1.1 车辆台账 ----------
@app.route('/api/vehicles', methods=['GET'])
def api_vehicles():
    """车辆列表，可按状态过滤"""
    status = request.args.get('status')
    with get_db() as db:
        if status:
            rows = db.execute('SELECT * FROM vehicles WHERE status=? ORDER BY plate_no', (status,)).fetchall()
        else:
            rows = db.execute('SELECT * FROM vehicles ORDER BY plate_no').fetchall()
        return jsonify([dict(r) for r in rows])


@app.route('/api/vehicles', methods=['POST'])
@login_required
def api_vehicles_create():
    """新增车辆"""
    g_ = require_admin()
    if g_:
        return g_
    data = request.get_json() or {}
    plate_no = data['plate_no']
    model = data.get('model', '')
    seats = data.get('seats', 5)
    with get_db() as db:
        cur = db.execute(
            'INSERT INTO vehicles (plate_no, model, seats) VALUES (?,?,?)',
            (plate_no, model, seats))
        db.commit()
        row = db.execute('SELECT * FROM vehicles WHERE id=?', (cur.lastrowid,)).fetchone()
        return jsonify(dict(row)), 201

@app.route('/api/vehicles/<int:vid>', methods=['PUT'])
@login_required
def api_vehicles_update(vid):
    """更新车辆信息（管理员限定）"""
    if g.current_user.get('role') != 'admin':
        return jsonify({'error': '仅管理员可编辑车辆信息'}), 403
    data = request.get_json(silent=True) or {}
    with get_db() as db:
        existing = db.execute('SELECT * FROM vehicles WHERE id=?', (vid,)).fetchone()
        if not existing:
            return jsonify({'error': '车辆不存在'}), 404
        if 'next_maintenance_mileage' in data:
            db.execute('UPDATE vehicles SET next_maintenance_mileage=? WHERE id=?',
                       (data['next_maintenance_mileage'], vid))
        # 通用字段更新
        for k in ['model','seats','plate_no','status','current_mileage']:
            if k in data:
                db.execute(f'UPDATE vehicles SET {k}=? WHERE id=?', (data[k], vid))
        db.commit()
        row = db.execute('SELECT * FROM vehicles WHERE id=?', (vid,)).fetchone()
        return jsonify(dict(row))


# ---------- 1.2 用车申请 ----------
@app.route('/api/vehicle/applications', methods=['GET'])
def api_vehicle_applications():
    """用车申请列表"""
    status = request.args.get('status')
    applicant_id = request.args.get('applicant_id')
    with get_db() as db:
        q = '''SELECT va.*, v.plate_no, v.model, u.real_name as applicant_name
               FROM vehicle_applications va
               LEFT JOIN vehicles v ON va.vehicle_id = v.id
               LEFT JOIN users u ON va.applicant_id = u.id WHERE 1=1'''
        params = []
        if status:
            q += ' AND va.status=?'; params.append(status)
        if applicant_id:
            q += ' AND va.applicant_id=?'; params.append(applicant_id)
        q += ' ORDER BY va.created_at DESC'
        rows = db.execute(q, params).fetchall()
        return jsonify([dict(r) for r in rows])


@app.route('/api/vehicle/applications', methods=['POST'])
@login_required
def api_vehicle_applications_create():
    """新建用车申请（支持完整/极简两种形态：移动端工单处置仅需填事由）"""
    data = request.get_json() or {}
    vehicle_id = data.get('vehicle_id')
    applicant_id = data.get('applicant_id') or (g.current_user['id'] if g.current_user else None)
    start_at = data.get('start_at')
    end_at = data.get('end_at')
    destination = data.get('destination', '')
    reason = data.get('reason', '')
    site_id = data.get('site_id')
    work_order_no = data.get('work_order_no', '')

    if not reason:
        return jsonify({'error': '用车事由不能为空'}), 400

    with get_db() as db:
        # 完整形态（含车辆/时间）才校验时间冲突；极简形态（仅事由）跳过
        if vehicle_id and applicant_id and start_at and end_at:
            conflict = db.execute(
                '''SELECT id FROM vehicle_applications
                   WHERE vehicle_id=? AND status!='cancelled'
                   AND NOT (end_at <= ? OR start_at >= ?)''',
                (vehicle_id, start_at, end_at)).fetchone()
            if conflict:
                return jsonify({'error': '车辆时间冲突', 'conflict_id': conflict['id']}), 409
            applicant_conflict = db.execute(
                '''SELECT id FROM vehicle_applications
                   WHERE applicant_id=? AND status!='cancelled'
                   AND NOT (end_at <= ? OR start_at >= ?)''',
                (applicant_id, start_at, end_at)).fetchone()
            if applicant_conflict:
                return jsonify({'error': '申请人时间冲突', 'conflict_id': applicant_conflict['id']}), 409

        cur = db.execute(
            'INSERT INTO vehicle_applications (vehicle_id, applicant_id, start_at, end_at, destination, reason, site_id, work_order_no) VALUES (?,?,?,?,?,?,?,?)',
            (vehicle_id, applicant_id, start_at, end_at, destination, reason, site_id, work_order_no))
        db.commit()
        row = db.execute('SELECT * FROM vehicle_applications WHERE id=?', (cur.lastrowid,)).fetchone()
        return jsonify(dict(row)), 201


@app.route('/api/vehicle/applications/<int:app_id>/approve', methods=['POST'])
def api_vehicle_application_approve(app_id):
    """主管审批用车申请"""
    data = request.get_json() or {}
    action = data.get('action', 'approve')
    approver_id = data.get('approver_id')
    reject_reason = data.get('reject_reason', '')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    status = 'approved' if action == 'approve' else 'rejected'
    with get_db() as db:
        db.execute(
            'UPDATE vehicle_applications SET status=?, approver_id=?, approved_at=?, reject_reason=? WHERE id=?',
            (status, approver_id, now, reject_reason, app_id))
        db.commit()
        row = db.execute('SELECT * FROM vehicle_applications WHERE id=?', (app_id,)).fetchone()
        if not row:
            return jsonify({'error': '申请不存在'}), 404
        return jsonify(dict(row))


# ---------- 1.3 出车/还车 ----------
@app.route('/api/vehicle/use-records', methods=['GET', 'POST'])
def api_vehicle_use_records():
    """出车/还车记录列表（GET）或出车登记（POST）"""
    if request.method == 'GET':
        vehicle_id = request.args.get('vehicle_id', type=int)
        with get_db() as db:
            q = '''SELECT r.*, v.plate_no, v.model, va.start_at, va.end_at, va.destination, va.reason,
                          u.real_name as applicant_name
                   FROM vehicle_use_records r
                   JOIN vehicle_applications va ON r.application_id = va.id
                   JOIN vehicles v ON va.vehicle_id = v.id
                   LEFT JOIN users u ON va.applicant_id = u.id
                   WHERE 1=1'''
            params = []
            if vehicle_id:
                q += ' AND v.id=?'; params.append(vehicle_id)
            q += ' ORDER BY r.id DESC'
            rows = db.execute(q, params).fetchall()
            return jsonify([dict(r) for r in rows])

    data = request.get_json() or {}
    application_id = data['application_id']
    start_mileage = data.get('start_mileage')
    with get_db() as db:
        cur = db.execute(
            'INSERT INTO vehicle_use_records (application_id, start_mileage) VALUES (?,?)',
            (application_id, start_mileage))
        db.execute('UPDATE vehicles SET status="in_use" WHERE id=(SELECT vehicle_id FROM vehicle_applications WHERE id=?)', (application_id,))
        db.commit()
        return jsonify({'id': cur.lastrowid}), 201


@app.route('/api/vehicle/use-records/<int:rec_id>/return', methods=['POST'])
def api_vehicle_use_record_return(rec_id):
    """还车登记"""
    data = request.get_json() or {}
    end_mileage = data.get('end_mileage')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as db:
        db.execute(
            'UPDATE vehicle_use_records SET end_mileage=?, returned_at=? WHERE id=?',
            (end_mileage, now, rec_id))
        rec = db.execute('SELECT application_id FROM vehicle_use_records WHERE id=?', (rec_id,)).fetchone()
        if rec:
            db.execute('UPDATE vehicles SET status="idle", current_mileage=? WHERE id=(SELECT vehicle_id FROM vehicle_applications WHERE id=?)', (end_mileage, rec['application_id']))
        db.commit()
        return jsonify({'ok': True})


# ---------- 1.4 加油 ----------
@app.route('/api/vehicle/refueling', methods=['GET', 'POST'])
def api_vehicle_refueling():
    """加油记录列表（GET）或登记加油（POST）"""
    if request.method == 'GET':
        vehicle_id = request.args.get('vehicle_id', type=int)
        with get_db() as db:
            q = '''SELECT r.*, v.plate_no, v.model
                     FROM vehicle_refueling_records r
                     JOIN vehicles v ON r.vehicle_id = v.id
                     WHERE 1=1'''
            params = []
            if vehicle_id:
                q += ' AND r.vehicle_id=?'; params.append(vehicle_id)
            q += ' ORDER BY r.refuel_at DESC'
            rows = db.execute(q, params).fetchall()
            return jsonify([dict(r) for r in rows])

    data = request.get_json() or {}
    vehicle_id = data['vehicle_id']
    liters = data['liters']
    amount = data.get('amount')
    mileage_at = data.get('mileage_at')
    remark = data.get('remark', '')
    with get_db() as db:
        cur = db.execute(
            'INSERT INTO vehicle_refueling_records (vehicle_id, liters, amount, mileage_at, remark) VALUES (?,?,?,?,?)',
            (vehicle_id, liters, amount, mileage_at, remark))
        db.commit()
        rid = cur.lastrowid
        return jsonify({'ok': True, 'id': rid}), 201


# ---------- 1.5 保养 ----------
@app.route('/api/vehicle/maintenance', methods=['GET', 'POST'])
def api_vehicle_maintenance():
    """保养记录列表（GET）或登记保养（POST）"""
    if request.method == 'GET':
        vehicle_id = request.args.get('vehicle_id', type=int)
        with get_db() as db:
            q = '''SELECT r.*, v.plate_no, v.model
                     FROM vehicle_maintenance_records r
                     JOIN vehicles v ON r.vehicle_id = v.id
                     WHERE 1=1'''
            params = []
            if vehicle_id:
                q += ' AND r.vehicle_id=?'; params.append(vehicle_id)
            q += ' ORDER BY r.maint_at DESC'
            rows = db.execute(q, params).fetchall()
            return jsonify([dict(r) for r in rows])

    data = request.get_json() or {}
    vehicle_id = data['vehicle_id']
    maint_type = data['maint_type']
    mileage_at = data.get('mileage_at')
    items = data.get('items', '')
    cost = data.get('cost')
    next_maint_mileage = data.get('next_maint_mileage')
    with get_db() as db:
        cur = db.execute(
            'INSERT INTO vehicle_maintenance_records (vehicle_id, maint_type, mileage_at, items, cost, next_maint_mileage) VALUES (?,?,?,?,?,?)',
            (vehicle_id, maint_type, mileage_at, items, cost, next_maint_mileage))
        db.execute(
            'UPDATE vehicles SET last_maintenance_at=datetime("now","localtime"), next_maintenance_mileage=? WHERE id=?',
            (next_maint_mileage or 5000, vehicle_id))
        db.commit()
        mid = cur.lastrowid
        return jsonify({'ok': True, 'id': mid}), 201


# ---------- 4.1 周巡检计划 ----------
@app.route('/api/weekly-plans', methods=['GET'])
def api_weekly_plans():
    """周计划列表，可按用户/周筛选"""
    user_id = request.args.get('user_id')
    week_start = request.args.get('week_start')
    status = request.args.get('status')
    with get_db() as db:
        q = '''SELECT wp.*, u.real_name as user_name
               FROM weekly_inspection_plans wp
               LEFT JOIN users u ON wp.user_id = u.id WHERE 1=1'''
        params = []
        if user_id:
            q += ' AND wp.user_id=?'; params.append(user_id)
        if week_start:
            q += ' AND wp.week_start=?'; params.append(week_start)
        if status:
            q += ' AND wp.status=?'; params.append(status)
        q += ' ORDER BY wp.week_start DESC'
        rows = db.execute(q, params).fetchall()
        for r in rows:
            r = dict(r)
            try:
                import json as _json
                r['plan_data'] = _json.loads(r['plan_data']) if r.get('plan_data') else {}
            except:
                pass
        return jsonify([dict(r) for r in rows])


@app.route('/api/weekly-plans', methods=['POST'])
def api_weekly_plans_create():
    """新建/提交周计划"""
    data = request.get_json() or {}
    user_id = data['user_id']
    week_start = data['week_start']
    plan_data = data.get('plan_data', {})
    vehicle_id = data.get('vehicle_id')
    submit = data.get('submit', False)
    remarks = data.get('remarks', '')

    import json as _json
    plan_json = _json.dumps(plan_data, ensure_ascii=False) if isinstance(plan_data, dict) else plan_data

    status = 'submitted' if submit else 'draft'
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    with get_db() as db:
        cur = db.execute(
            '''INSERT INTO weekly_inspection_plans (user_id, week_start, plan_data, vehicle_id, status, remarks, submitted_at)
               VALUES (?,?,?,?,?,?,?)''',
            (user_id, week_start, plan_json, vehicle_id, status, remarks, now if submit else None))
        db.commit()
        # 如果勾选了车辆且提交，自动创建用车申请
        if submit and vehicle_id:
            for day, site_ids in plan_data.items():
                if site_ids and isinstance(site_ids, list) and len(site_ids) > 0:
                    # 简单处理：为整周创建一条用车申请
                    db.execute(
                        '''INSERT INTO vehicle_applications (vehicle_id, applicant_id, start_at, end_at, destination, reason, status)
                           VALUES (?,?,?,?,?,"周巡检用车","approved")''',
                        (vehicle_id, user_id, week_start + ' 08:00:00', week_start + ' 18:00:00', '巡检'))
                    break
        db.commit()
        row = db.execute('SELECT * FROM weekly_inspection_plans WHERE id=?', (cur.lastrowid,)).fetchone()
        r = dict(row)
        try: r['plan_data'] = _json.loads(r['plan_data']) if r.get('plan_data') else {}
        except: pass
        return jsonify(r), 201


@app.route('/api/weekly-plans/<int:plan_id>/approve', methods=['POST'])
def api_weekly_plan_approve(plan_id):
    """主管审批周计划"""
    data = request.get_json() or {}
    approver_id = data.get('approver_id')
    action = data.get('action', 'approve')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if action == 'submit':
        status = 'submitted'
        with get_db() as db:
            db.execute(
                'UPDATE weekly_inspection_plans SET status=?, submitted_at=? WHERE id=?',
                (status, now, plan_id))
            db.commit()
        return jsonify({'ok': True, 'status': status})
    status = 'approved' if action == 'approve' else ('rejected' if action == 'reject' else 'draft')
    with get_db() as db:
        db.execute(
            'UPDATE weekly_inspection_plans SET status=?, approver_id=?, approved_at=? WHERE id=?',
            (status, approver_id, now if action == 'approve' else None, plan_id))
        db.commit()
        return jsonify({'ok': True, 'status': status})


# ===================== 巡检计划调度层（周/月/季/年统一） =====================
# 设计依据：《巡检计划体系设计》——调度层（排程+资源协调+审核）
# plan_schedules 为调度层计划；审批通过后生成 insp_plans/insp_plan_items（执行层）。
# 注：表名不叫 inspection_schedules（该名已被"检查项到期排程"表占用）。

_PS_FREQ_CN = {'weekly': '周检', 'monthly': '月检', 'quarterly': '季检', 'yearly': '年检'}
_PS_EDITABLE = ('draft', 'rejected')  # 仅这两个状态可编辑/重新提交


def _ps_parse_row(row):
    """plan_schedules 行 → dict，解析 JSON 字段"""
    r = dict(row)
    for f, empty in (('plan_data', {}), ('vehicle_days', {}), ('spare_parts', []),
                     ('work_order_ids', []), ('previous_plan_data', None)):
        v = r.get(f)
        if v:
            try:
                r[f] = json.loads(v)
            except Exception:
                r[f] = empty if empty is not None else v
        else:
            r[f] = empty if empty is not None else None
    return r


def _ps_check_vehicle_conflicts(db, user_id, vehicle_days, exclude_schedule_id=None):
    """用车冲突检测（三层交叉）：
    1. vehicle_applications 中他人已占用的时间段
    2. 他人计划（submitted/approved）的 vehicle_days 同日同车
    3. 车辆自身处于维保状态
    返回冲突列表 [{date, vehicle_id, plate_no, reason}]"""
    conflicts = []
    if not vehicle_days:
        return conflicts
    for date_str, vid in (vehicle_days or {}).items():
        if not vid:
            continue
        veh = db.execute("SELECT plate_no, status FROM vehicles WHERE id=?", (vid,)).fetchone()
        plate = veh['plate_no'] if veh else f'车辆{vid}'
        if veh and veh['status'] == 'maintenance':
            conflicts.append({'date': date_str, 'vehicle_id': vid, 'plate_no': plate,
                              'reason': f'{plate} 处于维保中，不可安排'})
            continue
        # 1. 他人用车申请占用当天
        occ = db.execute("""
            SELECT va.id, u.real_name FROM vehicle_applications va
            LEFT JOIN users u ON va.applicant_id = u.id
            WHERE va.vehicle_id=? AND va.status IN ('pending','approved')
              AND va.applicant_id IS NOT NULL AND va.applicant_id != ?
              AND date(va.start_at) <= ? AND date(va.end_at) >= ?
            LIMIT 1
        """, (vid, user_id, date_str, date_str)).fetchone()
        if occ:
            conflicts.append({'date': date_str, 'vehicle_id': vid, 'plate_no': plate,
                              'reason': f'{plate} 当天已被{occ["real_name"] or "他人"}预约'})
            continue
        # 2. 他人计划的 vehicle_days 占用
        others = db.execute("""
            SELECT id, user_id, vehicle_days, period_start, period_end FROM plan_schedules
            WHERE status IN ('submitted','approved') AND user_id != ?
              AND period_start <= ? AND period_end >= ?
        """, (user_id, date_str, date_str)).fetchall()
        hit = None
        for o in others:
            if exclude_schedule_id and o['id'] == exclude_schedule_id:
                continue
            try:
                ovd = json.loads(o['vehicle_days'] or '{}')
            except Exception:
                ovd = {}
            if str(ovd.get(date_str)) == str(vid):
                hit = o
                break
        if hit:
            ou = db.execute("SELECT real_name FROM users WHERE id=?", (hit['user_id'],)).fetchone()
            conflicts.append({'date': date_str, 'vehicle_id': vid, 'plate_no': plate,
                              'reason': f'{plate} 当天已被{ou["real_name"] if ou else "他人"}的计划安排'})
    return conflicts


def _ps_validate(db, user_id, schedule_type, period_start, period_end,
                 plan_data, vehicle_days, exclude_schedule_id=None):
    """计划校验：日期范围 + 周巡检全覆盖 + 用车冲突。
    errors 阻断提交；warnings 仅提示（审批端同样可见，由人拍板）。"""
    errors, warnings = [], []
    pd = plan_data or {}
    # 日期范围校验
    for d in pd.keys():
        if d < str(period_start) or d > str(period_end):
            errors.append(f'日期 {d} 超出周期范围（{period_start} ~ {period_end}）')
    # 周巡检全覆盖校验
    if schedule_type == 'weekly':
        user_sites = [r['site_id'] for r in
                      db.execute("SELECT site_id FROM user_sites WHERE user_id=?", (user_id,)).fetchall()]
        planned = set()
        for day_data in pd.values():
            if isinstance(day_data, dict):
                planned.update(day_data.get('sites') or [])
        missing = [sid for sid in user_sites if sid not in planned]
        if missing:
            names = [r['name'] for r in db.execute(
                f"SELECT name FROM sites WHERE id IN ({','.join('?'*len(missing))})", missing).fetchall()]
            warnings.append(f'本周未覆盖：{"、".join(names)}（周巡检要求全覆盖，请确认是否遗漏）')
    # 用车冲突
    for c in _ps_check_vehicle_conflicts(db, user_id, vehicle_days, exclude_schedule_id):
        errors.append(f'{c["date"]} {c["reason"]}')
    # 路线折返检测（仅警告，不阻断）
    for d in sorted(pd.keys()):
        day_data = pd[d]
        if not isinstance(day_data, dict):
            continue
        site_ids = day_data.get('sites') or []
        if len(site_ids) < 3:
            continue
        ph = ','.join('?' * len(site_ids))
        rows = db.execute(f"SELECT id, name, gps_lat, gps_lng FROM sites WHERE id IN ({ph})", site_ids).fetchall()
        coord_map = {r['id']: (r['gps_lat'], r['gps_lng'], r['name']) for r in rows}
        coords = [(sid, coord_map.get(sid)) for sid in site_ids if coord_map.get(sid) and coord_map[sid][0] and coord_map[sid][1]]
        for i in range(len(coords) - 2):
            _, a = coords[i]
            _, b = coords[i + 1]
            _, c = coords[i + 2]
            dist_ab = ((a[0]-b[0])**2 + (a[1]-b[1])**2) ** 0.5
            dist_ac = ((a[0]-c[0])**2 + (a[1]-c[1])**2) ** 0.5
            if dist_ac < dist_ab * 0.85:  # C 比 B 更靠近 A（留 15% 容差避免误报）
                saved_deg = dist_ab - dist_ac
                saved_km = saved_deg * 111  # 粗略：1度≈111km
                warnings.append(
                    f'{d} 路线折返：{a[2]}→{b[2]}→{c[2]}，'
                    f'"{c[2]}"比"{b[2]}"更靠近"{a[2]}"，调整顺序约省 {saved_km:.1f}km')
        # 可执行性底线：按每站90分钟、站间45分钟粗估一个8小时工作日。
        # 这是审批提示而非替代现场判断，避免看似顺路但实际无法完成的排程。
        count = len(site_ids)
        estimated_minutes = count * 90 + max(0, count - 1) * 45
        if estimated_minutes > 480:
            warnings.append(f'{d} 预计作业{estimated_minutes // 60}小时{estimated_minutes % 60}分，超过8小时作业窗口，建议拆分或说明例外')
    return {'ok': not errors, 'errors': errors, 'warnings': warnings}


def _ps_coverage_exception_required(validation, schedule_type, reason):
    """周检漏站不是普通提醒：必须由排程人说明例外原因，再交由审批人显式放行。"""
    missing = schedule_type == 'weekly' and any('本周未覆盖' in w for w in validation['warnings'])
    if missing and not (reason or '').strip():
        return '周巡检存在未覆盖站点，请填写例外原因后再提交'
    return None


def require_reviewer():
    """材料/影像/结果审核门禁：审核员可审核，不具备排程与资源审批权限。"""
    u = g.get('current_user')
    # 兼容已部署库中的 inspector 岗位。它在业务上承担审核员职责，权限边界不变。
    if not u or u.get('role') not in ('admin', 'manager', 'reviewer', 'inspector'):
        return jsonify({'success': False, 'error': '需要审核员、主管或管理员权限'}), 403
    return None


def _ps_record_event(db, schedule_id, version, event_type, operator_id, payload):
    db.execute("""INSERT INTO plan_schedule_events
        (schedule_id, version, event_type, operator_id, payload) VALUES (?,?,?,?,?)""",
        (schedule_id, version, event_type, operator_id,
         json.dumps(payload or {}, ensure_ascii=False)))


def _ps_reserve_parts(db, schedule):
    """审批时预留库存，不直接扣减。库存扣减必须等现场领用/消耗确认。"""
    try:
        parts = json.loads(schedule['spare_parts'] or '[]')
    except Exception:
        parts = []
    reserved = 0
    for part in parts:
        if not isinstance(part, dict) or not part.get('part_id'):
            continue
        part_id, qty = int(part['part_id']), max(1, int(part.get('quantity') or 1))
        inv = db.execute("SELECT quantity FROM spare_parts_inventory WHERE id=?", (part_id,)).fetchone()
        already = db.execute("""SELECT COALESCE(SUM(reserved_quantity-issued_quantity),0) AS n
            FROM plan_resource_reservations WHERE part_id=? AND status IN ('reserved','issued')""", (part_id,)).fetchone()['n']
        if not inv or int(inv['quantity'] or 0) - int(already or 0) < qty:
            raise ValueError(f'备件#{part_id}可预留库存不足')
        db.execute("""INSERT INTO plan_resource_reservations
            (schedule_id, part_id, planned_quantity, reserved_quantity, status)
            VALUES (?,?,?,?, 'reserved')""", (schedule['id'], part_id, qty, qty))
        reserved += 1
    return reserved


def _ps_site_scores(db, site_ids):
    """站点优先级评分：未关工单（等级+滞留天数）+ 待处理告警 + 试剂临期。
    评分逻辑透明可解释，供排程建议与审批风险预警共用。"""
    scores = {}
    details = {}
    today = datetime.now()
    for sid in site_ids:
        score, reasons = 0, []
        # 未关工单
        orders = db.execute("""
            SELECT id, title, level, created_at FROM work_orders
            WHERE site_id=? AND status NOT IN ('closed')
        """, (sid,)).fetchall()
        for o in orders:
            base = {'urgent': 30, 'normal': 10}.get(o['level'], 10)
            try:
                age = (today - datetime.strptime(str(o['created_at'])[:10], '%Y-%m-%d')).days
            except Exception:
                age = 0
            overdue = 10 if age >= 7 else 0
            score += base + overdue
            reasons.append(f'未关工单「{o["title"]}」（{"紧急" if o["level"] == "urgent" else "普通"}，已{age}天）')
        # 待处理告警
        alerts = db.execute("SELECT level, metric FROM alerts WHERE site_id=? AND status='pending'", (sid,)).fetchall()
        for a in alerts:
            score += {'red': 20, 'orange': 12, 'yellow': 6}.get(a['level'], 5)
            reasons.append(f'{ALERT_LEVEL_LABEL.get(a["level"], a["level"])}告警（{a["metric"] or "指标异常"}）待处理')
        # 试剂临期
        try:
            invs = db.execute("""
                SELECT ri.*, r.name as reagent_name FROM reagent_inventory ri
                LEFT JOIN reagents r ON ri.reagent_id = r.id WHERE ri.site_id=?
            """, (sid,)).fetchall()
            for inv in invs:
                st = compute_reagent_status(dict(inv))
                if st.get('status') in ('临期', '已过期', '低余量'):
                    score += 8
                    reasons.append(f'试剂{inv["reagent_name"] or ""}{st["status"]}（约剩{st.get("remaining_days")}天）')
        except Exception:
            pass
        scores[sid] = score
        details[sid] = reasons
    return scores, details


def _ps_generate_tasks(db, schedule):
    """审批通过后生成执行任务：每个有安排的日期生成一个 insp_plan（直接 active，
    因为计划已审批，不再二次审批），按站点类型+频次匹配模板展开检查项。幂等。"""
    plan_data = json.loads(schedule['plan_data'] or '{}')
    user = db.execute("SELECT real_name FROM users WHERE id=?", (schedule['user_id'],)).fetchone()
    op_name = user['real_name'] if user else str(schedule['user_id'])
    freq_cn = _PS_FREQ_CN.get(schedule['schedule_type'], '巡检')
    created, total_items = 0, 0
    for date_str in sorted(plan_data.keys()):
        day_data = plan_data[date_str] or {}
        site_ids = day_data.get('sites') if isinstance(day_data, dict) else day_data
        if not site_ids:
            continue
        # 同来源同日期已有可执行包时不重复生成；被替代/取消的历史包仍可保留审计。
        dup = db.execute("SELECT id FROM insp_plans WHERE plan_schedule_id=? AND generate_date=? AND status != 'cancelled'",
                         (schedule['id'], date_str)).fetchone()
        if dup:
            continue
        plan_name = f"{op_name}·{freq_cn}-{date_str.replace('-', '')}"
        cur = db.execute("""
            INSERT INTO insp_plans (plan_name, assignee, assignee_id, period, generate_date, status,
                                    plan_schedule_id, schedule_version, plan_snapshot)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (plan_name, op_name, schedule['user_id'], schedule['schedule_type'],
              date_str, 'active', schedule['id'], schedule['version'],
              json.dumps({date_str: day_data}, ensure_ascii=False)))
        plan_id = cur.lastrowid
        for sid in site_ids:
            site = db.execute("SELECT type FROM sites WHERE id=?", (sid,)).fetchone()
            site_type = site['type'] if site else ''
            # 按站点类型 + 频次匹配模板检查项
            items = db.execute("""
                SELECT iti.item_name, iti.category, t.id as template_id
                FROM inspection_configs c
                JOIN inspection_templates t ON c.template_id = t.id
                JOIN inspection_template_items iti ON iti.template_id = t.id
                WHERE c.site_type=? AND c.is_active=1 AND t.status='active' AND t.frequency=?
                ORDER BY iti.sort_order, iti.id
            """, (site_type, schedule['schedule_type'])).fetchall()
            if items:
                for it in items:
                    db.execute("""
                        INSERT INTO insp_plan_items (plan_id, site_id, template_id, item_name, category, frequency)
                        VALUES (?,?,?,?,?,?)
                    """, (plan_id, sid, it['template_id'], it['item_name'],
                          it['category'], schedule['schedule_type']))
                    total_items += 1
            else:
                # 无匹配模板时放一个兜底项，避免空计划
                db.execute("""
                    INSERT INTO insp_plan_items (plan_id, site_id, item_name, category, frequency)
                    VALUES (?,?,?,?,?)
                """, (plan_id, sid, '常规检查', '常规', schedule['schedule_type']))
                total_items += 1
        created += 1
    return created, total_items


def _ps_auto_flow(db, schedule):
    """审批通过后的自动流转：生成任务 + 锁定车辆 + 备件预留 + 通知。
    在审批事务内调用（db 由外部传入并统一 commit）。"""
    result = {'plans_created': 0, 'items_created': 0, 'vehicle_locked': 0, 'parts_reserved': 0}
    # 1. 生成执行任务
    if not schedule['tasks_generated']:
        pc, ti = _ps_generate_tasks(db, schedule)
        result['plans_created'], result['items_created'] = pc, ti
        db.execute("UPDATE plan_schedules SET tasks_generated=1 WHERE id=?", (schedule['id'],))
    # 2. 锁定车辆（每天一条 approved 用车申请，幂等跳过已存在的）
    try:
        vehicle_days = json.loads(schedule['vehicle_days'] or '{}')
    except Exception:
        vehicle_days = {}
    for date_str, vid in vehicle_days.items():
        if not vid:
            continue
        exists = db.execute("""
            SELECT id FROM vehicle_applications
            WHERE vehicle_id=? AND applicant_id=? AND date(start_at)=? AND reason LIKE ?
            LIMIT 1
        """, (vid, schedule['user_id'], date_str, f'%计划#{schedule["id"]}%')).fetchone()
        if exists:
            continue
        db.execute("""
            INSERT INTO vehicle_applications (vehicle_id, applicant_id, start_at, end_at, destination, reason, status)
            VALUES (?,?,?,?,?,?,?)
        """, (vid, schedule['user_id'], f'{date_str} 08:00:00', f'{date_str} 18:00:00',
              '巡检', f'巡检计划#{schedule["id"]}用车', 'approved'))
        result['vehicle_locked'] += 1
    # 3. 备件只预留；现场确认领用时再扣库存，避免计划变更造成虚假出库。
    result['parts_reserved'] = _ps_reserve_parts(db, schedule)
    # 4. 通知排程人
    _create_notification(schedule['user_id'], 'plan_schedule', schedule['id'],
                         f'巡检计划已通过（{_PS_FREQ_CN.get(schedule["schedule_type"], "巡检")}）',
                         f'已生成{result["plans_created"]}个巡检任务，请按计划执行。', db=db)
    return result


def _ps_rebuild_tasks_on_change(db, schedule):
    """变更通过后的任务同步：保留历史，取消未执行项，新增日期生成新执行包。"""
    sid = schedule['id']
    kept, cancelled = 0, 0
    plan_data = json.loads(schedule['plan_data'] or '{}')
    existing = db.execute("SELECT id, generate_date FROM insp_plans WHERE plan_schedule_id=? AND status != 'cancelled'", (sid,)).fetchall()
    for ep in existing:
        day_data = plan_data.get(ep['generate_date']) or {}
        desired_sites = set(day_data.get('sites') or []) if isinstance(day_data, dict) else set(day_data or [])
        started = db.execute(
            "SELECT id FROM insp_plan_items WHERE plan_id=? AND result IS NOT NULL LIMIT 1",
            (ep['id'],)).fetchone()
        if started:
            # 已开始的日执行包继续保留；只取消被移出的、尚未执行的站点检查项。
            if desired_sites:
                placeholders = ','.join('?' * len(desired_sites))
                cur = db.execute(f"UPDATE insp_plan_items SET execution_status='cancelled' WHERE plan_id=? AND result IS NULL AND site_id NOT IN ({placeholders}) AND COALESCE(execution_status,'active')='active'", [ep['id'], *desired_sites])
            else:
                cur = db.execute("UPDATE insp_plan_items SET execution_status='cancelled' WHERE plan_id=? AND result IS NULL", (ep['id'],))
            cancelled += cur.rowcount
            db.execute("UPDATE insp_plans SET schedule_version=? WHERE id=?", (schedule['version'], ep['id']))
            kept += 1
            continue
        # 未开始日包仅作废，不物理删除，保留原审批与变更证据。
        db.execute("UPDATE insp_plan_items SET execution_status='cancelled' WHERE plan_id=?", (ep['id'],))
        db.execute("UPDATE insp_plans SET status='cancelled' WHERE id=?", (ep['id'],))
        cancelled += 1
    fresh = db.execute("SELECT * FROM plan_schedules WHERE id=?", (sid,)).fetchone()
    created, items = _ps_generate_tasks(db, fresh)
    db.execute("UPDATE plan_schedules SET tasks_generated=1 WHERE id=?", (sid,))
    return {'kept': kept, 'cancelled': cancelled, 'plans_created': created, 'items_created': items}


@app.route('/api/plan-schedules/overview')
def api_plan_schedules_overview():
    """管理者团队执行总览：计划安排、现场执行、异常与工单放在同一人员行中。
    非管理者只返回本人，避免为查看总览额外建设第二套看板。"""
    u = g.current_user
    today = datetime.now().strftime('%Y-%m-%d')
    with get_db() as db:
        where, params = "u.role IN ('operator','inspector')", []
        if u['role'] not in ('admin', 'manager', 'reviewer'):
            where += ' AND u.id=?'
            params.append(u['id'])
        people = db.execute(f"SELECT u.id, u.real_name, u.role FROM users u WHERE {where} ORDER BY u.real_name", params).fetchall()
        rows = []
        for person in people:
            uid = person['id']
            current = db.execute("""SELECT COUNT(*) AS n FROM plan_schedules
                WHERE user_id=? AND status='approved' AND period_start<=? AND period_end>=?""", (uid, today, today)).fetchone()['n']
            pending = db.execute("""SELECT COUNT(*) AS n FROM plan_schedules
                WHERE user_id=? AND status IN ('submitted','change_submitted')""", (uid,)).fetchone()['n']
            execution = db.execute("""SELECT COUNT(*) AS total,
                    COALESCE(SUM(CASE WHEN i.result IS NOT NULL THEN 1 ELSE 0 END),0) AS completed,
                    COALESCE(SUM(CASE WHEN i.result='abnormal' THEN 1 ELSE 0 END),0) AS abnormal
                FROM insp_plans p LEFT JOIN insp_plan_items i ON i.plan_id=p.id
                WHERE p.assignee_id=? AND p.generate_date=? AND p.status IN ('active','completed')
                  AND COALESCE(i.execution_status,'active')='active'""", (uid, today)).fetchone()
            open_orders = db.execute("""SELECT COUNT(*) AS n FROM work_orders
                WHERE assignee=? AND status NOT IN ('closed','resolved')""", (person['real_name'],)).fetchone()['n']
            overdue = db.execute("""SELECT COUNT(*) AS n FROM insp_plans p
                WHERE p.assignee_id=? AND p.generate_date<? AND p.status='active'
                  AND EXISTS (SELECT 1 FROM insp_plan_items i WHERE i.plan_id=p.id
                              AND i.result IS NULL AND COALESCE(i.execution_status, 'active')='active')""",
                (uid, today)).fetchone()['n']
            coverage_exceptions = db.execute("""SELECT COUNT(*) AS n FROM plan_schedules
                WHERE user_id=? AND status IN ('submitted','change_submitted')
                  AND COALESCE(coverage_exception_reason, '') != ''""", (uid,)).fetchone()['n']
            snapshots = db.execute("""SELECT validation_snapshot FROM plan_schedules
                WHERE user_id=? AND status IN ('submitted','change_submitted')
                  AND validation_snapshot IS NOT NULL""", (uid,)).fetchall()
            resource_blocks = 0
            for snapshot in snapshots:
                try:
                    errors = json.loads(snapshot['validation_snapshot'] or '{}').get('errors') or []
                    resource_blocks += sum(1 for error in errors if '车辆' in error or '库存' in error)
                except (TypeError, ValueError):
                    continue
            total, completed = int(execution['total'] or 0), int(execution['completed'] or 0)
            rows.append({
                'user_id': uid, 'real_name': person['real_name'], 'role': person['role'],
                'approved_schedules': current, 'pending_schedules': pending,
                'today_items': total, 'completed_items': completed,
                'completion_rate': round(completed / total * 100, 1) if total else 0,
                'abnormal_items': int(execution['abnormal'] or 0), 'open_workorders': int(open_orders or 0),
                'overdue_executions': int(overdue or 0),
                'coverage_exceptions': int(coverage_exceptions or 0),
                'resource_blocks': resource_blocks,
            })
        summary = {
            'people': len(rows), 'on_schedule': sum(1 for r in rows if r['approved_schedules']),
            'pending_approval': sum(r['pending_schedules'] for r in rows),
            'today_items': sum(r['today_items'] for r in rows),
            'completed_items': sum(r['completed_items'] for r in rows),
            'abnormal_items': sum(r['abnormal_items'] for r in rows),
            'open_workorders': sum(r['open_workorders'] for r in rows),
            'overdue_executions': sum(r['overdue_executions'] for r in rows),
            'coverage_exceptions': sum(r['coverage_exceptions'] for r in rows),
            'resource_blocks': sum(r['resource_blocks'] for r in rows),
        }
        summary['completion_rate'] = round(summary['completed_items'] / summary['today_items'] * 100, 1) if summary['today_items'] else 0
        return jsonify({'date': today, 'summary': summary, 'people': rows})


@app.route('/api/plan-schedules', methods=['GET'])
def api_plan_schedules_list():
    """巡检计划列表（调度层）。管理者看全部，一线只看自己。
    筛选：user_id / schedule_type / status / period_start（周期重叠过滤）"""
    u = g.current_user
    user_id = request.args.get('user_id')
    schedule_type = request.args.get('schedule_type')
    status = request.args.get('status')
    period_start = request.args.get('period_start')
    attention = request.args.get('attention', '')
    if attention not in ('', 'overdue', 'coverage', 'resource'):
        return jsonify({'error': '无效的关注项筛选'}), 400
    q = '''SELECT ps.*, u.real_name as user_name, u2.real_name as approver_name
           FROM plan_schedules ps
           LEFT JOIN users u ON ps.user_id = u.id
           LEFT JOIN users u2 ON ps.approver_id = u2.id WHERE 1=1'''
    params = []
    if u['role'] not in ('admin', 'manager', 'reviewer'):
        q += ' AND ps.user_id=?'
        params.append(u['id'])
    if user_id:
        q += ' AND ps.user_id=?'; params.append(user_id)
    if schedule_type:
        q += ' AND ps.schedule_type=?'; params.append(schedule_type)
    if status:
        q += ' AND ps.status=?'; params.append(status)
    if attention == 'overdue':
        # 计划本身已审批不代表按时执行。仅展示仍有有效未完成项的逾期执行包。
        q += ''' AND EXISTS (
            SELECT 1 FROM insp_plans ip WHERE ip.plan_schedule_id=ps.id
              AND ip.generate_date < date('now','localtime') AND ip.status='active'
              AND EXISTS (SELECT 1 FROM insp_plan_items pi WHERE pi.plan_id=ip.id
                          AND pi.result IS NULL AND COALESCE(pi.execution_status, 'active')='active')
        )'''
    elif attention == 'coverage':
        q += " AND ps.status IN ('submitted','change_submitted') AND COALESCE(ps.coverage_exception_reason, '') != ''"
    if period_start:
        # 周期重叠：计划的周期与查询周期相交
        q += ' AND ps.period_start <= date(?, "+6 days") AND ps.period_end >= ?'
        params.append(period_start, period_start)
    q += ' ORDER BY ps.period_start DESC, ps.created_at DESC'
    with get_db() as db:
        rows = [_ps_parse_row(r) for r in db.execute(q, params).fetchall()]
    if attention == 'resource':
        # 资源冲突以提交时保存的校验快照为准，避免仅凭当前库存造成误判。
        filtered = []
        for row in rows:
            try:
                errors = (row.get('validation_snapshot') or {}).get('errors') or []
            except AttributeError:
                errors = []
            if any('车辆' in error or '库存' in error for error in errors):
                filtered.append(row)
        rows = filtered
    for r in rows:
        # 附加展示用汇总：天数/站点数
        days = [d for d, v in (r.get('plan_data') or {}).items()
                if isinstance(v, dict) and (v.get('sites'))]
        site_set = set()
        for d, v in (r.get('plan_data') or {}).items():
            if isinstance(v, dict):
                site_set.update(v.get('sites') or [])
        r['day_count'] = len(days)
        r['site_count'] = len(site_set)
        r['schedule_type_cn'] = _PS_FREQ_CN.get(r['schedule_type'], r['schedule_type'])
        if attention == 'overdue':
            r['attention_reason'] = '存在逾期且未完成的执行包'
        elif attention == 'coverage':
            r['attention_reason'] = '周巡检漏站，等待管理者确认例外原因'
        elif attention == 'resource':
            r['attention_reason'] = '车辆或库存校验未通过'
    return jsonify(rows)


@app.route('/api/plan-schedules', methods=['POST'])
def api_plan_schedules_create():
    """新建巡检计划（草稿或直接提交）。
    body: {user_id?, schedule_type, period_start, period_end, plan_data, vehicle_days?, spare_parts?, work_order_ids?, remarks?, submit?}"""
    u = g.current_user
    data = request.get_json(silent=True) or {}
    schedule_type = data.get('schedule_type', 'weekly')
    if schedule_type not in _PS_FREQ_CN:
        return jsonify({'error': f'无效的计划类型：{schedule_type}'}), 400
    period_start = data.get('period_start')
    period_end = data.get('period_end')
    if not period_start or not period_end:
        return jsonify({'error': '缺少周期起止日期'}), 400
    plan_data = data.get('plan_data') or {}
    if not isinstance(plan_data, dict):
        return jsonify({'error': 'plan_data 必须是 日期→安排 的对象'}), 400
    # 非管理者只能给自己排程
    user_id = data.get('user_id') or u['id']
    if u['role'] not in ('admin', 'manager') and int(user_id) != u['id']:
        return jsonify({'error': '只能为自己创建巡检计划'}), 403
    vehicle_days = data.get('vehicle_days') or {}
    spare_parts = data.get('spare_parts') or []
    work_order_ids = data.get('work_order_ids') or []
    coverage_exception_reason = (data.get('coverage_exception_reason') or '').strip()
    submit = bool(data.get('submit'))
    with get_db() as db:
        # 同用户同类型周期重叠检查（防止重复排程）
        overlap = db.execute("""
            SELECT id, period_start, period_end FROM plan_schedules
            WHERE user_id=? AND schedule_type=? AND status NOT IN ('rejected','archived')
              AND period_start <= ? AND period_end >= ?
        """, (user_id, schedule_type, period_end, period_start)).fetchone()
        if overlap:
            return jsonify({'error': f'该周期已有计划（{overlap["period_start"]} ~ {overlap["period_end"]}），请勿重复创建'}), 409
        v = _ps_validate(db, user_id, schedule_type, period_start, period_end, plan_data, vehicle_days)
        if submit and not v['ok']:
            return jsonify({'error': '；'.join(v['errors']), 'validation': v}), 400
        coverage_error = _ps_coverage_exception_required(v, schedule_type, coverage_exception_reason)
        if submit and coverage_error:
            return jsonify({'error': coverage_error, 'validation': v}), 400
        status = 'submitted' if submit else 'draft'
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cur = db.execute("""
            INSERT INTO plan_schedules
                (user_id, schedule_type, period_start, period_end, plan_data, vehicle_days,
                 spare_parts, work_order_ids, status, remarks, submitted_at, coverage_exception_reason, validation_snapshot)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (user_id, schedule_type, period_start, period_end,
              json.dumps(plan_data, ensure_ascii=False), json.dumps(vehicle_days, ensure_ascii=False),
              json.dumps(spare_parts, ensure_ascii=False), json.dumps(work_order_ids, ensure_ascii=False),
              status, data.get('remarks', ''), now if submit else None, coverage_exception_reason,
              json.dumps(v, ensure_ascii=False) if submit else None))
        _ps_record_event(db, cur.lastrowid, 1, 'submitted' if submit else 'created', u['id'],
                         {'validation': v, 'coverage_exception_reason': coverage_exception_reason})
        db.commit()
        sid = cur.lastrowid
        if submit:
            # 通知审批者
            approvers = db.execute("SELECT id FROM users WHERE role IN ('admin','manager')").fetchall()
            for ap in approvers:
                _create_notification(ap['id'], 'plan_schedule', sid,
                                     f'有新的巡检计划待审批', f'{_PS_FREQ_CN.get(schedule_type, "巡检")}计划（{period_start} ~ {period_end}）已提交，请审批。', db=db)
            db.commit()
        row = db.execute("""
            SELECT ps.*, u.real_name as user_name FROM plan_schedules ps
            LEFT JOIN users u ON ps.user_id=u.id WHERE ps.id=?""", (sid,)).fetchone()
        return jsonify(_ps_parse_row(row)), 201


@app.route('/api/plan-schedules/<int:sid>', methods=['GET'])
def api_plan_schedules_detail(sid):
    """计划详情：含站点信息映射（名称/经纬度，供前端渲染站点卡与路线）"""
    with get_db() as db:
        row = db.execute("""
            SELECT ps.*, u.real_name as user_name, u2.real_name as approver_name
            FROM plan_schedules ps
            LEFT JOIN users u ON ps.user_id=u.id
            LEFT JOIN users u2 ON ps.approver_id=u2.id WHERE ps.id=?""", (sid,)).fetchone()
        if not row:
            return jsonify({'error': '计划不存在'}), 404
        r = _ps_parse_row(row)
        site_ids = set()
        for day_data in (r.get('plan_data') or {}).values():
            if isinstance(day_data, dict):
                site_ids.update(day_data.get('sites') or [])
        site_map = {}
        if site_ids:
            for s in db.execute(
                    f"SELECT id, name, code, gps_lat as lat, gps_lng as lng, type, status FROM sites WHERE id IN ({','.join('?'*len(site_ids))})",
                    list(site_ids)).fetchall():
                site_map[s['id']] = dict(s)
        r['site_map'] = site_map
        # 出发前信息：移动端不能要求一线人员再切换车辆/库存/工单页面核对。
        vehicle_ids = {int(v) for v in (r.get('vehicle_days') or {}).values() if str(v).isdigit()}
        vehicle_map = {}
        if vehicle_ids:
            for vehicle in db.execute(
                    f"SELECT id, plate_no, model, status FROM vehicles WHERE id IN ({','.join('?' * len(vehicle_ids))})",
                    list(vehicle_ids)).fetchall():
                vehicle_map[vehicle['id']] = dict(vehicle)
        r['vehicle_map'] = vehicle_map

        r['resource_parts'] = [dict(part) for part in db.execute("""
            SELECT pr.part_id, pr.planned_quantity, pr.reserved_quantity, pr.issued_quantity,
                   pr.used_quantity, pr.returned_quantity, pr.status, spi.part_name, spi.part_code, spi.unit
            FROM plan_resource_reservations pr
            LEFT JOIN spare_parts_inventory spi ON spi.id=pr.part_id
            WHERE pr.schedule_id=? ORDER BY pr.id
        """, (sid,)).fetchall()]

        work_order_ids = [int(order_id) for order_id in (r.get('work_order_ids') or [])
                          if str(order_id).isdigit()]
        r['linked_workorders'] = []
        if work_order_ids:
            r['linked_workorders'] = [dict(order) for order in db.execute(
                f"""SELECT w.id, w.order_no, w.title, w.status, w.level, s.name AS site_name
                    FROM work_orders w LEFT JOIN sites s ON s.id=w.site_id
                    WHERE w.id IN ({','.join('?' * len(work_order_ids))})
                    ORDER BY CASE w.level WHEN 'critical' THEN 1 WHEN 'urgent' THEN 2 ELSE 3 END, w.created_at""",
                work_order_ids).fetchall()]
        # 关联的已生成执行任务
        r['generated_plans'] = [dict(p) for p in db.execute(
            "SELECT id, plan_name, generate_date, status, completion_rate FROM insp_plans WHERE plan_schedule_id=?",
            (sid,)).fetchall()]
        return jsonify(r)


@app.route('/api/plan-schedules/<int:sid>', methods=['PUT'])
def api_plan_schedules_update(sid):
    """修改计划（draft/rejected/modifying 状态；本人或管理员）。
    draft/rejected 修改后回到 draft；modifying（变更中）修改后保持 modifying。"""
    u = g.current_user
    data = request.get_json(silent=True) or {}
    with get_db() as db:
        row = db.execute("SELECT * FROM plan_schedules WHERE id=?", (sid,)).fetchone()
        if not row:
            return jsonify({'error': '计划不存在'}), 404
        if row['user_id'] != u['id'] and u['role'] not in ('admin', 'manager'):
            return jsonify({'error': '只能修改自己的计划'}), 403
        if row['status'] not in ('draft', 'rejected', 'modifying'):
            return jsonify({'error': f'当前状态（{row["status"]}）不可修改'}), 400
        plan_data = data.get('plan_data', json.loads(row['plan_data'] or '{}'))
        vehicle_days = data.get('vehicle_days', json.loads(row['vehicle_days'] or '{}'))
        spare_parts = data.get('spare_parts', json.loads(row['spare_parts'] or '[]'))
        work_order_ids = data.get('work_order_ids', json.loads(row['work_order_ids'] or '[]'))
        remarks = data.get('remarks', row['remarks'])
        coverage_exception_reason = (data.get('coverage_exception_reason', row['coverage_exception_reason'] or '') or '').strip()
        # 变更中保持 modifying；其余回到 draft
        new_status = 'modifying' if row['status'] == 'modifying' else 'draft'
        db.execute("""
            UPDATE plan_schedules SET plan_data=?, vehicle_days=?, spare_parts=?, work_order_ids=?,
                   remarks=?, coverage_exception_reason=?, status=?, reject_reason=NULL WHERE id=?
        """, (json.dumps(plan_data, ensure_ascii=False), json.dumps(vehicle_days, ensure_ascii=False),
              json.dumps(spare_parts, ensure_ascii=False), json.dumps(work_order_ids, ensure_ascii=False),
              remarks, coverage_exception_reason, new_status, sid))
        _ps_record_event(db, sid, row['version'], 'updated', u['id'], {'status': new_status})
        db.commit()
        return jsonify({'success': True, 'id': sid, 'status': new_status})


@app.route('/api/plan-schedules/<int:sid>/submit', methods=['POST'])
def api_plan_schedules_submit(sid):
    """提交审批：draft/rejected → submitted；modifying（变更）→ change_submitted 且版本+1。
    带校验：errors 阻断，warnings 放行但记录。"""
    u = g.current_user
    with get_db() as db:
        row = db.execute("SELECT * FROM plan_schedules WHERE id=?", (sid,)).fetchone()
        if not row:
            return jsonify({'error': '计划不存在'}), 404
        if row['user_id'] != u['id'] and u['role'] not in ('admin', 'manager'):
            return jsonify({'error': '只能提交自己的计划'}), 403
        is_change = row['status'] == 'modifying'
        if row['status'] not in ('draft', 'rejected', 'modifying'):
            return jsonify({'error': f'当前状态（{row["status"]}）不可提交'}), 400
        plan_data = json.loads(row['plan_data'] or '{}')
        vehicle_days = json.loads(row['vehicle_days'] or '{}')
        v = _ps_validate(db, row['user_id'], row['schedule_type'],
                         row['period_start'], row['period_end'], plan_data, vehicle_days)
        if not v['ok']:
            return jsonify({'error': '；'.join(v['errors']), 'validation': v}), 400
        coverage_error = _ps_coverage_exception_required(v, row['schedule_type'], row['coverage_exception_reason'])
        if coverage_error:
            return jsonify({'error': coverage_error, 'validation': v}), 400
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        if is_change:
            db.execute("""
                UPDATE plan_schedules SET status='change_submitted', submitted_at=?,
                       version=version+1, reject_reason=NULL, validation_snapshot=? WHERE id=?
            """, (now, json.dumps(v, ensure_ascii=False), sid))
            new_status = 'change_submitted'
            notif_title, notif_body = '巡检计划变更待审', \
                f'{_PS_FREQ_CN.get(row["schedule_type"], "巡检")}计划（{row["period_start"]} ~ {row["period_end"]}）发起变更（原因：{row["change_reason"] or "未填"}），请审核。'
        else:
            db.execute("UPDATE plan_schedules SET status='submitted', submitted_at=?, reject_reason=NULL, validation_snapshot=? WHERE id=?",
                       (now, json.dumps(v, ensure_ascii=False), sid))
            new_status = 'submitted'
            notif_title, notif_body = '有新的巡检计划待审批', \
                f'{_PS_FREQ_CN.get(row["schedule_type"], "巡检")}计划（{row["period_start"]} ~ {row["period_end"]}）已提交，请审批。'
        approvers = db.execute("SELECT id FROM users WHERE role IN ('admin','manager')").fetchall()
        for ap in approvers:
            _create_notification(ap['id'], 'plan_schedule', sid, notif_title, notif_body, db=db)
        _ps_record_event(db, sid, row['version'] + (1 if is_change else 0),
                         'change_submitted' if is_change else 'submitted', u['id'],
                         {'validation': v, 'coverage_exception_reason': row['coverage_exception_reason']})
        db.commit()
        return jsonify({'success': True, 'id': sid, 'status': new_status, 'validation': v})


@app.route('/api/plan-schedules/<int:sid>/approve', methods=['POST'])
def api_plan_schedules_approve(sid):
    """审批通过（admin/manager）：状态流转 + 自动生成执行任务 + 锁车 + 备件出库 + 通知"""
    denied = require_approver()
    if denied:
        return denied
    u = g.current_user
    with get_db() as db:
        row = db.execute("SELECT * FROM plan_schedules WHERE id=?", (sid,)).fetchone()
        if not row:
            return jsonify({'error': '计划不存在'}), 404
        if row['status'] not in ('submitted', 'change_submitted'):
            return jsonify({'error': f'当前状态（{row["status"]}）不可审批'}), 400
        is_change = row['status'] == 'change_submitted'
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        # 先校验预留可用性，避免计划已经标记通过却因备件不足而无法完成资源协调。
        if not is_change:
            try:
                _ps_reserve_parts(db, row)
                db.execute("DELETE FROM plan_resource_reservations WHERE schedule_id=? AND status='reserved'", (sid,))
            except ValueError as e:
                return jsonify({'error': str(e)}), 400
        db.execute("""
            UPDATE plan_schedules SET status='approved', approver_id=?, approved_at=?, reject_reason=NULL
            WHERE id=?
        """, (u['id'], now, sid))
        fresh = db.execute("SELECT * FROM plan_schedules WHERE id=?", (sid,)).fetchone()
        if is_change:
            # 变更后释放尚未领用的预留，并按新清单重新预留；已领用记录不回滚。
            db.execute("UPDATE plan_resource_reservations SET status='released', updated_at=datetime('now','localtime') WHERE schedule_id=? AND status='reserved'", (sid,))
            _ps_reserve_parts(db, fresh)
            # 取消今天及未来的旧用车预约，再按新日期/车辆锁定；历史出车申请保持原样。
            db.execute("UPDATE vehicle_applications SET status='cancelled' WHERE applicant_id=? AND date(start_at)>=date('now','localtime') AND reason LIKE ? AND status='approved'",
                       (fresh['user_id'], f'%计划#{sid}%'))
            try:
                changed_vehicle_days = json.loads(fresh['vehicle_days'] or '{}')
            except Exception:
                changed_vehicle_days = {}
            for date_str, vid in changed_vehicle_days.items():
                if not vid or date_str < datetime.now().strftime('%Y-%m-%d'):
                    continue
                db.execute("""INSERT INTO vehicle_applications
                    (vehicle_id, applicant_id, start_at, end_at, destination, reason, status)
                    VALUES (?,?,?,?,?,?,?)""", (vid, fresh['user_id'], f'{date_str} 08:00:00',
                    f'{date_str} 18:00:00', '巡检', f'巡检计划#{sid}用车（v{fresh["version"]}）', 'approved'))
            # 变更审批通过：保留已执行内容，取消未执行旧项并生成新增日执行包。
            flow = _ps_rebuild_tasks_on_change(db, fresh)
            _create_notification(fresh['user_id'], 'plan_schedule', sid, '计划变更已通过',
                                 f'变更原因：{fresh["change_reason"] or "—"}。任务已按新计划同步'
                                 f'（保留{flow["kept"]}个、重建{flow["plans_created"]}个）。', db=db)
        else:
            flow = _ps_auto_flow(db, fresh)
        _ps_record_event(db, sid, fresh['version'], 'approved', u['id'],
                         {'is_change': is_change, 'coverage_exception_reason': fresh['coverage_exception_reason']})
        db.commit()
        return jsonify({'success': True, 'id': sid, 'status': 'approved', 'is_change': is_change, **flow})


@app.route('/api/plan-schedules/<int:sid>/reject', methods=['POST'])
def api_plan_schedules_reject(sid):
    """审批退回（admin/manager）：必须填写原因，通知排程人"""
    denied = require_approver()
    if denied:
        return denied
    u = g.current_user
    data = request.get_json(silent=True) or {}
    reason = (data.get('reason') or data.get('reject_reason') or '').strip()
    if not reason:
        return jsonify({'error': '请填写退回原因'}), 400
    with get_db() as db:
        row = db.execute("SELECT * FROM plan_schedules WHERE id=?", (sid,)).fetchone()
        if not row:
            return jsonify({'error': '计划不存在'}), 404
        if row['status'] not in ('submitted', 'change_submitted'):
            return jsonify({'error': f'当前状态（{row["status"]}）不可退回'}), 400
        if row['status'] == 'change_submitted':
            # 驳回变更：回滚到变更前的已批准版本，计划整体仍保持 approved
            db.execute("""
                UPDATE plan_schedules SET status='approved', approver_id=?, reject_reason=?,
                       plan_data=COALESCE(previous_plan_data, plan_data),
                       vehicle_days=COALESCE(previous_vehicle_days, vehicle_days),
                       previous_plan_data=NULL, previous_vehicle_days=NULL, change_reason=NULL
                WHERE id=?
            """, (u['id'], reason, sid))
            _create_notification(row['user_id'], 'plan_schedule', sid, '计划变更被驳回',
                                 f'原因：{reason}。已恢复原计划，请继续按原计划执行。', db=db)
            db.commit()
            return jsonify({'success': True, 'id': sid, 'status': 'approved', 'rolled_back': True})
        db.execute("UPDATE plan_schedules SET status='rejected', approver_id=?, reject_reason=? WHERE id=?",
                   (u['id'], reason, sid))
        _create_notification(row['user_id'], 'plan_schedule', sid, '巡检计划被退回',
                             f'原因：{reason}。请修改后重新提交。', db=db)
        db.commit()
        return jsonify({'success': True, 'id': sid, 'status': 'rejected'})


@app.route('/api/plan-schedules/<int:sid>/parts/issue', methods=['POST'])
def api_plan_schedule_issue_parts(sid):
    """现场领料确认：将已预留备件转为实际出库，库存此时才扣减。"""
    u = g.current_user
    data = request.get_json(silent=True) or {}
    requested = data.get('items') or []
    with get_db() as db:
        schedule = db.execute("SELECT * FROM plan_schedules WHERE id=?", (sid,)).fetchone()
        if not schedule:
            return jsonify({'error': '计划不存在'}), 404
        if schedule['user_id'] != u['id'] and u['role'] not in ('admin', 'manager'):
            return jsonify({'error': '只能确认自己计划的领料'}), 403
        if schedule['status'] != 'approved':
            return jsonify({'error': '仅已批准计划可领料'}), 400
        issued = 0
        for item in requested:
            part_id, qty = int(item.get('part_id') or 0), int(item.get('quantity') or 0)
            if not part_id or qty <= 0:
                continue
            reservation = db.execute("""SELECT * FROM plan_resource_reservations
                WHERE schedule_id=? AND part_id=? AND status='reserved' ORDER BY id LIMIT 1""", (sid, part_id)).fetchone()
            if not reservation or qty > reservation['reserved_quantity'] - reservation['issued_quantity']:
                return jsonify({'error': f'备件#{part_id}没有足够的已预留数量'}), 400
            inv = db.execute("SELECT quantity FROM spare_parts_inventory WHERE id=?", (part_id,)).fetchone()
            if not inv or inv['quantity'] < qty:
                return jsonify({'error': f'备件#{part_id}库存不足'}), 400
            db.execute("UPDATE spare_parts_inventory SET quantity=quantity-?, updated_at=datetime('now','localtime') WHERE id=?", (qty, part_id))
            db.execute("UPDATE plan_resource_reservations SET issued_quantity=issued_quantity+?, status='issued', updated_at=datetime('now','localtime') WHERE id=?", (qty, reservation['id']))
            db.execute("INSERT INTO inventory_logs (part_id, type, quantity, ref_type, ref_id, operator, remark) VALUES (?,?,?,?,?,?,?)",
                       (part_id, 'out', qty, 'plan_schedule', sid, u.get('real_name') or u.get('username'), f'巡检计划#{sid}现场领用'))
            issued += qty
        _ps_record_event(db, sid, schedule['version'], 'parts_issued', u['id'], {'items': requested})
        db.commit()
        return jsonify({'success': True, 'issued_quantity': issued})


@app.route('/api/plan-schedules/<int:sid>/request-change', methods=['POST'])
def api_plan_schedules_request_change(sid):
    """发起变更（approved → modifying）：车辆故障/突发事件等场景，运维修改已批准计划。
    快照当前 plan_data/vehicle_days 到 previous_* 字段，变更被驳回时可回滚。"""
    u = g.current_user
    data = request.get_json(silent=True) or {}
    reason = (data.get('change_reason') or data.get('reason') or '').strip()
    if not reason:
        return jsonify({'error': '请填写变更原因'}), 400
    with get_db() as db:
        row = db.execute("SELECT * FROM plan_schedules WHERE id=?", (sid,)).fetchone()
        if not row:
            return jsonify({'error': '计划不存在'}), 404
        if row['user_id'] != u['id'] and u['role'] not in ('admin', 'manager'):
            return jsonify({'error': '只能变更自己的计划'}), 403
        if row['status'] != 'approved':
            return jsonify({'error': f'当前状态（{row["status"]}）不可发起变更，仅已通过的计划可变更'}), 400
        db.execute("""
            UPDATE plan_schedules SET status='modifying', change_reason=?,
                   previous_plan_data=?, previous_vehicle_days=?
            WHERE id=?
        """, (reason, row['plan_data'], row['vehicle_days'], sid))
        # 通知审批人有变更待处理
        approvers = db.execute("SELECT id FROM users WHERE role IN ('admin','manager')").fetchall()
        for ap in approvers:
            _create_notification(ap['id'], 'plan_schedule', sid, '巡检计划发起变更',
                                 f'原因：{reason}。修改后将重新提交审核。', db=db)
        db.commit()
        return jsonify({'success': True, 'id': sid, 'status': 'modifying'})


@app.route('/api/plan-schedules/validate', methods=['POST'])
def api_plan_schedules_validate():
    """实时校验（排程编辑过程中调用）：全覆盖/日期范围/用车冲突。
    body: {user_id?, schedule_type, period_start, period_end, plan_data, vehicle_days, exclude_schedule_id?}"""
    u = g.current_user
    data = request.get_json(silent=True) or {}
    user_id = data.get('user_id') or u['id']
    with get_db() as db:
        v = _ps_validate(db, user_id,
                         data.get('schedule_type', 'weekly'),
                         data.get('period_start'), data.get('period_end'),
                         data.get('plan_data') or {}, data.get('vehicle_days') or {},
                         data.get('exclude_schedule_id'))
        return jsonify(v)


@app.route('/api/plan-schedules/suggestions', methods=['GET'])
def api_plan_schedules_suggestions():
    """智能建议（排程端实时提示 / 审批端风险预警，一套数据双端复用）。
    ?user_id=1 或 ?site_ids=3,7,12。返回各站点的工单顺路/告警复核/试剂临期建议与优先级评分。"""
    u = g.current_user
    site_ids_arg = request.args.get('site_ids')
    user_id = request.args.get('user_id')
    with get_db() as db:
        if site_ids_arg:
            try:
                site_ids = [int(x) for x in site_ids_arg.split(',') if x.strip()]
            except ValueError:
                return jsonify({'error': 'site_ids 格式错误'}), 400
        else:
            uid = user_id or u['id']
            site_ids = [r['site_id'] for r in
                        db.execute("SELECT site_id FROM user_sites WHERE user_id=?", (uid,)).fetchall()]
        if not site_ids:
            return jsonify({'suggestions': [], 'site_scores': {}, 'site_reasons': {}})
        site_names = {r['id']: r['name'] for r in db.execute(
            f"SELECT id, name FROM sites WHERE id IN ({','.join('?'*len(site_ids))})", site_ids).fetchall()}
        scores, reasons = _ps_site_scores(db, site_ids)
        suggestions = []
        for sid in site_ids:
            sname = site_names.get(sid, f'站点{sid}')
            # 工单顺路
            orders = db.execute("""
                SELECT id, title, level, created_at FROM work_orders
                WHERE site_id=? AND status NOT IN ('closed')
            """, (sid,)).fetchall()
            for o in orders:
                try:
                    age = (datetime.now() - datetime.strptime(str(o['created_at'])[:10], '%Y-%m-%d')).days
                except Exception:
                    age = 0
                suggestions.append({
                    'type': 'work_order', 'site_id': sid, 'site_name': sname, 'ref_id': o['id'],
                    'level': o['level'],
                    'text': f'{sname}有未关工单「{o["title"]}」（{"紧急" if o["level"] == "urgent" else "普通"}，已{age}天），建议顺路处理并优先安排'})
            # 告警复核
            alerts = db.execute("SELECT id, level, metric FROM alerts WHERE site_id=? AND status='pending'",
                                (sid,)).fetchall()
            for a in alerts:
                suggestions.append({
                    'type': 'alert', 'site_id': sid, 'site_name': sname, 'ref_id': a['id'],
                    'level': a['level'],
                    'text': f'{sname}有{ALERT_LEVEL_LABEL.get(a["level"], "")}告警（{a["metric"] or "指标异常"}）待现场复核'})
            # 试剂临期
            try:
                invs = db.execute("""
                    SELECT ri.*, r.name as reagent_name FROM reagent_inventory ri
                    LEFT JOIN reagents r ON ri.reagent_id = r.id WHERE ri.site_id=?
                """, (sid,)).fetchall()
                for inv in invs:
                    st = compute_reagent_status(dict(inv))
                    if st.get('status') in ('临期', '已过期', '低余量'):
                        suggestions.append({
                            'type': 'reagent', 'site_id': sid, 'site_name': sname, 'ref_id': inv['id'],
                            'level': 'warning',
                            'text': f'{sname}试剂{inv["reagent_name"] or ""}{st["status"]}（约剩{st.get("remaining_days")}天），建议本次携带更换'})
            except Exception:
                pass
        # 高危优先建议：评分最高的站点
        if scores:
            top_sid = max(scores, key=lambda k: scores[k])
            if scores[top_sid] >= 30:
                suggestions.append({
                    'type': 'priority', 'site_id': top_sid,
                    'site_name': site_names.get(top_sid, f'站点{top_sid}'),
                    'level': 'high', 'ref_id': None,
                    'text': f'{site_names.get(top_sid, "")}优先级最高（{"；".join(reasons.get(top_sid, []))}），建议排在周期前几天'})
        return jsonify({
            'suggestions': suggestions,
            'site_scores': {str(k): v for k, v in scores.items()},
            'site_reasons': {str(k): v for k, v in reasons.items()},
        })


# ---------- 5.1 阈值规则 ----------
@app.route('/api/threshold-rules', methods=['GET'])
def api_threshold_rules():
    """阈值规则列表"""
    site_id = request.args.get('site_id')
    metric = request.args.get('metric')
    with get_db() as db:
        q = 'SELECT * FROM threshold_rules WHERE 1=1'
        params = []
        if site_id:
            q += ' AND (site_id=? OR scope="global")'; params.append(site_id)
        if metric:
            q += ' AND (metric=? OR scope="global")'; params.append(metric)
        q += ' ORDER BY scope, metric'
        rows = db.execute(q, params).fetchall()
        result = []
        for r in rows:
            r = dict(r)
            try:
                import json as _json
                r['conditions'] = _json.loads(r['conditions']) if r.get('conditions') else {}
            except:
                pass
            result.append(r)
        return jsonify(result)


@app.route('/api/threshold-rules', methods=['POST'])
def api_threshold_rules_create():
    """新增阈值规则"""
    data = request.get_json() or {}
    import json as _json
    with get_db() as db:
        cur = db.execute(
            '''INSERT INTO threshold_rules (name, scope, site_id, metric, rule_type, conditions, severity, created_by)
               VALUES (?,?,?,?,?,?,?,?)''',
            (data['name'], data.get('scope', 'metric'), data.get('site_id'), data.get('metric'),
             data['rule_type'], _json.dumps(data.get('conditions', {}), ensure_ascii=False),
             data.get('severity', 'warning'), data.get('created_by')))
        db.commit()
        row = db.execute('SELECT * FROM threshold_rules WHERE id=?', (cur.lastrowid,)).fetchone()
        r = dict(row)
        try: r['conditions'] = _json.loads(r['conditions']) if r.get('conditions') else {}
        except: pass
        return jsonify(r), 201


# ---------- 告警规则引擎配置（前台可编辑阈值，持久化到 DB）----------
DEFAULT_ALERT_RULES = [
    {'id': 'rule_data_gap',     'metric': 'data_gap',     'metric_label': '数据缺失', 'description': '监测数据连续缺失超过设定时间触发告警', 'enabled': 1, 'flow_type': 'auto',   'unit': '分钟', 'thresholds': '{"blue":30,"yellow":60,"orange":120,"red":240}'},
    {'id': 'rule_data_freeze',  'metric': 'data_freeze',  'metric_label': '数据冻结', 'description': '监测数据长时间保持不变（疑似传感器故障）', 'enabled': 1, 'flow_type': 'manual', 'unit': '条',   'thresholds': '{"blue":4,"yellow":6,"orange":10,"red":20}'},
    {'id': 'rule_data_spike',   'metric': 'data_spike',   'metric_label': '数据突变', 'description': '监测数据短时间内变化幅度超过阈值',     'enabled': 1, 'flow_type': 'manual', 'unit': '%',   'thresholds': '{"blue":15,"yellow":30,"orange":50,"red":80}'},
    {'id': 'rule_device_status','metric': 'device_status','metric_label': '设备离线', 'description': '设备心跳超时判定为离线状态',               'enabled': 1, 'flow_type': 'auto',   'unit': '分钟', 'thresholds': '{"blue":10,"yellow":30,"orange":60,"red":120}'},
    {'id': 'rule_arrival_rate', 'metric': 'arrival_rate', 'metric_label': '到报率',   'description': '站点数据到报率低于设定阈值',                 'enabled': 1, 'flow_type': 'manual', 'unit': '%',   'thresholds': '{"blue":95,"yellow":90,"orange":80,"red":70}', 'is_reversed': 1},
]

def ensure_alert_rules_config():
    with get_db() as db:
        count = db.execute('SELECT COUNT(*) FROM alert_rule_config').fetchone()[0]
        if count == 0:
            for r in DEFAULT_ALERT_RULES:
                db.execute(
                    '''INSERT OR REPLACE INTO alert_rule_config
                       (id, metric, metric_label, description, enabled, flow_type, unit, thresholds, is_reversed)
                       VALUES (?,?,?,?,?,?,?,?,?)''',
                    (r['id'], r['metric'], r['metric_label'], r['description'],
                     r['enabled'], r['flow_type'], r['unit'], r['thresholds'],
                     r.get('is_reversed', 0)))
            db.commit()

@app.route('/api/alert-rules', methods=['GET'])
@login_required
def api_alert_rules_get():
    """获取告警规则引擎配置"""
    with get_db() as db:
        rows = db.execute('SELECT * FROM alert_rule_config ORDER BY id').fetchall()
        import json as _j
        result = []
        for r in rows:
            d = dict(r)
            try: d['thresholds'] = _j.loads(d['thresholds']) if d.get('thresholds') else {}
            except: pass
            d['enabled'] = bool(d['enabled'])
            d['isReversed'] = bool(d.get('is_reversed', 0))
            d['flowType'] = d.get('flow_type', 'auto')  # 前端用 camelCase
            d['metricLabel'] = d.get('metric_label', '')
            result.append(d)
        return jsonify(result)

@app.route('/api/alert-rules/<rule_id>', methods=['PUT'])
@login_required
def api_alert_rules_update(rule_id):
    """更新告警规则阈值/启用状态"""
    data = request.get_json(silent=True) or {}
    with get_db() as db:
        existing = db.execute('SELECT * FROM alert_rule_config WHERE id=?', (rule_id,)).fetchone()
        if not existing:
            return jsonify({'error': '规则不存在'}), 404
        import json as _j
        thresholds = data.get('thresholds')
        if thresholds is not None:
            db.execute('UPDATE alert_rule_config SET thresholds=? WHERE id=?',
                       (_j.dumps(thresholds, ensure_ascii=False), rule_id))
        if 'enabled' in data:
            db.execute('UPDATE alert_rule_config SET enabled=? WHERE id=?',
                       (1 if data['enabled'] else 0, rule_id))
        db.commit()
        row = db.execute('SELECT * FROM alert_rule_config WHERE id=?', (rule_id,)).fetchone()
        d = dict(row)
        try: d['thresholds'] = _j.loads(d['thresholds']) if d.get('thresholds') else {}
        except: pass
        d['enabled'] = bool(d['enabled'])
        d['isReversed'] = bool(d.get('is_reversed', 0))
        d['flowType'] = d.get('flow_type', 'auto')
        d['metricLabel'] = d.get('metric_label', '')
        return jsonify(d)


# ---------- 新增：异常主动上报 ----------
@app.route('/api/manual-reports', methods=['GET'])
def api_manual_reports():
    """人工上报列表（按站点范围隔离）"""
    status = request.args.get('status')
    site_id = request.args.get('site_id')
    allowed = _filter_site_ids()
    site_id_int = int(site_id) if site_id and site_id.isdigit() else None
    with get_db() as db:
        q = '''SELECT mr.*, u.real_name as reporter_name, s.name as site_name
               FROM manual_reports mr
               LEFT JOIN users u ON mr.reporter_id = u.id
               LEFT JOIN sites s ON mr.site_id = s.id WHERE 1=1'''
        params = []
        if status:
            q += ' AND mr.status=?'; params.append(status)
        if site_id_int is not None:
            if allowed is not None and site_id_int not in allowed:
                return jsonify([])
            q += ' AND mr.site_id=?'; params.append(site_id_int)
        elif allowed is not None:
            ph = ','.join('?' * len(allowed))
            q += f' AND mr.site_id IN ({ph})'; params.extend(allowed)
        q += ' ORDER BY mr.reported_at DESC'
        rows = db.execute(q, params).fetchall()
        return jsonify([dict(r) for r in rows])


@app.route('/api/manual-reports', methods=['POST'])
def api_manual_reports_create():
    """移动端一键上报异常，自动生成工单（site_id 受当前用户站点范围约束）"""
    data = request.get_json() or {}
    site_id = data.get('site_id')
    allowed = _filter_site_ids()
    if allowed is not None and (not site_id or int(site_id) not in allowed):
        return jsonify({'error': '无权限上报非本人负责站点的异常'}), 403
    report_type = data['report_type']
    description = data.get('description', '')
    photo_urls = data.get('photo_urls', [])
    gps_lat = data.get('gps_lat')
    gps_lng = data.get('gps_lng')
    # 上报人必须来自已认证的请求上下文，不能信任客户端传入的 user id。
    reporter_id = g.current_user['id']

    import json as _json
    photos_json = _json.dumps(photo_urls, ensure_ascii=False)
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    with get_db() as db:
        # 写上报记录
        cur = db.execute(
            'INSERT INTO manual_reports (site_id, report_type, description, photo_urls, gps_lat, gps_lng, reporter_id, reported_at) VALUES (?,?,?,?,?,?,?,?)',
            (site_id, report_type, description, photos_json, gps_lat, gps_lng, reporter_id, now))
        report_id = cur.lastrowid

        # 自动生成工单（与机器告警统一队列）
        report_type_map = {'sensory':'感官异常','equipment':'设备异常','environment':'环境异常','violation':'违规操作','pollution':'污染事件'}
        event_type = report_type_map.get(report_type, report_type)
        order_no = 'MR' + now[:10].replace('-','') + str(report_id).zfill(4)
        cur2 = db.execute(
            '''INSERT INTO work_orders (order_no, site_id, source, event_type, level, title, description, status)
               VALUES (?,?,"manual_report",?,"normal",?,?,"pending")''',
            (order_no, site_id, event_type, f'【人工上报】{event_type}', description))

        # 写入告警队列（与机器告警同表同级别，去重：30秒内同站点同类型同描述不重复创建）
        level_map = {'sensory':'yellow','equipment':'orange','environment':'orange','violation':'red','pollution':'red'}
        alert_level = level_map.get(report_type, 'yellow')
        alert_msg = f'【人工上报】{event_type}: {description[:200]}'
        existing = db.execute(
            "SELECT id FROM alerts WHERE site_id=? AND message=? AND created_at > datetime('now','-30 seconds')",
            (site_id, alert_msg)).fetchone()
        if not existing:
            db.execute(
                '''INSERT INTO alerts (site_id, metric, value, level, message, status, related_order_no, flow_type)
                   VALUES (?,?,?,?,?,'pending',?,'manual')''',
                (site_id, 'manual_report', 0, alert_level, alert_msg, order_no))

        # 关联工单号回写
        db.execute('UPDATE manual_reports SET order_no=?, status="dispatched" WHERE id=?', (order_no, report_id))
        db.commit()

        row = db.execute('SELECT * FROM manual_reports WHERE id=?', (report_id,)).fetchone()
        return jsonify(dict(row)), 201


# ============================================================
# 三级数据质量审核（仪器自动 → 平台智能 → 人工复核）
# ============================================================

# GB3838-2002 III类水标准（与 migrate_v2.sql 一致）
_METRIC_BOUNDS = {
    'ph': (6.0, 9.0), 'cod': (0, 40.0), 'ammonia': (0, 5.0),
    'total_phosphorus': (0, 1.0), 'total_nitrogen': (0, 2.0),
    'dissolved_oxygen': (3.0, 100.0), 'turbidity': (0, 50.0),
    'water_temp': (-5.0, 45.0),
}

# 多指标关联规则：(关联指标, 期望趋势一致性)
# direction: 1=同向变化(同升同降), -1=反向变化(一升一降)
_METRIC_CORRELATIONS = {
    'ph': [('dissolved_oxygen', 1)],   # pH↑→DO↑；pH↓→DO↓
    'dissolved_oxygen': [('ph', 1), ('turbidity', -1), ('water_temp', -1)],  # DO↑→浊度↓, DO↑→水温↓
    'turbidity': [('dissolved_oxygen', -1)],  # 浊度↑→DO↓
    'ammonia': [('dissolved_oxygen', -1)],    # 氨氮↑→DO↓
}

# 指标名映射（sensor_data 中的值用于 SQL 查历史）
_METRIC_LABELS = {
    'ph': 'pH', 'cod': 'COD', 'ammonia': '氨氮',
    'total_phosphorus': '总磷', 'total_nitrogen': '总氮',
    'dissolved_oxygen': '溶解氧', 'turbidity': '浊度', 'water_temp': '水温',
}


# ---------- 从 sensor_data 生成待审核记录 ----------
@app.route('/api/data-reviews/generate', methods=['POST'])
def api_data_reviews_generate():
    data = request.get_json() or {}
    hours = data.get('hours', 24)
    site_id = data.get('site_id')
    metric = data.get('metric')
    date_from = data.get('date_from', (datetime.now() - timedelta(hours=hours)).strftime('%Y-%m-%d %H:%M:%S'))
    with get_db() as db:
        q = '''SELECT sd.site_id, sd.metric, sd.value, sd.recorded_at
               FROM sensor_data sd WHERE sd.recorded_at >= ?'''
        params = [date_from]
        if site_id:
            q += ' AND sd.site_id=?'; params.append(site_id)
        if metric:
            q += ' AND sd.metric=?'; params.append(metric)
        q += ' ORDER BY sd.recorded_at DESC LIMIT 5000'
        rows = db.execute(q, params).fetchall()
        inserted = 0
        for r in rows:
            existed = db.execute(
                'SELECT id FROM data_reviews WHERE site_id=? AND metric=? AND value=? AND recorded_at=?',
                (r['site_id'], r['metric'], r['value'], r['recorded_at'])).fetchone()
            if existed:
                continue
            db.execute(
                'INSERT INTO data_reviews (site_id, metric, value, recorded_at, status) VALUES (?,?,?,?,"pending")',
                (r['site_id'], r['metric'], r['value'], r['recorded_at']))
            inserted += 1
        db.commit()
        return jsonify({'generated': inserted, 'scanned': len(rows)})


# ---------- 审核列表（多维筛选 + 分页） ----------
@app.route('/api/data-reviews', methods=['GET'])
def api_data_reviews_list():
    status = request.args.get('status')
    site_id = request.args.get('site_id')
    metric = request.args.get('metric')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    level = request.args.get('level')  # 1=pending/auto_reviewed; 2=auto/smart; 3=smart/manual
    page = int(request.args.get('page', 1))
    per_page = min(int(request.args.get('per_page', 50)), 200)
    with get_db() as db:
        q = '''SELECT dr.*, s.name as site_name, s.code as site_code
               FROM data_reviews dr JOIN sites s ON dr.site_id = s.id WHERE 1=1'''
        params = []
        if status:
            q += ' AND dr.status=?'; params.append(status)
        if site_id:
            q += ' AND dr.site_id=?'; params.append(site_id)
        if metric:
            q += ' AND dr.metric=?'; params.append(metric)
        if date_from:
            q += ' AND dr.recorded_at >= ?'; params.append(date_from)
        if date_to:
            q += ' AND dr.recorded_at <= ?'; params.append(date_to)
        if level == '1':
            q += " AND dr.status IN ('pending','auto_reviewed')"
        elif level == '2':
            q += " AND dr.status IN ('auto_reviewed','smart_reviewed')"
        elif level == '3':
            q += " AND dr.status IN ('smart_reviewed','manual_reviewed')"
        q += ' ORDER BY dr.recorded_at DESC LIMIT ? OFFSET ?'
        params.extend([per_page, (page - 1) * per_page])
        items = [dict(r) for r in db.execute(q, params).fetchall()]
        cq = 'SELECT COUNT(*) FROM data_reviews dr WHERE 1=1'
        cparams = []
        if status: cq += ' AND dr.status=?'; cparams.append(status)
        if site_id: cq += ' AND dr.site_id=?'; cparams.append(site_id)
        if metric: cq += ' AND dr.metric=?'; cparams.append(metric)
        if date_from: cq += ' AND dr.recorded_at >= ?'; cparams.append(date_from)
        if date_to: cq += ' AND dr.recorded_at <= ?'; cparams.append(date_to)
        if level == '1':
            cq += " AND dr.status IN ('pending','auto_reviewed')"
        elif level == '2':
            cq += " AND dr.status IN ('auto_reviewed','smart_reviewed')"
        elif level == '3':
            cq += " AND dr.status IN ('smart_reviewed','manual_reviewed')"
        total = db.execute(cq, cparams).fetchone()[0]
        return jsonify({'items': items, 'total': total, 'page': page, 'per_page': per_page})


# ---------- L1: 仪器自动审核（已整合到 auto_data_review 定时任务） ----------
@app.route('/api/data-reviews/auto-review', methods=['POST'])
def api_data_reviews_auto_review():
    """手动触发 L1，实际调用 auto_data_review 的 L1 逻辑"""
    # 此路由保留向后兼容，直接调度 auto_data_review 中的统计函数
    # 前端已删手动按钮，仅 API 客户端还可能调用
    return jsonify({'info': 'L1 已集成到后台定时任务（每10分钟），请使用定时审核', 'auto': True})


# ---------- L2: 平台智能审核（统计窗口 + Z-score） ----------
@app.route('/api/data-reviews/smart-review', methods=['POST'])
def api_data_reviews_smart_review():
    data = request.get_json() or {}
    batch_size = data.get('batch_size', 500)
    z_threshold = data.get('z_threshold', 3.0)  # 3-sigma
    with get_db() as db:
        rows = db.execute(
            "SELECT * FROM data_reviews WHERE status='auto_reviewed' ORDER BY recorded_at DESC LIMIT ?",
            (batch_size,)).fetchall()
        passed = suspicious = 0
        for r in rows:
            # 取该站点该指标近 7 天的均值/标准差作基准
            stats = db.execute(
                '''SELECT AVG(value) as mu, COUNT(*) as n FROM sensor_data
                   WHERE site_id=? AND metric=?
                   AND recorded_at < ?
                   AND recorded_at > datetime(?, '-7 day')''',
                (r['site_id'], r['metric'], r['recorded_at'], r['recorded_at'])).fetchone()
            mu, n = stats['mu'] or 0, stats['n']
            if n < 20 or mu is None:
                # 样本不足，跳过智能判断：视为正常，自动归档（不进 L3）
                db.execute(
                    "UPDATE data_reviews SET smart_result='pass', smart_score=0, status='archived' WHERE id=?",
                    (r['id'],))
                passed += 1
                continue
            sigma_row = db.execute(
                '''SELECT AVG(value*value) as e2 FROM sensor_data
                   WHERE site_id=? AND metric=? AND recorded_at < ?
                   AND recorded_at > datetime(?, '-7 day')''',
                (r['site_id'], r['metric'], r['recorded_at'], r['recorded_at'])).fetchone()
            import math
            var = (sigma_row['e2'] or 0) - mu * mu
            sigma = math.sqrt(max(var, 0))
            if sigma == 0:
                sigma = 0.001
            z = abs((r['value'] - mu) / sigma)
            if z > z_threshold:
                db.execute(
                    "UPDATE data_reviews SET smart_result='suspicious', smart_score=?, status='smart_reviewed' WHERE id=?",
                    (round(z, 3), r['id']))
                suspicious += 1
                _notify_review_l3(db, r)
            else:
                db.execute(
                    "UPDATE data_reviews SET smart_result='pass', smart_score=?, status='archived' WHERE id=?",
                    (round(z, 3), r['id']))
                passed += 1
        db.commit()
        return jsonify({'processed': len(rows), 'passed': passed, 'suspicious': suspicious, 'z_threshold': z_threshold})


# ---------- L3: 人工复核（单条/批量） ----------
@app.route('/api/data-reviews/<int:review_id>/manual-review', methods=['POST'])
def api_data_reviews_manual_review(review_id):
    data = request.get_json() or {}
    action = data.get('action', 'approve')  # approve/reject
    reviewer_id = data.get('reviewer_id')
    reason = data.get('reason', '')
    conclusion = data.get('conclusion')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    result = 'approved' if action == 'approve' else 'rejected'
    new_status = 'archived'  # 人工复核即最终判定，统一归档
    with get_db() as db:
        cur = db.execute("SELECT site_id, metric FROM data_reviews WHERE id=?", (review_id,)).fetchone()
        db.execute(
            "UPDATE data_reviews SET manual_result=?, manual_reason=?, reviewer_id=?, reviewed_at=?, status=? WHERE id=?",
            (result, (conclusion or reason), reviewer_id, now, new_status, review_id))
        # === 闭环：人工判定结论 → 联动办结关联告警（场景2：管理员抢先消除）===
        if conclusion:
            _link_review_alert(db, cur['site_id'], cur['metric'])
            _resolve_linked_alert(db, cur['site_id'], cur['metric'], conclusion, review_id=review_id)
        db.commit()
        return jsonify({'ok': True, 'result': result, 'status': new_status})


@app.route('/api/data-reviews/batch-manual-review', methods=['POST'])
def api_data_reviews_batch_manual_review():
    data = request.get_json() or {}
    ids = data.get('ids', [])
    action = data.get('action', 'approve')
    reviewer_id = data.get('reviewer_id')
    reason = data.get('reason', '')
    conclusion = data.get('conclusion')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    result = 'approved' if action == 'approve' else 'rejected'
    new_status = 'archived'  # 人工复核即最终判定，统一归档
    with get_db() as db:
        for rid in ids:
            db.execute(
                "UPDATE data_reviews SET manual_result=?, manual_reason=?, reviewer_id=?, reviewed_at=?, status=? WHERE id=?",
                (result, reason, reviewer_id, now, new_status, rid))
        db.commit()
        return jsonify({'ok': True, 'count': len(ids), 'result': result})


# ---------- 审核统计 ----------
@app.route('/api/data-reviews/stats', methods=['GET'])
def api_data_reviews_stats():
    """各级审核统计 + 按指标分组"""
    with get_db() as db:
        # 各级数量
        rows = db.execute(
            '''SELECT status, COUNT(*) as n FROM data_reviews GROUP BY status''').fetchall()
        by_status = {r['status']: r['n'] for r in rows}
        total = sum(by_status.values())
        # 待处理 = 尚未归档（仍需各级审核）的条数，而非全部历史总数
        archived = by_status.get('archived', 0)
        pending_total = total - archived
        # 按指标分组
        rows = db.execute(
            '''SELECT metric, COUNT(*) as n,
                      SUM(CASE WHEN auto_result='reject' THEN 1 ELSE 0 END) as auto_reject,
                      SUM(CASE WHEN smart_result='suspicious' THEN 1 ELSE 0 END) as smart_sus
               FROM data_reviews GROUP BY metric ORDER BY n DESC''').fetchall()
        by_metric = [dict(r) for r in rows]
        # 按站点
        rows = db.execute(
            '''SELECT dr.site_id as site_id, s.name as site_name, COUNT(*) as n,
                      SUM(CASE WHEN dr.status='archived' THEN 1 ELSE 0 END) as archived
               FROM data_reviews dr JOIN sites s ON dr.site_id = s.id
               GROUP BY dr.site_id ORDER BY n DESC LIMIT 10''').fetchall()
        by_site = [dict(r) for r in rows]
        return jsonify({
            'total': pending_total,
            'by_status': by_status,
            'by_metric': by_metric,
            'by_site': by_site,
            'pass_rate': round((by_status.get('archived', 0) / max(total, 1)) * 100, 1)
        })


# ---------- 异常编码参考（Step 2）----------
@app.route('/api/anomaly-codes', methods=['GET'])
def api_anomaly_codes():
    """获取异常编码参考表"""
    category = request.args.get('category')
    with get_db() as db:
        if category:
            rows = db.execute('SELECT * FROM anomaly_codes WHERE category=? AND is_active=1 ORDER BY code', (category,)).fetchall()
        else:
            rows = db.execute('SELECT * FROM anomaly_codes WHERE is_active=1 ORDER BY code').fetchall()
        return jsonify([dict(r) for r in rows])


# ---------- 告警升级配置（Step 3）----------
@app.route('/api/alert-escalation-config', methods=['GET'])
def api_alert_escalation_config_get():
    """获取告警升级配置"""
    with get_db() as db:
        rows = db.execute('SELECT * FROM alert_escalation_config ORDER BY id').fetchall()
        return jsonify([dict(r) for r in rows])


@app.route('/api/alert-escalation-config', methods=['PUT'])
def api_alert_escalation_config_update():
    """更新某级别升级配置"""
    data = request.get_json() or {}
    level = data.get('level')
    if not level:
        return jsonify({'error': '缺少 level'}), 400
    with get_db() as db:
        updates = []
        params = []
        for field in ['sla_minutes', 'auto_workorder', 'notify_type', 'escalate_to_level', 'is_active']:
            if field in data:
                updates.append(f'{field}=?')
                params.append(data[field])
        if not updates:
            return jsonify({'error': '无字段更新'}), 400
        params.append(level)
        db.execute(f'UPDATE alert_escalation_config SET {",".join(updates)} WHERE level=?', params)
        db.commit()
        row = db.execute('SELECT * FROM alert_escalation_config WHERE level=?', (level,)).fetchone()
        return jsonify(dict(row) if row else {'ok': True})


# ---------- 规则模板与覆盖分析（Step 5）----------
_RULE_TEMPLATES = {
    'gb3838_iii': {'name': 'GB3838-2002 III类水', 'description': '国家地表水III类标准',
        'rules': [
            {'metric':'ph','conditions':{'min':6.0,'max':9.0},'rule_type':'static','severity':'warning'},
            {'metric':'ammonia','conditions':{'max':1.0},'rule_type':'static','severity':'warning'},
            {'metric':'total_phosphorus','conditions':{'max':0.2},'rule_type':'static','severity':'warning'},
            {'metric':'total_nitrogen','conditions':{'max':1.0},'rule_type':'static','severity':'warning'},
            {'metric':'dissolved_oxygen','conditions':{'min':5.0},'rule_type':'static','severity':'warning'},
            {'metric':'cod','conditions':{'max':20.0},'rule_type':'static','severity':'warning'},
    ]},
    'spc_3sigma': {'name': 'SPC 3σ 动态', 'description': '基于历史均值±3σ生成动态阈值', 'rules': []},
    'trend_alarm': {'name': '趋势预警', 'description': '连续N周期同向变化检测', 'rules': []},
}
@app.route('/api/rule-templates', methods=['GET'])
def api_rule_templates():
    return jsonify([{'id': k, **v} for k, v in _RULE_TEMPLATES.items()])
@app.route('/api/rule-templates/<template_id>/apply', methods=['POST'])
def api_rule_template_apply(template_id):
    data = request.get_json() or {}
    scope = data.get('scope', 'global')
    template = _RULE_TEMPLATES.get(template_id)
    if not template: return jsonify({'error': '模板不存在'}), 404
    created = 0
    with get_db() as db:
        for rule in template['rules']:
            exists = db.execute('SELECT id FROM threshold_rules WHERE metric=? AND rule_type=? AND scope=?',
                               (rule['metric'], rule['rule_type'], scope)).fetchone()
            if exists: continue
            db.execute('INSERT INTO threshold_rules (name, scope, metric, rule_type, conditions, severity) VALUES (?,?,?,?,?,?)',
                      (f'{template["name"]}-{rule["metric"]}', scope, rule['metric'],
                       rule['rule_type'], _json.dumps(rule['conditions'], ensure_ascii=False), rule['severity']))
            created += 1
        db.commit()
    return jsonify({'applied': created, 'template': template_id, 'scope': scope})
@app.route('/api/rule-coverage', methods=['GET'])
def api_rule_coverage():
    """规则覆盖分析"""
    all_metrics = ['ph','ammonia','cod','total_phosphorus','total_nitrogen','dissolved_oxygen','turbidity','water_temp']
    with get_db() as db:
        rows = db.execute('SELECT scope, metric, COUNT(*) as n FROM threshold_rules WHERE enabled=1 GROUP BY scope, metric').fetchall()
        covered = set((r['scope'], r['metric']) for r in rows if r['n'] > 0)
        uncovered = [m for m in all_metrics if ('global', m) not in covered and ('metric', m) not in covered]
        return jsonify({
            'coverage': [dict(r) for r in rows],
            'total_metrics': len(all_metrics),
            'covered': len(set(r['metric'] for r in rows)),
            'uncovered': uncovered,
            'total_rules': db.execute('SELECT COUNT(*) FROM threshold_rules WHERE enabled=1').fetchone()[0],
        })


# ---------- 趋势预测（Step 6）----------
import math as _math
@app.route('/api/prediction/trend', methods=['GET'])
def api_prediction_trend():
    """短期趋势预测：返回历史数据 + 移动平均 + 线性回归预测"""
    site_id = request.args.get('site_id', type=int)
    metric = request.args.get('metric')
    hours = int(request.args.get('hours', 48))
    forecast_steps = int(request.args.get('forecast_steps', 12))  # 预测未来12个点
    if not site_id or not metric:
        return jsonify({'error': '需要 site_id 和 metric'}), 400
    date_from = (datetime.now() - timedelta(hours=hours)).strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as db:
        rows = db.execute(
            'SELECT recorded_at, value FROM sensor_data WHERE site_id=? AND metric=? AND recorded_at>=? ORDER BY recorded_at',
            (site_id, metric, date_from)).fetchall()
    if not rows or len(rows) < 10:
        return jsonify({'error': '历史数据不足', 'count': len(rows) if rows else 0}), 422
    actual = [{'time': r['recorded_at'], 'value': r['value']} for r in rows]
    vals = [r['value'] for r in rows]
    # 移动平均（窗口=6）
    ma = []
    for i in range(len(vals)):
        window = vals[max(0, i-5):i+1]
        ma.append({'time': rows[i]['recorded_at'], 'value': round(sum(window)/len(window), 3)})
    # 线性回归（用最近 N/2 个点拟合趋势）
    n = max(len(vals) // 2, 12)
    recent = vals[-n:]
    x_mean = (n - 1) / 2
    y_mean = sum(recent) / n
    num = den = 0
    for i, y in enumerate(recent):
        num += (i - x_mean) * (y - y_mean)
        den += (i - x_mean) ** 2
    slope = num / den if den != 0 else 0
    intercept = y_mean - slope * x_mean
    # 预测
    last_time = rows[-1]['recorded_at']
    try:
        from datetime import timedelta as _td
        import json as _json
        last_dt = datetime.strptime(last_time, '%Y-%m-%d %H:%M:%S')
    except:
        last_dt = datetime.now()
    interval = (datetime.strptime(rows[-1]['recorded_at'], '%Y-%m-%d %H:%M:%S') -
                datetime.strptime(rows[-2]['recorded_at'], '%Y-%m-%d %H:%M:%S')).total_seconds() if len(rows) >= 2 else 3600
    forecast = []
    for i in range(1, forecast_steps + 1):
        pred_val = intercept + slope * (n + i - 1)
        pred_time = (last_dt + timedelta(seconds=interval * i)).strftime('%Y-%m-%d %H:%M:%S')
        forecast.append({'time': pred_time, 'value': round(max(pred_val, 0), 3)})
    upper = [{'time': f['time'], 'value': round(f['value'] * 1.15, 3)} for f in forecast]
    lower = [{'time': f['time'], 'value': round(max(f['value'] * 0.85, 0), 3)} for f in forecast]
    return jsonify({
        'site_id': site_id, 'metric': metric,
        'actual': actual, 'moving_average': ma,
        'forecast': forecast, 'upper': upper, 'lower': lower,
        'slope': round(slope, 5), 'r_squared': round(num * num / (den * sum((y - y_mean)**2 for y in recent) + 0.001), 3),
        'metadata': {'data_points': len(rows), 'forecast_steps': forecast_steps, 'interval_seconds': interval},
    })


# (已移除：下方重复的未鉴权 /api/notifications GET/POST 路由；
#  鉴权版 get_notifications 见 L5881，按 user['id'] 返回本人通知，与前端字段对齐)


# ---------- 试剂总览看板数据 ----------
@app.route('/api/reagent-dashboard', methods=['GET'])
def api_reagent_dashboard():
    """返回试剂总览统计数据：汇总+告警+用量趋势（按站点范围隔离）"""
    allowed = _filter_site_ids()
    sc = f' AND site_id IN ({",".join("?" * len(allowed))})' if allowed is not None else ''
    scp = list(allowed) if allowed is not None else []
    with get_db() as db:
        # 试剂种类
        total_types = db.execute('SELECT COUNT(*) FROM reagents').fetchone()[0]
        # 有库存的站点数
        sites_with_inventory = db.execute(f'SELECT COUNT(DISTINCT site_id) FROM reagent_inventory{sc}', scp).fetchone()[0]
        # 低库存站点数
        low_stock = db.execute(f'SELECT COUNT(*) FROM reagent_alerts WHERE handled=0 AND alert_type="low_stock"{sc}', scp).fetchone()[0]
        # 未处理告警列表
        alerts = [dict(r) for r in db.execute(f'''
            SELECT ra.*, r.name as reagent_name, s.name as site_name
               FROM reagent_alerts ra JOIN reagents r ON ra.reagent_id = r.id
               LEFT JOIN sites s ON ra.site_id = s.id
               WHERE ra.handled=0{sc} ORDER BY ra.alert_at DESC LIMIT 30''', scp).fetchall()]
        # 本月用量（按试剂汇总）
        month_start = datetime.now().strftime('%Y-%m-01 00:00:00')
        usage_by_reagent = [dict(r) for r in db.execute(
            '''SELECT ru.reagent_id, r.name as reagent_name, SUM(ru.used_qty) as total_used
               FROM reagent_usage ru JOIN reagents r ON ru.reagent_id = r.id
               WHERE ru.used_at >= ? GROUP BY ru.reagent_id ORDER BY total_used DESC LIMIT 10''',
            (month_start,)).fetchall()]
        # 近30天用量趋势
        thirty_ago = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        usage_trend = [dict(r) for r in db.execute(
            '''SELECT DATE(used_at) as d, SUM(used_qty) as qty FROM reagent_usage
               WHERE used_at >= ? GROUP BY DATE(used_at) ORDER BY d''',
            (thirty_ago,)).fetchall()]
        # 低库存站点列表
        low_stock_sites = [dict(r) for r in db.execute(f'''
            SELECT ri.site_id, s.name as site_name, ri.reagent_id, r.name as reagent_name,
                      ri.current_qty, ri.low_stock_threshold
               FROM reagent_inventory ri JOIN reagents r ON ri.reagent_id = r.id
               JOIN sites s ON ri.site_id = s.id
               WHERE ri.current_qty <= ri.low_stock_threshold{sc}
               ORDER BY ri.current_qty ASC LIMIT 20''', scp).fetchall()]
        return jsonify({
            'total_types': total_types,
            'sites_with_inventory': sites_with_inventory,
            'low_stock_count': low_stock,
            'alerts': alerts,
            'usage_by_reagent': usage_by_reagent,
            'usage_trend': usage_trend,
            'low_stock_sites': low_stock_sites,
        })


# ---------- 试剂预警总览（跨站，预警中心用）----------
@app.route('/api/reagent-overview', methods=['GET'])
def api_reagent_overview():
    """跨站试剂状态总览：返回所有站点的试剂及其计算状态。
    一线/管理者只需看「剩余可用天数低于阈值」或「低余量」的试剂。
    """
    status_filter = request.args.get('status')  # 可选：临期/低余量/已过期
    allowed = _filter_site_ids()
    with get_db() as db:
        rows = db.execute(
            '''SELECT ri.*, r.name as reagent_name, r.unit, r.shelf_life_days,
                      s.name as site_name, s.id as site_id
               FROM reagent_inventory ri
               JOIN reagents r ON ri.reagent_id = r.id
               JOIN sites s ON ri.site_id = s.id
               ORDER BY s.name, r.name''').fetchall()
        out = []
        for r in rows:
            if allowed is not None and r['site_id'] not in allowed:
                continue
            d = dict(r)
            d.update(compute_reagent_status(d))
            if status_filter and d['status'] != status_filter:
                continue
            out.append(d)
        # 仅返回需要关注的状态（临期/低余量/已过期），未设置/正常不占列表
        concern = [d for d in out if d['status'] in ('临期', '低余量', '已过期')]
        return jsonify({
            'total': len(out),
            'concern_count': len(concern),
            'items': concern,
        })


# ---------- 异常溯源链（Step 4）----------
import json as _json
@app.route('/api/anomaly-traceability', methods=['POST'])
def api_anomaly_traceability_create():
    """创建/更新溯源记录（含自动数据快照）"""
    data = request.get_json() or {}
    review_id = data.get('review_id')
    alert_id = data.get('alert_id')
    work_order_id = data.get('work_order_id')
    code = data.get('code', '')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as db:
        existing = db.execute('SELECT id FROM anomaly_traceability WHERE review_id=?', (review_id,)).fetchone()
        if existing:
            db.execute('''UPDATE anomaly_traceability SET root_cause=?, symptom=?, treatment=?, impact=?, updated_at=? WHERE id=?''',
                       (data.get('root_cause',''), data.get('symptom',''), data.get('treatment',''), data.get('impact',''), now, existing['id']))
            row = db.execute('SELECT * FROM anomaly_traceability WHERE id=?', (existing['id'],)).fetchone()
        else:
            snapshot = {}
            if review_id:
                dr = db.execute('SELECT * FROM data_reviews WHERE id=?', (review_id,)).fetchone()
                if dr:
                    snapshot['review'] = dict(dr)
                    site = db.execute('SELECT name FROM sites WHERE id=?', (dr['site_id'],)).fetchone()
                    snapshot['review']['site_name'] = site['name'] if site else ''
                    shadow = db.execute('SELECT * FROM device_shadows WHERE site_id=? ORDER BY last_data_time DESC LIMIT 1', (dr['site_id'],)).fetchone()
                    if shadow: snapshot['device'] = dict(shadow)
                    maint = db.execute('SELECT * FROM maintenance_plans WHERE site_id=? ORDER BY created_at DESC LIMIT 1', (dr['site_id'],)).fetchone()
                    if maint: snapshot['maintenance'] = dict(maint)
            cur = db.execute('''INSERT INTO anomaly_traceability (review_id, alert_id, work_order_id, code, root_cause, symptom, treatment, impact, data_snapshot) VALUES (?,?,?,?,?,?,?,?,?)''',
                           (review_id, alert_id, work_order_id, code, data.get('root_cause',''), data.get('symptom',''), data.get('treatment',''), data.get('impact',''), _json.dumps(snapshot, ensure_ascii=False)))
            row = db.execute('SELECT * FROM anomaly_traceability WHERE id=?', (cur.lastrowid,)).fetchone()
        db.commit()
        r = dict(row)
        try: r['data_snapshot'] = _json.loads(r['data_snapshot']) if r.get('data_snapshot') else {}
        except: pass
        return jsonify(r), 201


@app.route('/api/anomaly-traceability/chain', methods=['GET'])
def api_anomaly_traceability_chain():
    """获取完整溯源链：数据快照+设备日志+运维记录+周边站点+编码参考"""
    review_id = request.args.get('review_id')
    alert_id = request.args.get('alert_id')
    if not review_id and not alert_id:
        return jsonify({'error': '需提供 review_id 或 alert_id'}), 400
    with get_db() as db:
        trace, dr = None, None
        if review_id:
            trace = db.execute('SELECT * FROM anomaly_traceability WHERE review_id=?', (review_id,)).fetchone()
            dr = db.execute('SELECT dr.*, s.name as site_name FROM data_reviews dr LEFT JOIN sites s ON dr.site_id=s.id WHERE dr.id=?', (review_id,)).fetchone()
        elif alert_id:
            trace = db.execute('SELECT * FROM anomaly_traceability WHERE alert_id=?', (alert_id,)).fetchone()
        result = {'trace': dict(trace) if trace else {}, 'review': dict(dr) if dr else {},
                  'device': None, 'maintenance': None, 'nearby_sites': [], 'codes': []}
        if dr:
            sid, rt = dr['site_id'], dr['recorded_at']
            s = db.execute('SELECT * FROM device_shadows WHERE site_id=? ORDER BY last_data_time DESC LIMIT 1', (sid,)).fetchone()
            if s: result['device'] = dict(s)
            m = db.execute('SELECT * FROM maintenance_plans WHERE site_id=? ORDER BY created_at DESC LIMIT 1', (sid,)).fetchone()
            if m: result['maintenance'] = dict(m)
            n = db.execute('''SELECT s.name, sd.metric, sd.value, sd.recorded_at FROM sensor_data sd
                              JOIN sites s ON sd.site_id = s.id WHERE sd.metric=? AND sd.recorded_at>=datetime(?,'-10 minutes')
                              AND sd.recorded_at<=datetime(?,'+10 minutes') AND sd.site_id!=? ORDER BY sd.recorded_at DESC LIMIT 5''',
                          (dr['metric'], rt, rt, sid)).fetchall()
            result['nearby_sites'] = [dict(x) for x in n] if n else []
            c = db.execute('SELECT code, title, suggestion FROM anomaly_codes WHERE is_active=1').fetchall()
            result['codes'] = [dict(x) for x in c]
        if trace and trace['data_snapshot']:
            try: result['trace']['data_snapshot'] = _json.loads(trace['data_snapshot'])
            except: pass
        return jsonify(result)


# ---------- 自动数据审核（后台定时任务：生成待审 → L1 → L2）----------
def auto_data_review():
    """定时任务：自动生成待审数据→L1自动审核→L2智能审核"""
    try:
        with app.app_context():
            with get_db() as db:
                # 1. 生成最近4小时的待审
                date_from = (datetime.now() - timedelta(hours=4)).strftime('%Y-%m-%d %H:%M:%S')
                rows = db.execute(
                    '''SELECT sd.site_id, sd.metric, sd.value, sd.recorded_at
                       FROM sensor_data sd WHERE sd.recorded_at >= ? ORDER BY sd.recorded_at DESC LIMIT 2000''',
                    (date_from,)).fetchall()
                inserted = 0
                for r in rows:
                    existed = db.execute(
                        'SELECT id FROM data_reviews WHERE site_id=? AND metric=? AND value=? AND recorded_at=?',
                        (r['site_id'], r['metric'], r['value'], r['recorded_at'])).fetchone()
                    if not existed:
                        db.execute(
                            'INSERT INTO data_reviews (site_id, metric, value, recorded_at, status) VALUES (?,?,?,?,"pending")',
                            (r['site_id'], r['metric'], r['value'], r['recorded_at']))
                        inserted += 1
                if inserted:
                    print(f'[AutoReview] 生成 {inserted} 条待审')

                # 2. L1 四级自动审核管线 -----------------------------------------
                # 管线1→量程校验（GB3838 上下限+负值+NULL）
                # 管线2→趋势突变校验（Z-score vs 近7天历史，>5σ 标记疑）
                # 管线3→关联校验（多指标逻辑一致性，如pH↑→DO↑）
                # 管线4→僵死校验（连续多笔值完全不变，怀疑传感器冻结）
                pending = db.execute("SELECT * FROM data_reviews WHERE status='pending'").fetchall()
                for r in pending:
                    v, m, sid, rid = r['value'], r['metric'], r['site_id'], r['id']
                    rt = r['recorded_at']
                    reasons = []  # 累加各级校验结果
                    # ----- L1-1: 量程校验（一级否决，不通过直接标记 reject）-----
                    range_fail = None
                    if m in _METRIC_BOUNDS:
                        lo, hi = _METRIC_BOUNDS[m]
                        if v is None or v < 0:
                            range_fail = '【Q-001】值缺失或为负'
                        elif v < lo:
                            range_fail = f'【Q-001】低于量程下限 {lo}'
                        elif v > hi:
                            range_fail = f'【Q-001】超出量程上限 {hi}'
                    elif v is None:
                        range_fail = '【Q-001】值缺失'
                    if range_fail:
                        db.execute(
                            "UPDATE data_reviews SET auto_result='reject', auto_reason=?, status='archived' WHERE id=?",
                            (range_fail, rid))
                        continue  # 量程不合格，直接归档，不进 L3

                    # ----- L1-2: 趋势突变校验（Z-score > 5 标记疑点）-----
                    trend_reason = None
                    import math
                    stats = db.execute(
                        '''SELECT AVG(value) as mu, AVG(value*value) as e2, COUNT(*) as n FROM sensor_data
                           WHERE site_id=? AND metric=? AND recorded_at < ? AND recorded_at > datetime(?,'-7 day')''',
                        (sid, m, rt, rt)).fetchone()
                    if stats and stats['mu'] and stats['n'] and stats['n'] >= 20:
                        mu, e2 = stats['mu'], stats['e2'] or 0
                        sigma = math.sqrt(max(e2 - mu * mu, 0))
                        if sigma == 0: sigma = 0.001
                        z_spike = abs((v - mu) / sigma)
                        if z_spike > 5.0:
                            trend_reason = f'【Q-002】趋势突变 Z={z_spike:.1f}'
                        elif z_spike > 3.0:
                            trend_reason = f'【Q-002】趋势偏离 Z={z_spike:.1f}'
                    if trend_reason:
                        reasons.append(trend_reason)

                    # ----- L1-3: 关联校验（多指标逻辑一致性）-----
                    corr_reasons = []
                    if m in _METRIC_CORRELATIONS:
                        for peer_m, direction in _METRIC_CORRELATIONS[m]:
                            # 取同期（±5分钟）的关联指标值
                            peer = db.execute(
                                '''SELECT value FROM sensor_data
                                   WHERE site_id=? AND metric=? AND recorded_at >= datetime(?,'-5 minutes')
                                   AND recorded_at <= datetime(?,'+5 minutes') ORDER BY ABS(julianday(recorded_at)-julianday(?)) LIMIT 1''',
                                (sid, peer_m, rt, rt, rt)).fetchone()
                            if peer and peer['value'] is not None:
                                pv = peer['value']
                                # 简单判断：用本单位历史均值作基准判断升降
                                peer_stats = db.execute(
                                    'SELECT AVG(value) as mu FROM sensor_data WHERE site_id=? AND metric=? AND recorded_at > datetime(?,"-7 day") AND recorded_at < ?',
                                    (sid, peer_m, rt, rt)).fetchone()
                                if peer_stats and peer_stats['mu']:
                                    p_mu = peer_stats['mu']
                                    v_dir = 'up' if v > (stats['mu'] if stats and stats['mu'] else v) else 'down'
                                    p_dir = 'up' if pv > p_mu else 'down'
                                    if direction == 1 and v_dir != p_dir:
                                        corr_reasons.append(f'【Q-003】{_METRIC_LABELS.get(m,m)}↑但{_METRIC_LABELS.get(peer_m,peer_m)}↓（应同向）')
                                    elif direction == -1 and v_dir == p_dir:
                                        corr_reasons.append(f'【Q-003】{_METRIC_LABELS.get(m,m)}↑且{_METRIC_LABELS.get(peer_m,peer_m)}↑（应反向）')
                    reasons.extend(corr_reasons)

                    # ----- L1-4: 僵死校验（连续 5+ 笔值完全一样）-----
                    frozen_reason = None
                    frozen_count = db.execute(
                        '''SELECT COUNT(*) as n FROM sensor_data
                           WHERE site_id=? AND metric=? AND value=? AND recorded_at >= datetime(?,'-2 hours')
                           AND recorded_at < ?''',
                        (sid, m, v, rt, rt)).fetchone()
                    if frozen_count and frozen_count['n'] >= 5:
                        frozen_reason = f'【Q-004】数据僵死: 2小时内{frozen_count["n"]}笔值相同'
                    if frozen_reason:
                        reasons.append(frozen_reason)

                    # ----- 汇总结果 -----
                    if reasons:
                        db.execute(
                            "UPDATE data_reviews SET auto_result='suspicious', auto_reason=?, status='auto_reviewed' WHERE id=?",
                            ('; '.join(reasons), rid))
                    else:
                        db.execute(
                            "UPDATE data_reviews SET auto_result='pass', auto_reason='L1通过', status='auto_reviewed' WHERE id=?",
                            (rid,))

                # 3. L2 智能审核（Z-score）
                import math
                l2_items = db.execute("SELECT * FROM data_reviews WHERE status='auto_reviewed'").fetchall()
                for r in l2_items:
                    stats = db.execute(
                        '''SELECT AVG(value) as mu, AVG(value*value) as e2, COUNT(*) as n FROM sensor_data
                           WHERE site_id=? AND metric=? AND recorded_at < ? AND recorded_at > datetime(?,'-7 day')''',
                        (r['site_id'], r['metric'], r['recorded_at'], r['recorded_at'])).fetchone()
                    if stats and stats['mu'] and stats['n'] and stats['n'] >= 20:
                        mu, e2 = stats['mu'], stats['e2'] or 0
                        sigma = math.sqrt(max(e2 - mu * mu, 0))
                        if sigma == 0: sigma = 0.001
                        z = abs((r['value'] - mu) / sigma)
                        if z > 3.0:
                            db.execute(
                                "UPDATE data_reviews SET smart_result='suspicious', smart_score=?, status='smart_reviewed' WHERE id=?",
                                (round(z, 3), r['id']))
                        else:
                            # L1+L2 均通过：自动归档，不再进入 L3 人工复核队列
                            db.execute(
                                "UPDATE data_reviews SET smart_result='pass', smart_score=?, status='archived' WHERE id=?",
                                (round(z, 3), r['id']))
                    else:
                        # 样本不足，跳过智能判断：视为正常，自动归档（不进 L3）
                        db.execute(
                            "UPDATE data_reviews SET smart_result='pass', smart_score=0, status='archived' WHERE id=?",
                            (r['id'],))
                db.commit()
    except Exception as e:
        print(f'[AutoReview] 自动审核出错（非致命）: {e}')


def check_device_offline():
    """定时任务：检测设备离线状态并生成告警（替代被注释的 simulator）"""
    try:
        db = _scheduler_db()
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        # 30分钟无数据 = 设备离线
        threshold = (datetime.now() - timedelta(minutes=30)).strftime('%Y-%m-%d %H:%M:%S')

        # 1. 检测设备离线
        stale_devices = db.execute(
            "SELECT ds.id, ds.site_id, ds.status as dev_status, ds.device_name, s.status as site_status, s.name as site_name "
            "FROM device_shadows ds JOIN sites s ON ds.site_id = s.id "
            "WHERE (ds.last_data_time IS NULL OR ds.last_data_time < ?) AND ds.status != 'offline'",
            (threshold,)
        ).fetchall()
        for dev in stale_devices:
            dev = dict(dev)
            sid = dev['site_id']
            # 标记离线
            db.execute("UPDATE device_shadows SET status='offline' WHERE id=?", (dev['id'],))
            # 创建告警（如果5分钟内同站点同类型未创建）
            existing = db.execute(
                "SELECT id FROM alerts WHERE site_id=? AND metric='device_status' AND status='pending' AND created_at > datetime('now','-5 minutes')",
                (sid,)).fetchone()
            if not existing:
                sname = dev.get('site_name', '') or f'站点{sid}'
                dname = dev.get('device_name', '')
                create_alert_internal(db, sid, 'device_status', 0, 'yellow',
                    f'{sname} 设备离线超30分钟（{dname}），请及时处理')

        # 2. 检测数据缺失（站点有设备但所有设备都已离线）
        for dev in stale_devices:
            dev = dict(dev)
            sid = dev['site_id']
            online_count = db.execute(
                "SELECT COUNT(*) FROM device_shadows WHERE site_id=? AND status='online'",
                (sid,)).fetchone()[0]
            if online_count == 0:
                db.execute("UPDATE sites SET status='offline' WHERE id=?", (sid,))
                existing_gap = db.execute(
                    "SELECT id FROM alerts WHERE site_id=? AND metric='data_gap' AND status='pending'",
                    (sid,)).fetchone()
                if not existing_gap:
                    sname = dev.get('site_name', f'站点#{sid}')
                    create_alert_internal(db, sid, 'data_gap', 30, 'orange',
                        f'{sname} 所有设备离线，数据缺失超过30分钟')

        db.commit()
        if stale_devices:
            print(f'[DeviceCheck] {len(stale_devices)} 设备离线, 已生成告警')

        # 3. 兜底：已离线但缺告警的设备（seed 遗留）
        missing_alerts = db.execute(
            "SELECT ds.id, ds.site_id, ds.device_name, s.name as site_name "
            "FROM device_shadows ds JOIN sites s ON ds.site_id = s.id "
            "WHERE ds.status='offline' AND NOT EXISTS ("
            "  SELECT 1 FROM alerts WHERE site_id=ds.site_id AND metric='device_status' AND status='pending'"
            ")").fetchall()
        for dev in missing_alerts:
            dev = dict(dev)
            sname = dev.get('site_name') or '站点' + str(dev['site_id'])
            dname = dev.get('device_name') or ''
            create_alert_internal(db, dev['site_id'], 'device_status', 0, 'yellow',
                f'{sname} 设备离线（{dname}）')
        if missing_alerts:
            db.commit()
            print(f'[DeviceCheck] 补了 {len(missing_alerts)} 条缺告警')
    except Exception as e:
        print(f'[DeviceCheck] 错误: {e}')
    finally:
        try: db.close()
        except: pass


def alert_escalation_check():
    """定时任务：每5分钟扫描未处理告警，按SLA自动升级"""
    try:
        with app.app_context():
            with get_db() as db:
                configs = {r['level']: r for r in db.execute(
                    'SELECT * FROM alert_escalation_config WHERE is_active=1').fetchall()}
                active = db.execute(
                    "SELECT * FROM alerts WHERE status NOT IN ('resolved','closed')").fetchall()
                now_dt = datetime.now()
                now = now_dt.strftime('%Y-%m-%d %H:%M:%S')
                for alert in active:
                    cfg = configs.get(alert['level'])
                    if not cfg or not cfg['sla_minutes']:
                        continue
                    if not alert['created_at']:
                        continue
                    created = datetime.strptime(alert['created_at'], '%Y-%m-%d %H:%M:%S')
                    elapsed = (now_dt - created).total_seconds() / 60.0
                    if elapsed <= cfg['sla_minutes']:
                        continue
                    # 升级
                    next_lev = cfg['escalate_to_level'] if cfg['escalate_to_level'] and cfg['escalate_to_level'] != 'None' else None
                    if next_lev:
                        db.execute(
                            '''UPDATE alerts SET level=?, urge_count=urge_count+1, last_urged_at=?
                               WHERE id=? AND status NOT IN ('resolved','closed')''',
                            (next_lev, now, alert['id']))
                    # 自动生成工单：对同一站点+同一测项，未闭环前只保留一个升级工单，避免重复
                    if cfg['auto_workorder'] and not alert['related_order_no']:
                        order_no = 'AL' + now[:10].replace('-', '') + str(alert['id']).zfill(4)
                        site_id = alert['site_id'] if alert['site_id'] else 0
                        event_type = alert['metric'] or '告警升级'
                        metric_label = METRIC_LABEL.get(alert['metric'], alert['metric'])
                        if site_id:
                            site = db.execute('SELECT name, manager FROM sites WHERE id=?', (site_id,)).fetchone()
                            site_name = site['name'] if site else ''
                            assignee = _station_operator(site_id) if site else ''
                        else:
                            site_name = ''
                            assignee = ''
                        # 去重：同一站点+同一测项已存在未闭环的升级工单，则不再创建
                        existing = db.execute(
                            "SELECT order_no FROM work_orders WHERE source='escalation' AND site_id=? AND event_type=? AND status NOT IN ('closed','resolved') LIMIT 1",
                            (site_id, event_type)).fetchone()
                        if existing:
                            print(f'[Escalation] 告警#{alert["id"]} 已存在升级工单 {existing["order_no"]}，跳过重复创建')
                            continue
                        db.execute(
                            '''INSERT INTO work_orders (order_no, site_id, source, event_type, level, title, description, status, assignee, created_at, related_alert_id)
                               VALUES (?,?,'escalation',?,'urgent',?,?,'pending',?,?,?)''',
                            (order_no, site_id, event_type,
                             f'【告警升级】{metric_label}: {alert["message"]}'[:200],
                             f'源自告警#{alert["id"]}升级，原级别{ALERT_LEVEL_LABEL.get(alert["level"], alert["level"])}→{ALERT_LEVEL_LABEL.get(next_lev, next_lev)}'[:300],
                             assignee, now, alert['id']))
                        db.execute('UPDATE alerts SET related_order_no=? WHERE id=?', (order_no, alert['id']))
                        print(f'[Escalation] 告警#{alert["id"]} 升级 {alert["level"]}→{next_lev}，已生成工单{order_no}')
                    else:
                        print(f'[Escalation] 告警#{alert["id"]} {alert["level"]}→{next_lev}（超SLA {elapsed:.0f}min）')
                db.commit()
    except Exception as e:
        import traceback
        print(f'[Escalation] 升级出错（非致命）: {e}')
        traceback.print_exc()


if __name__ == '__main__':
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    init_db()
    ensure_alert_rules_config()
    migrate_spare_parts_inventory_columns()
    migrate_workorder_flow_columns()
    migrate_vehicle_applications_nullable()
    migrate_plan_schedules()
    migrate_reagent_qc()
    seed_data()
    seed_inspections()
    seed_alerts()
    seed_maintenance()
    seed_maintenance_templates()
    seed_users()
    seed_vehicles()
    seed_inspection_v2()
    seed_param_thresholds()
    demo_seed_enabled = os.environ.get('ENABLE_DEMO_SEED') == '1'
    if demo_seed_enabled:
        try:
            seed_abnormal_scenarios()
        except Exception as e:
            print(f'[Seed] 异常场景注入失败（可忽略）: {e}')
    migrate_alerts_messages()
    migrate_alert_flow()
    fix_site_river()
    if demo_seed_enabled and os.environ.get('SKIP_BACKFILL') != '1':
        backfill_history(72)
        # 生成初始数据（仅演示环境）
        for _ in range(6):
            try:
                generate_sensor_data()
                time.sleep(0.3)
            except Exception as e:
                print(f'[Seed] 初始数据生成跳过: {e}')
        try:
            from seed_demo import generate as demo_generate
            demo_generate()
        except Exception as e:
            import traceback
            print(f"[Seed] 演示数据生成失败: {e}")
            traceback.print_exc()
    else:
        print('[Seed] 演示数据初始化已关闭（ENABLE_DEMO_SEED=1 可显式启用）')
    # 清理过期已办结告警（保留最近7天，档案卡片需要7天历史）
    # 清理过期已办结告警（保留最近7天，档案卡片需要7天历史）
    try:
        with get_db() as db:
            total = db.execute("SELECT COUNT(*) as c FROM alerts WHERE status='resolved' AND created_at < datetime('now','-7 day')").fetchone()['c']
            if total > 50:
                db.execute("DELETE FROM alerts WHERE status='resolved' AND created_at < datetime('now','-7 day')")
                db.commit()
                print(f"[Cleanup] 清理过期已办结告警: 删除{total}条（保留近7天）")
    except Exception as e:
        print(f"[Cleanup] 告警清理跳过: {e}")
    # 模拟器已关闭（仅保留静态异常场景数据用于演示）
    # if os.environ.get('SKIP_SIMULATOR') != '1':
    #     scheduler.add_job(generate_sensor_data, 'interval', seconds=60, id='simulator')
    #     print("[Server] 数据仿真器已启动（每60秒），SKIP_SIMULATOR=1 可关闭")
    # else:
    print("[Server] 数据仿真器已关闭（静态演示模式）")
    # 每30分钟更新天气
    scheduler.add_job(fetch_real_weather, 'interval', minutes=30, id='weather_updater')
    # 每10分钟自动数据审核（L1 + L2）
    scheduler.add_job(auto_data_review, 'interval', minutes=10, id='auto_data_review')
    # 每5分钟设备离线检测
    scheduler.add_job(check_device_offline, 'interval', minutes=5, id='device_offline_check')
    # 启动后立即运行一次（不等待5分钟）
    check_device_offline()
    # 每5分钟告警升级检查
    scheduler.add_job(alert_escalation_check, 'interval', minutes=5, id='alert_escalation')

    # ===== 可选: SL651 国家水站协议接收器 =====
    # 环境变量 ENABLE_SL651=1 时启动
    if os.environ.get('ENABLE_SL651') == '1':
        import threading
        from sl651_server import run_server as sl651_run, get_local_ip, DEFAULT_PORT as SL651_PORT
        _t = threading.Thread(target=sl651_run, args=(SL651_PORT,), daemon=True)
        _t.start()
        local_ip = get_local_ip()
        print(f"[SL651] 国家水站协议接收器已启动（ENABLE_SL651=1）")
        print(f"[SL651] 请在科蓝平台配置转发: TCP → {local_ip}:{SL651_PORT}")
    else:
        print("[SL651] 跳过（ENABLE_SL651=1 可启用国家水站协议接收器）")

    print("[Server] 水利运维智慧运营平台 启动成功!")
    print("[Server] API: http://localhost:5000/api/health")
    app.run(host='0.0.0.0', port=5000, debug=False, threaded=True)
